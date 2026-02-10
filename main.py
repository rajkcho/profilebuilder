"""
Orbital — M&A Intelligence Platform

Professional-grade company research platform with Sky.money-inspired UI.
Generates an 8-slide investment-banker-grade PowerPoint tear sheet.

Run:  streamlit run main.py

v5.0 - Full Feature Suite:
- Watchlist with session persistence & notes
- Excel/CSV export for all financial data
- DCF Valuation with sensitivity analysis
- Quick Compare mode with correlation matrix
- Merger Analysis with deal books
- Technical Analysis (RSI, MACD, Bollinger Bands)
- Options overview with put/call ratio
- Dividend analysis & financial health scorecard
- Institutional ownership breakdown
- Market sentiment gauge & sector heatmap
- Earnings calendar & news feed
- Search history & keyboard shortcuts
- Enhanced visualizations & print-friendly styles
"""

import streamlit as st
import plotly.graph_objects as go
import plotly.express as px
import pandas as pd
import numpy as np
import os
import random
import time
import json
import io
from datetime import datetime, timedelta
from dotenv import load_dotenv

load_dotenv()

from data_engine import (
    fetch_company_data, fetch_peer_data,
    format_number, format_pct, format_multiple,
    calculate_piotroski_score, calculate_intrinsic_value, get_key_ratios_summary
)
from ai_insights import generate_insights, generate_merger_insights
from pptx_generator import generate_presentation, generate_deal_book
from merger_analysis import MergerAssumptions, calculate_pro_forma, build_football_field
from comps_analysis import run_comps_analysis, generate_comps_table, format_comps_for_display, CompsAnalysis
import yfinance as yf

# ══════════════════════════════════════════════════════════════
# WATCHLIST MANAGEMENT
# ══════════════════════════════════════════════════════════════
def _init_watchlist():
    """Initialize watchlist in session state."""
    if "watchlist" not in st.session_state:
        st.session_state.watchlist = []
    if "watchlist_data" not in st.session_state:
        st.session_state.watchlist_data = {}

def _add_to_watchlist(ticker: str):
    """Add a ticker to the watchlist."""
    _init_watchlist()
    ticker = ticker.upper().strip()
    if ticker and ticker not in st.session_state.watchlist:
        st.session_state.watchlist.append(ticker)
        # Fetch basic data
        info = _quick_ticker_lookup(ticker)
        if info.get("valid"):
            st.session_state.watchlist_data[ticker] = info
        return True
    return False

def _remove_from_watchlist(ticker: str):
    """Remove a ticker from the watchlist."""
    _init_watchlist()
    ticker = ticker.upper().strip()
    if ticker in st.session_state.watchlist:
        st.session_state.watchlist.remove(ticker)
        st.session_state.watchlist_data.pop(ticker, None)
        return True
    return False

def _is_in_watchlist(ticker: str) -> bool:
    """Check if ticker is in watchlist."""
    _init_watchlist()
    return ticker.upper().strip() in st.session_state.watchlist

def _get_watchlist() -> list:
    """Get current watchlist."""
    _init_watchlist()
    return st.session_state.watchlist

# ══════════════════════════════════════════════════════════════
# VMS SCREENER UNIVERSE & DATA
# ══════════════════════════════════════════════════════════════
VMS_UNIVERSE = [
    {"ticker": "OTEX.TO", "name": "Open Text Corp", "vertical": "Legal Tech", "geo": "North America"},
    {"ticker": "DSGX", "name": "Descartes Systems", "vertical": "Transportation", "geo": "North America"},
    {"ticker": "CSGP", "name": "CoStar Group", "vertical": "Real Estate Tech", "geo": "North America"},
    {"ticker": "ENV", "name": "Envestnet", "vertical": "Financial Services Tech", "geo": "North America"},
    {"ticker": "PCTY", "name": "Paylocity", "vertical": "Financial Services Tech", "geo": "North America"},
    {"ticker": "PAYC", "name": "Paycom Software", "vertical": "Financial Services Tech", "geo": "North America"},
    {"ticker": "GWRE", "name": "Guidewire Software", "vertical": "Financial Services Tech", "geo": "North America"},
    {"ticker": "BSY", "name": "Bentley Systems", "vertical": "Construction Tech", "geo": "North America"},
    {"ticker": "ALRM", "name": "Alarm.com", "vertical": "Real Estate Tech", "geo": "North America"},
    {"ticker": "CCSI", "name": "Consensus Cloud", "vertical": "Healthcare IT", "geo": "North America"},
    {"ticker": "POWI", "name": "Power Integrations", "vertical": "Utilities", "geo": "North America"},
    {"ticker": "TYLT.TO", "name": "Tyler Technologies", "vertical": "GovTech", "geo": "North America"},
    {"ticker": "TYL", "name": "Tyler Technologies", "vertical": "GovTech", "geo": "North America"},
    {"ticker": "SSNC", "name": "SS&C Technologies", "vertical": "Financial Services Tech", "geo": "North America"},
    {"ticker": "APPF", "name": "AppFolio", "vertical": "Real Estate Tech", "geo": "North America"},
    {"ticker": "PRGS", "name": "Progress Software", "vertical": "Education Tech", "geo": "North America"},
    {"ticker": "IIIV", "name": "i3 Verticals", "vertical": "Financial Services Tech", "geo": "North America"},
    {"ticker": "RNG", "name": "RingCentral", "vertical": "Utilities", "geo": "North America"},
    {"ticker": "ADSK", "name": "Autodesk", "vertical": "Construction Tech", "geo": "North America"},
    {"ticker": "POWL", "name": "Powell Industries", "vertical": "Utilities", "geo": "North America"},
]

@st.cache_data(ttl=600, show_spinner=False)
def _fetch_vms_screening_data(tickers_tuple):
    """Fetch financial data for VMS screening candidates."""
    results = []
    for item in tickers_tuple:
        try:
            t = yf.Ticker(item["ticker"])
            info = t.info or {}
            revenue = info.get("totalRevenue", 0) or 0
            ebitda = info.get("ebitda", 0) or 0
            rev_growth = info.get("revenueGrowth", None)
            ev = info.get("enterpriseValue", 0) or 0
            ebitda_margin = (ebitda / revenue * 100) if revenue > 0 else 0
            ev_rev = (ev / revenue) if revenue > 0 else 0
            ev_ebitda = (ev / ebitda) if ebitda > 0 else 0
            rev_growth_pct = (rev_growth * 100) if rev_growth is not None else 0
            results.append({
                "Company": item["name"],
                "Ticker": item["ticker"],
                "Vertical": item["vertical"],
                "Geography": item["geo"],
                "Revenue ($M)": round(revenue / 1e6, 1),
                "EBITDA Margin (%)": round(ebitda_margin, 1),
                "Revenue Growth (%)": round(rev_growth_pct, 1),
                "EV/Revenue": round(ev_rev, 2),
                "EV/EBITDA": round(ev_ebitda, 1),
            })
        except Exception:
            pass
    return results

# ══════════════════════════════════════════════════════════════
# SEARCH HISTORY
# ══════════════════════════════════════════════════════════════
def _init_search_history():
    """Initialize search history in session state."""
    if "search_history" not in st.session_state:
        st.session_state.search_history = []

def _add_to_search_history(ticker: str):
    """Add a ticker to search history."""
    _init_search_history()
    ticker = ticker.upper().strip()
    if ticker:
        # Remove if already exists (to move to front)
        if ticker in st.session_state.search_history:
            st.session_state.search_history.remove(ticker)
        # Add to front
        st.session_state.search_history.insert(0, ticker)
        # Keep only last 10
        st.session_state.search_history = st.session_state.search_history[:10]

def _get_search_history() -> list:
    """Get search history."""
    _init_search_history()
    return st.session_state.search_history

# ══════════════════════════════════════════════════════════════
# MARKET INDICES OVERVIEW
# ══════════════════════════════════════════════════════════════
@st.cache_data(ttl=300, show_spinner=False)
def _fetch_market_indices() -> list:
    """Fetch major market indices data."""
    indices = [
        ("^GSPC", "S&P 500"),
        ("^DJI", "Dow Jones"),
        ("^IXIC", "NASDAQ"),
        ("^RUT", "Russell 2000"),
        ("^VIX", "VIX"),
        ("^GSPTSE", "TSX"),
    ]
    
    results = []
    for symbol, name in indices:
        try:
            tk = yf.Ticker(symbol)
            info = tk.info or {}
            price = info.get("regularMarketPrice") or info.get("previousClose") or 0
            change = info.get("regularMarketChange") or 0
            change_pct = info.get("regularMarketChangePercent") or 0
            results.append({
                "symbol": symbol,
                "name": name,
                "price": price,
                "change": change,
                "change_pct": change_pct,
            })
        except Exception:
            continue
    
    return results

def _render_market_ticker(indices: list):
    """Render a scrolling market ticker."""
    if not indices:
        return
    
    ticker_items = []
    for idx in indices:
        color = "#10B981" if idx["change_pct"] >= 0 else "#EF4444"
        arrow = "▲" if idx["change_pct"] >= 0 else "▼"
        ticker_items.append(
            f'<span style="margin-right:2rem;">'
            f'<span style="color:#E0DCF5; font-weight:600;">{idx["name"]}</span> '
            f'<span style="color:{color};">{idx["price"]:,.2f} {arrow} {idx["change_pct"]:+.2f}%</span>'
            f'</span>'
        )
    
    # Duplicate for seamless scroll
    ticker_html = "".join(ticker_items) * 2
    
    st.markdown(
        f'<div style="overflow:hidden; background:rgba(107,92,231,0.05); '
        f'border-top:1px solid rgba(107,92,231,0.15); border-bottom:1px solid rgba(107,92,231,0.15); '
        f'padding:0.5rem 0; margin-bottom:1rem;">'
        f'<div style="display:inline-block; white-space:nowrap; animation:ticker-scroll 30s linear infinite;">'
        f'{ticker_html}'
        f'</div></div>',
        unsafe_allow_html=True,
    )

# ══════════════════════════════════════════════════════════════
# SCREENER - Quick filter for stocks by criteria
# ══════════════════════════════════════════════════════════════
SECTOR_TICKERS = {
    "Technology": ["AAPL", "MSFT", "GOOGL", "META", "NVDA", "AMZN", "CRM", "ADBE", "INTC", "AMD", 
                   "CSCO", "ORCL", "IBM", "QCOM", "TXN", "AVGO", "NOW", "SHOP", "SNOW", "PLTR"],
    "Healthcare": ["JNJ", "UNH", "PFE", "ABBV", "MRK", "TMO", "ABT", "LLY", "BMY", "AMGN",
                   "GILD", "MDT", "CVS", "CI", "ISRG", "VRTX", "REGN", "ZTS", "BDX", "EW"],
    "Financials": ["JPM", "BAC", "WFC", "GS", "MS", "BLK", "C", "AXP", "SCHW", "USB",
                   "PNC", "TFC", "COF", "BK", "STT", "SPGI", "CME", "ICE", "MMC", "AON"],
    "Consumer": ["WMT", "PG", "KO", "PEP", "COST", "NKE", "MCD", "SBUX", "TGT", "HD",
                 "LOW", "TJX", "DG", "DLTR", "ROST", "YUM", "CMG", "DPZ", "EL", "CL"],
    "Industrials": ["CAT", "HON", "UNP", "UPS", "BA", "RTX", "DE", "GE", "LMT", "MMM",
                    "EMR", "ETN", "ITW", "PH", "ROK", "FDX", "CSX", "NSC", "WM", "RSG"],
    "Energy": ["XOM", "CVX", "COP", "SLB", "EOG", "OXY", "MPC", "VLO", "PSX", "DVN",
               "HAL", "BKR", "FANG", "HES", "PXD", "KMI", "WMB", "OKE", "TRGP", "LNG"],
}

@st.cache_data(ttl=300, show_spinner=False)
def _fetch_top_movers() -> dict:
    """Fetch top gainers and losers from major stocks."""
    major_stocks = [
        "AAPL", "MSFT", "GOOGL", "AMZN", "META", "NVDA", "TSLA", "BRK-B", "JPM", "V",
        "JNJ", "WMT", "PG", "MA", "UNH", "HD", "DIS", "PYPL", "NFLX", "ADBE",
        "CRM", "INTC", "AMD", "CSCO", "PFE", "ABT", "KO", "PEP", "MRK", "VZ"
    ]
    
    results = []
    for ticker in major_stocks[:20]:
        try:
            tk = yf.Ticker(ticker)
            info = tk.info or {}
            change_pct = info.get("regularMarketChangePercent") or 0
            price = info.get("currentPrice") or info.get("regularMarketPrice") or 0
            results.append({
                "ticker": ticker,
                "name": info.get("shortName", ticker)[:20],
                "price": price,
                "change_pct": change_pct,
            })
        except Exception:
            continue
    
    # Sort by change percentage
    gainers = sorted([r for r in results if r["change_pct"] > 0], 
                     key=lambda x: x["change_pct"], reverse=True)[:5]
    losers = sorted([r for r in results if r["change_pct"] < 0], 
                    key=lambda x: x["change_pct"])[:5]
    
    return {"gainers": gainers, "losers": losers}

# ══════════════════════════════════════════════════════════════
# SPARKLINE - Mini inline charts
# ══════════════════════════════════════════════════════════════
def _render_sparkline_svg(values: list, color: str = "#6B5CE7", width: int = 80, height: int = 24) -> str:
    """Generate an SVG sparkline from a list of values."""
    if not values or len(values) < 2:
        return ""
    
    # Normalize values to fit in the height
    min_val = min(values)
    max_val = max(values)
    range_val = max_val - min_val if max_val != min_val else 1
    
    # Calculate points
    step = width / (len(values) - 1)
    points = []
    for i, v in enumerate(values):
        x = i * step
        y = height - ((v - min_val) / range_val * (height - 4) + 2)  # 2px padding
        points.append(f"{x:.1f},{y:.1f}")
    
    path = " ".join(points)
    
    # Determine if trend is up or down
    trend_color = "#10B981" if values[-1] > values[0] else "#EF4444" if values[-1] < values[0] else color
    
    return (
        f'<svg width="{width}" height="{height}" viewBox="0 0 {width} {height}" '
        f'style="display:inline-block; vertical-align:middle;">'
        f'<polyline fill="none" stroke="{trend_color}" stroke-width="2" '
        f'stroke-linecap="round" stroke-linejoin="round" points="{path}" '
        f'style="animation: sparklinePulse 2s ease-in-out infinite;"/>'
        f'<circle cx="{(len(values)-1)*step:.1f}" cy="{height - ((values[-1] - min_val) / range_val * (height - 4) + 2):.1f}" '
        f'r="3" fill="{trend_color}"/>'
        f'</svg>'
    )

# ══════════════════════════════════════════════════════════════
# STATUS BADGES - Colored status indicators
# ══════════════════════════════════════════════════════════════
def _render_status_badge(text: str, status: str = "neutral") -> str:
    """Render a colored status badge."""
    colors = {
        "positive": ("#10B981", "rgba(16,185,129,0.15)"),
        "negative": ("#EF4444", "rgba(239,68,68,0.15)"),
        "warning": ("#F5A623", "rgba(245,166,35,0.15)"),
        "neutral": ("#8A85AD", "rgba(138,133,173,0.15)"),
        "info": ("#6B5CE7", "rgba(107,92,231,0.15)"),
    }
    text_color, bg_color = colors.get(status, colors["neutral"])
    
    return (
        f'<span style="display:inline-block; padding:0.25rem 0.6rem; border-radius:12px; '
        f'font-size:0.7rem; font-weight:600; letter-spacing:0.5px; '
        f'color:{text_color}; background:{bg_color}; '
        f'animation: badgePop 0.3s ease-out;">{text}</span>'
    )

# ══════════════════════════════════════════════════════════════
# KEYBOARD SHORTCUTS OVERLAY
# ══════════════════════════════════════════════════════════════
def _render_keyboard_shortcuts():
    """Render keyboard shortcuts help section."""
    shortcuts = [
        ("⌘/Ctrl + K", "Quick search"),
        ("⌘/Ctrl + B", "Toggle sidebar"),
        ("⌘/Ctrl + D", "Download PPTX"),
        ("⌘/Ctrl + E", "Export Excel"),
        ("⌘/Ctrl + W", "Add to watchlist"),
        ("?", "Show shortcuts"),
    ]
    
    st.markdown(
        '<div style="background:rgba(107,92,231,0.05); border:1px solid rgba(107,92,231,0.15); '
        'border-radius:12px; padding:1rem; margin:1rem 0;">'
        '<div style="font-size:0.75rem; font-weight:700; color:#9B8AFF; text-transform:uppercase; '
        'letter-spacing:1px; margin-bottom:0.8rem;">⌨️ Keyboard Shortcuts</div>',
        unsafe_allow_html=True,
    )
    
    for key, desc in shortcuts:
        st.markdown(
            f'<div style="display:flex; justify-content:space-between; padding:0.3rem 0; '
            f'border-bottom:1px solid rgba(255,255,255,0.05);">'
            f'<kbd style="background:rgba(0,0,0,0.3); padding:0.2rem 0.5rem; border-radius:4px; '
            f'font-family:monospace; font-size:0.75rem; color:#E0DCF5;">{key}</kbd>'
            f'<span style="color:#B8B3D7; font-size:0.8rem;">{desc}</span>'
            f'</div>',
            unsafe_allow_html=True,
        )
    
    st.markdown('</div>', unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════
# METRIC CARD WITH SPARKLINE
# ══════════════════════════════════════════════════════════════
def _render_metric_with_sparkline(label: str, value: str, sparkline_data: list = None, 
                                   delta: str = None, delta_color: str = None):
    """Render a metric card with optional sparkline."""
    sparkline_html = ""
    if sparkline_data and len(sparkline_data) >= 2:
        sparkline_html = _render_sparkline_svg(sparkline_data)
    
    delta_html = ""
    if delta:
        d_color = delta_color or ("#10B981" if delta.startswith("+") else "#EF4444")
        delta_html = f'<span style="color:{d_color}; font-size:0.75rem; margin-left:0.5rem;">{delta}</span>'
    
    st.markdown(
        f'<div style="background:rgba(255,255,255,0.04); border:1px solid rgba(107,92,231,0.15); '
        f'border-radius:12px; padding:1rem; position:relative; overflow:hidden;">'
        f'<div style="font-size:0.7rem; font-weight:600; color:#8A85AD; text-transform:uppercase; '
        f'letter-spacing:0.5px; margin-bottom:0.3rem;">{label}</div>'
        f'<div style="display:flex; align-items:center; justify-content:space-between;">'
        f'<div style="font-size:1.3rem; font-weight:700; color:#E0DCF5;">{value}{delta_html}</div>'
        f'{sparkline_html}'
        f'</div></div>',
        unsafe_allow_html=True,
    )

def _render_movers_cards(movers: dict):
    """Render top gainers and losers cards."""
    if not movers or (not movers.get("gainers") and not movers.get("losers")):
        return
    
    st.markdown(
        '<div style="display:grid; grid-template-columns:1fr 1fr; gap:1rem; margin-top:1rem;">',
        unsafe_allow_html=True,
    )
    
    col1, col2 = st.columns(2)
    
    with col1:
        st.markdown(
            '<div style="background:rgba(16,185,129,0.08); border:1px solid rgba(16,185,129,0.2); '
            'border-radius:12px; padding:1rem;">'
            '<div style="font-size:0.75rem; font-weight:700; color:#10B981; text-transform:uppercase; '
            'letter-spacing:1px; margin-bottom:0.8rem;">🚀 Top Gainers</div>',
            unsafe_allow_html=True,
        )
        for stock in movers.get("gainers", [])[:5]:
            st.markdown(
                f'<div style="display:flex; justify-content:space-between; padding:0.3rem 0; '
                f'border-bottom:1px solid rgba(255,255,255,0.05);">'
                f'<span style="color:#E0DCF5; font-weight:600;">{stock["ticker"]}</span>'
                f'<span style="color:#10B981; font-weight:700;">+{stock["change_pct"]:.2f}%</span>'
                f'</div>',
                unsafe_allow_html=True,
            )
        st.markdown('</div>', unsafe_allow_html=True)
    
    with col2:
        st.markdown(
            '<div style="background:rgba(239,68,68,0.08); border:1px solid rgba(239,68,68,0.2); '
            'border-radius:12px; padding:1rem;">'
            '<div style="font-size:0.75rem; font-weight:700; color:#EF4444; text-transform:uppercase; '
            'letter-spacing:1px; margin-bottom:0.8rem;">📉 Top Losers</div>',
            unsafe_allow_html=True,
        )
        for stock in movers.get("losers", [])[:5]:
            st.markdown(
                f'<div style="display:flex; justify-content:space-between; padding:0.3rem 0; '
                f'border-bottom:1px solid rgba(255,255,255,0.05);">'
                f'<span style="color:#E0DCF5; font-weight:600;">{stock["ticker"]}</span>'
                f'<span style="color:#EF4444; font-weight:700;">{stock["change_pct"]:.2f}%</span>'
                f'</div>',
                unsafe_allow_html=True,
            )
        st.markdown('</div>', unsafe_allow_html=True)
    
    st.markdown('</div>', unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════
# EARNINGS CALENDAR
# ══════════════════════════════════════════════════════════════
@st.cache_data(ttl=1800, show_spinner=False)
def _fetch_earnings_calendar() -> list:
    """Fetch upcoming earnings for popular tickers."""
    watchlist_tickers = ["AAPL", "MSFT", "GOOGL", "AMZN", "META", "NVDA", "TSLA", "JPM",
                         "V", "MA", "CRM", "ADBE", "NFLX", "DIS", "PYPL", "SQ", "SHOP",
                         "RY.TO", "TD.TO", "BNS.TO", "CSU.TO", "BAM.TO"]
    earnings = []
    now = datetime.now()
    
    for ticker in watchlist_tickers[:15]:  # Limit API calls
        try:
            tk = yf.Ticker(ticker)
            cal = tk.calendar
            if cal is not None and not (isinstance(cal, pd.DataFrame) and cal.empty):
                if isinstance(cal, dict):
                    ed = cal.get("Earnings Date")
                    if ed:
                        # ed can be a list of dates or a single date
                        if isinstance(ed, list) and len(ed) > 0:
                            earn_date = ed[0]
                        else:
                            earn_date = ed
                        if hasattr(earn_date, 'date'):
                            earn_date = earn_date.date() if hasattr(earn_date, 'date') else earn_date
                        earnings.append({
                            "ticker": ticker,
                            "date": str(earn_date),
                            "estimate_eps": cal.get("Earnings Average"),
                            "revenue_est": cal.get("Revenue Average"),
                        })
                elif isinstance(cal, pd.DataFrame):
                    if "Earnings Date" in cal.columns or "Earnings Date" in cal.index:
                        try:
                            ed = cal.loc["Earnings Date"] if "Earnings Date" in cal.index else None
                            if ed is not None:
                                earn_date = ed.iloc[0] if hasattr(ed, 'iloc') else ed
                                earnings.append({
                                    "ticker": ticker,
                                    "date": str(earn_date),
                                    "estimate_eps": None,
                                    "revenue_est": None,
                                })
                        except Exception:
                            pass
        except Exception:
            continue
    
    # Sort by date
    earnings.sort(key=lambda x: x["date"])
    return earnings


def _render_earnings_calendar(earnings: list):
    """Render an earnings calendar widget TradingView-style."""
    if not earnings:
        return
    
    st.markdown(
        '<div style="background:rgba(107,92,231,0.05); border:1px solid rgba(107,92,231,0.15); '
        'border-radius:16px; padding:1.5rem; margin-top:1rem;">'
        '<div style="display:flex; align-items:center; gap:0.5rem; margin-bottom:1rem;">'
        '<span style="font-size:1.2rem;">📅</span>'
        '<span style="font-size:0.8rem; font-weight:700; color:#9B8AFF; text-transform:uppercase; '
        'letter-spacing:1.5px;">Upcoming Earnings</span>'
        '</div>',
        unsafe_allow_html=True,
    )
    
    # Group by date
    from itertools import groupby
    for date_str, group in groupby(earnings[:12], key=lambda x: x["date"]):
        try:
            dt = datetime.strptime(date_str[:10], "%Y-%m-%d")
            date_label = dt.strftime("%b %d, %Y")
            day_name = dt.strftime("%A")
        except Exception:
            date_label = date_str
            day_name = ""
        
        st.markdown(
            f'<div style="font-size:0.7rem; color:#6B5CE7; font-weight:600; margin-top:0.8rem; '
            f'margin-bottom:0.3rem; padding-bottom:0.2rem; border-bottom:1px solid rgba(107,92,231,0.15);">'
            f'{day_name} — {date_label}</div>',
            unsafe_allow_html=True,
        )
        
        for item in group:
            eps_str = ""
            if item.get("estimate_eps"):
                eps_str = f'<span style="color:#8A85AD; font-size:0.65rem;"> Est EPS: ${item["estimate_eps"]:.2f}</span>'
            st.markdown(
                f'<div style="display:flex; justify-content:space-between; align-items:center; '
                f'padding:0.35rem 0.5rem; border-radius:6px; margin:0.15rem 0; '
                f'background:rgba(255,255,255,0.02); transition:background 0.2s;">'
                f'<span style="color:#E0DCF5; font-weight:600; font-size:0.8rem;">{item["ticker"]}</span>'
                f'{eps_str}'
                f'</div>',
                unsafe_allow_html=True,
            )
    
    st.markdown('</div>', unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════
# NEWS FEED
# ══════════════════════════════════════════════════════════════
@st.cache_data(ttl=900, show_spinner=False)
def _fetch_news_feed(tickers: list = None) -> list:
    """Fetch recent news for given tickers or market-wide."""
    if not tickers:
        tickers = ["AAPL", "MSFT", "NVDA", "GOOGL", "TSLA"]
    
    all_news = []
    seen_titles = set()
    
    for ticker in tickers[:5]:
        try:
            tk = yf.Ticker(ticker)
            news = tk.news or []
            for item in news[:3]:
                title = item.get("title", "")
                if title and title not in seen_titles:
                    seen_titles.add(title)
                    all_news.append({
                        "title": title,
                        "publisher": item.get("publisher", ""),
                        "link": item.get("link", ""),
                        "ticker": ticker,
                        "published": item.get("providerPublishTime", 0),
                        "type": item.get("type", ""),
                    })
        except Exception:
            continue
    
    # Sort by publish time descending
    all_news.sort(key=lambda x: x.get("published", 0), reverse=True)
    return all_news[:15]


def _render_news_feed(news: list):
    """Render a news feed widget."""
    if not news:
        return
    
    st.markdown(
        '<div style="background:rgba(107,92,231,0.05); border:1px solid rgba(107,92,231,0.15); '
        'border-radius:16px; padding:1.5rem; margin-top:1rem;">'
        '<div style="display:flex; align-items:center; gap:0.5rem; margin-bottom:1rem;">'
        '<span style="font-size:1.2rem;">📰</span>'
        '<span style="font-size:0.8rem; font-weight:700; color:#9B8AFF; text-transform:uppercase; '
        'letter-spacing:1.5px;">Market News</span>'
        '</div>',
        unsafe_allow_html=True,
    )
    
    for item in news[:10]:
        # Time ago
        time_str = ""
        if item.get("published"):
            try:
                delta = datetime.now() - datetime.fromtimestamp(item["published"])
                if delta.days > 0:
                    time_str = f'{delta.days}d ago'
                elif delta.seconds > 3600:
                    time_str = f'{delta.seconds // 3600}h ago'
                else:
                    time_str = f'{delta.seconds // 60}m ago'
            except Exception:
                pass
        
        link_html = f'href="{item["link"]}" target="_blank"' if item.get("link") else ""
        
        st.markdown(
            f'<a {link_html} style="display:block; padding:0.6rem 0.5rem; border-radius:8px; '
            f'margin:0.2rem 0; background:rgba(255,255,255,0.02); text-decoration:none; '
            f'transition:background 0.2s; border-bottom:1px solid rgba(255,255,255,0.03);">'
            f'<div style="font-size:0.78rem; color:#E0DCF5; font-weight:500; line-height:1.35;">{item["title"]}</div>'
            f'<div style="display:flex; justify-content:space-between; margin-top:0.25rem;">'
            f'<span style="font-size:0.65rem; color:#6B5CE7; font-weight:600;">{item["ticker"]}</span>'
            f'<span style="font-size:0.6rem; color:#8A85AD;">{item.get("publisher", "")} · {time_str}</span>'
            f'</div>'
            f'</a>',
            unsafe_allow_html=True,
        )
    
    st.markdown('</div>', unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════
# WATCHLIST NOTES
# ══════════════════════════════════════════════════════════════
def _init_watchlist_notes():
    """Initialize watchlist notes in session state."""
    if "watchlist_notes" not in st.session_state:
        st.session_state.watchlist_notes = {}


def _set_watchlist_note(ticker: str, note: str):
    """Set a note for a watchlist ticker."""
    _init_watchlist_notes()
    st.session_state.watchlist_notes[ticker.upper().strip()] = note


def _get_watchlist_note(ticker: str) -> str:
    """Get note for a watchlist ticker."""
    _init_watchlist_notes()
    return st.session_state.watchlist_notes.get(ticker.upper().strip(), "")


# ══════════════════════════════════════════════════════════════
# FEAR & GREED INDEX (Simplified Market Sentiment)
# ══════════════════════════════════════════════════════════════
@st.cache_data(ttl=1800, show_spinner=False)
def _calculate_market_sentiment() -> dict:
    """Calculate a simplified market sentiment score based on market data."""
    try:
        spy = yf.Ticker("SPY")
        hist = spy.history(period="1mo")
        if hist.empty:
            return {"score": 50, "label": "Neutral", "color": "#F59E0B"}
        
        # Simple sentiment factors
        current_price = hist["Close"].iloc[-1]
        sma_20 = hist["Close"].rolling(20).mean().iloc[-1]
        
        # Price vs SMA
        price_signal = (current_price / sma_20 - 1) * 100  # % above/below 20-day SMA
        
        # Recent momentum (5-day return)
        if len(hist) >= 5:
            momentum = (hist["Close"].iloc[-1] / hist["Close"].iloc[-5] - 1) * 100
        else:
            momentum = 0
        
        # Volatility (higher = more fearful)
        volatility = hist["Close"].pct_change().std() * 100
        vol_signal = max(0, 2 - volatility) * 25  # Lower vol = higher score
        
        # Combine signals (0-100 scale)
        raw_score = 50 + (price_signal * 5) + (momentum * 3) + (vol_signal - 25)
        score = max(0, min(100, raw_score))
        
        if score >= 80:
            return {"score": round(score), "label": "Extreme Greed", "color": "#10B981"}
        elif score >= 60:
            return {"score": round(score), "label": "Greed", "color": "#34D399"}
        elif score >= 40:
            return {"score": round(score), "label": "Neutral", "color": "#F59E0B"}
        elif score >= 20:
            return {"score": round(score), "label": "Fear", "color": "#F97316"}
        else:
            return {"score": round(score), "label": "Extreme Fear", "color": "#EF4444"}
    except Exception:
        return {"score": 50, "label": "Neutral", "color": "#F59E0B"}


def _render_sentiment_gauge(sentiment: dict):
    """Render a CNN-style fear & greed gauge."""
    score = sentiment["score"]
    label = sentiment["label"]
    color = sentiment["color"]
    
    # SVG gauge
    angle = (score / 100) * 180 - 90  # -90 to 90 degrees
    
    gauge_svg = f'''
    <div style="background:rgba(107,92,231,0.05); border:1px solid rgba(107,92,231,0.15);
        border-radius:16px; padding:1.5rem; margin-top:1rem; text-align:center;">
        <div style="display:flex; align-items:center; justify-content:center; gap:0.5rem; margin-bottom:1rem;">
            <span style="font-size:1.2rem;">🎯</span>
            <span style="font-size:0.8rem; font-weight:700; color:#9B8AFF; text-transform:uppercase;
            letter-spacing:1.5px;">Market Sentiment</span>
        </div>
        <svg viewBox="0 0 200 120" width="200" height="120" style="margin:0 auto; display:block;">
            <!-- Background arc -->
            <path d="M 20 100 A 80 80 0 0 1 180 100" fill="none" stroke="rgba(255,255,255,0.1)" stroke-width="12" stroke-linecap="round"/>
            <!-- Colored segments -->
            <path d="M 20 100 A 80 80 0 0 1 52 40" fill="none" stroke="#EF4444" stroke-width="12" stroke-linecap="round" opacity="0.6"/>
            <path d="M 52 40 A 80 80 0 0 1 100 20" fill="none" stroke="#F97316" stroke-width="12" stroke-linecap="round" opacity="0.6"/>
            <path d="M 100 20 A 80 80 0 0 1 148 40" fill="none" stroke="#F59E0B" stroke-width="12" stroke-linecap="round" opacity="0.6"/>
            <path d="M 148 40 A 80 80 0 0 1 180 100" fill="none" stroke="#10B981" stroke-width="12" stroke-linecap="round" opacity="0.6"/>
            <!-- Needle -->
            <line x1="100" y1="100" x2="{100 + 60 * np.cos(np.radians(angle + 180))}" y2="{100 + 60 * np.sin(np.radians(angle + 180))}"
                stroke="{color}" stroke-width="3" stroke-linecap="round"/>
            <circle cx="100" cy="100" r="5" fill="{color}"/>
        </svg>
        <div style="font-size:2rem; font-weight:800; color:{color}; margin-top:0.5rem;">{score}</div>
        <div style="font-size:0.85rem; font-weight:700; color:{color}; text-transform:uppercase; letter-spacing:1px;">{label}</div>
        <div style="font-size:0.6rem; color:#8A85AD; margin-top:0.3rem;">Based on SPY momentum, trend & volatility</div>
    </div>
    '''
    st.markdown(gauge_svg, unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════
# SECTOR HEATMAP (TradingView-inspired)
# ══════════════════════════════════════════════════════════════
@st.cache_data(ttl=900, show_spinner=False)
def _fetch_sector_performance() -> list:
    """Fetch sector ETF performance for heatmap."""
    sector_etfs = {
        "Technology": "XLK",
        "Healthcare": "XLV",
        "Financials": "XLF",
        "Consumer Disc.": "XLY",
        "Consumer Stpl.": "XLP",
        "Energy": "XLE",
        "Industrials": "XLI",
        "Materials": "XLB",
        "Utilities": "XLU",
        "Real Estate": "XLRE",
        "Comm. Services": "XLC",
    }
    
    results = []
    for name, etf in sector_etfs.items():
        try:
            tk = yf.Ticker(etf)
            info = tk.info or {}
            change = info.get("regularMarketChangePercent") or 0
            results.append({"name": name, "etf": etf, "change_pct": change})
        except Exception:
            results.append({"name": name, "etf": etf, "change_pct": 0})
    
    return results


def _render_sector_heatmap(sectors: list):
    """Render a TradingView-style sector heatmap."""
    if not sectors:
        return
    
    st.markdown(
        '<div style="background:rgba(107,92,231,0.05); border:1px solid rgba(107,92,231,0.15); '
        'border-radius:16px; padding:1.5rem; margin-top:1rem;">'
        '<div style="display:flex; align-items:center; gap:0.5rem; margin-bottom:1rem;">'
        '<span style="font-size:1.2rem;">🗺️</span>'
        '<span style="font-size:0.8rem; font-weight:700; color:#9B8AFF; text-transform:uppercase; '
        'letter-spacing:1.5px;">Sector Performance</span>'
        '</div>'
        '<div style="display:grid; grid-template-columns:repeat(4, 1fr); gap:0.5rem;">',
        unsafe_allow_html=True,
    )
    
    for sector in sectors:
        pct = sector["change_pct"]
        if pct >= 2:
            bg = "rgba(16,185,129,0.35)"
        elif pct >= 0.5:
            bg = "rgba(16,185,129,0.2)"
        elif pct >= 0:
            bg = "rgba(16,185,129,0.08)"
        elif pct >= -0.5:
            bg = "rgba(239,68,68,0.08)"
        elif pct >= -2:
            bg = "rgba(239,68,68,0.2)"
        else:
            bg = "rgba(239,68,68,0.35)"
        
        color = "#10B981" if pct >= 0 else "#EF4444"
        arrow = "▲" if pct >= 0 else "▼"
        
        st.markdown(
            f'<div style="background:{bg}; border-radius:10px; padding:0.7rem 0.5rem; text-align:center; '
            f'border:1px solid rgba(255,255,255,0.05);">'
            f'<div style="font-size:0.65rem; color:#C4BFE0; font-weight:600; white-space:nowrap; '
            f'overflow:hidden; text-overflow:ellipsis;">{sector["name"]}</div>'
            f'<div style="font-size:0.95rem; font-weight:800; color:{color}; margin-top:0.2rem;">'
            f'{arrow} {pct:+.2f}%</div>'
            f'</div>',
            unsafe_allow_html=True,
        )
    
    st.markdown('</div></div>', unsafe_allow_html=True)


@st.cache_data(ttl=600, show_spinner=False)
def _screen_sector(sector: str, sort_by: str = "market_cap", top_n: int = 10) -> list:
    """Screen stocks in a sector by various criteria."""
    tickers = SECTOR_TICKERS.get(sector, [])
    results = []
    
    for ticker in tickers[:20]:  # Limit API calls
        try:
            tk = yf.Ticker(ticker)
            info = tk.info or {}
            results.append({
                "ticker": ticker,
                "name": info.get("shortName", ticker)[:30],
                "price": info.get("currentPrice") or info.get("regularMarketPrice") or 0,
                "market_cap": info.get("marketCap") or 0,
                "pe_ratio": info.get("trailingPE") or 0,
                "change_pct": info.get("regularMarketChangePercent") or 0,
            })
        except Exception:
            continue
    
    # Sort by specified criteria
    if sort_by == "market_cap":
        results.sort(key=lambda x: x["market_cap"], reverse=True)
    elif sort_by == "pe_ratio":
        results.sort(key=lambda x: x["pe_ratio"] if x["pe_ratio"] > 0 else 9999)
    elif sort_by == "change_pct":
        results.sort(key=lambda x: x["change_pct"], reverse=True)
    
    return results[:top_n]

# ══════════════════════════════════════════════════════════════
# EXCEL/CSV EXPORT UTILITIES
# ══════════════════════════════════════════════════════════════
def _export_to_excel(cd) -> bytes:
    """Export company data to Excel with multiple sheets."""
    output = io.BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Summary sheet
        summary_data = {
            'Metric': ['Company Name', 'Ticker', 'Sector', 'Industry', 'Market Cap', 
                      'Enterprise Value', 'Current Price', 'P/E Ratio', 'Forward P/E',
                      'EV/EBITDA', 'EV/Revenue', 'Gross Margin', 'Operating Margin',
                      'Net Margin', 'ROE', 'ROA', 'Debt/Equity', 'Current Ratio'],
            'Value': [cd.name, cd.ticker, cd.sector, cd.industry, 
                     format_number(cd.market_cap, currency_symbol=cd.currency_symbol),
                     format_number(cd.enterprise_value, currency_symbol=cd.currency_symbol),
                     f"{cd.currency_symbol}{cd.current_price:,.2f}",
                     format_multiple(cd.trailing_pe), format_multiple(cd.forward_pe),
                     format_multiple(cd.ev_to_ebitda), format_multiple(cd.ev_to_revenue),
                     format_pct(cd.gross_margins), format_pct(cd.operating_margins),
                     format_pct(cd.profit_margins), format_pct(cd.return_on_equity),
                     format_pct(cd.return_on_assets), format_multiple(cd.debt_to_equity),
                     format_multiple(cd.current_ratio)]
        }
        pd.DataFrame(summary_data).to_excel(writer, sheet_name='Summary', index=False)
        
        # Income Statement
        if cd.revenue is not None and len(cd.revenue) > 0:
            income_data = {'Year': [str(idx.year) if hasattr(idx, 'year') else str(idx) for idx in cd.revenue.index]}
            income_data['Revenue'] = cd.revenue.values
            if cd.gross_profit is not None:
                income_data['Gross Profit'] = cd.gross_profit.values[:len(cd.revenue)]
            if cd.operating_income is not None:
                income_data['Operating Income'] = cd.operating_income.values[:len(cd.revenue)]
            if cd.net_income is not None:
                income_data['Net Income'] = cd.net_income.values[:len(cd.revenue)]
            if cd.ebitda is not None:
                income_data['EBITDA'] = cd.ebitda.values[:len(cd.revenue)]
            pd.DataFrame(income_data).to_excel(writer, sheet_name='Income Statement', index=False)
        
        # Balance Sheet
        if cd.total_assets is not None and len(cd.total_assets) > 0:
            bs_data = {'Year': [str(idx.year) if hasattr(idx, 'year') else str(idx) for idx in cd.total_assets.index]}
            bs_data['Total Assets'] = cd.total_assets.values
            if cd.total_equity is not None:
                bs_data['Total Equity'] = cd.total_equity.values[:len(cd.total_assets)]
            if cd.total_debt is not None:
                bs_data['Total Debt'] = cd.total_debt.values[:len(cd.total_assets)]
            if cd.cash_and_equivalents is not None:
                bs_data['Cash'] = cd.cash_and_equivalents.values[:len(cd.total_assets)]
            pd.DataFrame(bs_data).to_excel(writer, sheet_name='Balance Sheet', index=False)
        
        # Cash Flow
        if cd.operating_cashflow_series is not None and len(cd.operating_cashflow_series) > 0:
            cf_data = {'Year': [str(idx.year) if hasattr(idx, 'year') else str(idx) for idx in cd.operating_cashflow_series.index]}
            cf_data['Operating Cash Flow'] = cd.operating_cashflow_series.values
            if cd.capital_expenditure is not None:
                cf_data['CapEx'] = cd.capital_expenditure.values[:len(cd.operating_cashflow_series)]
            if cd.free_cashflow_series is not None:
                cf_data['Free Cash Flow'] = cd.free_cashflow_series.values[:len(cd.operating_cashflow_series)]
            pd.DataFrame(cf_data).to_excel(writer, sheet_name='Cash Flow', index=False)
        
        # Key Ratios
        try:
            ratios_data = {
                'Ratio': ['Gross Margin', 'Operating Margin', 'Net Margin', 'ROE', 'ROA', 'ROIC',
                          'Current Ratio', 'Quick Ratio', 'Debt/Equity', 'Interest Coverage',
                          'Asset Turnover', 'Inventory Turnover', 'Revenue Growth', 'Earnings Growth',
                          'Dividend Yield', 'Payout Ratio', 'Beta'],
                'Value': [
                    format_pct(cd.gross_margins), format_pct(cd.operating_margins), format_pct(cd.profit_margins),
                    format_pct(cd.return_on_equity), format_pct(cd.return_on_assets),
                    format_pct(getattr(cd, 'return_on_invested_capital', None)),
                    format_multiple(cd.current_ratio), format_multiple(getattr(cd, 'quick_ratio', None)),
                    format_multiple(cd.debt_to_equity), format_multiple(getattr(cd, 'interest_coverage', None)),
                    format_multiple(getattr(cd, 'asset_turnover', None)),
                    format_multiple(getattr(cd, 'inventory_turnover', None)),
                    format_pct(cd.revenue_growth), format_pct(getattr(cd, 'earnings_growth', None)),
                    format_pct(cd.dividend_yield), format_pct(getattr(cd, 'payout_ratio', None)),
                    f"{cd.beta:.2f}" if cd.beta else "N/A",
                ]
            }
            pd.DataFrame(ratios_data).to_excel(writer, sheet_name='Key Ratios', index=False)
        except Exception:
            pass

        # Valuation Multiples
        try:
            val_data = {
                'Multiple': ['Trailing P/E', 'Forward P/E', 'PEG Ratio', 'Price/Book', 'Price/Sales',
                             'EV/Revenue', 'EV/EBITDA', 'EV/EBIT', 'Price/FCF', 'Dividend Yield'],
                'Value': [
                    format_multiple(cd.trailing_pe), format_multiple(cd.forward_pe),
                    format_multiple(getattr(cd, 'peg_ratio', None)), format_multiple(cd.price_to_book),
                    format_multiple(getattr(cd, 'price_to_sales', None)),
                    format_multiple(cd.ev_to_revenue), format_multiple(cd.ev_to_ebitda),
                    format_multiple(getattr(cd, 'ev_to_ebit', None)),
                    format_multiple(getattr(cd, 'price_to_fcf', None)),
                    format_pct(cd.dividend_yield),
                ]
            }
            pd.DataFrame(val_data).to_excel(writer, sheet_name='Valuation Multiples', index=False)
        except Exception:
            pass

        # Peer Comparison
        if cd.peer_data:
            peer_df = pd.DataFrame(cd.peer_data)
            peer_df.to_excel(writer, sheet_name='Peer Comparison', index=False)

        # Format headers on all sheets
        try:
            from openpyxl.styles import Font, PatternFill, Alignment, numbers
            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_fill = PatternFill(start_color="6B5CE7", end_color="6B5CE7", fill_type="solid")
            for ws in writer.book.worksheets:
                for cell in ws[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center")
                # Auto-width columns
                for col in ws.columns:
                    max_len = max((len(str(c.value or "")) for c in col), default=10)
                    ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 30)
        except Exception:
            pass
    
    return output.getvalue()


def _export_comps_to_excel(comps_analysis) -> bytes:
    """Export comps analysis to Excel with conditional formatting."""
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        comps_df = generate_comps_table(comps_analysis)
        comps_df.to_excel(writer, sheet_name='Comps Table', index=False)
        # Summary sheet
        tc = comps_analysis.target_comps
        summary = {
            'Metric': ['Target Company', 'Ticker', 'Sector', 'Industry', 'Peers Analyzed',
                       'Median EV/Revenue', 'Median EV/EBITDA', 'Median P/E'],
            'Value': [tc.name, tc.ticker, tc.sector, tc.industry, len(comps_analysis.peers),
                      f"{comps_analysis.median_ev_revenue:.1f}x" if hasattr(comps_analysis, 'median_ev_revenue') else "N/A",
                      f"{comps_analysis.median_ev_ebitda:.1f}x" if hasattr(comps_analysis, 'median_ev_ebitda') else "N/A",
                      f"{comps_analysis.median_pe:.1f}x" if hasattr(comps_analysis, 'median_pe') else "N/A"]
        }
        pd.DataFrame(summary).to_excel(writer, sheet_name='Summary', index=False)
        # Format
        try:
            from openpyxl.styles import Font, PatternFill, Alignment
            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_fill = PatternFill(start_color="6B5CE7", end_color="6B5CE7", fill_type="solid")
            for ws in writer.book.worksheets:
                for cell in ws[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center")
                for col in ws.columns:
                    max_len = max((len(str(c.value or "")) for c in col), default=10)
                    ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 30)
        except Exception:
            pass
    return output.getvalue()


def _export_dcf_to_excel(dcf_cd, dcf_result, assumptions) -> bytes:
    """Export DCF model to Excel."""
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Assumptions
        cs = dcf_cd.currency_symbol
        assumptions_data = {
            'Parameter': ['Company', 'Ticker', 'Current Price', 'Shares Outstanding',
                         'FCF Growth Rate', 'Terminal Growth Rate', 'Discount Rate (WACC)',
                         'Projection Years'],
            'Value': [dcf_cd.name, dcf_cd.ticker, f"{cs}{dcf_cd.current_price:,.2f}",
                     f"{dcf_result.get('shares_outstanding', 0):,.0f}",
                     f"{assumptions['growth_rate']*100:.1f}%", f"{assumptions['terminal_growth']*100:.1f}%",
                     f"{assumptions['discount_rate']*100:.1f}%", str(assumptions['years'])]
        }
        pd.DataFrame(assumptions_data).to_excel(writer, sheet_name='Assumptions', index=False)
        # Projected FCF
        years = list(range(1, dcf_result.get("projection_years", 5) + 1))
        proj_data = {
            'Year': years,
            'Projected FCF': dcf_result.get("projected_fcf", []),
            'PV of FCF': dcf_result.get("pv_fcf", []),
        }
        pd.DataFrame(proj_data).to_excel(writer, sheet_name='Projections', index=False)
        # Valuation Summary
        val_data = {
            'Component': ['PV of Projected FCFs', 'Terminal Value', 'PV of Terminal Value',
                         'Enterprise Value', 'Less: Net Debt', 'Equity Value',
                         'Implied Share Price', 'Current Price', 'Upside/Downside'],
            'Value': [
                f"{cs}{dcf_result.get('pv_fcf_total', 0):,.0f}",
                f"{cs}{dcf_result.get('terminal_value', 0):,.0f}",
                f"{cs}{dcf_result.get('pv_terminal_value', 0):,.0f}",
                f"{cs}{dcf_result.get('enterprise_value', 0):,.0f}",
                f"{cs}{dcf_result.get('net_debt', 0):,.0f}",
                f"{cs}{dcf_result.get('equity_value', 0):,.0f}",
                f"{cs}{dcf_result.get('implied_share_price', 0):,.2f}",
                f"{cs}{dcf_result.get('current_price', 0):,.2f}",
                f"{dcf_result.get('upside_pct', 0):+.1f}%",
            ]
        }
        pd.DataFrame(val_data).to_excel(writer, sheet_name='Valuation', index=False)
        # Format
        try:
            from openpyxl.styles import Font, PatternFill, Alignment
            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_fill = PatternFill(start_color="6B5CE7", end_color="6B5CE7", fill_type="solid")
            for ws in writer.book.worksheets:
                for cell in ws[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center")
                for col in ws.columns:
                    max_len = max((len(str(c.value or "")) for c in col), default=10)
                    ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 30)
        except Exception:
            pass
    return output.getvalue()

def _export_to_csv(cd) -> str:
    """Export key metrics to CSV."""
    data = {
        'Metric': ['Company', 'Ticker', 'Price', 'Market Cap', 'P/E', 'EV/EBITDA', 
                  'Gross Margin', 'Net Margin', 'ROE', 'Debt/Equity'],
        'Value': [cd.name, cd.ticker, cd.current_price, cd.market_cap, 
                 cd.trailing_pe, cd.ev_to_ebitda, cd.gross_margins, 
                 cd.profit_margins, cd.return_on_equity, cd.debt_to_equity]
    }
    return pd.DataFrame(data).to_csv(index=False)

# ══════════════════════════════════════════════════════════════
# DCF VALUATION MODULE
# ══════════════════════════════════════════════════════════════
def _calculate_dcf(cd, growth_rate: float = 0.05, terminal_growth: float = 0.025, 
                   discount_rate: float = 0.10, projection_years: int = 5) -> dict:
    """
    Calculate a simple DCF valuation.
    
    Args:
        cd: Company data object
        growth_rate: Revenue/FCF growth rate for projection period
        terminal_growth: Perpetual growth rate for terminal value
        discount_rate: WACC or required return
        projection_years: Number of years to project
    
    Returns:
        dict with DCF valuation results
    """
    # Get base FCF
    base_fcf = None
    if cd.free_cashflow_series is not None and len(cd.free_cashflow_series) > 0:
        base_fcf = cd.free_cashflow_series.iloc[0]
    elif cd.operating_cashflow_series is not None and cd.capital_expenditure is not None:
        base_fcf = cd.operating_cashflow_series.iloc[0] + cd.capital_expenditure.iloc[0]  # CapEx is negative
    
    if base_fcf is None or base_fcf <= 0:
        return {"error": "Insufficient FCF data for DCF valuation"}
    
    # Project FCF
    projected_fcf = []
    current_fcf = base_fcf
    for year in range(1, projection_years + 1):
        current_fcf = current_fcf * (1 + growth_rate)
        projected_fcf.append(current_fcf)
    
    # Calculate PV of projected FCF
    pv_fcf = []
    for i, fcf in enumerate(projected_fcf):
        pv = fcf / ((1 + discount_rate) ** (i + 1))
        pv_fcf.append(pv)
    
    # Terminal Value (Gordon Growth Model)
    terminal_fcf = projected_fcf[-1] * (1 + terminal_growth)
    terminal_value = terminal_fcf / (discount_rate - terminal_growth)
    pv_terminal = terminal_value / ((1 + discount_rate) ** projection_years)
    
    # Enterprise Value
    dcf_enterprise_value = sum(pv_fcf) + pv_terminal
    
    # Equity Value
    net_debt = (cd.total_debt.iloc[0] if cd.total_debt is not None and len(cd.total_debt) > 0 else 0) - \
               (cd.cash_and_equivalents.iloc[0] if cd.cash_and_equivalents is not None and len(cd.cash_and_equivalents) > 0 else 0)
    
    equity_value = dcf_enterprise_value - net_debt
    
    # Per Share Value
    shares_outstanding = cd.shares_outstanding or (cd.market_cap / cd.current_price if cd.current_price > 0 else 1)
    implied_share_price = equity_value / shares_outstanding if shares_outstanding > 0 else 0
    
    # Upside/Downside
    upside = ((implied_share_price / cd.current_price) - 1) * 100 if cd.current_price > 0 else 0
    
    return {
        "base_fcf": base_fcf,
        "projected_fcf": projected_fcf,
        "pv_fcf": pv_fcf,
        "terminal_value": terminal_value,
        "pv_terminal": pv_terminal,
        "enterprise_value": dcf_enterprise_value,
        "net_debt": net_debt,
        "equity_value": equity_value,
        "shares_outstanding": shares_outstanding,
        "implied_share_price": implied_share_price,
        "current_price": cd.current_price,
        "upside_pct": upside,
        "growth_rate": growth_rate,
        "terminal_growth": terminal_growth,
        "discount_rate": discount_rate,
        "projection_years": projection_years,
    }

def _build_dcf_chart(dcf_result: dict, currency_symbol: str = "$", key: str = "dcf_chart"):
    """Build a visualization for DCF results."""
    if "error" in dcf_result:
        st.warning(dcf_result["error"])
        return
    
    years = list(range(1, dcf_result["projection_years"] + 1))
    
    fig = go.Figure()
    
    # Projected FCF bars
    fig.add_trace(go.Bar(
        x=[f"Year {y}" for y in years],
        y=dcf_result["projected_fcf"],
        name="Projected FCF",
        marker=dict(color="rgba(107,92,231,0.7)", line=dict(color="rgba(255,255,255,0.15)", width=1)),
        text=[format_number(v, currency_symbol=currency_symbol) for v in dcf_result["projected_fcf"]],
        textposition="outside",
        textfont=dict(size=9, color="#B8B3D7"),
    ))
    
    # PV of FCF line
    fig.add_trace(go.Scatter(
        x=[f"Year {y}" for y in years],
        y=dcf_result["pv_fcf"],
        name="PV of FCF",
        mode="lines+markers",
        line=dict(color="#10B981", width=3),
        marker=dict(size=8, line=dict(color="#fff", width=1.5)),
    ))
    
    fig.update_layout(
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="rgba(0,0,0,0)",
        font=dict(family="Inter", size=14, color="#B8B3D7"),
        height=400,
        margin=dict(t=40, b=40, l=60, r=60),
        xaxis=dict(tickfont=dict(size=10, color="#8A85AD"), showgrid=False),
        yaxis=dict(tickfont=dict(size=9, color="#8A85AD"), gridcolor="rgba(107,92,231,0.1)", griddash="dot"),
        legend=dict(font=dict(size=10, color="#B8B3D7"), orientation="h", yanchor="bottom", y=1.02),
        barmode="group",
    )
    
    st.plotly_chart(fig, use_container_width=True, key=key)

def _build_dcf_sensitivity(cd, base_dcf: dict, key: str = "dcf_sensitivity"):
    """Build a sensitivity analysis table for DCF valuation."""
    if "error" in base_dcf:
        return
    
    # Growth rate sensitivities (columns)
    growth_rates = [0.03, 0.05, 0.08, 0.10, 0.12]
    # Discount rate sensitivities (rows)
    discount_rates = [0.08, 0.09, 0.10, 0.11, 0.12]
    
    sensitivity_data = []
    for dr in discount_rates:
        row = {"WACC": f"{dr*100:.0f}%"}
        for gr in growth_rates:
            result = _calculate_dcf(
                cd,
                growth_rate=gr,
                terminal_growth=base_dcf["terminal_growth"],
                discount_rate=dr,
                projection_years=base_dcf["projection_years"]
            )
            if "error" not in result:
                row[f"{gr*100:.0f}% Growth"] = f"${result['implied_share_price']:,.2f}"
            else:
                row[f"{gr*100:.0f}% Growth"] = "N/A"
        sensitivity_data.append(row)
    
    sens_df = pd.DataFrame(sensitivity_data)
    sens_df = sens_df.set_index("WACC")
    
    # Style the dataframe - highlight cells above/below current price
    current_price = base_dcf["current_price"]
    
    def color_cells(val):
        if val == "N/A":
            return "background-color: rgba(138,133,173,0.1)"
        try:
            price = float(val.replace("$", "").replace(",", ""))
            if price > current_price * 1.1:
                return "background-color: rgba(16,185,129,0.2); color: #10B981"
            elif price < current_price * 0.9:
                return "background-color: rgba(239,68,68,0.2); color: #EF4444"
            else:
                return "background-color: rgba(245,166,35,0.15); color: #F5A623"
        except:
            return ""
    
    styled_df = sens_df.style.applymap(color_cells)
    
    st.dataframe(styled_df, use_container_width=True, height=250)
    
    st.markdown(
        '<div style="font-size:0.7rem; color:#8A85AD; margin-top:0.5rem;">'
        '🟢 Green: >10% upside | 🟡 Yellow: ±10% of current | 🔴 Red: >10% downside'
        '</div>',
        unsafe_allow_html=True,
    )

def _build_terminal_value_sensitivity(cd, base_dcf: dict, key: str = "tv_sensitivity"):
    """Build terminal growth vs WACC sensitivity chart."""
    if "error" in base_dcf:
        return
    
    terminal_rates = [0.015, 0.020, 0.025, 0.030, 0.035]
    discount_rates = [0.08, 0.10, 0.12]
    
    fig = go.Figure()
    colors = ["#6B5CE7", "#E8638B", "#10B981"]
    
    for i, dr in enumerate(discount_rates):
        prices = []
        for tr in terminal_rates:
            result = _calculate_dcf(
                cd,
                growth_rate=base_dcf["growth_rate"],
                terminal_growth=tr,
                discount_rate=dr,
                projection_years=base_dcf["projection_years"]
            )
            if "error" not in result:
                prices.append(result["implied_share_price"])
            else:
                prices.append(0)
        
        fig.add_trace(go.Scatter(
            x=[f"{r*100:.1f}%" for r in terminal_rates],
            y=prices,
            mode="lines+markers",
            name=f"WACC {dr*100:.0f}%",
            line=dict(color=colors[i], width=3),
            marker=dict(size=8),
        ))
    
    # Add current price reference line
    fig.add_hline(
        y=base_dcf["current_price"],
        line_dash="dash",
        line_color="rgba(255,255,255,0.3)",
        annotation_text=f"Current: ${base_dcf['current_price']:,.2f}",
        annotation_position="right",
    )
    
    fig.update_layout(
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="rgba(0,0,0,0)",
        font=dict(family="Inter", size=14, color="#B8B3D7"),
        height=350,
        margin=dict(t=40, b=40, l=60, r=80),
        xaxis=dict(title="Terminal Growth Rate", tickfont=dict(size=10, color="#8A85AD")),
        yaxis=dict(title="Implied Share Price", tickfont=dict(size=9, color="#8A85AD"), 
                  gridcolor="rgba(107,92,231,0.1)", griddash="dot", tickprefix="$"),
        legend=dict(font=dict(size=10, color="#B8B3D7"), orientation="h", yanchor="bottom", y=1.02),
    )
    
    st.plotly_chart(fig, use_container_width=True, key=key)

# ══════════════════════════════════════════════════════════════
# STOCK PRICE PERFORMANCE CHART
# ══════════════════════════════════════════════════════════════
def _build_price_performance_chart(tickers: list, period: str = "1y", key: str = "price_perf"):
    """Build a normalized price performance chart for multiple tickers."""
    if not tickers:
        return
    
    fig = go.Figure()
    colors = ["#6B5CE7", "#E8638B", "#10B981", "#F5A623", "#3B82F6", 
              "#8B5CF6", "#EC4899", "#14B8A6", "#F59E0B", "#6366F1"]
    
    for i, ticker in enumerate(tickers[:10]):
        try:
            tk = yf.Ticker(ticker)
            hist = tk.history(period=period)
            if hist.empty:
                continue
            
            # Normalize to 100 at start
            normalized = (hist["Close"] / hist["Close"].iloc[0]) * 100
            
            fig.add_trace(go.Scatter(
                x=hist.index,
                y=normalized,
                mode="lines",
                name=ticker,
                line=dict(color=colors[i % len(colors)], width=2),
            ))
        except Exception:
            continue
    
    # Add 100 reference line
    fig.add_hline(y=100, line_dash="dot", line_color="rgba(255,255,255,0.2)")
    
    fig.update_layout(
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="rgba(0,0,0,0)",
        font=dict(family="Inter", size=14, color="#B8B3D7"),
        height=400,
        margin=dict(t=40, b=40, l=60, r=60),
        xaxis=dict(tickfont=dict(size=9, color="#8A85AD"), showgrid=False),
        yaxis=dict(title="Indexed (100 = Start)", tickfont=dict(size=9, color="#8A85AD"), 
                  gridcolor="rgba(107,92,231,0.1)", griddash="dot"),
        legend=dict(font=dict(size=10, color="#B8B3D7"), orientation="h", yanchor="bottom", y=1.02),
        hovermode="x unified",
    )
    
    st.plotly_chart(fig, use_container_width=True, key=key)

# ══════════════════════════════════════════════════════════════
# QUICK COMPARE - Side-by-side company comparison
# ══════════════════════════════════════════════════════════════
@st.cache_data(ttl=600, show_spinner=False)
def _fetch_comparison_data(tickers: list) -> list:
    """Fetch data for multiple companies for comparison."""
    results = []
    for ticker in tickers:
        try:
            cd = fetch_company_data(ticker)
            results.append(cd)
        except Exception:
            continue
    return results

def _build_comparison_table(companies: list) -> pd.DataFrame:
    """Build a comparison table for multiple companies."""
    if not companies:
        return pd.DataFrame()
    
    metrics = [
        ("Ticker", lambda cd: cd.ticker),
        ("Name", lambda cd: cd.name[:25] + "..." if len(cd.name) > 25 else cd.name),
        ("Price", lambda cd: f"{cd.currency_symbol}{cd.current_price:,.2f}"),
        ("Market Cap", lambda cd: format_number(cd.market_cap, currency_symbol=cd.currency_symbol)),
        ("P/E", lambda cd: format_multiple(cd.trailing_pe)),
        ("Fwd P/E", lambda cd: format_multiple(cd.forward_pe)),
        ("EV/EBITDA", lambda cd: format_multiple(cd.ev_to_ebitda)),
        ("EV/Revenue", lambda cd: format_multiple(cd.ev_to_revenue)),
        ("Gross Margin", lambda cd: format_pct(cd.gross_margins)),
        ("Op Margin", lambda cd: format_pct(cd.operating_margins)),
        ("Net Margin", lambda cd: format_pct(cd.profit_margins)),
        ("ROE", lambda cd: format_pct(cd.return_on_equity)),
        ("ROA", lambda cd: format_pct(cd.return_on_assets)),
        ("Debt/Equity", lambda cd: format_multiple(cd.debt_to_equity)),
        ("Dividend Yield", lambda cd: format_pct(cd.dividend_yield) if cd.dividend_yield else "N/A"),
    ]
    
    data = {}
    for metric_name, getter in metrics:
        data[metric_name] = []
        for cd in companies:
            try:
                data[metric_name].append(getter(cd))
            except Exception:
                data[metric_name].append("N/A")
    
    return pd.DataFrame(data)

def _build_comparison_radar(companies: list, key: str = "compare_radar"):
    """Build a radar chart comparing multiple companies."""
    if len(companies) < 2:
        return
    
    metrics = ["P/E", "EV/EBITDA", "Gross Margin", "ROE", "Debt/Equity"]
    
    fig = go.Figure()
    
    colors = ["#6B5CE7", "#E8638B", "#10B981", "#F5A623", "#3B82F6"]
    
    for i, cd in enumerate(companies[:5]):  # Max 5 companies
        values = []
        for metric in metrics:
            if metric == "P/E":
                val = cd.trailing_pe if cd.trailing_pe and cd.trailing_pe > 0 else 0
            elif metric == "EV/EBITDA":
                val = cd.ev_to_ebitda if cd.ev_to_ebitda and cd.ev_to_ebitda > 0 else 0
            elif metric == "Gross Margin":
                val = (cd.gross_margins or 0) * 100
            elif metric == "ROE":
                val = (cd.return_on_equity or 0) * 100
            elif metric == "Debt/Equity":
                val = cd.debt_to_equity if cd.debt_to_equity else 0
            values.append(val)
        
        # Normalize values to 0-100 scale
        max_vals = [50, 30, 100, 50, 200]  # Reasonable max for each metric
        norm_values = [min(v / m * 100, 120) for v, m in zip(values, max_vals)]
        
        fig.add_trace(go.Scatterpolar(
            r=norm_values + [norm_values[0]],
            theta=metrics + [metrics[0]],
            fill='toself',
            name=cd.ticker,
            fillcolor=f"rgba({int(colors[i][1:3], 16)},{int(colors[i][3:5], 16)},{int(colors[i][5:7], 16)},0.1)",
            line=dict(color=colors[i], width=2),
            marker=dict(size=6),
        ))
    
    fig.update_layout(
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="rgba(0,0,0,0)",
        font=dict(family="Inter", size=12, color="#B8B3D7"),
        polar=dict(
            radialaxis=dict(visible=True, range=[0, 120], tickfont=dict(size=8, color="#8A85AD"),
                           gridcolor="rgba(107,92,231,0.1)"),
            angularaxis=dict(tickfont=dict(size=10, color="#8A85AD"),
                            gridcolor="rgba(107,92,231,0.08)"),
            bgcolor="rgba(0,0,0,0)",
        ),
        showlegend=True,
        height=450,
        margin=dict(t=50, b=50, l=70, r=70),
        legend=dict(font=dict(size=11, color="#B8B3D7")),
    )
    
    st.plotly_chart(fig, use_container_width=True, key=key)

# ══════════════════════════════════════════════════════════════
# SECTOR SCREENING
# ══════════════════════════════════════════════════════════════
# ══════════════════════════════════════════════════════════════
# SECTOR → PEER MAPPING (for Smart Peer Suggestions)
# ══════════════════════════════════════════════════════════════
SECTOR_PEER_MAP = {
    "Technology": ["AAPL", "MSFT", "GOOGL", "META", "NVDA", "CRM", "ADBE", "ORCL", "IBM", "INTC", "AMD", "CSCO", "TXN", "AVGO", "NOW"],
    "Communication Services": ["GOOGL", "META", "NFLX", "DIS", "CMCSA", "T", "VZ", "TMUS", "CHTR", "EA"],
    "Consumer Cyclical": ["AMZN", "TSLA", "HD", "NKE", "MCD", "SBUX", "TGT", "LOW", "TJX", "BKNG"],
    "Consumer Defensive": ["WMT", "PG", "KO", "PEP", "COST", "PM", "CL", "MDLZ", "KHC", "GIS"],
    "Healthcare": ["JNJ", "UNH", "PFE", "ABBV", "MRK", "TMO", "ABT", "LLY", "BMY", "AMGN", "GILD", "ISRG"],
    "Financial Services": ["JPM", "BAC", "WFC", "GS", "MS", "BLK", "C", "AXP", "SCHW", "USB", "PNC", "BK"],
    "Financials": ["JPM", "BAC", "WFC", "GS", "MS", "BLK", "C", "AXP", "SCHW", "USB"],
    "Energy": ["XOM", "CVX", "COP", "SLB", "EOG", "OXY", "MPC", "VLO", "PSX", "DVN", "HES", "HAL"],
    "Industrials": ["HON", "UNP", "UPS", "CAT", "DE", "RTX", "BA", "LMT", "GE", "MMM", "WM", "ETN"],
    "Basic Materials": ["LIN", "APD", "ECL", "SHW", "DD", "NEM", "FCX", "NUE", "DOW", "PPG"],
    "Real Estate": ["AMT", "PLD", "CCI", "EQIX", "SPG", "O", "WELL", "DLR", "PSA", "AVB"],
    "Utilities": ["NEE", "DUK", "SO", "D", "AEP", "SRE", "EXC", "XEL", "ED", "WEC"],
    "Software—Infrastructure": ["MSFT", "ORCL", "CRM", "NOW", "ADBE", "INTU", "PLTR", "SNOW", "DDOG", "NET"],
    "Software—Application": ["CRM", "ADBE", "WDAY", "TEAM", "ZS", "CRWD", "OKTA", "MDB", "HUBS", "VEEV"],
    "Semiconductors": ["NVDA", "AMD", "INTC", "QCOM", "AVGO", "TXN", "MU", "AMAT", "LRCX", "KLAC"],
    "Internet Content & Information": ["GOOGL", "META", "SNAP", "PINS", "TWTR", "ZG", "IAC", "TTGT"],
    "Biotechnology": ["AMGN", "GILD", "REGN", "VRTX", "BIIB", "MRNA", "BNTX", "ILMN", "SGEN", "ALNY"],
    "Drug Manufacturers": ["JNJ", "PFE", "ABBV", "MRK", "LLY", "BMY", "AZN", "NVS", "GSK", "SNY"],
    "Banks—Diversified": ["JPM", "BAC", "WFC", "C", "USB", "PNC", "TFC", "FITB", "CFG", "KEY"],
    "Insurance": ["BRK-B", "PRU", "MET", "AFL", "AIG", "TRV", "ALL", "CB", "HIG", "PGR"],
}

POPULAR_TICKERS = {
    "Technology": ["AAPL", "MSFT", "GOOGL", "META", "NVDA", "AMZN", "CRM", "ADBE", "INTC", "AMD"],
    "Healthcare": ["JNJ", "UNH", "PFE", "ABBV", "MRK", "TMO", "ABT", "LLY", "BMY", "AMGN"],
    "Financials": ["JPM", "BAC", "WFC", "GS", "MS", "BLK", "C", "AXP", "SCHW", "USB"],
    "Consumer": ["WMT", "PG", "KO", "PEP", "COST", "NKE", "MCD", "SBUX", "TGT", "HD"],
    "Energy": ["XOM", "CVX", "COP", "SLB", "EOG", "OXY", "MPC", "VLO", "PSX", "DVN"],
    "Canadian": ["RY.TO", "TD.TO", "BNS.TO", "BMO.TO", "CM.TO", "ENB.TO", "CNQ.TO", "SU.TO", "TRP.TO", "BCE.TO"],
    "VMS/Software": ["DDOG", "SNOW", "PLTR", "NET", "CRWD", "ZS", "OKTA", "MDB", "TEAM", "ESTC"],
}

# ── Quick Ticker Lookup (for sidebar previews) ───────────────
@st.cache_data(ttl=300, show_spinner=False)
def _quick_ticker_lookup(ticker: str) -> dict:
    """Lightweight ticker lookup for sidebar preview cards."""
    if not ticker or len(ticker) < 1:
        return {}
    try:
        tk = yf.Ticker(ticker)
        info = tk.info or {}
        name = info.get("shortName") or info.get("longName") or ticker
        price = info.get("currentPrice") or info.get("regularMarketPrice")
        currency = info.get("currency", "USD")
        market_cap = info.get("marketCap")
        change_pct = info.get("regularMarketChangePercent")
        w52_high = info.get("fiftyTwoWeekHigh")
        w52_low = info.get("fiftyTwoWeekLow")
        return {
            "name": name,
            "price": price,
            "currency": currency,
            "market_cap": market_cap,
            "change_pct": change_pct,
            "52w_high": w52_high,
            "52w_low": w52_low,
            "valid": True,
        }
    except Exception:
        return {"valid": False}

# ── Page Config ──────────────────────────────────────────────
st.set_page_config(
    page_title="Orbital — M&A Intelligence",
    page_icon="https://img.icons8.com/fluency/48/combo-chart.png",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ── Generate star box-shadow strings (deterministic seed) ──────
random.seed(42)
def _gen_stars(count, spread=2000):
    return ", ".join(f"{random.randint(0,spread)}px {random.randint(0,spread)}px #FFF" for _ in range(count))
_STARS1 = _gen_stars(80)
_STARS2 = _gen_stars(50)
_STARS3 = _gen_stars(30)

# ── Chart visual helpers ──────────────────────────────────
_CHART_LAYOUT_BASE = dict(
    paper_bgcolor="rgba(0,0,0,0)",
    plot_bgcolor="rgba(0,0,0,0)",
    font=dict(family="Inter", size=14, color="#B8B3D7"),
    title=dict(text=""),  # Explicitly no title (prevents "undefined" rendering)
    hoverlabel=dict(
        bgcolor="rgba(11,14,26,0.98)",
        bordercolor="rgba(107,92,231,0.5)",
        font=dict(size=14, color="#fff", family="Inter"),
    ),
    hovermode="x unified",
    dragmode="zoom",  # Enable zoom by default
    modebar=dict(
        bgcolor="rgba(0,0,0,0)",
        color="#6B5CE7",
        activecolor="#9B8AFF",
    ),
)

def _apply_space_grid(fig, show_x_grid=False, show_y_grid=True):
    """Apply purple-tinted dot grid for space-coordinate look."""
    if show_y_grid:
        fig.update_yaxes(gridcolor="rgba(107,92,231,0.1)", griddash="dot")
    if show_x_grid:
        fig.update_xaxes(gridcolor="rgba(107,92,231,0.1)", griddash="dot")

def _glow_line_traces(fig, x, y, color, name, width=2.5, glow_width=8, yaxis=None):
    """Add a neon glow effect: wide transparent underlay + sharp main line."""
    # Parse hex color to rgba for glow
    glow_color = color
    if color.startswith("#") and len(color) == 7:
        r, g, b = int(color[1:3], 16), int(color[3:5], 16), int(color[5:7], 16)
        glow_color = f"rgba({r},{g},{b},0.15)"
    # Glow underlay
    fig.add_trace(go.Scatter(
        x=x, y=y, mode="lines", name=name,
        line=dict(color=glow_color, width=glow_width),
        showlegend=False, hoverinfo="skip",
        yaxis=yaxis,
    ))
    # Main line
    fig.add_trace(go.Scatter(
        x=x, y=y, mode="lines+markers", name=name,
        line=dict(color=color, width=width),
        marker=dict(size=7, line=dict(color="#fff", width=1.5)),
        yaxis=yaxis,
    ))

# ── Global animated starfield (visible on ALL pages) ──────────
st.markdown(
    '<div class="global-starfield">'
    '<div class="global-star-1">&#8203;</div>'
    '<div class="global-star-2">&#8203;</div>'
    '<div class="global-star-3">&#8203;</div>'
    '<div class="global-nebula">&#8203;</div>'
    '</div>',
    unsafe_allow_html=True,
)

# ══════════════════════════════════════════════════════════════
# COMPREHENSIVE CUSTOM CSS — Immersive space theme
# ══════════════════════════════════════════════════════════════
st.markdown(f"""
<style>
/* ── GLOBAL ──────────────────────────────────────────────── */
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800;900&display=swap');

html, body, [class*="css"] {{
    font-family: 'Inter', -apple-system, BlinkMacSystemFont, sans-serif;
}}

[data-testid="stApp"] {{
    background: linear-gradient(170deg, #020515, #0B0E1A, #151933, #1a1040) !important;
}}

.block-container {{
    padding-top: 0 !important;
    padding-bottom: 2rem;
    max-width: 1400px;
    position: relative;
    z-index: 1;
}}

/* ── GLOBAL STARFIELD (fixed behind all content) ──────── */
.global-starfield {{
    position: fixed; top: 0; left: 0; right: 0; bottom: 0;
    z-index: 0; pointer-events: none; overflow: hidden;
}}
.global-star-1 {{
    position: absolute; top: 0; left: 0; width: 1px; height: 1px;
    box-shadow: {_STARS1};
    opacity: 0.4;
    animation: starDrift1 150s linear infinite;
}}
.global-star-1::after {{
    content: ''; position: absolute; top: 2000px; left: 0;
    width: 1px; height: 1px;
    box-shadow: {_STARS1};
}}
.global-star-2 {{
    position: absolute; top: 0; left: 0; width: 1.5px; height: 1.5px;
    box-shadow: {_STARS2};
    opacity: 0.5;
    animation: starDrift2 100s linear infinite;
}}
.global-star-2::after {{
    content: ''; position: absolute; top: 2000px; left: 0;
    width: 1.5px; height: 1.5px;
    box-shadow: {_STARS2};
}}
.global-star-3 {{
    position: absolute; top: 0; left: 0; width: 2px; height: 2px;
    box-shadow: {_STARS3};
    opacity: 0.6;
    animation: starDrift3 75s linear infinite;
}}
.global-star-3::after {{
    content: ''; position: absolute; top: 2000px; left: 0;
    width: 2px; height: 2px;
    box-shadow: {_STARS3};
}}
.global-nebula {{
    position: absolute; top: 0; left: 0; right: 0; bottom: 0;
    background:
        radial-gradient(ellipse at 30% 40%, rgba(107,92,231,0.06) 0%, transparent 50%),
        radial-gradient(ellipse at 70% 60%, rgba(232,99,139,0.04) 0%, transparent 50%);
    animation: nebulaPulse 30s ease-in-out infinite;
}}

/* ── GLOBAL TEXT OVERRIDES FOR NATIVE STREAMLIT ELEMENTS ─ */
[data-testid="stAppViewContainer"] {{ color: #E0DCF5; }}
[data-testid="stAlert"] {{ background: rgba(255,255,255,0.05) !important; border: 1px solid rgba(255,255,255,0.1) !important; color: #E0DCF5 !important; }}
[data-testid="stAlert"] p {{ color: #E0DCF5 !important; }}
[data-testid="stExpanderDetails"] {{ background: rgba(255,255,255,0.02) !important; }}

/* ── ANIMATIONS (15+ keyframes) ────────────────────────── */
@keyframes ticker-scroll {{
    from {{ transform: translateX(0); }}
    to {{ transform: translateX(-50%); }}
}}
@keyframes fadeInUp {{
    from {{ opacity: 0; transform: translateY(30px) scale(0.98); }}
    to {{ opacity: 1; transform: translateY(0) scale(1); }}
}}
@keyframes fadeInScale {{
    from {{ opacity: 0; transform: scale(0.9); }}
    to {{ opacity: 1; transform: scale(1); }}
}}
@keyframes starDrift1 {{
    from {{ transform: translateY(0); }}
    to {{ transform: translateY(-2000px); }}
}}
@keyframes starDrift2 {{
    from {{ transform: translateY(0); }}
    to {{ transform: translateY(-2000px); }}
}}
@keyframes starDrift3 {{
    from {{ transform: translateY(0); }}
    to {{ transform: translateY(-2000px); }}
}}
@keyframes shootingStar {{
    0% {{ transform: translate(0, 0) rotate(-45deg); opacity: 0; }}
    5% {{ opacity: 1; }}
    40% {{ opacity: 1; }}
    100% {{ transform: translate(-600px, 600px) rotate(-45deg); opacity: 0; }}
}}
@keyframes nebulaPulse {{
    0%, 100% {{ opacity: 0.4; transform: scale(1); }}
    50% {{ opacity: 0.7; transform: scale(1.05); }}
}}
@keyframes float1 {{
    0%, 100% {{ transform: translate(0, 0); }}
    25% {{ transform: translate(15px, -20px); }}
    50% {{ transform: translate(-10px, -35px); }}
    75% {{ transform: translate(-20px, -10px); }}
}}
@keyframes float2 {{
    0%, 100% {{ transform: translate(0, 0); }}
    25% {{ transform: translate(-20px, 15px); }}
    50% {{ transform: translate(10px, 25px); }}
    75% {{ transform: translate(25px, -15px); }}
}}
@keyframes float3 {{
    0%, 100% {{ transform: translate(0, 0); }}
    33% {{ transform: translate(20px, -25px); }}
    66% {{ transform: translate(-15px, 20px); }}
}}
@keyframes float4 {{
    0%, 100% {{ transform: translate(0, 0); }}
    20% {{ transform: translate(-15px, -10px); }}
    40% {{ transform: translate(10px, -30px); }}
    60% {{ transform: translate(25px, -5px); }}
    80% {{ transform: translate(-5px, 15px); }}
}}
@keyframes titleGlow {{
    0%, 100% {{ opacity: 0.3; transform: scale(1); }}
    50% {{ opacity: 0.6; transform: scale(1.1); }}
}}
@keyframes gradientShift {{
    0% {{ background-position: 0% 50%; }}
    50% {{ background-position: 100% 50%; }}
    100% {{ background-position: 0% 50%; }}
}}
@keyframes shimmerLine {{
    0% {{ background-position: -200% 0; }}
    100% {{ background-position: 200% 0; }}
}}
@keyframes gentlePulse {{
    0%, 100% {{ opacity: 1; }}
    50% {{ opacity: 0.8; }}
}}
@keyframes glowPulse {{
    0%, 100% {{ box-shadow: 0 0 5px rgba(107,92,231,0.3); }}
    50% {{ box-shadow: 0 0 15px rgba(107,92,231,0.6); }}
}}
@keyframes twinkle {{
    0%, 100% {{ opacity: 0.3; }}
    50% {{ opacity: 1; }}
}}
@keyframes pulse-glow {{
    0%, 100% {{ opacity: 0.6; }}
    50% {{ opacity: 1; }}
}}
@keyframes shimmer {{
    0% {{ background-position: -200% 0; }}
    100% {{ background-position: 200% 0; }}
}}
@keyframes borderGlow {{
    0%, 100% {{ border-color: rgba(107,92,231,0.3); }}
    50% {{ border-color: rgba(155,138,255,0.6); }}
}}
@keyframes rocketLaunch {{
    0% {{ transform: translateY(0) scale(1); opacity: 1; }}
    60% {{ transform: translateY(-120px) scale(0.9); opacity: 0.8; }}
    100% {{ transform: translateY(-300px) scale(0.6); opacity: 0; }}
}}
@keyframes flameFlicker {{
    0%, 100% {{ transform: scaleY(1) scaleX(1); opacity: 0.9; }}
    25% {{ transform: scaleY(1.3) scaleX(0.85); opacity: 1; }}
    50% {{ transform: scaleY(0.8) scaleX(1.15); opacity: 0.85; }}
    75% {{ transform: scaleY(1.15) scaleX(0.9); opacity: 1; }}
}}
@keyframes exhaustTrail {{
    0% {{ opacity: 0.6; transform: translateY(0); }}
    100% {{ opacity: 0; transform: translateY(40px); }}
}}
@keyframes missionPulse {{
    0%, 100% {{ box-shadow: 0 0 8px rgba(107,92,231,0.2); }}
    50% {{ box-shadow: 0 0 20px rgba(107,92,231,0.5), 0 0 40px rgba(107,92,231,0.15); }}
}}
@keyframes checkPop {{
    0% {{ transform: scale(0); }}
    60% {{ transform: scale(1.25); }}
    100% {{ transform: scale(1); }}
}}
@keyframes progressGlow {{
    0% {{ background-position: -200% 0; }}
    100% {{ background-position: 200% 0; }}
}}
@keyframes spin {{
    from {{ transform: rotate(0deg); }}
    to {{ transform: rotate(360deg); }}
}}
@keyframes slideInLeft {{
    from {{ opacity: 0; transform: translateX(-20px); }}
    to {{ opacity: 1; transform: translateX(0); }}
}}
@keyframes slideInRight {{
    from {{ opacity: 0; transform: translateX(20px); }}
    to {{ opacity: 1; transform: translateX(0); }}
}}
@keyframes countUp {{
    from {{ opacity: 0; transform: translateY(8px); }}
    to {{ opacity: 1; transform: translateY(0); }}
}}
@keyframes sparklinePulse {{
    0%, 100% {{ stroke-opacity: 0.8; }}
    50% {{ stroke-opacity: 1; }}
}}
@keyframes numberGrow {{
    from {{ transform: scale(0.5); opacity: 0; }}
    to {{ transform: scale(1); opacity: 1; }}
}}
@keyframes badgePop {{
    0% {{ transform: scale(0); }}
    70% {{ transform: scale(1.1); }}
    100% {{ transform: scale(1); }}
}}
@keyframes borderShimmer {{
    0% {{ background-position: 0% 0%; }}
    100% {{ background-position: 100% 100%; }}
}}
@keyframes cardReveal {{
    from {{ opacity: 0; transform: translateY(15px) scale(0.98); }}
    to {{ opacity: 1; transform: none; }}
}}
@keyframes pulseRing {{
    0% {{ transform: scale(1); opacity: 0.6; }}
    100% {{ transform: scale(1.5); opacity: 0; }}
}}
@keyframes sb-fill {{
    from {{ max-width: 0; }}
    to {{ max-width: 100%; }}
}}
@keyframes sb-btn-pulse {{
    0%, 100% {{ box-shadow: 0 4px 20px rgba(107,92,231,0.3); }}
    50% {{ box-shadow: 0 4px 30px rgba(107,92,231,0.55); }}
}}
@keyframes orbBreath1 {{
    0%, 100% {{ filter: blur(80px) hue-rotate(0deg); }}
    50% {{ filter: blur(80px) hue-rotate(30deg); }}
}}
@keyframes orbBreath4 {{
    0%, 100% {{ filter: blur(90px) hue-rotate(0deg); }}
    50% {{ filter: blur(90px) hue-rotate(30deg); }}
}}
@keyframes bounceIn {{
    0%   {{ opacity: 0; transform: scale(0.85) translateY(30px); }}
    50%  {{ opacity: 1; transform: scale(1.03) translateY(-5px); }}
    70%  {{ transform: scale(0.98) translateY(2px); }}
    100% {{ opacity: 1; transform: scale(1) translateY(0); }}
}}
@keyframes slideUpBounce {{
    0%   {{ opacity: 0; transform: translateY(40px); }}
    60%  {{ opacity: 1; transform: translateY(-8px); }}
    80%  {{ transform: translateY(3px); }}
    100% {{ transform: translateY(0); }}
}}
@keyframes chartGlow {{
    0%, 100% {{ box-shadow: 0 2px 15px rgba(107,92,231,0.15); }}
    50%      {{ box-shadow: 0 8px 35px rgba(107,92,231,0.3); }}
}}
/* Elastic bounce for chart containers — chartscss.org inspired */
@keyframes chartBounceIn {{
    0%   {{ transform: scale(0.92) translateY(20px); opacity: 0; }}
    40%  {{ transform: scale(1.03) translateY(-4px); opacity: 1; }}
    60%  {{ transform: scaleY(0.97) scaleX(1.02); }}
    80%  {{ transform: scaleY(1.01) scaleX(0.99); }}
    100% {{ transform: scale(1); }}
}}
/* Glow pulse on chart data */
@keyframes dataGlowPulse {{
    0%, 100% {{ box-shadow: none; }}
    50%      {{ box-shadow: 0 0 4px 0 rgba(107,92,231,0.4), 0 0 20px 5px rgba(107,92,231,0.15); }}
}}
/* Scanner keyframes (profile loading) */
@keyframes scannerSweep {{
    0%, 100% {{ transform: rotate(-15deg); }}
    50%      {{ transform: rotate(15deg); }}
}}
@keyframes scannerLock {{
    0%   {{ transform: scale(1); filter: drop-shadow(0 0 12px rgba(6,182,212,0.5)); }}
    50%  {{ transform: scale(1.15); filter: drop-shadow(0 0 25px rgba(16,185,129,0.7)); }}
    100% {{ transform: scale(1); filter: drop-shadow(0 0 15px rgba(16,185,129,0.5)); }}
}}
@keyframes scannerBeamSweep {{
    0%   {{ transform: scaleX(0.3) rotate(-20deg); opacity: 0.3; }}
    50%  {{ transform: scaleX(1) rotate(0deg); opacity: 0.8; }}
    100% {{ transform: scaleX(0.3) rotate(20deg); opacity: 0.3; }}
}}
@keyframes scannerRingPulse {{
    0%   {{ transform: translate(-50%, -50%) scale(1); opacity: 0.4; }}
    100% {{ transform: translate(-50%, -50%) scale(2); opacity: 0; }}
}}
@keyframes scannerPhasePulse {{
    0%, 100% {{ box-shadow: 0 0 8px rgba(6,182,212,0.2); }}
    50%      {{ box-shadow: 0 0 20px rgba(6,182,212,0.5), 0 0 40px rgba(6,182,212,0.15); }}
}}

/* ── ORBITAL LOGO ANIMATIONS ───────────────────────────── */
@keyframes orbitRotate {{
    0%   {{ transform: rotate(0deg); }}
    100% {{ transform: rotate(360deg); }}
}}
@keyframes orbitRotateReverse {{
    0%   {{ transform: rotate(360deg); }}
    100% {{ transform: rotate(0deg); }}
}}
@keyframes orbitPulse {{
    0%, 100% {{ opacity: 0.4; transform: scale(1); }}
    50%      {{ opacity: 1; transform: scale(1.1); }}
}}
@keyframes particleGlow {{
    0%, 100% {{ box-shadow: 0 0 4px currentColor, 0 0 8px currentColor; }}
    50%      {{ box-shadow: 0 0 10px currentColor, 0 0 20px currentColor, 0 0 30px currentColor; }}
}}
@keyframes coreGlow {{
    0%, 100% {{ box-shadow: 0 0 15px rgba(107,92,231,0.6), 0 0 30px rgba(107,92,231,0.3); }}
    50%      {{ box-shadow: 0 0 25px rgba(107,92,231,0.9), 0 0 50px rgba(107,92,231,0.5), 0 0 80px rgba(107,92,231,0.2); }}
}}
@keyframes ringFlash {{
    0%, 90%, 100% {{ opacity: 0.3; }}
    95%           {{ opacity: 1; }}
}}

/* ── ORBITAL LOGO COMPONENT ────────────────────────────── */
.orbital-logo {{
    position: relative;
    width: 140px; height: 140px;
    display: inline-flex;
    align-items: center;
    justify-content: center;
    flex-shrink: 0;
}}
.orbital-text {{
    position: absolute;
    font-size: 1.4rem;
    font-weight: 900;
    letter-spacing: 3px;
    color: #fff;
    z-index: 5;
    text-shadow: 0 0 25px rgba(107,92,231,1), 0 0 50px rgba(107,92,231,0.6), 0 0 80px rgba(107,92,231,0.3);
    font-family: 'Inter', 'Arial Black', sans-serif;
}}
.orbital-ring {{
    position: absolute;
    border: 2.5px solid transparent;
    border-radius: 50%;
}}
.orbital-ring-1 {{
    width: 70px; height: 70px;
    border-top-color: #6B5CE7;
    border-right-color: rgba(107,92,231,0.4);
    border-bottom-color: rgba(107,92,231,0.1);
    animation: orbitRotate 3s linear infinite;
}}
.orbital-ring-2 {{
    width: 100px; height: 100px;
    border-top-color: #E8638B;
    border-right-color: rgba(232,99,139,0.3);
    animation: orbitRotateReverse 5s linear infinite;
}}
.orbital-ring-3 {{
    width: 130px; height: 130px;
    border-top-color: #9B8AFF;
    border-left-color: rgba(155,138,255,0.2);
    animation: orbitRotate 7s linear infinite, ringFlash 4s ease-in-out infinite;
}}
.orbital-particle {{
    position: absolute;
    width: 10px; height: 10px;
    border-radius: 50%;
}}
.orbital-particle-1 {{
    background: #6B5CE7;
    color: #6B5CE7;
    top: 5px; left: 50%;
    transform: translateX(-50%);
    animation: particleGlow 1.5s ease-in-out infinite;
}}
.orbital-particle-2 {{
    background: #E8638B;
    color: #E8638B;
    bottom: 14px; right: 14px;
    animation: particleGlow 1.5s ease-in-out infinite 0.5s;
}}
.orbital-particle-3 {{
    background: #10B981;
    color: #10B981;
    bottom: 14px; left: 14px;
    animation: particleGlow 1.5s ease-in-out infinite 1s;
}}

/* Small orbital logo for sidebar */
.orbital-logo-sm {{
    width: 70px; height: 70px;
}}
.orbital-logo-sm .orbital-text {{
    font-size: 0.6rem;
    letter-spacing: 1.5px;
    font-weight: 900;
}}
.orbital-logo-sm .orbital-ring-1 {{ width: 34px; height: 34px; border-width: 2px; }}
.orbital-logo-sm .orbital-ring-2 {{ width: 48px; height: 48px; border-width: 2px; }}
.orbital-logo-sm .orbital-ring-3 {{ width: 64px; height: 64px; border-width: 2px; }}
.orbital-logo-sm .orbital-particle {{ width: 5px; height: 5px; }}
.orbital-logo-sm .orbital-particle-1 {{ top: 3px; }}
.orbital-logo-sm .orbital-particle-2 {{ bottom: 6px; right: 6px; }}
.orbital-logo-sm .orbital-particle-3 {{ bottom: 6px; left: 6px; }}

/* Large orbital logo for splash */
.orbital-logo-lg {{
    width: 200px; height: 200px;
}}
.orbital-logo-lg .orbital-text {{
    font-size: 1.8rem;
    letter-spacing: 4px;
    font-weight: 900;
    text-shadow: 0 0 30px rgba(107,92,231,1), 0 0 60px rgba(107,92,231,0.7), 0 0 100px rgba(107,92,231,0.4);
}}
.orbital-logo-lg .orbital-ring-1 {{ width: 100px; height: 100px; border-width: 3px; }}
.orbital-logo-lg .orbital-ring-2 {{ width: 145px; height: 145px; border-width: 3px; }}
.orbital-logo-lg .orbital-ring-3 {{ width: 190px; height: 190px; border-width: 3px; }}
.orbital-logo-lg .orbital-particle {{ width: 12px; height: 12px; }}
.orbital-logo-lg .orbital-particle-1 {{ top: 5px; }}
.orbital-logo-lg .orbital-particle-2 {{ bottom: 18px; right: 18px; }}
.orbital-logo-lg .orbital-particle-3 {{ bottom: 18px; left: 18px; }}

/* Orbital brand container */
.orbital-brand {{
    display: flex;
    flex-direction: column;
    align-items: center;
    gap: 0.5rem;
}}
.orbital-tagline {{
    font-size: 1rem;
    color: #A8A3C7;
    font-weight: 500;
    margin-top: -0.3rem;
    letter-spacing: 0.5px;
}}

/* ── Deal Terms & Consideration animations ─────────────── */
@keyframes dealCardPulse {{
    0%, 100% {{ box-shadow: 0 0 15px rgba(107,92,231,0.2), inset 0 0 20px rgba(107,92,231,0.05); }}
    50%      {{ box-shadow: 0 0 30px rgba(107,92,231,0.4), inset 0 0 40px rgba(107,92,231,0.1); }}
}}
@keyframes dealIconSpin {{
    0%   {{ transform: rotate(0deg) scale(1); }}
    50%  {{ transform: rotate(10deg) scale(1.1); }}
    100% {{ transform: rotate(0deg) scale(1); }}
}}
@keyframes dealRowSlide {{
    from {{ opacity: 0; transform: translateX(-20px); }}
    to   {{ opacity: 1; transform: translateX(0); }}
}}
@keyframes barFillLeft {{
    from {{ width: 0; }}
    to   {{ width: var(--fill-pct); }}
}}
@keyframes barFillRight {{
    from {{ width: 0; }}
    to   {{ width: var(--fill-pct); }}
}}
@keyframes pfRowReveal {{
    from {{ opacity: 0; transform: translateY(10px); }}
    to   {{ opacity: 1; transform: translateY(0); }}
}}
@keyframes suBarGrow {{
    from {{ transform: scaleX(0); }}
    to   {{ transform: scaleX(1); }}
}}
@keyframes valueCountUp {{
    from {{ opacity: 0; transform: scale(0.8); }}
    to   {{ opacity: 1; transform: scale(1); }}
}}

/* ── DEAL TERMS CONSIDERATION CARD ─────────────────────── */
.deal-consideration-card {{
    background: linear-gradient(145deg, rgba(107,92,231,0.08), rgba(232,99,139,0.04));
    border: 1px solid rgba(107,92,231,0.25);
    border-radius: 20px;
    padding: 1.5rem;
    position: relative;
    overflow: hidden;
    animation: bounceIn 0.6s cubic-bezier(0.34, 1.56, 0.64, 1) both,
               dealCardPulse 3s ease-in-out 0.5s infinite;
}}
.deal-consideration-card::before {{
    content: '';
    position: absolute; top: -50%; left: -50%; width: 200%; height: 200%;
    background: radial-gradient(circle at 30% 30%, rgba(107,92,231,0.06) 0%, transparent 50%),
                radial-gradient(circle at 70% 70%, rgba(232,99,139,0.04) 0%, transparent 50%);
    animation: nebulaPulse 15s ease-in-out infinite;
    pointer-events: none;
}}
.deal-consideration-card .deal-header {{
    font-size: 0.7rem; font-weight: 700; color: #9B8AFF;
    text-transform: uppercase; letter-spacing: 1.5px;
    margin-bottom: 1rem;
    display: flex; align-items: center; gap: 0.5rem;
}}
.deal-consideration-card .deal-header-icon {{
    font-size: 1rem;
    animation: dealIconSpin 3s ease-in-out infinite;
}}
.deal-consideration-row {{
    display: flex; justify-content: space-between; align-items: center;
    padding: 0.8rem 1rem;
    margin: 0.4rem 0;
    background: rgba(255,255,255,0.03);
    border-radius: 12px;
    border-left: 3px solid;
    animation: dealRowSlide 0.5s ease-out both;
    transition: all 0.25s ease;
}}
.deal-consideration-row:hover {{
    background: rgba(107,92,231,0.08);
    transform: translateX(5px);
}}
.deal-consideration-row.cash {{ border-left-color: #10B981; }}
.deal-consideration-row.stock {{ border-left-color: #6B5CE7; }}
.deal-consideration-row.offer {{ border-left-color: #E8638B; }}
.deal-consideration-row .deal-label {{
    font-size: 0.8rem; color: #8A85AD; font-weight: 600;
    display: flex; align-items: center; gap: 0.4rem;
}}
.deal-consideration-row .deal-label .emoji {{ font-size: 1.1rem; }}
.deal-consideration-row .deal-value {{
    font-size: 1rem; font-weight: 700; color: #E0DCF5;
    animation: valueCountUp 0.6s ease-out both;
}}
.deal-consideration-row .deal-sub {{
    font-size: 0.7rem; color: #8A85AD; margin-top: 2px;
}}

/* ── PRO FORMA FINANCIALS TABLE ────────────────────────── */
.pf-table-wrapper {{
    background: linear-gradient(145deg, rgba(107,92,231,0.06), rgba(16,185,129,0.02));
    border: 1px solid rgba(107,92,231,0.2);
    border-radius: 20px;
    padding: 1.5rem;
    overflow: hidden;
    animation: bounceIn 0.6s cubic-bezier(0.34, 1.56, 0.64, 1) both;
}}
.pf-table {{
    width: 100%; border-collapse: separate; border-spacing: 0;
}}
.pf-table th {{
    background: rgba(107,92,231,0.12);
    color: #9B8AFF; font-size: 0.75rem;
    text-transform: uppercase; letter-spacing: 1px;
    padding: 0.8rem 1rem; font-weight: 700;
    border-bottom: 2px solid rgba(107,92,231,0.25);
}}
.pf-table th:first-child {{ border-radius: 12px 0 0 0; }}
.pf-table th:last-child {{ border-radius: 0 12px 0 0; background: linear-gradient(135deg, rgba(16,185,129,0.2), rgba(107,92,231,0.12)); }}
.pf-table td {{
    padding: 0.7rem 1rem; font-size: 0.85rem; color: #C8C3E3;
    border-bottom: 1px solid rgba(255,255,255,0.05);
    animation: pfRowReveal 0.4s ease-out both;
}}
.pf-table tr:nth-child(1) td {{ animation-delay: 0.1s; }}
.pf-table tr:nth-child(2) td {{ animation-delay: 0.15s; }}
.pf-table tr:nth-child(3) td {{ animation-delay: 0.2s; }}
.pf-table tr:nth-child(4) td {{ animation-delay: 0.25s; }}
.pf-table tr:nth-child(5) td {{ animation-delay: 0.3s; }}
.pf-table td:first-child {{
    font-weight: 700; color: #B8B3D7;
    border-left: 3px solid rgba(107,92,231,0.3);
    background: rgba(107,92,231,0.03);
}}
.pf-table td:last-child {{
    font-weight: 700; color: #10B981;
    background: linear-gradient(90deg, transparent, rgba(16,185,129,0.08));
}}
.pf-table tr:hover td {{
    background: rgba(107,92,231,0.06);
}}
.pf-table tr:last-child td {{ border-bottom: none; }}
.pf-table tr:last-child td:first-child {{ border-radius: 0 0 0 12px; }}
.pf-table tr:last-child td:last-child {{ border-radius: 0 0 12px 0; }}
.pf-adj {{ color: #F5A623 !important; font-style: italic; }}

/* ── SOURCES & USES VISUAL ─────────────────────────────── */
.su-container {{
    display: grid; grid-template-columns: 1fr 1fr; gap: 2rem;
    animation: bounceIn 0.6s cubic-bezier(0.34, 1.56, 0.64, 1) both;
}}
.su-panel {{
    background: linear-gradient(145deg, rgba(107,92,231,0.05), rgba(0,0,0,0.2));
    border: 1px solid rgba(107,92,231,0.2);
    border-radius: 20px;
    padding: 1.5rem;
    position: relative;
    overflow: hidden;
}}
.su-panel::before {{
    content: '';
    position: absolute; top: 0; left: 0; right: 0; height: 3px;
    border-radius: 20px 20px 0 0;
}}
.su-panel.sources::before {{ background: linear-gradient(90deg, #10B981, #6B5CE7); }}
.su-panel.uses::before {{ background: linear-gradient(90deg, #E8638B, #F5A623); }}
.su-panel-header {{
    font-size: 0.85rem; font-weight: 700; text-transform: uppercase;
    letter-spacing: 1.5px; margin-bottom: 1rem;
    display: flex; align-items: center; gap: 0.6rem;
}}
.su-panel.sources .su-panel-header {{ color: #10B981; }}
.su-panel.uses .su-panel-header {{ color: #E8638B; }}
.su-panel-header .su-icon {{ font-size: 1.2rem; }}
.su-row {{
    margin: 0.6rem 0;
    animation: dealRowSlide 0.4s ease-out both;
}}
.su-row:nth-child(2) {{ animation-delay: 0.1s; }}
.su-row:nth-child(3) {{ animation-delay: 0.15s; }}
.su-row:nth-child(4) {{ animation-delay: 0.2s; }}
.su-row:nth-child(5) {{ animation-delay: 0.25s; }}
.su-row-header {{
    display: flex; justify-content: space-between; align-items: center;
    margin-bottom: 0.3rem;
}}
.su-row-label {{ font-size: 0.8rem; color: #B8B3D7; }}
.su-row-value {{ font-size: 0.9rem; font-weight: 700; color: #E0DCF5; }}
.su-bar {{
    height: 8px; border-radius: 4px;
    background: rgba(255,255,255,0.08);
    overflow: hidden;
}}
.su-bar-fill {{
    height: 100%; border-radius: 4px;
    transform-origin: left;
    animation: suBarGrow 0.8s ease-out both;
}}
.su-panel.sources .su-bar-fill {{ background: linear-gradient(90deg, #10B981, #6B5CE7); }}
.su-panel.uses .su-bar-fill {{ background: linear-gradient(90deg, #E8638B, #F5A623); }}
.su-row.total {{
    margin-top: 1rem; padding-top: 1rem;
    border-top: 2px solid rgba(107,92,231,0.2);
}}
.su-row.total .su-row-label {{ font-weight: 700; color: #E0DCF5; }}
.su-row.total .su-row-value {{ font-size: 1.1rem; }}
.su-row.total .su-bar {{ height: 12px; }}
.su-row.total .su-bar-fill {{
    background: linear-gradient(90deg, #6B5CE7, #9B8AFF);
    box-shadow: 0 0 15px rgba(107,92,231,0.5);
}}

/* ── SIDEBAR ─────────────────────────────────────────────── */
section[data-testid="stSidebar"] {{
    background: linear-gradient(180deg, #0B0E1A 0%, #10132A 50%, #151933 100%);
    border-right: 1px solid rgba(107,92,231,0.2);
    box-shadow: 4px 0 30px rgba(107,92,231,0.08);
    min-width: 340px !important;
}}
section[data-testid="stSidebar"] > div:first-child {{
    padding: 1rem 1.5rem !important;
}}
section[data-testid="stSidebar"] * {{
    color: #C8C3E3 !important;
}}
/* Hide default radio labels */
section[data-testid="stSidebar"] .stRadio > label {{
    display: none !important;
}}
section[data-testid="stSidebar"] .stRadio > div {{
    flex-direction: column !important;
    gap: 4px !important;
    background: rgba(107,92,231,0.06);
    border-radius: 14px;
    padding: 6px;
    border: 1px solid rgba(107,92,231,0.15);
}}
section[data-testid="stSidebar"] .stRadio > div > label {{
    margin: 0 !important;
    padding: 0.55rem 0.8rem !important;
    border-radius: 10px !important;
    font-weight: 600 !important;
    font-size: 0.78rem !important;
    text-align: left !important;
    transition: all 0.2s ease !important;
    cursor: pointer !important;
    background: transparent !important;
    border: 1px solid transparent !important;
}}
section[data-testid="stSidebar"] .stRadio > div > label:hover {{
    background: rgba(107,92,231,0.08) !important;
    border-color: rgba(107,92,231,0.2) !important;
}}
section[data-testid="stSidebar"] .stRadio > div > label[data-checked="true"] {{
    background: linear-gradient(135deg, #6B5CE7 0%, #9B8AFF 100%) !important;
    box-shadow: 0 4px 15px rgba(107,92,231,0.4) !important;
    border-color: transparent !important;
}}
section[data-testid="stSidebar"] .stRadio > div > label[data-checked="true"] span,
section[data-testid="stSidebar"] .stRadio > div > label[data-checked="true"] p {{
    color: #fff !important;
}}
section[data-testid="stSidebar"] .stTextInput > div > div > input {{
    background: rgba(107,92,231,0.08);
    border: 1px solid rgba(107,92,231,0.3);
    border-radius: 12px;
    color: #fff !important;
    font-weight: 700;
    font-size: 1.2rem;
    letter-spacing: 3px;
    text-align: center;
    padding: 0.9rem;
    text-transform: uppercase;
}}
section[data-testid="stSidebar"] .stTextInput > div > div > input:focus {{
    border-color: #6B5CE7;
    box-shadow: 0 0 20px rgba(107,92,231,0.4);
}}
section[data-testid="stSidebar"] .stTextInput > div > div > input::placeholder {{
    color: #6B5CE7 !important;
    opacity: 0.5;
    letter-spacing: 1px;
    font-size: 0.85rem;
}}
section[data-testid="stSidebar"] .stButton > button {{
    background: linear-gradient(135deg, #6B5CE7 0%, #9B8AFF 100%) !important;
    color: #fff !important;
    font-weight: 700 !important;
    border: none !important;
    border-radius: 14px !important;
    padding: 0.9rem 2rem !important;
    font-size: 1rem !important;
    letter-spacing: 0.5px;
    box-shadow: 0 4px 20px rgba(107,92,231,0.3);
    animation: sb-btn-pulse 2s ease-in-out infinite;
    margin-top: 0.5rem !important;
}}
section[data-testid="stSidebar"] .stButton > button:hover {{
    transform: translateY(-2px);
    box-shadow: 0 8px 30px rgba(107,92,231,0.5);
}}
/* ── SIDEBAR SELECTBOX ──────────────────────────────── */
section[data-testid="stSidebar"] .stSelectbox > div > div {{
    background: rgba(107,92,231,0.08) !important;
    border: 1px solid rgba(107,92,231,0.25) !important;
    border-radius: 10px !important;
    color: #E0DCF5 !important;
    font-size: 0.82rem !important;
}}
section[data-testid="stSidebar"] .stSelectbox > div > div:hover {{
    border-color: rgba(107,92,231,0.5) !important;
}}
section[data-testid="stSidebar"] .stSelectbox > label {{
    color: #8A85AD !important;
    font-size: 0.72rem !important;
    font-weight: 600 !important;
    text-transform: uppercase !important;
    letter-spacing: 0.8px !important;
}}
/* ── SIDEBAR SLIDER ──────────────────────────────── */
section[data-testid="stSidebar"] .stSlider > div > div > div {{
    color: #6B5CE7 !important;
}}
section[data-testid="stSidebar"] .stSlider > label {{
    color: #8A85AD !important;
    font-size: 0.72rem !important;
    font-weight: 600 !important;
}}
/* ── SIDEBAR NUMBER INPUT ──────────────────────────── */
section[data-testid="stSidebar"] .stNumberInput > div > div > input {{
    background: rgba(107,92,231,0.08) !important;
    border: 1px solid rgba(107,92,231,0.25) !important;
    border-radius: 10px !important;
    color: #E0DCF5 !important;
}}
section[data-testid="stSidebar"] .stNumberInput > label {{
    color: #8A85AD !important;
    font-size: 0.72rem !important;
    font-weight: 600 !important;
}}
/* ── SIDEBAR CHECKBOX ──────────────────────────────── */
section[data-testid="stSidebar"] .stCheckbox > label {{
    color: #B8B3D7 !important;
    font-size: 0.78rem !important;
}}
/* ── SIDEBAR EXPANDER ──────────────────────────────── */
section[data-testid="stSidebar"] .streamlit-expanderHeader {{
    background: rgba(107,92,231,0.06) !important;
    border-radius: 10px !important;
    color: #B8B3D7 !important;
    font-size: 0.82rem !important;
    font-weight: 600 !important;
}}
section[data-testid="stSidebar"] hr {{
    border-color: rgba(107,92,231,0.2) !important;
}}
/* Company preview card */
.sb-company-card {{
    background: linear-gradient(135deg, rgba(107,92,231,0.12), rgba(232,99,139,0.05));
    border: 1px solid rgba(107,92,231,0.25);
    border-radius: 16px;
    padding: 0.9rem 1rem;
    margin: 0.6rem 0;
    display: flex;
    align-items: center;
    gap: 0.9rem;
    animation: cardReveal 0.5s ease-out;
    transition: all 0.3s ease;
}}
.sb-company-card:hover {{
    border-color: rgba(107,92,231,0.5);
    box-shadow: 0 4px 20px rgba(107,92,231,0.2);
    transform: translateY(-2px);
}}
.sb-logo-fallback {{
    width: 44px;
    height: 44px;
    border-radius: 12px;
    background: linear-gradient(135deg, #6B5CE7, #9B8AFF);
    display: flex;
    align-items: center;
    justify-content: center;
    font-size: 1.4rem;
    font-weight: 800;
    color: #fff !important;
    text-transform: uppercase;
    flex-shrink: 0;
    box-shadow: 0 2px 8px rgba(107,92,231,0.3);
}}
.sb-company-info {{
    flex: 1;
    min-width: 0;
}}
.sb-company-name {{
    font-size: 0.9rem;
    font-weight: 700;
    color: #fff !important;
    white-space: nowrap;
    overflow: hidden;
    text-overflow: ellipsis;
    margin: 0;
    line-height: 1.3;
}}
.sb-company-ticker {{
    font-size: 0.7rem;
    color: #9B8AFF !important;
    font-weight: 600;
    letter-spacing: 1px;
}}
.sb-company-price {{
    text-align: right;
    flex-shrink: 0;
}}
.sb-company-price-value {{
    font-size: 1rem;
    font-weight: 800;
    color: #fff !important;
}}
.sb-company-price-change {{
    font-size: 0.7rem;
    font-weight: 600;
}}
.sb-company-price-change.up {{ color: #10B981 !important; }}
.sb-company-price-change.down {{ color: #EF4444 !important; }}
.sb-company-invalid {{
    background: rgba(239,68,68,0.1);
    border: 1px solid rgba(239,68,68,0.3);
    border-radius: 12px;
    padding: 0.6rem 0.9rem;
    margin: 0.5rem 0;
    font-size: 0.75rem;
    color: #EF4444 !important;
    text-align: center;
}}
/* Role label styling */
.sb-role-label {{
    font-size: 0.65rem;
    font-weight: 700;
    text-transform: uppercase;
    letter-spacing: 2px;
    color: #6B5CE7 !important;
    margin-bottom: 0.3rem;
    display: block;
}}
.sb-role-label.acquirer {{ color: #9B8AFF !important; }}
.sb-role-label.target {{ color: #E8638B !important; }}
}}

/* ── SIDEBAR SECTIONS (merger mode) ─────────────────────── */
.sb-section {{
    background: linear-gradient(135deg, rgba(107,92,231,0.1), rgba(232,99,139,0.03));
    border-left: 3px solid #6B5CE7;
    border-radius: 0 8px 8px 0;
    padding: 0.45rem 0.75rem;
    margin: 0.9rem 0 0.4rem 0;
    font-size: 0.65rem;
    font-weight: 700;
    text-transform: uppercase;
    letter-spacing: 1.8px;
    color: #A8A3C7 !important;
    animation: slideInLeft 0.4s ease-out both;
}}
.sb-section-icon {{
    color: #9B8AFF !important;
    margin-right: 0.3rem;
    font-size: 0.55rem;
}}
section[data-testid="stSidebar"] .stSlider [data-baseweb="slider"] [role="slider"] {{
    background: #9B8AFF !important;
    border-color: #6B5CE7 !important;
    box-shadow: 0 0 8px rgba(107,92,231,0.4);
    width: 14px !important; height: 14px !important;
}}
section[data-testid="stSidebar"] .stSlider label p {{
    font-size: 0.72rem !important;
    color: #8A85AD !important;
}}
.sb-split-bar {{
    display: flex;
    height: 8px;
    border-radius: 4px;
    overflow: hidden;
    margin: 0.5rem 0 0.3rem 0;
    background: rgba(255,255,255,0.05);
}}
.sb-split-cash {{
    background: linear-gradient(90deg, #6B5CE7, #9B8AFF);
    border-radius: 4px 0 0 4px;
    transition: width 0.4s ease;
    overflow: hidden;
    animation: sb-fill 0.6s ease-out;
}}
.sb-split-stock {{
    background: linear-gradient(90deg, #E8638B, #F5A4BD);
    border-radius: 0 4px 4px 0;
    transition: width 0.4s ease;
    overflow: hidden;
    animation: sb-fill 0.6s ease-out;
}}
.sb-split-labels {{
    display: flex;
    justify-content: space-between;
    font-size: 0.65rem;
    font-weight: 600;
    margin-top: 0.15rem;
}}
.sb-split-labels .cash-label {{ color: #9B8AFF !important; }}
.sb-split-labels .stock-label {{ color: #E8638B !important; }}
.sb-divider {{
    height: 1px;
    border: none;
    margin: 0.6rem 0;
    background: linear-gradient(90deg, transparent, rgba(107,92,231,0.3), transparent);
}}

/* ── HERO / HEADER (profile view) ──────────────────────── */
.hero-header {{
    background: linear-gradient(135deg, #050816 0%, #0B0E1A 40%, #151933 100%);
    border-radius: 20px;
    padding: 2rem 2.5rem;
    margin-bottom: 1.5rem;
    border-bottom: 3px solid rgba(107,92,231,0.5);
    box-shadow: 0 8px 40px rgba(11,14,26,0.4);
    position: relative;
    overflow: hidden;
}}
.hero-header::before {{
    content: '';
    position: absolute;
    top: 0; left: 0; right: 0; bottom: 0;
    background: transparent;
    box-shadow: {_gen_stars(50, 800)};
    width: 1px; height: 1px;
    opacity: 0.5;
    pointer-events: none;
}}
.hero-header::after {{
    content: '';
    position: absolute;
    top: 0; left: 0; right: 0; bottom: 0;
    background: radial-gradient(ellipse at 20% 50%, rgba(107,92,231,0.1) 0%, transparent 60%),
                radial-gradient(ellipse at 80% 30%, rgba(232,99,139,0.06) 0%, transparent 50%);
    pointer-events: none;
}}
.hero-title {{
    font-size: 2.2rem;
    font-weight: 800;
    color: #ffffff;
    margin: 0;
    letter-spacing: -0.5px;
    position: relative; z-index: 1;
}}
.hero-accent {{ color: #9B8AFF; }}
.hero-sub {{
    font-size: 1rem;
    color: #A8A3C7;
    margin-top: 0.3rem;
    font-weight: 400;
    position: relative; z-index: 1;
}}
.hero-tagline {{
    display: inline-block;
    background: rgba(107,92,231,0.15);
    color: #9B8AFF;
    padding: 0.3rem 0.8rem;
    border-radius: 20px;
    font-size: 0.75rem;
    font-weight: 600;
    letter-spacing: 1px;
    text-transform: uppercase;
    margin-top: 0.5rem;
    position: relative; z-index: 1;
}}

/* ── COMPANY HEADER CARD ─────────────────────────────────── */
.company-card {{
    background: linear-gradient(135deg, #050816 0%, #0B0E1A 50%, #151933 100%);
    border-radius: 20px;
    padding: 1.8rem 2rem;
    margin-bottom: 1.2rem;
    border-left: 4px solid;
    border-image: linear-gradient(180deg, #6B5CE7, #E8638B) 1;
    box-shadow: 0 4px 30px rgba(11,14,26,0.3);
    position: relative;
    overflow: hidden;
    animation: cardReveal 0.6s ease-out both;
}}
.company-card::before {{
    content: '';
    position: absolute;
    top: 0; left: 0; right: 0; bottom: 0;
    background: radial-gradient(ellipse at 80% 20%, rgba(107,92,231,0.08) 0%, transparent 60%);
    pointer-events: none;
}}
.company-card::after {{
    content: '';
    position: absolute;
    width: 80px; height: 80px; border-radius: 50%;
    background: rgba(107,92,231,0.06);
    filter: blur(40px);
    top: -20px; right: 40px;
    animation: float1 20s ease-in-out infinite;
    pointer-events: none;
}}
.company-name {{
    font-size: 1.8rem;
    font-weight: 800;
    color: #ffffff;
    margin: 0;
    letter-spacing: -0.3px;
}}
.company-meta {{
    font-size: 0.85rem;
    color: #A8A3C7;
    margin-top: 0.25rem;
}}
.company-meta span {{ color: #9B8AFF; font-weight: 600; }}
.price-tag {{ font-size: 1.5rem; font-weight: 700; margin: 0; }}
.price-up {{ color: #10B981; }}
.price-down {{ color: #EF4444; }}
.price-change {{
    font-size: 0.85rem; font-weight: 600;
    padding: 0.15rem 0.5rem; border-radius: 6px;
    display: inline-block; margin-left: 0.5rem;
}}
.change-up {{ background: rgba(16,185,129,0.15); color: #10B981; }}
.change-down {{ background: rgba(239,68,68,0.15); color: #EF4444; }}

/* ── SECTION STYLING ─────────────────────────────────────── */
.section-header {{
    display: flex; align-items: center; gap: 0.8rem;
    margin: 2.5rem 0 1rem 0; padding-bottom: 0.6rem;
    position: relative;
    animation: slideUpBounce 0.5s cubic-bezier(0.34, 1.56, 0.64, 1) both;
}}
.section-header::after {{
    content: '';
    position: absolute; bottom: 0; left: 0; right: 0; height: 2px;
    background: linear-gradient(90deg, #6B5CE7, #E8638B, transparent);
    animation: glowPulse 3s ease-in-out infinite;
    border-radius: 2px;
}}
.section-header h3 {{
    font-size: 1.3rem; font-weight: 800; color: #E0DCF5; margin: 0; letter-spacing: -0.3px;
}}
.section-header .accent-bar {{
    width: 5px; height: 26px; background: linear-gradient(180deg, #6B5CE7, #E8638B); border-radius: 3px;
}}

/* ── GRADIENT DIVIDER ────────────────────────────────────── */
.gradient-divider {{
    height: 1px; border: none; margin: 1.5rem 0;
    background: linear-gradient(90deg, transparent, rgba(107,92,231,0.3), rgba(232,99,139,0.2), transparent);
}}

/* ── METRIC CARDS ────────────────────────────────────────── */
div[data-testid="stMetric"] {{
    background: rgba(255,255,255,0.04);
    backdrop-filter: blur(12px); -webkit-backdrop-filter: blur(12px);
    border: 1px solid rgba(107,92,231,0.15);
    border-radius: 16px;
    padding: 1rem 1.2rem;
    box-shadow: 0 4px 15px rgba(107,92,231,0.1);
    position: relative;
    overflow: hidden;
    animation: slideUpBounce 0.6s cubic-bezier(0.34, 1.56, 0.64, 1) both;
}}
div[data-testid="stMetric"]::before {{
    content: '';
    position: absolute; top: 0; left: 0; right: 0; height: 3px;
    background: linear-gradient(90deg, #6B5CE7, #9B8AFF, #E8638B);
    opacity: 0; transition: opacity 0.3s ease;
}}
div[data-testid="stMetric"]:hover {{
    border-color: rgba(107,92,231,0.4);
    box-shadow: 0 10px 30px rgba(107,92,231,0.25);
    transform: translateY(-5px);
}}
div[data-testid="stMetric"]:hover::before {{
    opacity: 1;
}}
div[data-testid="stMetric"] label {{
    font-size: 0.75rem !important; font-weight: 600 !important;
    text-transform: uppercase; letter-spacing: 0.8px; color: #8A85AD !important;
}}
div[data-testid="stMetric"] div[data-testid="stMetricValue"] {{
    font-size: 1.25rem !important; font-weight: 700 !important; color: #E0DCF5 !important;
}}

/* ── TABS ────────────────────────────────────────────────── */
.stTabs [data-baseweb="tab-list"] {{
    gap: 0; background: rgba(255,255,255,0.05); border-radius: 12px; padding: 4px;
}}
.stTabs [data-baseweb="tab"] {{
    border-radius: 10px; font-weight: 600; font-size: 0.82rem;
    padding: 0.5rem 1.2rem; color: #8A85AD;
}}
.stTabs [data-baseweb="tab"][aria-selected="true"] {{
    background: linear-gradient(135deg, #6B5CE7, #9B8AFF);
    color: #ffffff;
    box-shadow: 0 2px 12px rgba(107,92,231,0.4);
}}
.stTabs [data-baseweb="tab-highlight"] {{ display: none; }}
.stTabs [data-baseweb="tab-border"] {{ display: none; }}

/* ── EXPANDERS ───────────────────────────────────────────── */
.streamlit-expanderHeader {{
    font-weight: 600 !important; font-size: 0.95rem !important;
    color: #E0DCF5 !important; background: rgba(255,255,255,0.05);
    border: 1px solid rgba(255,255,255,0.1); border-radius: 12px;
}}

/* ── DATAFRAMES ──────────────────────────────────────────── */
.stDataFrame {{
    border: 1px solid rgba(255,255,255,0.1); border-radius: 12px; overflow: hidden;
}}

/* ── DOWNLOAD BUTTON ─────────────────────────────────────── */
.stDownloadButton > button {{
    background: linear-gradient(135deg, #6B5CE7, #E8638B, #F5A623) !important;
    background-size: 200% 200% !important;
    animation: gradientShift 6s ease infinite !important;
    color: white !important; font-weight: 700 !important;
    border: none !important; border-radius: 14px !important;
    padding: 0.8rem 2rem !important; font-size: 1rem !important;
    width: 100% !important; transition: all 0.3s ease;
    box-shadow: 0 4px 25px rgba(107,92,231,0.3);
}}
.stDownloadButton > button:hover {{
    transform: translateY(-2px);
    box-shadow: 0 8px 35px rgba(107,92,231,0.5);
}}

/* ── NEWS CARDS ──────────────────────────────────────────── */
.news-item {{
    padding: 0.65rem 0; border-bottom: 1px solid rgba(255,255,255,0.1);
    transition: background 0.15s;
}}
.news-item:hover {{ background: rgba(255,255,255,0.03); }}
.news-title {{
    font-weight: 600; color: #E0DCF5; font-size: 0.88rem; text-decoration: none;
}}
.news-title:hover {{ color: #9B8AFF; }}
.news-pub {{ font-size: 0.72rem; color: #8A85AD; font-weight: 500; }}

/* ── PILLS ──────────────────────────────────────────────── */
.pill {{
    display: inline-block; padding: 0.2rem 0.7rem; border-radius: 20px;
    font-size: 0.72rem; font-weight: 600; letter-spacing: 0.5px;
}}
.pill-purple {{ background: rgba(107,92,231,0.12); color: #6B5CE7; }}
.pill-dark {{ background: rgba(26,29,46,0.08); color: #1A1D2E; }}
.pill-green {{ background: rgba(16,185,129,0.12); color: #10B981; }}

/* ── PLOTLY CHARTS ──────────────────────────────────────── */
.stPlotlyChart {{
    border: 1px solid rgba(107,92,231,0.25);
    border-radius: 20px;
    overflow: hidden;
    box-shadow: 0 8px 32px rgba(107,92,231,0.18);
    backdrop-filter: blur(12px); -webkit-backdrop-filter: blur(12px);
    background: rgba(107,92,231,0.04);
    animation: chartBounceIn 0.8s cubic-bezier(0.34, 1.56, 0.64, 1) both;
    transition: all 0.35s cubic-bezier(0.34, 1.56, 0.64, 1);
    filter: saturate(0.9);
    padding: 0.5rem;
}}
.stPlotlyChart:hover {{
    border-color: rgba(107,92,231,0.6);
    box-shadow: 0 16px 48px rgba(107,92,231,0.35), 0 0 80px rgba(107,92,231,0.1);
    transform: translateY(-4px) scale(1.008);
    filter: saturate(1.15);
}}
/* Ensure chart modebar is visible and styled */
.stPlotlyChart .modebar {{
    top: 10px !important;
    right: 10px !important;
}}
.stPlotlyChart .modebar-btn {{
    font-size: 16px !important;
}}

/* ── RADIO BUTTONS ──────────────────────────────────────── */
.stRadio > div {{ gap: 0.3rem; }}
.stRadio > div > label {{
    background: rgba(255,255,255,0.05); border-radius: 8px; padding: 0.3rem 1rem;
    font-weight: 600; font-size: 0.8rem; border: 1px solid rgba(255,255,255,0.1); color: #B8B3D7;
}}
.stRadio > div > label[data-checked="true"] {{
    background: linear-gradient(135deg, #6B5CE7, #9B8AFF); color: #ffffff;
}}

/* ── SCROLLBAR ──────────────────────────────────────────── */
::-webkit-scrollbar {{ width: 6px; height: 6px; }}
::-webkit-scrollbar-track {{ background: rgba(255,255,255,0.03); border-radius: 10px; }}
::-webkit-scrollbar-thumb {{ background: rgba(107,92,231,0.4); border-radius: 10px; }}
::-webkit-scrollbar-thumb:hover {{ background: #9B8AFF; }}

/* ── SPINNER ────────────────────────────────────────────── */
.stSpinner > div > div {{ border-top-color: #6B5CE7 !important; }}

/* ── HIDE BRANDING ──────────────────────────────────────── */
#MainMenu {{ visibility: hidden; }}
footer {{ visibility: hidden; }}
header {{ visibility: hidden; }}

/* ── LOADING SKELETON ──────────────────────────────────── */
@keyframes shimmer {{
    0% {{ background-position: -200% 0; }}
    100% {{ background-position: 200% 0; }}
}}
.skeleton {{
    background: linear-gradient(90deg, rgba(107,92,231,0.05) 25%, rgba(107,92,231,0.12) 50%, rgba(107,92,231,0.05) 75%);
    background-size: 200% 100%;
    animation: shimmer 1.5s ease-in-out infinite;
    border-radius: 8px;
}}
.skeleton-text {{ height: 14px; margin-bottom: 8px; width: 80%; }}
.skeleton-card {{ height: 100px; margin-bottom: 12px; }}
.skeleton-chart {{ height: 200px; margin-bottom: 12px; }}

/* ── CARD HOVER EFFECTS ──────────────────────────────── */
.hover-lift {{
    transition: transform 0.2s ease, box-shadow 0.2s ease;
}}
.hover-lift:hover {{
    transform: translateY(-2px);
    box-shadow: 0 8px 24px rgba(107,92,231,0.15);
}}

/* ── CUSTOM FOOTER ──────────────────────────────────── */
.orbital-footer {{
    margin-top: 3rem;
    padding: 2rem 1rem;
    text-align: center;
    border-top: 1px solid rgba(107,92,231,0.15);
}}
.orbital-footer-brand {{
    font-size: 1.1rem;
    font-weight: 800;
    background: linear-gradient(135deg, #6B5CE7 0%, #E8638B 100%);
    -webkit-background-clip: text;
    -webkit-text-fill-color: transparent;
    letter-spacing: 2px;
}}
.orbital-footer-links {{
    margin-top: 0.5rem;
    display: flex;
    justify-content: center;
    gap: 1.5rem;
}}
.orbital-footer-links a {{
    color: #8A85AD;
    text-decoration: none;
    font-size: 0.7rem;
    font-weight: 600;
    letter-spacing: 0.5px;
    transition: color 0.2s;
}}
.orbital-footer-links a:hover {{
    color: #6B5CE7;
}}
.orbital-footer-version {{
    font-size: 0.6rem;
    color: #5A567A;
    margin-top: 0.5rem;
}}

/* ── PRINT STYLES ──────────────────────────────────── */
@media print {{
    .stSidebar, .stToolbar, .orbital-footer, header, #MainMenu {{ display: none !important; }}
    .main .block-container {{ padding: 0 !important; max-width: 100% !important; }}
    * {{ color: #333 !important; background: white !important; }}
}}

/* ── PRICE DISPLAY BAR ──────────────────────────────────── */
.price-bar {{
    border-radius: 14px; padding: 1rem 1.5rem; margin-bottom: 1rem;
    display: flex; gap: 1.5rem; align-items: center;
    backdrop-filter: blur(12px); -webkit-backdrop-filter: blur(12px);
}}

/* ── MERGER CHART WRAPPER ──────────────────────────────── */
.merger-chart-wrapper {{
    background: linear-gradient(135deg, rgba(107,92,231,0.06), rgba(232,99,139,0.03));
    border: 1px solid rgba(107,92,231,0.18);
    border-radius: 24px; padding: 2rem; margin: 1rem 0 1.5rem 0;
    animation: bounceIn 0.7s cubic-bezier(0.34, 1.56, 0.64, 1) both,
               chartGlow 4s ease-in-out 1s infinite;
}}

/* ── PRECEDENT & INSIDER TABLES ────────────────────────── */
.precedent-table, .insider-table {{
    width: 100%; border-collapse: separate; border-spacing: 0;
    border-radius: 14px; overflow: hidden;
    animation: bounceIn 0.7s cubic-bezier(0.34, 1.56, 0.64, 1) both;
}}
.precedent-table th, .insider-table th {{
    background: rgba(107,92,231,0.15); color: #9B8AFF; font-size: 0.7rem;
    text-transform: uppercase; letter-spacing: 1px;
    padding: 0.7rem 0.8rem; font-weight: 700;
    border-bottom: 2px solid rgba(107,92,231,0.3);
}}
.precedent-table td, .insider-table td {{
    padding: 0.55rem 0.8rem; font-size: 0.8rem; color: #C8C3E3;
    border-bottom: 1px solid rgba(255,255,255,0.05);
}}
.precedent-table tr:hover td, .insider-table tr:hover td {{
    background: rgba(107,92,231,0.08);
}}

/* ── NEWS SENTIMENT CARDS ──────────────────────────────── */
.news-card {{
    background: rgba(255,255,255,0.03);
    border: 1px solid rgba(255,255,255,0.08);
    border-radius: 12px; padding: 0.8rem 1rem; margin-bottom: 0.6rem;
    animation: slideUpBounce 0.5s cubic-bezier(0.34, 1.56, 0.64, 1) both;
    transition: all 0.25s ease;
}}
.news-card:hover {{
    border-color: rgba(107,92,231,0.3);
    transform: translateY(-2px);
    box-shadow: 0 6px 20px rgba(107,92,231,0.15);
}}
.news-sentiment-bullish {{ border-left: 3px solid #10B981; }}
.news-sentiment-bearish {{ border-left: 3px solid #EF4444; }}
.news-sentiment-neutral {{ border-left: 3px solid #8A85AD; }}

/* ── EARNINGS SURPRISE CHART CARD ─────────────────────── */
.earnings-beat {{ color: #10B981; font-weight: 700; }}
.earnings-miss {{ color: #EF4444; font-weight: 700; }}

/* ── PROFILE CHART WRAPPER ───────────────────────────── */
.profile-chart-wrapper {{
    background: linear-gradient(135deg, rgba(107,92,231,0.05), rgba(6,182,212,0.03));
    border: 1px solid rgba(107,92,231,0.15);
    border-radius: 24px; padding: 2rem; margin: 1rem 0 1.5rem 0;
    position: relative; overflow: hidden;
    animation: chartBounceIn 0.8s cubic-bezier(0.34, 1.56, 0.64, 1) both,
               chartGlow 4s ease-in-out 1s infinite;
}}
.profile-chart-wrapper::before {{
    content: '';
    position: absolute; top: -50%; left: -50%; width: 200%; height: 200%;
    background: radial-gradient(circle at 30% 40%, rgba(107,92,231,0.03) 0%, transparent 50%),
                radial-gradient(circle at 70% 60%, rgba(6,182,212,0.02) 0%, transparent 50%);
    pointer-events: none;
    animation: nebulaPulse 20s ease-in-out infinite;
}}

/* ── SCANNER LOADING (profile mode) ──────────────────── */
.scanner-control {{
    background: linear-gradient(170deg, #020515 0%, #0B0E1A 40%, #151933 100%);
    border-radius: 24px;
    padding: 2.5rem;
    min-height: 360px;
    position: relative;
    overflow: hidden;
    animation: fadeInScale 0.5s ease-out both;
    border: 1px solid rgba(6,182,212,0.2);
}}
.scanner-control::before {{
    content: '';
    position: absolute; top: 0; left: 0; right: 0; bottom: 0;
    background:
        radial-gradient(ellipse at 25% 30%, rgba(6,182,212,0.08) 0%, transparent 55%),
        radial-gradient(ellipse at 75% 70%, rgba(59,130,246,0.05) 0%, transparent 55%);
    pointer-events: none;
}}
.scanner-control::after {{
    content: '';
    position: absolute; top: 0; left: 0; right: 0; bottom: 0;
    background: transparent;
    box-shadow: {_gen_stars(40, 600)};
    width: 1px; height: 1px;
    opacity: 0.4;
    pointer-events: none;
}}
.scanner-dish-container {{
    text-align: center;
    height: 100px;
    position: relative;
    z-index: 1;
    display: flex;
    flex-direction: column;
    align-items: center;
    justify-content: center;
}}
.scanner-dish {{
    font-size: 3.5rem;
    filter: drop-shadow(0 0 12px rgba(6,182,212,0.5));
    position: relative;
    z-index: 2;
}}
.scanner-dish-scanning {{
    animation: scannerSweep 2s ease-in-out infinite;
}}
.scanner-dish-locked {{
    animation: scannerLock 1.5s ease-in-out infinite;
}}
.scanner-beam {{
    width: 60px; height: 3px;
    background: linear-gradient(90deg, transparent, rgba(6,182,212,0.6), transparent);
    border-radius: 2px;
    margin: -6px auto 0 auto;
    animation: scannerBeamSweep 1.5s ease-in-out infinite;
}}
.scanner-ring {{
    position: absolute;
    top: 50%; left: 50%;
    width: 40px; height: 40px;
    border-radius: 50%;
    border: 1px solid rgba(6,182,212,0.3);
    transform: translate(-50%, -50%);
    animation: scannerRingPulse 2s ease-out infinite;
}}
.scanner-ring-2 {{
    animation-delay: 0.7s;
}}
.scanner-ring-3 {{
    animation-delay: 1.4s;
}}
/* Cyan accent overrides for scanner */
.scanner-control .phase-indicator-active {{
    border-color: rgba(6,182,212,0.5);
    color: #06B6D4;
}}
.scanner-control .phase-indicator-active::after {{
    border-top-color: #06B6D4;
}}
.scanner-control .mission-phase-active {{
    background: rgba(6,182,212,0.1);
    border-color: rgba(6,182,212,0.25);
    animation: scannerPhasePulse 2s ease-in-out infinite;
}}
.scanner-control .mission-progress-fill {{
    background: linear-gradient(90deg, #06B6D4, #3B82F6, #6B5CE7, #06B6D4);
    background-size: 200% 100%;
}}
.scanner-control .mission-progress-fill::after {{
    box-shadow: 0 0 10px rgba(6,182,212,0.8), 0 0 20px rgba(6,182,212,0.4);
}}
.scanner-ticker {{
    text-align: center;
    margin-top: 1.2rem;
    padding-top: 1rem;
    border-top: 1px solid rgba(255,255,255,0.06);
    position: relative;
    z-index: 1;
}}
.scanner-ticker span {{
    font-size: 1rem;
    font-weight: 800;
    color: #06B6D4;
    letter-spacing: 3px;
    text-shadow: 0 0 15px rgba(6,182,212,0.4);
}}
.scanner-dots {{
    display: inline-block;
    margin-left: 4px;
}}
.scanner-dots span {{
    display: inline-block;
    width: 4px; height: 4px;
    border-radius: 50%;
    background: #06B6D4;
    margin: 0 2px;
    animation: gentlePulse 1.5s ease-in-out infinite;
}}
.scanner-dots span:nth-child(2) {{ animation-delay: 0.3s; }}
.scanner-dots span:nth-child(3) {{ animation-delay: 0.6s; }}
</style>
""", unsafe_allow_html=True)

# ── Space-specific CSS (starfield, nebula, orbs, glass cards) ──
st.markdown(f"""
<style>
/* ── SPLASH HERO ────────────────────────────────────────── */
.splash-hero {{
    background: transparent;
    border-radius: 0; padding: 5rem 3rem 4rem; text-align: center;
    margin: -1rem calc(-50vw + 50%); width: 100vw;
    position: relative; overflow: hidden;
    min-height: 90vh;
}}

/* Star Layer 1 — small distant stars */
.star-layer-1 {{
    position: absolute; top: 0; left: 0; width: 1px; height: 1px;
    box-shadow: {_STARS1};
    opacity: 0.6;
    animation: starDrift1 150s linear infinite;
}}
.star-layer-1::after {{
    content: ''; position: absolute; top: 2000px; left: 0;
    width: 1px; height: 1px;
    box-shadow: {_STARS1};
}}

/* Star Layer 2 — medium stars */
.star-layer-2 {{
    position: absolute; top: 0; left: 0; width: 1.5px; height: 1.5px;
    box-shadow: {_STARS2};
    opacity: 0.8;
    animation: starDrift2 100s linear infinite;
}}
.star-layer-2::after {{
    content: ''; position: absolute; top: 2000px; left: 0;
    width: 1.5px; height: 1.5px;
    box-shadow: {_STARS2};
}}

/* Star Layer 3 — large close stars */
.star-layer-3 {{
    position: absolute; top: 0; left: 0; width: 2px; height: 2px;
    box-shadow: {_STARS3};
    opacity: 1.0;
    animation: starDrift3 75s linear infinite;
}}
.star-layer-3::after {{
    content: ''; position: absolute; top: 2000px; left: 0;
    width: 2px; height: 2px;
    box-shadow: {_STARS3};
}}

/* Nebula overlay */
.nebula-overlay {{
    position: absolute; top: 0; left: 0; right: 0; bottom: 0;
    background:
        radial-gradient(ellipse at 20% 50%, rgba(107,92,231,0.15) 0%, transparent 50%),
        radial-gradient(ellipse at 75% 20%, rgba(232,99,139,0.1) 0%, transparent 45%),
        radial-gradient(ellipse at 50% 80%, rgba(59,130,246,0.08) 0%, transparent 50%),
        radial-gradient(ellipse at 90% 70%, rgba(45,195,195,0.06) 0%, transparent 40%);
    animation: nebulaPulse 30s ease-in-out infinite;
    pointer-events: none;
}}

/* Floating luminous orbs */
.orb {{
    position: absolute;
    border-radius: 50%;
    pointer-events: none;
}}
.orb-1 {{
    width: 200px; height: 200px;
    background: rgba(107,92,231,0.12);
    filter: blur(80px);
    top: 10%; left: 5%;
    animation: float1 20s ease-in-out infinite, orbBreath1 10s ease-in-out infinite;
}}
.orb-2 {{
    width: 160px; height: 160px;
    background: rgba(232,99,139,0.1);
    filter: blur(70px);
    top: 60%; right: 10%;
    animation: float2 22s ease-in-out infinite;
}}
.orb-3 {{
    width: 120px; height: 120px;
    background: rgba(59,130,246,0.08);
    filter: blur(60px);
    top: 30%; right: 25%;
    animation: float3 18s ease-in-out infinite;
}}
.orb-4 {{
    width: 180px; height: 180px;
    background: rgba(155,138,255,0.08);
    filter: blur(90px);
    bottom: 10%; left: 30%;
    animation: float4 25s ease-in-out infinite, orbBreath4 12s ease-in-out 3s infinite;
}}
.orb-5 {{
    width: 100px; height: 100px;
    background: rgba(45,195,195,0.06);
    filter: blur(60px);
    top: 15%; right: 5%;
    animation: float2 19s ease-in-out infinite reverse;
}}

/* Shooting stars */
.shooting-star {{
    position: absolute;
    width: 120px; height: 1.5px;
    background: linear-gradient(90deg, rgba(255,255,255,0.8), transparent);
    border-radius: 50%;
    pointer-events: none;
    opacity: 0;
}}
.shooting-star-1 {{
    top: 15%; right: -120px;
    animation: shootingStar 8s ease-in-out 2s infinite;
}}
.shooting-star-2 {{
    top: 40%; right: -120px;
    animation: shootingStar 10s ease-in-out 5s infinite;
}}
.shooting-star-3 {{
    top: 25%; right: -120px;
    animation: shootingStar 12s ease-in-out 8s infinite;
}}
.shooting-star-4 {{
    top: 55%; right: -120px;
    animation: shootingStar 15s ease-in-out 11s infinite;
    transform: rotate(-8deg);
}}
.shooting-star-5 {{
    top: 8%; right: -120px;
    animation: shootingStar 20s ease-in-out 16s infinite;
    transform: rotate(5deg);
}}

/* Noise/grain overlay */
.noise-overlay {{
    position: absolute; top: 0; left: 0; right: 0; bottom: 0;
    opacity: 0.04;
    background-image: url("data:image/svg+xml,%3Csvg viewBox='0 0 256 256' xmlns='http://www.w3.org/2000/svg'%3E%3Cfilter id='n'%3E%3CfeTurbulence type='fractalNoise' baseFrequency='0.9' numOctaves='4' stitchTiles='stitch'/%3E%3C/filter%3E%3Crect width='100%25' height='100%25' filter='url(%23n)'/%3E%3C/svg%3E");
    pointer-events: none;
}}

/* Title glow halo */
.title-glow {{
    position: absolute;
    width: 400px; height: 200px;
    top: 50%; left: 50%;
    transform: translate(-50%, -70%);
    background: radial-gradient(ellipse, rgba(107,92,231,0.2) 0%, transparent 70%);
    animation: titleGlow 4s ease-in-out infinite;
    pointer-events: none;
}}

/* Content layer */
.splash-content {{
    position: relative; z-index: 10;
}}

.splash-title {{
    font-size: 4.5rem; font-weight: 900; color: #ffffff; margin: 0;
    letter-spacing: -2px; animation: fadeInUp 0.6s ease-out;
    text-shadow: 0 0 60px rgba(107,92,231,0.3);
}}
.splash-accent {{
    background: linear-gradient(135deg, #9B8AFF, #E8638B, #F5A623, #9B8AFF);
    background-size: 200% auto;
    -webkit-background-clip: text; -webkit-text-fill-color: transparent;
    background-clip: text;
    animation: shimmer 3s linear infinite;
}}
.splash-subtitle {{
    font-size: 1.2rem; color: #B8B3D7; margin-top: 0.8rem;
    font-weight: 300; animation: fadeInUp 0.8s ease-out;
    letter-spacing: 0.5px;
}}
.splash-stats {{
    display: flex; justify-content: center; gap: 3rem; margin-top: 2.5rem;
    animation: fadeInUp 1s ease-out;
}}
.splash-stat-value {{
    font-size: 1.8rem; font-weight: 800; color: #fff;
    animation: gentlePulse 3s ease-in-out infinite;
    position: relative;
}}
.splash-stat {{
    position: relative;
}}
.splash-stat:nth-child(1) .splash-stat-value {{ animation-delay: 0s; }}
.splash-stat:nth-child(2) .splash-stat-value {{ animation-delay: 0.5s; }}
.splash-stat:nth-child(3) .splash-stat-value {{ animation-delay: 1.0s; }}
.splash-stat-icon {{
    position: relative;
    display: inline-block;
}}
.splash-stat-icon::before {{
    content: '';
    position: absolute;
    inset: -6px;
    border-radius: 50%;
    border: 2px solid rgba(107,92,231,0.4);
    animation: pulseRing 2s ease-out infinite;
    pointer-events: none;
}}
.splash-stat-label {{
    font-size: 0.7rem; color: #A8A3C7; text-transform: uppercase;
    letter-spacing: 1px; font-weight: 500;
}}
.pill-row {{
    display: flex; justify-content: center; gap: 0.7rem; margin-top: 1.8rem;
    flex-wrap: wrap;
    animation: fadeInUp 1.2s ease-out;
}}
.feature-pill {{
    border: 1px solid rgba(107,92,231,0.3); border-radius: 24px;
    padding: 0.4rem 1.1rem; font-size: 0.75rem; font-weight: 600;
    color: #B8B3D7; background: rgba(107,92,231,0.06);
    backdrop-filter: blur(10px); -webkit-backdrop-filter: blur(10px);
    transition: all 0.3s ease;
}}
.feature-pill:hover {{
    border-color: rgba(155,138,255,0.6);
    box-shadow: 0 0 15px rgba(107,92,231,0.2);
    color: #fff;
}}

/* ── SPACE SECTION (dark container for glass cards) ──── */
.space-section {{
    background: rgba(11,14,26,0.5);
    border-radius: 0;
    padding: 2.5rem 3rem;
    margin: 0 calc(-50vw + 50%); width: 100vw;
    margin-bottom: 2rem;
    position: relative;
    overflow: hidden;
}}
.space-section-title {{
    font-size: 0.75rem; font-weight: 600; color: #A8A3C7;
    text-transform: uppercase; letter-spacing: 2px;
    text-align: center; margin-bottom: 1.5rem;
}}

/* ── GLASS STEP CARDS ──────────────────────────────────── */
.step-grid {{
    display: flex; gap: 1.2rem; margin: 0 0 2rem 0;
    position: relative;
}}
.step-card {{
    flex: 1;
    background: rgba(255,255,255,0.03);
    border: 1px solid rgba(255,255,255,0.08);
    border-radius: 18px; padding: 1.5rem; text-align: center;
    transition: all 0.3s ease;
    backdrop-filter: blur(20px); -webkit-backdrop-filter: blur(20px);
    position: relative; overflow: hidden;
    animation: fadeInUp 0.6s ease-out both;
}}
.step-card:nth-child(1) {{ animation-delay: 0.1s; }}
.step-card:nth-child(2) {{ animation-delay: 0.2s; }}
.step-card:nth-child(3) {{ animation-delay: 0.3s; }}
.step-card:nth-child(4) {{ animation-delay: 0.4s; }}
.step-card::before {{
    content: '';
    position: absolute; top: 0; left: 0; right: 0; bottom: 0;
    border-radius: 18px;
    padding: 1px;
    background: linear-gradient(135deg, rgba(107,92,231,0.3), rgba(232,99,139,0.1), transparent);
    -webkit-mask: linear-gradient(#fff 0 0) content-box, linear-gradient(#fff 0 0);
    -webkit-mask-composite: xor;
    mask-composite: exclude;
    opacity: 0; transition: opacity 0.3s;
    pointer-events: none;
}}
.step-card:hover {{
    border-color: rgba(107,92,231,0.3); transform: translateY(-4px);
    box-shadow: 0 8px 30px rgba(107,92,231,0.15);
}}
.step-card:hover::before {{ opacity: 1; }}
.step-num {{
    background: linear-gradient(135deg, #6B5CE7, #9B8AFF);
    color: #fff; width: 38px; height: 38px; border-radius: 50%;
    display: inline-flex; align-items: center; justify-content: center;
    font-weight: 800; font-size: 1rem; margin-bottom: 0.6rem;
    box-shadow: 0 4px 15px rgba(107,92,231,0.3);
}}
.step-label {{ font-size: 0.88rem; font-weight: 700; color: #E0DCF5; }}
.step-detail {{ font-size: 0.72rem; color: #8A85AD; margin-top: 0.3rem; }}

/* Connector lines between steps */
.step-connector {{
    position: absolute; top: 50%; height: 2px; z-index: 0;
    background: linear-gradient(90deg, rgba(107,92,231,0.2), rgba(232,99,139,0.2), rgba(107,92,231,0.2));
    background-size: 200% 100%;
    animation: shimmerLine 3s linear infinite;
}}

/* ── GLASS FEATURE CARDS ───────────────────────────────── */
.feature-grid {{
    display: grid; grid-template-columns: repeat(4, 1fr);
    gap: 1rem; margin-top: 0.5rem;
}}
.feature-card {{
    background: rgba(255,255,255,0.03);
    border: 1px solid rgba(255,255,255,0.08);
    border-radius: 18px; padding: 1.5rem 1.2rem; text-align: center;
    transition: all 0.3s ease;
    backdrop-filter: blur(20px); -webkit-backdrop-filter: blur(20px);
    animation: fadeInScale 0.6s ease-out both;
    position: relative; overflow: hidden;
}}
.feature-card:nth-child(1) {{ animation-delay: 0.05s; }}
.feature-card:nth-child(2) {{ animation-delay: 0.1s; }}
.feature-card:nth-child(3) {{ animation-delay: 0.15s; }}
.feature-card:nth-child(4) {{ animation-delay: 0.2s; }}
.feature-card:nth-child(5) {{ animation-delay: 0.25s; }}
.feature-card:nth-child(6) {{ animation-delay: 0.3s; }}
.feature-card:nth-child(7) {{ animation-delay: 0.35s; }}
.feature-card:nth-child(8) {{ animation-delay: 0.4s; }}
.feature-card::before {{
    content: '';
    position: absolute; top: 0; left: 0; right: 0; bottom: 0;
    border-radius: 18px;
    padding: 1px;
    background: linear-gradient(135deg, rgba(107,92,231,0.3), rgba(232,99,139,0.1), transparent);
    -webkit-mask: linear-gradient(#fff 0 0) content-box, linear-gradient(#fff 0 0);
    -webkit-mask-composite: xor;
    mask-composite: exclude;
    opacity: 0; transition: opacity 0.3s;
    pointer-events: none;
}}
.feature-card:hover {{
    border-color: rgba(107,92,231,0.3); transform: translateY(-4px);
    box-shadow: 0 8px 24px rgba(107,92,231,0.15);
}}
.feature-card:hover::before {{ opacity: 1; }}
.feature-icon {{ font-size: 2.2rem; margin-bottom: 0.5rem; }}
.feature-title {{ font-size: 0.88rem; font-weight: 700; color: #E0DCF5; margin-bottom: 0.3rem; }}
.feature-desc {{ font-size: 0.72rem; color: #8A85AD; line-height: 1.6; }}

/* ── MISSION CONTROL (Merger Loading) ──────────────────── */
.mission-control {{
    background: linear-gradient(170deg, #020515 0%, #0B0E1A 40%, #151933 100%);
    border-radius: 24px;
    padding: 2.5rem;
    min-height: 420px;
    position: relative;
    overflow: hidden;
    animation: fadeInScale 0.5s ease-out both;
    border: 1px solid rgba(107,92,231,0.2);
}}
.mission-control::before {{
    content: '';
    position: absolute; top: 0; left: 0; right: 0; bottom: 0;
    background:
        radial-gradient(ellipse at 25% 30%, rgba(107,92,231,0.08) 0%, transparent 55%),
        radial-gradient(ellipse at 75% 70%, rgba(232,99,139,0.05) 0%, transparent 55%);
    pointer-events: none;
}}
.mission-control::after {{
    content: '';
    position: absolute; top: 0; left: 0; right: 0; bottom: 0;
    background: transparent;
    box-shadow: {_gen_stars(40, 600)};
    width: 1px; height: 1px;
    opacity: 0.4;
    pointer-events: none;
}}
.mission-header {{
    text-align: center;
    margin-bottom: 1.5rem;
    position: relative;
    z-index: 1;
}}
.mission-title {{
    font-size: 1.1rem;
    font-weight: 800;
    color: #E0DCF5;
    text-transform: uppercase;
    letter-spacing: 3px;
    margin: 0;
    text-shadow: 0 0 20px rgba(107,92,231,0.4);
}}
.mission-subtitle {{
    font-size: 0.72rem;
    color: #8A85AD;
    margin-top: 0.3rem;
    text-transform: uppercase;
    letter-spacing: 2px;
}}
.rocket-container {{
    text-align: center;
    height: 120px;
    position: relative;
    z-index: 1;
    display: flex;
    flex-direction: column;
    align-items: center;
    justify-content: center;
}}
.rocket {{
    font-size: 3.5rem;
    filter: drop-shadow(0 0 12px rgba(107,92,231,0.5));
    position: relative;
    z-index: 2;
}}
.rocket-idle {{
    animation: float1 6s ease-in-out infinite;
}}
.rocket-launching {{
    animation: rocketLaunch 2s ease-in forwards;
}}
.rocket-flame {{
    font-size: 1.5rem;
    filter: drop-shadow(0 0 8px rgba(255,165,0,0.7));
    animation: flameFlicker 0.3s ease-in-out infinite;
    margin-top: -8px;
}}
.exhaust-trail {{
    width: 4px;
    height: 30px;
    background: linear-gradient(to bottom, rgba(255,165,0,0.4), rgba(107,92,231,0.2), transparent);
    filter: blur(2px);
    margin: 0 auto;
    animation: exhaustTrail 0.8s ease-out infinite;
}}
.mission-progress-track {{
    height: 6px;
    background: rgba(255,255,255,0.06);
    border-radius: 3px;
    margin: 1.2rem 0;
    overflow: hidden;
    position: relative;
    z-index: 1;
}}
.mission-progress-fill {{
    height: 100%;
    border-radius: 3px;
    background: linear-gradient(90deg, #6B5CE7, #9B8AFF, #E8638B, #6B5CE7);
    background-size: 200% 100%;
    animation: progressGlow 2s linear infinite;
    transition: width 0.6s ease;
    position: relative;
}}
.mission-progress-fill::after {{
    content: '';
    position: absolute;
    right: 0; top: 50%;
    transform: translateY(-50%);
    width: 10px; height: 10px;
    border-radius: 50%;
    background: #fff;
    box-shadow: 0 0 10px rgba(155,138,255,0.8), 0 0 20px rgba(107,92,231,0.4);
}}
.mission-phases {{
    display: flex;
    flex-direction: column;
    gap: 0.6rem;
    position: relative;
    z-index: 1;
}}
.mission-phase {{
    display: flex;
    align-items: center;
    gap: 0.8rem;
    padding: 0.6rem 1rem;
    border-radius: 12px;
    transition: all 0.3s ease;
}}
.mission-phase-active {{
    background: rgba(107,92,231,0.1);
    border: 1px solid rgba(107,92,231,0.25);
    animation: missionPulse 2s ease-in-out infinite;
}}
.mission-phase-complete {{
    background: rgba(16,185,129,0.06);
    border: 1px solid rgba(16,185,129,0.15);
}}
.mission-phase-pending {{
    opacity: 0.4;
    border: 1px solid transparent;
}}
.phase-indicator {{
    width: 28px; height: 28px;
    border-radius: 50%;
    display: flex;
    align-items: center;
    justify-content: center;
    font-size: 0.7rem;
    font-weight: 700;
    flex-shrink: 0;
}}
.phase-indicator-active {{
    border: 2px solid rgba(107,92,231,0.5);
    color: #9B8AFF;
    position: relative;
}}
.phase-indicator-active::after {{
    content: '';
    position: absolute;
    inset: -4px;
    border-radius: 50%;
    border: 2px solid transparent;
    border-top-color: #9B8AFF;
    animation: spin 1s linear infinite;
}}
.phase-indicator-complete {{
    background: rgba(16,185,129,0.2);
    color: #10B981;
    animation: checkPop 0.4s ease-out both;
}}
.phase-indicator-pending {{
    border: 1px solid rgba(255,255,255,0.15);
    color: #555;
}}
.phase-label {{
    font-size: 0.82rem;
    font-weight: 600;
    color: #E0DCF5;
}}
.phase-sublabel {{
    font-size: 0.68rem;
    color: #8A85AD;
    margin-top: 0.1rem;
}}
.mission-stats {{
    display: flex;
    justify-content: center;
    align-items: center;
    gap: 0.8rem;
    margin-top: 1.2rem;
    padding-top: 1rem;
    border-top: 1px solid rgba(255,255,255,0.06);
    position: relative;
    z-index: 1;
}}
.mission-stats span {{
    font-size: 0.85rem;
    font-weight: 700;
    color: #9B8AFF;
    letter-spacing: 1px;
}}
.mission-stats .mission-x {{
    font-size: 1.2rem;
    color: #E8638B;
    font-weight: 300;
}}
</style>
""", unsafe_allow_html=True)


# ── HELPER: Mission Control loading screen ───────────────────
def _render_mission_control(acquirer: str, target: str, current_phase: int, total_phases: int = 5) -> str:
    """Return HTML for the animated mission control loading UI."""
    phases = [
        (f"Acquiring Target Intel: {acquirer}", "Scanning financial databases..."),
        (f"Locking Signal: {target}", "Establishing secure data link..."),
        ("Mapping Sector Constellation", "Triangulating peer coordinates..."),
        ("Computing Orbital Mechanics", "Pro forma trajectory analysis..."),
        ("Synthesizing Mission Report", "AI engines generating insights..."),
    ]
    completed = min(current_phase, total_phases)
    pct = int((completed / total_phases) * 100)

    # Rocket state
    if current_phase >= total_phases:
        rocket_cls = "rocket rocket-launching"
        flame_html = ""
        exhaust_html = ""
    else:
        rocket_cls = "rocket rocket-idle"
        flame_html = '<div class="rocket-flame">🔥</div>'
        exhaust_html = '<div class="exhaust-trail"></div>'

    # Build phase rows
    phase_rows = ""
    for i, (label, sublabel) in enumerate(phases):
        if i < current_phase:
            row_cls = "mission-phase mission-phase-complete"
            ind_cls = "phase-indicator phase-indicator-complete"
            ind_content = "✓"
        elif i == current_phase:
            row_cls = "mission-phase mission-phase-active"
            ind_cls = "phase-indicator phase-indicator-active"
            ind_content = str(i + 1)
        else:
            row_cls = "mission-phase mission-phase-pending"
            ind_cls = "phase-indicator phase-indicator-pending"
            ind_content = str(i + 1)
        phase_rows += (
            f'<div class="{row_cls}">'
            f'<div class="{ind_cls}">{ind_content}</div>'
            f'<div><div class="phase-label">{label}</div>'
            f'<div class="phase-sublabel">{sublabel}</div></div>'
            f'</div>'
        )

    return (
        f'<div class="mission-control">'
        f'<div class="mission-header">'
        f'<div class="mission-title">Merger Analysis Mission Control</div>'
        f'<div class="mission-subtitle">Phase {completed} of {total_phases}</div>'
        f'</div>'
        f'<div class="rocket-container">'
        f'<div class="{rocket_cls}">🚀</div>'
        f'{flame_html}'
        f'{exhaust_html}'
        f'</div>'
        f'<div class="mission-progress-track">'
        f'<div class="mission-progress-fill" style="width:{pct}%;"></div>'
        f'</div>'
        f'<div class="mission-phases">{phase_rows}</div>'
        f'<div class="mission-stats">'
        f'<span>{acquirer}</span>'
        f'<span class="mission-x">×</span>'
        f'<span>{target}</span>'
        f'</div>'
        f'</div>'
    )


# ── HELPER: Profile scanner loading screen ────────────────────
def _render_profile_scanner(ticker: str, current_phase: int, total_phases: int = 3) -> str:
    """Return HTML for the animated scanner loading UI for company profiles."""
    phases = [
        ("Scanning Financial Databases", "Fetching market data & fundamentals..."),
        ("Analyzing Financials & Peers", "Comparing against sector peers..."),
        ("Generating AI Insights", "Building investment thesis..."),
    ]
    completed = min(current_phase, total_phases)
    pct = int((completed / total_phases) * 100)

    # Dish state
    if current_phase >= total_phases:
        dish_cls = "scanner-dish scanner-dish-locked"
        beam_html = ""
        ring_html = ""
    else:
        dish_cls = "scanner-dish scanner-dish-scanning"
        beam_html = '<div class="scanner-beam"></div>'
        ring_html = (
            '<div class="scanner-ring"></div>'
            '<div class="scanner-ring scanner-ring-2"></div>'
            '<div class="scanner-ring scanner-ring-3"></div>'
        )

    # Build phase rows (reuse mission-phase classes)
    phase_rows = ""
    for i, (label, sublabel) in enumerate(phases):
        if i < current_phase:
            row_cls = "mission-phase mission-phase-complete"
            ind_cls = "phase-indicator phase-indicator-complete"
            ind_content = "\u2713"
        elif i == current_phase:
            row_cls = "mission-phase mission-phase-active"
            ind_cls = "phase-indicator phase-indicator-active"
            ind_content = str(i + 1)
        else:
            row_cls = "mission-phase mission-phase-pending"
            ind_cls = "phase-indicator phase-indicator-pending"
            ind_content = str(i + 1)
        phase_rows += (
            f'<div class="{row_cls}">'
            f'<div class="{ind_cls}">{ind_content}</div>'
            f'<div><div class="phase-label">{label}</div>'
            f'<div class="phase-sublabel">{sublabel}</div></div>'
            f'</div>'
        )

    dots = '<span class="scanner-dots"><span></span><span></span><span></span></span>'

    return (
        f'<div class="scanner-control">'
        f'<div class="mission-header">'
        f'<div class="mission-title">Company Scanner</div>'
        f'<div class="mission-subtitle">Phase {completed} of {total_phases}</div>'
        f'</div>'
        f'<div class="scanner-dish-container">'
        f'<div class="{dish_cls}">\U0001F4E1</div>'
        f'{beam_html}'
        f'{ring_html}'
        f'</div>'
        f'<div class="mission-progress-track">'
        f'<div class="mission-progress-fill" style="width:{pct}%;"></div>'
        f'</div>'
        f'<div class="mission-phases">{phase_rows}</div>'
        f'<div class="scanner-ticker">'
        f'<span>{ticker}</span>{dots}'
        f'</div>'
        f'</div>'
    )


# ── HELPER: Section header with accent bar ──────────────────
def _section(title, icon=""):
    st.markdown(
        f'<div class="section-header">'
        f'<div class="accent-bar"></div>'
        f'<h3>{icon}  {title}</h3>'
        f'</div>',
        unsafe_allow_html=True,
    )


# ── HELPER: Gradient divider between sections ────────────────
def _divider():
    st.markdown('<div class="gradient-divider"></div>', unsafe_allow_html=True)


from contextlib import contextmanager
@contextmanager
def _safe_section(name=""):
    """Wrap a section in try/except to prevent one section from crashing the whole profile."""
    try:
        yield
    except Exception as e:
        st.warning(f"⚠️ {name} section encountered an error: {str(e)[:100]}")
        import traceback
        with st.expander("Show error details", expanded=False):
            st.code(traceback.format_exc())


# ── HELPER: Peer radar chart ────────────────────────────────
def _build_peer_radar_chart(cd):
    """Build a Plotly radar chart comparing target vs peer median."""
    if not cd.peer_data:
        return

    metrics = ["P/E", "Fwd P/E", "EV/EBITDA", "Gross Margin", "Op Margin", "ROE"]

    target_vals = [
        cd.trailing_pe, cd.forward_pe, cd.ev_to_ebitda,
        (cd.gross_margins or 0) * 100, (cd.operating_margins or 0) * 100,
        (cd.return_on_equity or 0) * 100,
    ]

    peer_keys = ["trailing_pe", "forward_pe", "ev_to_ebitda",
                 "gross_margins", "operating_margins", "return_on_equity"]
    pct_keys = {"gross_margins", "operating_margins", "return_on_equity"}

    peer_medians = []
    for key in peer_keys:
        vals = [p.get(key) for p in cd.peer_data if p.get(key) is not None]
        if key in pct_keys:
            vals = [v * 100 for v in vals]
        peer_medians.append(float(np.median(vals)) if vals else 0)

    # Normalize to 0-100 scale
    norm_target, norm_peer = [], []
    for t, p in zip(target_vals, peer_medians):
        t = t if t is not None else 0
        mx = max(abs(t), abs(p), 1)
        norm_target.append(min(t / mx * 100, 120))
        norm_peer.append(min(p / mx * 100, 120))

    fig = go.Figure()
    fig.add_trace(go.Scatterpolar(
        r=norm_target + [norm_target[0]],
        theta=metrics + [metrics[0]],
        fill='toself', name=cd.ticker,
        fillcolor='rgba(107,92,231,0.15)',
        line=dict(color='#6B5CE7', width=3),
        marker=dict(size=8, line=dict(color="#fff", width=1.5)),
    ))
    fig.add_trace(go.Scatterpolar(
        r=norm_peer + [norm_peer[0]],
        theta=metrics + [metrics[0]],
        fill='toself', name='Peer Median',
        fillcolor='rgba(232,99,139,0.08)',
        line=dict(color='#E8638B', width=3),
        marker=dict(size=7, line=dict(color="#fff", width=1.5)),
    ))
    fig.update_layout(
        **_CHART_LAYOUT_BASE,
        polar=dict(
            radialaxis=dict(visible=True, range=[0, 120], tickfont=dict(size=8, color="#8A85AD"),
                            gridcolor="rgba(107,92,231,0.1)"),
            angularaxis=dict(tickfont=dict(size=10, color="#8A85AD"),
                             gridcolor="rgba(107,92,231,0.08)"),
            bgcolor="rgba(0,0,0,0)",
        ),
        showlegend=True, height=520,
        margin=dict(t=50, b=50, l=70, r=70),
        legend=dict(font=dict(size=11, color="#B8B3D7")),
    )
    st.plotly_chart(fig, use_container_width=True)


# ── CHART: Revenue & Margins ──────────────────────────────────
def _build_revenue_margin_chart(cd, key="rev_margin"):
    """Revenue bars with gross/EBITDA/net margin lines on secondary y-axis."""
    if cd.revenue is None or len(cd.revenue) == 0:
        st.info("Revenue data not available for chart.")
        return
    rev = cd.revenue.dropna().sort_index()
    years = [idx.strftime("%Y") if hasattr(idx, "strftime") else str(idx) for idx in rev.index]
    n = len(years)
    # Progressive alpha — older bars dimmer, newest brightest
    bar_alphas = [0.35 + 0.45 * (i / max(n - 1, 1)) for i in range(n)]
    bar_colors = [f"rgba(107,92,231,{a:.2f})" for a in bar_alphas]
    fig = go.Figure()
    fig.add_trace(go.Bar(
        x=years, y=rev.values, name="Revenue",
        marker=dict(color=bar_colors, line=dict(color="rgba(255,255,255,0.15)", width=1)),
        text=[format_number(v, currency_symbol=cd.currency_symbol) for v in rev.values],
        textposition="outside", textfont=dict(size=9, color="#B8B3D7"),
    ))
    for series, name, color in [
        (cd.gross_margin_series, "Gross Margin", "#10B981"),
        (cd.ebitda_margin, "EBITDA Margin", "#E8638B"),
        (cd.net_margin_series, "Net Margin", "#F5A623"),
    ]:
        if series is not None and len(series) > 0:
            s = series.dropna().sort_index()
            yrs = [idx.strftime("%Y") if hasattr(idx, "strftime") else str(idx) for idx in s.index]
            _glow_line_traces(fig, yrs, s.values, color, name, yaxis="y2")
    fig.update_layout(
        **_CHART_LAYOUT_BASE,
        height=500, margin=dict(t=40, b=40, l=60, r=60),
        xaxis=dict(tickfont=dict(size=9, color="#8A85AD"), showgrid=False),
        yaxis=dict(title=dict(text="Revenue", font=dict(size=10, color="#8A85AD")),
                   tickfont=dict(size=9, color="#8A85AD")),
        yaxis2=dict(title=dict(text="Margin %", font=dict(size=10, color="#8A85AD")),
                    overlaying="y", side="right", showgrid=False,
                    tickfont=dict(size=9, color="#8A85AD"), ticksuffix="%"),
        legend=dict(font=dict(size=10, color="#B8B3D7"), orientation="h", yanchor="bottom", y=1.02),
        barmode="group",
    )
    _apply_space_grid(fig)
    st.plotly_chart(fig, use_container_width=True, key=key)


# ── CHART: Cash Flow ──────────────────────────────────────────
def _build_cashflow_chart(cd, key="cashflow"):
    """Grouped bars: OCF, CapEx (negative), FCF, dividends."""
    series_map = [
        (cd.operating_cashflow_series, "Operating CF", "#6B5CE7"),
        (cd.capital_expenditure, "CapEx", "#EF4444"),
        (cd.free_cashflow_series, "Free CF", "#10B981"),
        (cd.dividends_paid, "Dividends", "#F5A623"),
    ]
    has_data = any(s is not None and len(s) > 0 for s, _, _ in series_map)
    if not has_data:
        st.info("Cash flow data not available for chart.")
        return
    fig = go.Figure()
    for series, name, color in series_map:
        if series is not None and len(series) > 0:
            s = series.dropna().sort_index()
            years = [idx.strftime("%Y") if hasattr(idx, "strftime") else str(idx) for idx in s.index]
            nc = len(years)
            bar_alphas = [0.4 + 0.5 * (i / max(nc - 1, 1)) for i in range(nc)]
            # Parse hex to build progressive rgba
            r, g, b = int(color[1:3], 16), int(color[3:5], 16), int(color[5:7], 16)
            bar_cols = [f"rgba({r},{g},{b},{a:.2f})" for a in bar_alphas]
            fig.add_trace(go.Bar(
                x=years, y=s.values, name=name,
                marker=dict(color=bar_cols, line=dict(color="rgba(255,255,255,0.15)", width=1)),
            ))
    fig.add_hline(y=0, line_dash="dot", line_color="rgba(255,255,255,0.15)", line_width=1)
    fig.update_layout(
        **_CHART_LAYOUT_BASE,
        height=500, margin=dict(t=40, b=40, l=60, r=60),
        xaxis=dict(tickfont=dict(size=9, color="#8A85AD"), showgrid=False),
        yaxis=dict(title=dict(text="Amount", font=dict(size=10, color="#8A85AD")),
                   tickfont=dict(size=9, color="#8A85AD")),
        legend=dict(font=dict(size=10, color="#B8B3D7"), orientation="h", yanchor="bottom", y=1.02),
        barmode="group",
    )
    _apply_space_grid(fig)
    st.plotly_chart(fig, use_container_width=True, key=key)


# ── CHART: Balance Sheet ──────────────────────────────────────
def _build_balance_sheet_chart(cd, key="balance_sheet"):
    """Stacked equity+debt bars with cash and total assets overlay lines."""
    has_data = any(
        s is not None and len(s) > 0
        for s in [cd.total_equity, cd.total_debt, cd.cash_and_equivalents, cd.total_assets]
    )
    if not has_data:
        st.info("Balance sheet data not available for chart.")
        return
    fig = go.Figure()
    # Stacked bars: equity + debt with progressive alpha
    for series, name, base_rgba in [
        (cd.total_equity, "Equity", (107, 92, 231)),
        (cd.total_debt, "Debt", (239, 68, 68)),
    ]:
        if series is not None and len(series) > 0:
            s = series.dropna().sort_index()
            years = [idx.strftime("%Y") if hasattr(idx, "strftime") else str(idx) for idx in s.index]
            nc = len(years)
            bar_alphas = [0.3 + 0.45 * (i / max(nc - 1, 1)) for i in range(nc)]
            bar_colors = [f"rgba({base_rgba[0]},{base_rgba[1]},{base_rgba[2]},{a:.2f})" for a in bar_alphas]
            fig.add_trace(go.Bar(
                x=years, y=s.values, name=name,
                marker=dict(color=bar_colors, line=dict(color="rgba(255,255,255,0.15)", width=1)),
            ))
    # Overlay lines with glow
    for series, name, color in [
        (cd.cash_and_equivalents, "Cash", "#10B981"),
        (cd.total_assets, "Total Assets", "#F5A623"),
    ]:
        if series is not None and len(series) > 0:
            s = series.dropna().sort_index()
            years = [idx.strftime("%Y") if hasattr(idx, "strftime") else str(idx) for idx in s.index]
            _glow_line_traces(fig, years, s.values, color, name)
    fig.update_layout(
        **_CHART_LAYOUT_BASE,
        height=500, margin=dict(t=40, b=40, l=60, r=60),
        xaxis=dict(tickfont=dict(size=9, color="#8A85AD"), showgrid=False),
        yaxis=dict(tickfont=dict(size=9, color="#8A85AD")),
        legend=dict(font=dict(size=10, color="#B8B3D7"), orientation="h", yanchor="bottom", y=1.02),
        barmode="stack",
    )
    _apply_space_grid(fig)
    st.plotly_chart(fig, use_container_width=True, key=key)


# ── CHART: Peer Valuation Comparison ──────────────────────────
def _build_peer_valuation_chart(cd, key="peer_val"):
    """Horizontal grouped bars: company vs peer median on key multiples."""
    if not cd.peer_data:
        st.info("Peer data not available for valuation comparison.")
        return
    metrics = [
        ("P/E", "trailing_pe", cd.trailing_pe),
        ("Fwd P/E", "forward_pe", cd.forward_pe),
        ("EV/EBITDA", "ev_to_ebitda", cd.ev_to_ebitda),
        ("P/S", "price_to_sales", cd.price_to_sales),
    ]
    labels, company_vals, peer_vals = [], [], []
    for label, key, company_val in metrics:
        if company_val is None:
            continue
        peer_raw = [p.get(key) for p in cd.peer_data if p.get(key) is not None]
        if not peer_raw:
            continue
        labels.append(label)
        company_vals.append(company_val)
        peer_vals.append(float(np.median(peer_raw)))
    if not labels:
        st.info("Insufficient data for peer valuation chart.")
        return
    fig = go.Figure()
    fig.add_trace(go.Bar(
        y=labels, x=company_vals, orientation="h", name=cd.ticker,
        marker=dict(color="#6B5CE7", line=dict(color="rgba(255,255,255,0.15)", width=1)),
        text=[f"{v:.1f}x" for v in company_vals],
        textposition="outside", textfont=dict(size=10, color="#B8B3D7"),
    ))
    fig.add_trace(go.Bar(
        y=labels, x=peer_vals, orientation="h", name="Peer Median",
        marker=dict(color="#E8638B", line=dict(color="rgba(255,255,255,0.15)", width=1)),
        text=[f"{v:.1f}x" for v in peer_vals],
        textposition="outside", textfont=dict(size=10, color="#B8B3D7"),
    ))
    # Premium/discount annotations
    for i, (cv, pv) in enumerate(zip(company_vals, peer_vals)):
        if pv != 0:
            pct = (cv - pv) / abs(pv) * 100
            color = "#10B981" if pct < 0 else "#EF4444"
            sign = "+" if pct >= 0 else ""
            fig.add_annotation(
                y=labels[i], x=max(cv, pv) * 1.15,
                text=f"{sign}{pct:.0f}%", showarrow=False,
                font=dict(size=9, color=color, family="Inter"),
            )
    fig.update_layout(
        **_CHART_LAYOUT_BASE,
        height=400, margin=dict(t=40, b=30, l=90, r=90),
        xaxis=dict(tickfont=dict(size=9, color="#8A85AD")),
        yaxis=dict(tickfont=dict(size=10, color="#8A85AD"), autorange="reversed"),
        legend=dict(font=dict(size=10, color="#B8B3D7"), orientation="h", yanchor="bottom", y=1.02),
        barmode="group",
    )
    _apply_space_grid(fig, show_x_grid=True)
    st.plotly_chart(fig, use_container_width=True, key=key)


# ── CHART: Earnings Surprise ─────────────────────────────────
def _build_earnings_surprise_chart(cd, key="earnings_surprise"):
    """Color-coded bars: green for beats, red for misses."""
    if cd.earnings_dates is None or cd.earnings_dates.empty:
        st.info("Earnings data not available for surprise chart.")
        return
    df = cd.earnings_dates.copy()
    # Try to find EPS columns
    est_col = None
    act_col = None
    for c in df.columns:
        cl = str(c).lower()
        if "estimate" in cl or "eps estimate" in cl:
            est_col = c
        if "reported" in cl or "actual" in cl or "eps actual" in cl:
            act_col = c
    if est_col is None or act_col is None:
        st.info("Earnings surprise data not available.")
        return
    df = df.dropna(subset=[est_col, act_col])
    if df.empty:
        st.info("No earnings surprise data to display.")
        return
    df = df.head(8).sort_index()
    surprises = df[act_col].astype(float) - df[est_col].astype(float)
    labels = [f"{s:+.2f}" for s in surprises]
    dates = [idx.strftime("%b %Y") if hasattr(idx, "strftime") else str(idx) for idx in df.index]
    # Intensity-proportional alpha: bigger surprise = brighter
    max_abs = max(abs(s) for s in surprises) if len(surprises) > 0 else 1
    bar_colors = []
    for s in surprises:
        intensity = 0.4 + 0.6 * (abs(s) / max(max_abs, 0.01))
        if s >= 0:
            bar_colors.append(f"rgba(16,185,129,{intensity:.2f})")
        else:
            bar_colors.append(f"rgba(239,68,68,{intensity:.2f})")

    fig = go.Figure(go.Bar(
        x=dates, y=surprises.values,
        marker=dict(color=bar_colors, line=dict(color="rgba(255,255,255,0.15)", width=1)),
        text=labels, textposition="outside",
        textfont=dict(size=10, color="#B8B3D7"),
    ))
    fig.add_hline(y=0, line_dash="dot", line_color="rgba(255,255,255,0.15)", line_width=1)
    fig.update_layout(
        **_CHART_LAYOUT_BASE,
        height=400, margin=dict(t=40, b=40, l=60, r=40),
        xaxis=dict(tickfont=dict(size=9, color="#8A85AD"), showgrid=False),
        yaxis=dict(title=dict(text="EPS Surprise", font=dict(size=10, color="#8A85AD")),
                   tickfont=dict(size=9, color="#8A85AD")),
    )
    _apply_space_grid(fig)
    st.plotly_chart(fig, use_container_width=True, key=key)


# ── CHART: Accretion/Dilution Waterfall ───────────────────────
def _build_accretion_waterfall(pro_forma, key="accretion_waterfall"):
    """Waterfall chart showing EPS bridge from standalone to pro forma."""
    steps = pro_forma.waterfall_steps
    if not steps:
        st.info("Waterfall data not available.")
        return

    labels = [s["label"] for s in steps]
    values = [s["value"] for s in steps]
    types = [s["type"] for s in steps]

    # Build Plotly waterfall measure types
    measures = []
    for t in types:
        if t == "absolute":
            measures.append("absolute")
        elif t == "total":
            measures.append("total")
        else:
            measures.append("relative")

    colors = []
    for i, (v, t) in enumerate(zip(values, types)):
        if t == "absolute":
            colors.append("#6B5CE7")
        elif t == "total":
            colors.append("#9B8AFF" if v >= values[0] else "#EF4444")
        else:
            colors.append("#10B981" if v >= 0 else "#EF4444")

    # Determine totals marker outline
    totals_color = "#9B8AFF" if values[-1] >= values[0] else "#EF4444"
    fig = go.Figure(go.Waterfall(
        x=labels, y=values, measure=measures,
        text=[f"${v:.2f}" for v in values],
        textposition="outside",
        textfont=dict(size=10, color="#B8B3D7"),
        connector=dict(line=dict(color="rgba(107,92,231,0.2)", width=1, dash="dot")),
        increasing=dict(marker=dict(color="#10B981", line=dict(color="rgba(255,255,255,0.15)", width=1))),
        decreasing=dict(marker=dict(color="#EF4444", line=dict(color="rgba(255,255,255,0.15)", width=1))),
        totals=dict(marker=dict(color=totals_color, line=dict(color="#fff", width=1.5))),
    ))

    fig.update_layout(
        **_CHART_LAYOUT_BASE,
        height=600, margin=dict(t=40, b=40, l=60, r=60),
        xaxis=dict(tickfont=dict(size=10, color="#8A85AD"), showgrid=False),
        yaxis=dict(title=dict(text="EPS ($)", font=dict(size=10, color="#8A85AD")),
                   tickfont=dict(size=9, color="#8A85AD"),
                   tickprefix="$"),
    )
    _apply_space_grid(fig)
    st.plotly_chart(fig, use_container_width=True, key=key)


# ── CHART: Football Field Valuation ──────────────────────────
def _build_football_field_chart(football_field, currency_symbol="$", key="football_field"):
    """Horizontal range bars with offer price reference line."""
    offer_price = football_field.get("_offer_price", 0)
    methods = {k: v for k, v in football_field.items() if not k.startswith("_")}

    if not methods:
        st.info("Insufficient data for football field chart.")
        return

    labels = list(methods.keys())
    lows = [methods[m]["low"] for m in labels]
    highs = [methods[m]["high"] for m in labels]

    colors = ["#6B5CE7", "#10B981", "#F5A623", "#E8638B", "#3B82F6"]

    fig = go.Figure()
    for i, label in enumerate(labels):
        fig.add_trace(go.Bar(
            y=[label], x=[highs[i] - lows[i]],
            base=[lows[i]], orientation="h",
            marker=dict(
                color=colors[i % len(colors)], opacity=0.85,
                line=dict(color="rgba(255,255,255,0.15)", width=1),
            ),
            name=label,
            text=[f"{format_number(lows[i], currency_symbol=currency_symbol)} \u2014 {format_number(highs[i], currency_symbol=currency_symbol)}"],
            textposition="inside",
            textfont=dict(size=9, color="#fff"),
            hoverinfo="text",
            showlegend=False,
        ))

    if offer_price > 0:
        # Shaded band around offer price (+-5%)
        band_lo = offer_price * 0.95
        band_hi = offer_price * 1.05
        fig.add_vrect(
            x0=band_lo, x1=band_hi,
            fillcolor="rgba(239,68,68,0.06)", line_width=0,
        )
        fig.add_vline(
            x=offer_price, line_dash="dash", line_color="#EF4444", line_width=2,
            annotation_text=f"Offer: {format_number(offer_price, currency_symbol=currency_symbol)}",
            annotation_position="top",
            annotation_font=dict(size=10, color="#EF4444"),
        )

    fig.update_layout(
        **_CHART_LAYOUT_BASE,
        height=550, margin=dict(t=50, b=40, l=130, r=70),
        xaxis=dict(tickfont=dict(size=9, color="#8A85AD")),
        yaxis=dict(tickfont=dict(size=10, color="#8A85AD"), autorange="reversed"),
        barmode="stack",
    )
    _apply_space_grid(fig, show_x_grid=True)
    st.plotly_chart(fig, use_container_width=True, key=key)


# ── CHART: Deal Structure Donut ──────────────────────────────
def _build_deal_structure_donut(assumptions, key="deal_donut"):
    """Pie chart with hole showing cash/stock split."""
    # Pull the larger slice
    pull_vals = [0.05, 0] if assumptions.pct_cash >= assumptions.pct_stock else [0, 0.05]
    fig = go.Figure(go.Pie(
        labels=["Cash", "Stock"],
        values=[assumptions.pct_cash, assumptions.pct_stock],
        hole=0.55,
        pull=pull_vals,
        marker=dict(
            colors=["#6B5CE7", "#E8638B"],
            line=dict(color="#fff", width=1.5),
        ),
        textinfo="label+percent",
        textfont=dict(size=12, color="#fff"),
        hoverinfo="label+percent+value",
    ))
    fig.update_layout(
        **_CHART_LAYOUT_BASE,
        height=450, margin=dict(t=40, b=40, l=40, r=40),
        showlegend=False,
        annotations=[dict(text="Deal<br>Mix", x=0.5, y=0.5, font_size=14,
                         font_color="#E0DCF5", showarrow=False)],
    )
    st.plotly_chart(fig, use_container_width=True, key=key)


# ── CHART: Company Comparison Bars ───────────────────────────
def _build_company_comparison_bars(acq_cd, tgt_cd, key="company_compare"):
    """Grouped horizontal bars comparing acquirer vs target on key metrics."""
    metrics = []
    acq_vals = []
    tgt_vals = []

    for label, acq_v, tgt_v in [
        ("Gross Margin %", (acq_cd.gross_margins or 0) * 100, (tgt_cd.gross_margins or 0) * 100),
        ("Op Margin %", (acq_cd.operating_margins or 0) * 100, (tgt_cd.operating_margins or 0) * 100),
        ("Net Margin %", (acq_cd.profit_margins or 0) * 100, (tgt_cd.profit_margins or 0) * 100),
        ("ROE %", (acq_cd.return_on_equity or 0) * 100, (tgt_cd.return_on_equity or 0) * 100),
    ]:
        metrics.append(label)
        acq_vals.append(acq_v)
        tgt_vals.append(tgt_v)

    fig = go.Figure()
    fig.add_trace(go.Bar(
        y=metrics, x=acq_vals, orientation="h", name=acq_cd.ticker,
        marker=dict(color="#6B5CE7", line=dict(color="rgba(255,255,255,0.15)", width=1)),
        text=[f"{v:.1f}%" for v in acq_vals],
        textposition="outside", textfont=dict(size=10, color="#B8B3D7"),
    ))
    fig.add_trace(go.Bar(
        y=metrics, x=tgt_vals, orientation="h", name=tgt_cd.ticker,
        marker=dict(color="#E8638B", line=dict(color="rgba(255,255,255,0.15)", width=1)),
        text=[f"{v:.1f}%" for v in tgt_vals],
        textposition="outside", textfont=dict(size=10, color="#B8B3D7"),
    ))
    # Star annotation on winning metric
    for i, (av, tv) in enumerate(zip(acq_vals, tgt_vals)):
        winner_x = max(av, tv) * 1.12
        fig.add_annotation(
            y=metrics[i], x=winner_x,
            text="\u2605", showarrow=False,
            font=dict(size=10, color="#F5A623"),
        )
    fig.update_layout(
        **_CHART_LAYOUT_BASE,
        height=500, margin=dict(t=40, b=30, l=110, r=80),
        xaxis=dict(tickfont=dict(size=9, color="#8A85AD"), ticksuffix="%"),
        yaxis=dict(tickfont=dict(size=10, color="#8A85AD"), autorange="reversed"),
        legend=dict(font=dict(size=10, color="#B8B3D7"), orientation="h", yanchor="bottom", y=1.02),
        barmode="group",
    )
    _apply_space_grid(fig, show_x_grid=True)
    st.plotly_chart(fig, use_container_width=True, key=key)


# ── RENDER: SWOT Grid ─────────────────────────────────────────
def _render_swot_grid(swot):
    """2x2 CSS grid with color-coded SWOT cards."""
    if not swot:
        st.info("SWOT analysis not available.")
        return
    quadrants = [
        ("Strengths", swot.get("strengths", []), "#10B981", "rgba(16,185,129,0.08)", "rgba(16,185,129,0.25)"),
        ("Weaknesses", swot.get("weaknesses", []), "#EF4444", "rgba(239,68,68,0.08)", "rgba(239,68,68,0.25)"),
        ("Opportunities", swot.get("opportunities", []), "#6B5CE7", "rgba(107,92,231,0.08)", "rgba(107,92,231,0.25)"),
        ("Threats", swot.get("threats", []), "#F5A623", "rgba(245,166,35,0.08)", "rgba(245,166,35,0.25)"),
    ]
    html = '<div style="display:grid; grid-template-columns:1fr 1fr; gap:1rem;">'
    for title, items, color, bg, border_color in quadrants:
        bullets = "".join(
            f'<div style="font-size:0.84rem; color:#B8B3D7; line-height:1.7; padding:0.15rem 0;">&bull; {item}</div>'
            for item in items
        ) if items else '<div style="font-size:0.84rem; color:#8A85AD;">No data available</div>'
        html += (
            f'<div style="background:{bg}; border:1px solid {border_color}; border-radius:14px; padding:1.2rem;">'
            f'<div style="font-size:0.85rem; font-weight:700; color:{color}; margin-bottom:0.5rem; text-transform:uppercase; letter-spacing:0.5px;">{title}</div>'
            f'{bullets}'
            f'</div>'
        )
    html += '</div>'
    st.markdown(html, unsafe_allow_html=True)


# ── RENDER: Growth Outlook ────────────────────────────────────
def _render_growth_outlook(growth, cd):
    """Rating badge + thesis sub-sections + catalyst/risk columns."""
    if not growth:
        st.info("Growth outlook not available.")
        return
    rating = growth.get("growth_rating", "MODERATE")
    rating_colors = {"STRONG": "#10B981", "MODERATE": "#F5A623", "WEAK": "#EF4444"}
    rating_color = rating_colors.get(rating, "#8A85AD")
    rating_bg = {"STRONG": "rgba(16,185,129,0.12)", "MODERATE": "rgba(245,166,35,0.12)", "WEAK": "rgba(239,68,68,0.12)"}

    st.markdown(
        f'<div style="display:inline-block; background:{rating_bg.get(rating, "rgba(138,133,173,0.12)")}; '
        f'color:{rating_color}; padding:0.4rem 1.2rem; border-radius:20px; font-weight:700; '
        f'font-size:0.9rem; letter-spacing:1px; margin-bottom:1rem;">Growth Rating: {rating}</div>',
        unsafe_allow_html=True,
    )

    for key, title in [("revenue_thesis", "Revenue Thesis"), ("margin_thesis", "Margin Thesis"), ("earnings_path", "Earnings Path")]:
        text = growth.get(key, "")
        if text:
            # Clean bullet prefix
            clean = text.strip()
            if clean.startswith("- "):
                clean = clean[2:]
            st.markdown(
                f'<div style="margin-bottom:0.8rem;">'
                f'<div style="font-size:0.8rem; font-weight:700; color:#9B8AFF; text-transform:uppercase; letter-spacing:0.5px; margin-bottom:0.2rem;">{title}</div>'
                f'<div style="font-size:0.85rem; color:#B8B3D7; line-height:1.7;">{clean}</div>'
                f'</div>',
                unsafe_allow_html=True,
            )

    # Catalysts & Risks in two columns
    cat_col, risk_col = st.columns(2)
    with cat_col:
        st.markdown('<div style="font-size:0.8rem; font-weight:700; color:#10B981; text-transform:uppercase; letter-spacing:0.5px; margin-bottom:0.3rem;">Key Catalysts</div>', unsafe_allow_html=True)
        for item in growth.get("key_catalysts", []):
            st.markdown(f'<div style="font-size:0.84rem; color:#B8B3D7; line-height:1.7; padding:0.1rem 0;">&bull; {item}</div>', unsafe_allow_html=True)
    with risk_col:
        st.markdown('<div style="font-size:0.8rem; font-weight:700; color:#EF4444; text-transform:uppercase; letter-spacing:0.5px; margin-bottom:0.3rem;">Key Risks to Growth</div>', unsafe_allow_html=True)
        for item in growth.get("key_risks_to_growth", []):
            st.markdown(f'<div style="font-size:0.84rem; color:#B8B3D7; line-height:1.7; padding:0.1rem 0;">&bull; {item}</div>', unsafe_allow_html=True)


# ── RENDER: Capital Allocation ────────────────────────────────
def _render_capital_allocation(ca, cd):
    """Letter grade badge + border-left styled sub-section cards."""
    if not ca:
        st.info("Capital allocation analysis not available.")
        return
    grade = ca.get("capital_allocation_grade", "B")
    grade_colors = {"A": "#10B981", "B": "#6B5CE7", "C": "#F5A623", "D": "#EF4444"}
    grade_color = grade_colors.get(grade, "#8A85AD")
    grade_bg = {"A": "rgba(16,185,129,0.12)", "B": "rgba(107,92,231,0.12)", "C": "rgba(245,166,35,0.12)", "D": "rgba(239,68,68,0.12)"}

    st.markdown(
        f'<div style="display:inline-block; background:{grade_bg.get(grade, "rgba(138,133,173,0.12)")}; '
        f'color:{grade_color}; padding:0.4rem 1.2rem; border-radius:20px; font-weight:700; '
        f'font-size:0.9rem; letter-spacing:1px; margin-bottom:1rem;">Capital Allocation Grade: {grade}</div>',
        unsafe_allow_html=True,
    )

    sections = [
        ("Strategy Summary", ca.get("strategy_summary", ""), "#6B5CE7"),
        ("CapEx Assessment", ca.get("capex_assessment", ""), "#E8638B"),
        ("Shareholder Returns", ca.get("shareholder_returns", ""), "#10B981"),
        ("M&A Strategy", ca.get("ma_strategy", ""), "#F5A623"),
        ("Debt Management", ca.get("debt_management", ""), "#8A85AD"),
    ]
    for title, text, color in sections:
        if text:
            clean = text.strip()
            if clean.startswith("- "):
                clean = clean[2:]
            st.markdown(
                f'<div style="border-left:3px solid {color}; padding:0.6rem 0 0.6rem 1rem; margin-bottom:0.6rem;">'
                f'<div style="font-size:0.8rem; font-weight:700; color:{color}; text-transform:uppercase; letter-spacing:0.5px; margin-bottom:0.15rem;">{title}</div>'
                f'<div style="font-size:0.85rem; color:#B8B3D7; line-height:1.7;">{clean}</div>'
                f'</div>',
                unsafe_allow_html=True,
            )


# ── Sidebar Helper: Render Company Preview Card ───────────────
def _render_company_card(ticker: str, role: str = "") -> None:
    """Render a company preview card in the sidebar."""
    if not ticker or len(ticker) < 1:
        return

    info = _quick_ticker_lookup(ticker)
    if not info or not info.get("valid"):
        st.markdown(
            f'<div class="sb-company-invalid">⚠️ Could not find: {ticker}</div>',
            unsafe_allow_html=True,
        )
        return

    name = info.get("name", ticker)
    price = info.get("price")
    currency = info.get("currency", "USD")
    change_pct = info.get("change_pct")

    # Currency symbol
    curr_sym = {"USD": "$", "EUR": "€", "GBP": "£", "JPY": "¥", "CAD": "C$"}.get(currency, "$")

    # Price display
    price_str = f"{curr_sym}{price:,.2f}" if price else "—"

    # Change display
    if change_pct is not None:
        change_class = "up" if change_pct >= 0 else "down"
        change_str = f'<div class="sb-company-price-change {change_class}">{change_pct:+.2f}%</div>'
    else:
        change_str = ""

    # Logo - show ticker initial with gradient background (reliable, no external deps)
    initial = ticker[0] if ticker else "?"
    logo_html = f'<div class="sb-logo-fallback">{initial}</div>'

    # Role label
    role_html = f'<span class="sb-role-label {role.lower()}">{role}</span>' if role else ""

    st.markdown(
        f'{role_html}'
        f'<div class="sb-company-card">'
        f'{logo_html}'
        f'<div class="sb-company-info">'
        f'<div class="sb-company-name">{name}</div>'
        f'<div class="sb-company-ticker">{ticker}</div>'
        f'</div>'
        f'<div class="sb-company-price">'
        f'<div class="sb-company-price-value">{price_str}</div>'
        f'{change_str}'
        f'</div>'
        f'</div>',
        unsafe_allow_html=True,
    )


# ── Sidebar ──────────────────────────────────────────────────
with st.sidebar:
    # Animated Orbital Logo
    st.markdown(
        '<div style="text-align:center; padding: 1rem 0 0.5rem 0;">'
        '<div class="orbital-logo orbital-logo-sm" style="margin:0 auto;">'
        '<span class="orbital-text">ORBITAL</span>'
        '<div class="orbital-ring orbital-ring-1"></div>'
        '<div class="orbital-ring orbital-ring-2"></div>'
        '<div class="orbital-ring orbital-ring-3"></div>'
        '<div class="orbital-particle orbital-particle-1"></div>'
        '<div class="orbital-particle orbital-particle-2"></div>'
        '<div class="orbital-particle orbital-particle-3"></div>'
        '</div>'
        '<div style="font-size:0.6rem; color:#8A85AD; margin-top:0.5rem; letter-spacing:1.5px; text-transform:uppercase;">M&amp;A Intelligence</div>'
        '</div>',
        unsafe_allow_html=True,
    )

    st.markdown('<div style="height:0.5rem;"></div>', unsafe_allow_html=True)

    # Mode Toggle - Enhanced with more analysis modes
    st.markdown(
        '<div style="font-size:0.65rem; font-weight:700; color:#8A85AD; text-transform:uppercase; '
        'letter-spacing:1.5px; margin-bottom:0.3rem;">Analysis Mode</div>',
        unsafe_allow_html=True,
    )
    _mode_options = ["📊 Company Profile", "📈 Comps Analysis", "💹 DCF Valuation", "⚖️ Quick Compare", "🤝 Merger Analysis", "🔍 VMS Screener"]
    _mode_selection = st.radio(
        "Mode", 
        _mode_options, 
        horizontal=False, 
        label_visibility="collapsed"
    )
    # Strip emoji prefix for internal use
    analysis_mode = _mode_selection.split(" ", 1)[1] if " " in _mode_selection else _mode_selection

    st.markdown('<div style="height:0.8rem;"></div>', unsafe_allow_html=True)
    
    # ══════════════════════════════════════════════════════
    # WATCHLIST SECTION (shown in all modes)
    # ══════════════════════════════════════════════════════
    _init_watchlist()
    watchlist = _get_watchlist()
    
    if watchlist:
        with st.expander(f"📋 Watchlist ({len(watchlist)})", expanded=False):
            _init_watchlist_notes()

            # ── Watchlist Dashboard ──
            try:
                total_mcap = 0
                best_perf = ("", -999)
                worst_perf = ("", 999)
                for _wt in watchlist:
                    _wi = st.session_state.watchlist_data.get(_wt, {})
                    _wmc = _wi.get("market_cap", 0) or 0
                    total_mcap += _wmc
                    _wch = _wi.get("change_pct", 0) or 0
                    if _wch > best_perf[1]:
                        best_perf = (_wt, _wch)
                    if _wch < worst_perf[1]:
                        worst_perf = (_wt, _wch)
                _mcap_str = f"${total_mcap/1e9:,.1f}B" if total_mcap >= 1e9 else f"${total_mcap/1e6:,.0f}M" if total_mcap >= 1e6 else "N/A"
                st.markdown(
                    f'<div style="background:rgba(107,92,231,0.06); border-radius:8px; padding:0.6rem; margin-bottom:0.5rem; font-size:0.72rem;">'
                    f'<div style="color:#8A85AD; font-weight:700; margin-bottom:0.3rem;">📊 PORTFOLIO SNAPSHOT</div>'
                    f'<div style="color:#B8B3D7;">Total Mkt Cap: <b style="color:#E0DCF5;">{_mcap_str}</b></div>'
                    f'<div style="color:#10B981;">Best: <b>{best_perf[0]}</b> ({best_perf[1]:+.2f}%)</div>'
                    f'<div style="color:#EF4444;">Worst: <b>{worst_perf[0]}</b> ({worst_perf[1]:+.2f}%)</div>'
                    f'</div>',
                    unsafe_allow_html=True,
                )
            except Exception:
                pass

            for wl_ticker in watchlist:
                wl_col1, wl_col2 = st.columns([4, 1])
                with wl_col1:
                    wl_info = st.session_state.watchlist_data.get(wl_ticker, {})
                    wl_price = wl_info.get("price", 0)
                    wl_change = wl_info.get("change_pct", 0)
                    change_color = "#10B981" if wl_change and wl_change >= 0 else "#EF4444"

                    # 52-week high/low alert
                    _52h = wl_info.get("52w_high", 0) or 0
                    _52l = wl_info.get("52w_low", 0) or 0
                    _alert_html = ""
                    if wl_price and _52l and _52l > 0:
                        _pct_from_low = (wl_price - _52l) / _52l * 100
                        _pct_from_high = (_52h - wl_price) / _52h * 100 if _52h > 0 else 100
                        if _pct_from_low <= 5:
                            _alert_html = '<span style="font-size:0.6rem; color:#10B981; margin-left:4px;">🟢 Near 52w Low</span>'
                        elif _pct_from_high <= 5:
                            _alert_html = '<span style="font-size:0.6rem; color:#F59E0B; margin-left:4px;">🟡 Near 52w High</span>'

                    st.markdown(
                        f'<div style="display:flex; justify-content:space-between; align-items:center; padding:0.3rem 0;">'
                        f'<span style="font-weight:700; color:#E0DCF5;">{wl_ticker}{_alert_html}</span>'
                        f'<span style="color:{change_color}; font-size:0.8rem;">${wl_price:,.2f}</span>'
                        f'</div>',
                        unsafe_allow_html=True,
                    )
                with wl_col2:
                    if st.button("✕", key=f"remove_{wl_ticker}", help=f"Remove {wl_ticker}"):
                        _remove_from_watchlist(wl_ticker)
                        st.rerun()
                # Inline note for this ticker
                current_note = _get_watchlist_note(wl_ticker)
                new_note = st.text_input(
                    "Note", value=current_note, key=f"note_{wl_ticker}",
                    placeholder="Add a note...", label_visibility="collapsed"
                )
                if new_note != current_note:
                    _set_watchlist_note(wl_ticker, new_note)
    
    st.markdown('<div class="sb-divider"></div>', unsafe_allow_html=True)

    # ══════════════════════════════════════════════════════
    # MODE-SPECIFIC SIDEBAR CONTENT
    # ══════════════════════════════════════════════════════
    
    # Initialize all variables with defaults
    ticker_input = ""
    generate_btn = False
    acquirer_input = ""
    target_input = ""
    merger_btn = False
    merger_assumptions = MergerAssumptions()
    comps_ticker_input = ""
    comps_btn = False
    max_peers = 10
    include_saas = False
    dcf_ticker_input = ""
    dcf_btn = False
    dcf_growth_rate = 0.05
    dcf_terminal_growth = 0.025
    dcf_discount_rate = 0.10
    dcf_years = 5
    compare_tickers = []
    compare_btn = False

    if analysis_mode == "Company Profile":
        # ── Company Profile Mode ──
        st.markdown(
            '<div class="sb-section"><span class="sb-section-icon">📊</span> COMPANY</div>',
            unsafe_allow_html=True,
        )

        ticker_input = st.text_input(
            "Stock Ticker", value="", max_chars=10,
            placeholder="Enter ticker (e.g. AAPL)",
            label_visibility="collapsed",
        ).strip().upper()

        # Show company preview card
        if ticker_input:
            _render_company_card(ticker_input)
            # Add to watchlist button
            if not _is_in_watchlist(ticker_input):
                if st.button("⭐ Add to Watchlist", key="add_wl_profile", use_container_width=True):
                    _add_to_watchlist(ticker_input)
                    st.rerun()

        st.markdown('<div style="height:0.5rem;"></div>', unsafe_allow_html=True)
        generate_btn = st.button("🚀 Generate Profile", type="primary", use_container_width=True)
        
        # Search History
        search_history = _get_search_history()
        if search_history:
            st.markdown('<div class="sb-section"><span class="sb-section-icon">🕐</span> RECENT</div>', unsafe_allow_html=True)
            recent_cols = st.columns(5)
            for i, hist_ticker in enumerate(search_history[:5]):
                with recent_cols[i % 5]:
                    if st.button(hist_ticker, key=f"hist_{hist_ticker}", use_container_width=True):
                        st.session_state["load_ticker"] = hist_ticker
                        st.rerun()
        
        # Quick sector picks
        st.markdown('<div class="sb-section"><span class="sb-section-icon">🔥</span> QUICK PICKS</div>', unsafe_allow_html=True)
        sector_choice = st.selectbox("Sector", list(POPULAR_TICKERS.keys()), label_visibility="collapsed")
        selected_quick = st.selectbox("Popular Tickers", POPULAR_TICKERS[sector_choice], label_visibility="collapsed")
        if st.button("Load Ticker", key="load_quick"):
            st.session_state["load_ticker"] = selected_quick
            st.rerun()

    elif analysis_mode == "Comps Analysis":
        # ── Comps Analysis Mode ──
        st.markdown(
            '<div class="sb-section"><span class="sb-section-icon">📊</span> TARGET COMPANY</div>',
            unsafe_allow_html=True,
        )

        comps_ticker_input = st.text_input(
            "Stock Ticker", value="", max_chars=10,
            placeholder="Enter ticker (e.g. AAPL)",
            label_visibility="collapsed",
            key="comps_ticker"
        ).strip().upper()

        if comps_ticker_input:
            _render_company_card(comps_ticker_input)

        st.markdown('<div style="height:0.5rem;"></div>', unsafe_allow_html=True)
        
        # Comps settings
        st.markdown(
            '<div class="sb-section"><span class="sb-section-icon">⚙️</span> SETTINGS</div>',
            unsafe_allow_html=True,
        )
        
        max_peers = st.slider("Number of Peers", 5, 20, 10, 1)
        include_saas = st.checkbox("Include SaaS/Software peers", value=False)

        # ── Smart Peer Suggestions ──
        if comps_ticker_input:
            with st.expander("🧠 Smart Peer Suggestions", expanded=False):
                st.markdown(
                    '<div style="font-size:0.7rem; color:#B8B3D7; margin-bottom:0.5rem;">'
                    'Auto-suggested peers based on sector/industry. Accept or override.</div>',
                    unsafe_allow_html=True,
                )
                # Try to get sector for this ticker
                _sps_sector = None
                _sps_industry = None
                try:
                    _sps_info = _quick_ticker_lookup(comps_ticker_input)
                    _sps_sector = _sps_info.get("sector", "")
                    _sps_industry = _sps_info.get("industry", "")
                except Exception:
                    pass

                # Find matching peers from our map
                _sps_candidates = []
                if _sps_industry and _sps_industry in SECTOR_PEER_MAP:
                    _sps_candidates = SECTOR_PEER_MAP[_sps_industry]
                elif _sps_sector and _sps_sector in SECTOR_PEER_MAP:
                    _sps_candidates = SECTOR_PEER_MAP[_sps_sector]
                else:
                    # Try partial match
                    for _sk, _sv in SECTOR_PEER_MAP.items():
                        if _sps_sector and _sps_sector.lower() in _sk.lower():
                            _sps_candidates = _sv
                            break
                        if _sps_industry and _sps_industry.lower() in _sk.lower():
                            _sps_candidates = _sv
                            break

                # Remove the target itself
                _sps_suggestions = [t for t in _sps_candidates if t.upper() != comps_ticker_input.upper()][:8]

                if _sps_suggestions:
                    if _sps_sector:
                        st.markdown(f'<div style="font-size:0.65rem; color:#9B8AFF;">Sector: {_sps_sector} · Industry: {_sps_industry or "N/A"}</div>', unsafe_allow_html=True)

                    # Initialize session state for peer suggestions
                    _sps_key = f"smart_peers_{comps_ticker_input}"
                    if _sps_key not in st.session_state:
                        st.session_state[_sps_key] = _sps_suggestions

                    _sps_selected = st.multiselect(
                        "Suggested Peers",
                        options=_sps_suggestions,
                        default=st.session_state[_sps_key],
                        key=f"sps_ms_{comps_ticker_input}",
                        label_visibility="collapsed",
                    )
                    st.session_state[_sps_key] = _sps_selected

                    # Manual override
                    _sps_manual = st.text_input(
                        "Add custom peers (comma-separated)",
                        placeholder="TICKER1, TICKER2",
                        key=f"sps_manual_{comps_ticker_input}",
                        label_visibility="collapsed",
                    )
                    if _sps_manual:
                        _sps_custom = [t.strip().upper() for t in _sps_manual.split(",") if t.strip()]
                        _sps_selected = list(set(_sps_selected + _sps_custom))
                        st.session_state[_sps_key] = _sps_selected

                    st.markdown(
                        f'<div style="font-size:0.65rem; color:#8A85AD; margin-top:0.3rem;">'
                        f'{len(_sps_selected)} peers selected</div>',
                        unsafe_allow_html=True,
                    )
                else:
                    st.markdown(
                        '<div style="font-size:0.7rem; color:#8A85AD;">'
                        'No auto-suggestions available for this ticker. The comps engine will find peers automatically.'
                        '</div>',
                        unsafe_allow_html=True,
                    )
        
        st.markdown('<div style="height:0.5rem;"></div>', unsafe_allow_html=True)
        comps_btn = st.button("🔍 Run Comps Analysis", type="primary", use_container_width=True)

    elif analysis_mode == "DCF Valuation":
        # ── DCF Valuation Mode ──
        st.markdown(
            '<div class="sb-section"><span class="sb-section-icon">💰</span> TARGET COMPANY</div>',
            unsafe_allow_html=True,
        )
        
        dcf_ticker_input = st.text_input(
            "Stock Ticker", value="", max_chars=10,
            placeholder="Enter ticker (e.g. AAPL)",
            label_visibility="collapsed",
            key="dcf_ticker"
        ).strip().upper()
        
        if dcf_ticker_input:
            _render_company_card(dcf_ticker_input)
        
        st.markdown('<div style="height:0.5rem;"></div>', unsafe_allow_html=True)
        
        # DCF Assumptions
        st.markdown(
            '<div class="sb-section"><span class="sb-section-icon">📈</span> GROWTH ASSUMPTIONS</div>',
            unsafe_allow_html=True,
        )
        
        dcf_growth_rate = st.slider("FCF Growth Rate (%)", 0, 30, 8, 1, 
                                     help="Expected annual free cash flow growth rate") / 100
        dcf_terminal_growth = st.slider("Terminal Growth (%)", 0.0, 4.0, 2.5, 0.5,
                                        help="Long-term perpetuity growth rate (typically GDP growth ~2-3%)") / 100
        dcf_years = st.slider("Projection Years", 3, 10, 5, 1,
                              help="Number of years to project FCF before terminal value")
        
        st.markdown(
            '<div class="sb-section"><span class="sb-section-icon">🎯</span> DISCOUNT RATE</div>',
            unsafe_allow_html=True,
        )
        
        dcf_auto_wacc = st.checkbox("🤖 Auto-Calculate WACC", value=False,
                                     help="Compute WACC from CAPM and company financials")
        
        if dcf_auto_wacc and dcf_ticker_input:
            try:
                _wacc_cd = fetch_company_data(dcf_ticker_input)
                _w_beta = _wacc_cd.beta if _wacc_cd.beta else 1.0
                _w_rf = 0.0425  # 10Y Treasury
                _w_erp = 0.055  # Equity risk premium
                _w_ke = _w_rf + _w_beta * _w_erp  # CAPM
                
                # Cost of Debt
                _w_ie = float(_wacc_cd.interest_expense.iloc[0]) if _wacc_cd.interest_expense is not None and len(_wacc_cd.interest_expense) > 0 else 0
                _w_td = float(_wacc_cd.total_debt.iloc[0]) if _wacc_cd.total_debt is not None and len(_wacc_cd.total_debt) > 0 else 0
                _w_ie = abs(_w_ie)  # interest expense may be negative
                _w_kd = (_w_ie / _w_td) if _w_td > 0 else 0.05
                
                # Tax Rate
                _w_tp = float(_wacc_cd.tax_provision.iloc[0]) if _wacc_cd.tax_provision is not None and len(_wacc_cd.tax_provision) > 0 else 0
                _w_oi = float(_wacc_cd.operating_income.iloc[0]) if _wacc_cd.operating_income is not None and len(_wacc_cd.operating_income) > 0 else 0
                _w_pretax = _w_oi - _w_ie
                _w_tax_rate = (abs(_w_tp) / abs(_w_pretax)) if abs(_w_pretax) > 0 else 0.21
                _w_tax_rate = max(0, min(_w_tax_rate, 0.50))
                
                # Capital Structure
                _w_mcap = _wacc_cd.market_cap or 0
                _w_ev = _w_mcap + _w_td
                _w_we = (_w_mcap / _w_ev) if _w_ev > 0 else 0.7
                _w_wd = (_w_td / _w_ev) if _w_ev > 0 else 0.3
                
                _w_wacc = _w_we * _w_ke + _w_wd * _w_kd * (1 - _w_tax_rate)
                _w_wacc_pct = round(_w_wacc * 100, 1)
                
                st.markdown(
                    f'<div style="background:rgba(16,185,129,0.08); border:1px solid rgba(16,185,129,0.25); '
                    f'border-radius:10px; padding:0.6rem; margin:0.3rem 0; font-size:0.7rem; color:#B8B3D7; line-height:1.7;">'
                    f'<b style="color:#10B981;">Calculated WACC: {_w_wacc_pct:.1f}%</b><br>'
                    f'Ke={_w_ke*100:.1f}% (β={_w_beta:.2f}) · Kd={_w_kd*100:.1f}% · Tax={_w_tax_rate*100:.0f}%<br>'
                    f'E/V={_w_we*100:.0f}% · D/V={_w_wd*100:.0f}%'
                    f'</div>',
                    unsafe_allow_html=True,
                )
                # Store for WACC breakdown section
                st.session_state["_auto_wacc_data"] = {
                    "rf": _w_rf, "beta": _w_beta, "erp": _w_erp, "ke": _w_ke,
                    "kd": _w_kd, "tax_rate": _w_tax_rate, "we": _w_we, "wd": _w_wd,
                    "wacc": _w_wacc, "mcap": _w_mcap, "debt": _w_td,
                }
                
                # Suggested FCF Growth (historical CAGR)
                if _wacc_cd.free_cashflow_series is not None and len(_wacc_cd.free_cashflow_series) >= 2:
                    _fcf_vals = _wacc_cd.free_cashflow_series.dropna()
                    if len(_fcf_vals) >= 2:
                        _fcf_latest = float(_fcf_vals.iloc[0])
                        _fcf_oldest = float(_fcf_vals.iloc[-1])
                        _n_yrs = len(_fcf_vals) - 1
                        if _fcf_oldest > 0 and _fcf_latest > 0 and _n_yrs > 0:
                            _fcf_cagr = (_fcf_latest / _fcf_oldest) ** (1 / _n_yrs) - 1
                            st.markdown(
                                f'<div style="font-size:0.65rem; color:#8A85AD; margin:0.2rem 0;">'
                                f'📈 Suggested FCF Growth: <b style="color:#9B8AFF;">{_fcf_cagr*100:.1f}%</b> '
                                f'({_n_yrs}Y CAGR)</div>',
                                unsafe_allow_html=True,
                            )
                
                _auto_wacc_int = max(5, min(20, int(round(_w_wacc_pct))))
            except Exception:
                _auto_wacc_int = 10
        else:
            _auto_wacc_int = 10
        
        dcf_discount_rate = st.slider("WACC / Discount Rate (%)", 5, 20, _auto_wacc_int, 1,
                                      help="Weighted average cost of capital — higher = more conservative") / 100
        
        st.markdown('<div style="height:0.5rem;"></div>', unsafe_allow_html=True)
        dcf_btn = st.button("💹 Calculate DCF", type="primary", use_container_width=True)
        
        st.markdown(
            '<div style="font-size:0.65rem; color:#8A85AD; margin-top:0.5rem; line-height:1.6;">'
            '💡 <b>Tip:</b> Use historical growth rates and peer WACC as starting points. '
            'Higher risk = higher discount rate.'
            '</div>',
            unsafe_allow_html=True,
        )

    elif analysis_mode == "Quick Compare":
        # ── Quick Compare Mode ──
        st.markdown(
            '<div class="sb-section"><span class="sb-section-icon">⚖️</span> COMPANIES TO COMPARE</div>',
            unsafe_allow_html=True,
        )
        
        compare_input = st.text_area(
            "Enter tickers (comma-separated)",
            placeholder="AAPL, MSFT, GOOGL, META",
            height=80,
            label_visibility="collapsed",
            key="compare_input"
        )
        
        compare_tickers = [t.strip().upper() for t in compare_input.split(",") if t.strip()]
        
        if compare_tickers:
            st.markdown(
                f'<div style="font-size:0.75rem; color:#9B8AFF; margin:0.5rem 0;">'
                f'📊 Comparing {len(compare_tickers)} companies: {", ".join(compare_tickers[:5])}'
                f'{"..." if len(compare_tickers) > 5 else ""}'
                f'</div>',
                unsafe_allow_html=True,
            )
        
        st.markdown('<div style="height:0.5rem;"></div>', unsafe_allow_html=True)
        
        # Preset comparisons
        st.markdown(
            '<div class="sb-section"><span class="sb-section-icon">🔥</span> PRESET COMPARISONS</div>',
            unsafe_allow_html=True,
        )
        
        preset_options = {
            "FAANG": "META, AAPL, AMZN, NFLX, GOOGL",
            "Big Tech": "AAPL, MSFT, GOOGL, AMZN, META, NVDA",
            "Canadian Banks": "RY.TO, TD.TO, BNS.TO, BMO.TO, CM.TO",
            "Software/SaaS": "CRM, ADBE, NOW, WDAY, TEAM",
            "Semiconductors": "NVDA, AMD, INTC, QCOM, AVGO",
            "Healthcare Giants": "JNJ, UNH, PFE, ABBV, MRK",
        }
        
        preset_choice = st.selectbox("Load Preset", ["Custom"] + list(preset_options.keys()), label_visibility="collapsed")
        if preset_choice != "Custom":
            if st.button("Load Preset", key="load_preset"):
                st.session_state["compare_input"] = preset_options[preset_choice]
                st.rerun()
        
        st.markdown('<div style="height:0.5rem;"></div>', unsafe_allow_html=True)
        compare_btn = st.button("⚖️ Compare Companies", type="primary", use_container_width=True)

    elif analysis_mode == "Merger Analysis":
        # ── Merger Analysis Mode ──
        # Acquirer
        st.markdown(
            '<div class="sb-section"><span class="sb-section-icon">🏢</span> ACQUIRER</div>',
            unsafe_allow_html=True,
        )
        acquirer_input = st.text_input(
            "Acquirer", value="", max_chars=10,
            placeholder="Enter ticker (e.g. MSFT)",
            label_visibility="collapsed",
        ).strip().upper()
        if acquirer_input:
            _render_company_card(acquirer_input, "Acquirer")

        # Target
        st.markdown(
            '<div class="sb-section"><span class="sb-section-icon">🎯</span> TARGET</div>',
            unsafe_allow_html=True,
        )
        target_input = st.text_input(
            "Target", value="", max_chars=10,
            placeholder="Enter ticker (e.g. ATVI)",
            label_visibility="collapsed",
        ).strip().upper()
        if target_input:
            _render_company_card(target_input, "Target")

        # ── Section: Deal Structure ──
        st.markdown(
            '<div class="sb-section"><span class="sb-section-icon">💰</span> DEAL STRUCTURE</div>',
            unsafe_allow_html=True,
        )
        offer_premium = st.slider("Offer Premium (%)", 0, 100, 30, 5,
                                  help="Premium over current market price. Typical M&A premiums: 20-40%")
        pct_cash = st.slider("Cash Consideration (%)", 0, 100, 50, 5,
                             help="% of deal funded by cash (remainder is stock)")
        pct_stock = 100 - pct_cash
        st.markdown(
            f'<div class="sb-split-bar">'
            f'<div class="sb-split-cash" style="width:{pct_cash}%"></div>'
            f'<div class="sb-split-stock" style="width:{pct_stock}%"></div>'
            f'</div>'
            f'<div class="sb-split-labels">'
            f'<span class="cash-label">Cash {pct_cash}%</span>'
            f'<span class="stock-label">Stock {pct_stock}%</span>'
            f'</div>',
            unsafe_allow_html=True,
        )

        # ── Section: Synergies ──
        st.markdown(
            '<div class="sb-section"><span class="sb-section-icon">⚡</span> SYNERGIES</div>',
            unsafe_allow_html=True,
        )
        cost_syn = st.slider("Cost Synergies (% of Target SG&A)", 0, 30, 10, 1,
                             help="Expected cost savings from eliminating redundancies. 10-15% is typical")
        rev_syn = st.slider("Revenue Synergies (% of Target Rev)", 0, 10, 2, 1,
                            help="Expected revenue uplift from cross-selling, market access. Usually conservative (1-3%)")

        # ── Section: Financing & Fees ──
        st.markdown(
            '<div class="sb-section"><span class="sb-section-icon">🏦</span> FINANCING &amp; FEES</div>',
            unsafe_allow_html=True,
        )
        txn_fees = st.slider("Transaction Fees (%)", 0.5, 5.0, 2.0, 0.5,
                             help="Advisory, legal, and other transaction costs. Typical: 1-3%")
        adv_cost_of_debt = st.slider("Cost of Debt (%)", 2.0, 10.0, 5.0, 0.5,
                                     help="Interest rate on new debt to fund the acquisition")
        adv_tax_rate = st.slider("Tax Rate (%)", 10, 40, 25, 1,
                                 help="Corporate tax rate for calculating after-tax synergies and interest")

        merger_assumptions = MergerAssumptions(
            offer_premium_pct=offer_premium,
            pct_cash=pct_cash,
            pct_stock=pct_stock,
            cost_synergies_pct=cost_syn,
            revenue_synergies_pct=rev_syn,
            transaction_fees_pct=txn_fees,
            tax_rate=adv_tax_rate,
            cost_of_debt=adv_cost_of_debt,
        )

        st.markdown('<div style="height:0.8rem;"></div>', unsafe_allow_html=True)
        merger_btn = st.button("🚀 Analyze Deal", type="primary", use_container_width=True)

    # VMS Screener variables
    vms_screen_btn = False
    vms_rev_min = 1
    vms_rev_max = 500
    vms_ebitda_min = 0
    vms_growth_min = 0
    vms_industries = []
    vms_geographies = []

    if analysis_mode == "VMS Screener":
        # ── VMS Screener Mode ──
        st.markdown(
            '<div class="sb-section"><span class="sb-section-icon">🔍</span> SCREENING CRITERIA</div>',
            unsafe_allow_html=True,
        )
        st.markdown('<div style="font-size:0.7rem; color:#8A85AD; margin-bottom:0.5rem;">Revenue Range ($M)</div>', unsafe_allow_html=True)
        vms_rev_min = st.slider("Min Revenue ($M)", 1, 500, 1, 1, key="vms_rev_min_sl")
        vms_rev_max = st.slider("Max Revenue ($M)", 1, 500, 500, 1, key="vms_rev_max_sl")
        if vms_rev_min > vms_rev_max:
            vms_rev_min, vms_rev_max = vms_rev_max, vms_rev_min

        st.markdown('<div style="font-size:0.7rem; color:#8A85AD; margin-top:0.5rem;">Profitability & Growth</div>', unsafe_allow_html=True)
        vms_ebitda_min = st.slider("Min EBITDA Margin (%)", 0, 50, 10, 1, key="vms_ebitda_sl")
        vms_growth_min = st.slider("Min Revenue Growth (%)", -20, 50, 0, 1, key="vms_growth_sl")

        st.markdown(
            '<div class="sb-section"><span class="sb-section-icon">🏭</span> INDUSTRY FILTER</div>',
            unsafe_allow_html=True,
        )
        _vms_verticals = [
            "Healthcare IT", "GovTech", "Legal Tech", "Education Tech",
            "Real Estate Tech", "Construction Tech", "Utilities",
            "Transportation", "Agriculture Tech", "Financial Services Tech",
        ]
        vms_industries = st.multiselect("VMS Verticals", _vms_verticals, default=[], key="vms_ind_ms")

        st.markdown(
            '<div class="sb-section"><span class="sb-section-icon">🌍</span> GEOGRAPHY</div>',
            unsafe_allow_html=True,
        )
        vms_geographies = st.multiselect("Geography", ["North America", "Europe", "Asia-Pacific", "Other"], default=[], key="vms_geo_ms")

        st.markdown('<div style="height:0.8rem;"></div>', unsafe_allow_html=True)
        vms_screen_btn = st.button("🚀 Run Screen", type="primary", use_container_width=True)

    # ══════════════════════════════════════════════════════
    # 📊 FULL DATA EXPORT (available in all modes)
    # ══════════════════════════════════════════════════════
    st.markdown('<div class="sb-divider"></div>', unsafe_allow_html=True)
    with st.expander("📊 Full Data Export", expanded=False):
        st.markdown(
            '<div style="font-size:0.72rem; color:#B8B3D7; margin-bottom:0.5rem;">'
            'Export financial data for the last analyzed company as a multi-sheet Excel workbook.'
            '</div>',
            unsafe_allow_html=True,
        )
        if "last_cd" in st.session_state and st.session_state["last_cd"] is not None:
            _exp_cd = st.session_state["last_cd"]
            try:
                _exp_excel = _export_to_excel(_exp_cd)
                st.download_button(
                    label=f"📥 {_exp_cd.ticker} Full Export (.xlsx)",
                    data=_exp_excel,
                    file_name=f"{_exp_cd.ticker}_Full_Financial_Data.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                    key="sidebar_full_export",
                )
                st.markdown(
                    '<div style="font-size:0.6rem; color:#8A85AD; margin-top:0.3rem;">'
                    '6 sheets: Summary · Income Statement · Balance Sheet · Cash Flow · Key Ratios · Valuation Multiples'
                    '</div>',
                    unsafe_allow_html=True,
                )
            except Exception as _exp_e:
                st.warning(f"Export error: {_exp_e}")
        else:
            st.info("Analyze a company first to enable export.", icon="ℹ️")

    # ══════════════════════════════════════════════════════
    # ⌨️ KEYBOARD SHORTCUTS PANEL
    # ══════════════════════════════════════════════════════
    with st.expander("⌨️ Shortcuts & Reference", expanded=False):
        _shortcuts_data = [
            ("⌘/Ctrl + K", "Quick search (Streamlit native)"),
            ("⌘/Ctrl + B", "Toggle sidebar"),
            ("⌘/Ctrl + E", "Export data"),
            ("⌘/Ctrl + D", "Download report"),
            ("⌘/Ctrl + W", "Add to watchlist"),
            ("R", "Rerun app"),
            ("?", "Show this panel"),
        ]
        for _sk, _sd in _shortcuts_data:
            st.markdown(
                f'<div style="display:flex; justify-content:space-between; padding:0.25rem 0; '
                f'border-bottom:1px solid rgba(255,255,255,0.04);">'
                f'<kbd style="background:rgba(0,0,0,0.3); padding:0.15rem 0.4rem; border-radius:4px; '
                f'font-family:monospace; font-size:0.7rem; color:#E0DCF5;">{_sk}</kbd>'
                f'<span style="color:#B8B3D7; font-size:0.72rem;">{_sd}</span></div>',
                unsafe_allow_html=True,
            )
        st.markdown(
            '<div style="font-size:0.6rem; color:#6B6B80; margin-top:0.5rem; line-height:1.5;">'
            '💡 <b>Quick Actions:</b><br>'
            '• Use sidebar radio buttons to switch modes<br>'
            '• Click any watchlist ticker to load it<br>'
            '• Use preset comparisons for quick benchmarks<br>'
            '• All exports support .xlsx, .csv, and .json formats'
            '</div>',
            unsafe_allow_html=True,
        )

    # Sidebar Footer
    st.markdown('<div class="sb-divider" style="margin-top:1.5rem;"></div>', unsafe_allow_html=True)
    st.markdown(
        '<div style="text-align:center; padding: 0.5rem 0;">'
        '<div style="font-size:0.55rem; font-weight:800; background:linear-gradient(135deg, #6B5CE7, #E8638B); '
        '-webkit-background-clip:text; -webkit-text-fill-color:transparent; letter-spacing:2px; margin-bottom:0.3rem;">'
        'ORBITAL v5.0</div>'
        '<div style="font-size:0.55rem; color:#4B5563; letter-spacing:0.5px; line-height:1.8;">'
        'DATA: YAHOO FINANCE • CHARTS: PLOTLY<br>'
        'AI: OPENAI (OPT.) • LOGOS: CLEARBIT'
        '</div>'
        '<div style="margin-top:0.4rem;">'
        '<a href="https://github.com/rajkcho/profilebuilder" target="_blank" '
        'style="font-size:0.55rem; color:#6B5CE7; text-decoration:none; font-weight:600;">GitHub ↗</a>'
        '</div>'
        '</div>',
        unsafe_allow_html=True,
    )

# ── Main Area ────────────────────────────────────────────────
# Orbital animated logo HTML (text inside orbit)
def _orbital_logo(size="", text="ORBITAL"):
    size_class = f" orbital-logo-{size}" if size else ""
    return (
        f'<div class="orbital-logo{size_class}">'
        f'<span class="orbital-text">{text}</span>'
        '<div class="orbital-ring orbital-ring-1"></div>'
        '<div class="orbital-ring orbital-ring-2"></div>'
        '<div class="orbital-ring orbital-ring-3"></div>'
        '<div class="orbital-particle orbital-particle-1"></div>'
        '<div class="orbital-particle orbital-particle-2"></div>'
        '<div class="orbital-particle orbital-particle-3"></div>'
        '</div>'
    )

if analysis_mode == "Company Profile":
    st.markdown(
        '<div class="hero-header">'
        '<div class="orbital-brand">'
        f'{_orbital_logo()}'
        '<p class="orbital-tagline">Company Intelligence & Tear Sheet Generator</p>'
        '</div>'
        '<span class="hero-tagline">Powered by Live Market Data</span>'
        '</div>',
        unsafe_allow_html=True,
    )
else:
    st.markdown(
        '<div class="hero-header">'
        '<div class="orbital-brand">'
        f'{_orbital_logo()}'
        '<p class="orbital-tagline">M&A Simulator & Deal Book Generator</p>'
        '</div>'
        '<span class="hero-tagline">Powered by Live Market Data</span>'
        '</div>',
        unsafe_allow_html=True,
    )

if analysis_mode == "Company Profile" and generate_btn and ticker_input:
    # ── Data Fetching (with scanner loading animation) ───
    _scanner_slot = st.empty()

    try:
        _scanner_slot.markdown(_render_profile_scanner(ticker_input.upper(), 0), unsafe_allow_html=True)
    except Exception:
        pass  # Scanner rendering is non-critical

    try:
        cd = fetch_company_data(ticker_input)
        st.session_state["last_cd"] = cd
        # Add to search history on successful fetch
        _add_to_search_history(ticker_input)
    except Exception as e:
        _scanner_slot.empty()
        st.error(f"Failed to fetch data for **{ticker_input}**: {e}")
        st.stop()

    try:
        _scanner_slot.markdown(_render_profile_scanner(ticker_input.upper(), 1), unsafe_allow_html=True)
    except Exception:
        pass

    try:
        cd = fetch_peer_data(cd)
    except Exception:
        pass  # Peer data is non-critical

    try:
        _scanner_slot.markdown(_render_profile_scanner(ticker_input.upper(), 2), unsafe_allow_html=True)
    except Exception:
        pass

    try:
        cd = generate_insights(cd)
    except Exception as e:
        print(f"Insights generation warning: {e}")  # Non-fatal

    try:
        _scanner_slot.markdown(_render_profile_scanner(ticker_input.upper(), 3), unsafe_allow_html=True)
        time.sleep(1.2)
    except Exception:
        pass

    _scanner_slot.empty()

    cs = cd.currency_symbol  # shorthand
    _fetch_timestamp = datetime.now().strftime("%b %d, %Y at %H:%M UTC")

    # Data freshness indicator
    st.markdown(
        f'<div style="text-align:right; padding:0.3rem 1rem; margin-bottom:-0.5rem;">'
        f'<span style="font-size:0.6rem; color:#5A567A;">🔄 Data as of {_fetch_timestamp}</span>'
        f'</div>',
        unsafe_allow_html=True,
    )

    # ══════════════════════════════════════════════════════
    # 1. COMPANY HEADER CARD (with logo)
    # ══════════════════════════════════════════════════════
    _cd_price_change = getattr(cd, 'price_change', None) or 0
    _cd_price_change_pct = getattr(cd, 'price_change_pct', None) or 0
    chg_class = "price-up" if _cd_price_change >= 0 else "price-down"
    chg_badge = "change-up" if _cd_price_change >= 0 else "change-down"
    arrow = "&#9650;" if _cd_price_change >= 0 else "&#9660;"

    logo_html = ""
    if cd.logo_url:
        _ld = getattr(cd, 'logo_domain', '')
        logo_fallback = f"this.onerror=null; this.src='https://logo.clearbit.com/{_ld}';" if _ld else "this.style.display='none';"
        logo_html = (
            f'<img src="{cd.logo_url}" '
            f'style="width:52px; height:52px; border-radius:10px; object-fit:contain; '
            f'background:white; padding:4px; margin-right:1.2rem; flex-shrink:0;" '
            f'onerror="{logo_fallback}">'
        )

    st.markdown(
        f'<div class="company-card">'
        f'<div style="display:flex; align-items:center; position:relative;">'
        f'{logo_html}'
        f'<div>'
        f'<p class="company-name">{cd.name}</p>'
        f'<p class="company-meta"><span>{cd.ticker}</span> &nbsp;&middot;&nbsp; {cd.exchange} &nbsp;&middot;&nbsp; {cd.sector} &rarr; {cd.industry}</p>'
        f'</div>'
        f'</div>'
        f'<div style="display:flex; align-items:baseline; gap:1rem; margin-top:0.8rem; position:relative;">'
        f'<p class="price-tag {chg_class}">{cs}{getattr(cd, "current_price", 0):,.2f}</p>'
        f'<span class="price-change {chg_badge}">{arrow} {_cd_price_change:+.2f} ({_cd_price_change_pct:+.2f}%)</span>'
        f'<span style="font-size:0.75rem; color:#A8A3C7; margin-left:0.5rem;">{cd.currency_code}</span>'
        f'</div>'
        f'</div>',
        unsafe_allow_html=True,
    )

    # ══════════════════════════════════════════════════════
    # 2. PROMINENT PRICE / VOLUME DISPLAY
    # ══════════════════════════════════════════════════════
    price_color = "#10B981" if _cd_price_change >= 0 else "#EF4444"
    price_bg = "rgba(16,185,129,0.05)" if _cd_price_change >= 0 else "rgba(239,68,68,0.05)"

    st.markdown(
        f'<div class="price-bar" style="background:{price_bg}; border:1px solid {"rgba(16,185,129,0.15)" if _cd_price_change >= 0 else "rgba(239,68,68,0.15)"};">'
        f'<div style="flex:1;">'
        f'<div style="font-size:0.65rem; font-weight:600; color:#8A85AD; text-transform:uppercase; letter-spacing:1px;">Current Price</div>'
        f'<div style="font-size:2rem; font-weight:800; color:{price_color};">'
        f'{cs}{getattr(cd, "current_price", 0):,.2f}'
        f'<span style="font-size:0.9rem; margin-left:0.5rem;">{arrow} {_cd_price_change:+.2f} ({_cd_price_change_pct:+.2f}%)</span></div>'
        f'</div>'
        f'<div style="flex:0 0 180px; text-align:center; border-left:1px solid rgba(255,255,255,0.1); padding-left:1rem;">'
        f'<div style="font-size:0.65rem; font-weight:600; color:#8A85AD; text-transform:uppercase; letter-spacing:1px;">Volume</div>'
        f'<div style="font-size:1.3rem; font-weight:700; color:#E0DCF5;">{format_number(getattr(cd, "volume", None), prefix="", decimals=0)}</div>'
        f'<div style="font-size:0.6rem; color:#8A85AD;">Avg: {format_number(getattr(cd, "avg_volume", None), prefix="", decimals=0)}</div>'
        f'</div>'
        f'<div style="flex:0 0 220px; text-align:center; border-left:1px solid rgba(255,255,255,0.1); padding-left:1rem;">'
        f'<div style="font-size:0.65rem; font-weight:600; color:#8A85AD; text-transform:uppercase; letter-spacing:1px;">52W Range</div>'
        f'<div style="font-size:1.1rem; font-weight:600; color:#E0DCF5;">'
        f'{cs}{getattr(cd, "fifty_two_week_low", 0):,.2f} &mdash; {cs}{getattr(cd, "fifty_two_week_high", 0):,.2f}</div>'
        f'</div>'
        f'</div>',
        unsafe_allow_html=True,
    )

    # Quick KPI strip with sparklines
    def _make_sparkline(data, color="#6B5CE7", height=40):
        """Create a minimal sparkline figure."""
        if data is None or (hasattr(data, '__len__') and len(data) < 2):
            return None
        try:
            import plotly.graph_objects as _sp_go
            vals = list(data) if not isinstance(data, list) else data
            vals = [v for v in vals if v is not None and not (isinstance(v, float) and (v != v))]
            if len(vals) < 2:
                return None
            fig = _sp_go.Figure(_sp_go.Scatter(
                y=vals, mode='lines', line=dict(color=color, width=1.5),
                fill='tozeroy', fillcolor=color.replace(')', ',0.1)').replace('rgb', 'rgba') if 'rgb' in color else f"rgba(107,92,231,0.1)",
            ))
            fig.update_layout(
                height=height, margin=dict(t=0, b=0, l=0, r=0),
                paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
                xaxis=dict(visible=False), yaxis=dict(visible=False),
                showlegend=False,
            )
            return fig
        except Exception:
            return None

    def _yoy_delta(series):
        """Return YoY delta string from a Series (most recent vs prior)."""
        if series is None or not hasattr(series, 'iloc') or len(series) < 2:
            return None
        try:
            curr = float(series.iloc[0])
            prev = float(series.iloc[1])
            if prev == 0:
                return None
            pct = (curr / prev - 1) * 100
            return f"{pct:+.1f}% YoY"
        except Exception:
            return None

    k1, k2, k3, k4, k5, k6 = st.columns(6)
    k1.metric("Market Cap", format_number(getattr(cd, 'market_cap', None), currency_symbol=cs))
    k2.metric("Enterprise Value", format_number(getattr(cd, 'enterprise_value', None), currency_symbol=cs))

    _rev_val = "N/A"
    _rev_delta = None
    if cd.revenue is not None and len(cd.revenue) > 0:
        _rev_val = format_number(cd.revenue.iloc[0], currency_symbol=cs)
        _rev_delta = _yoy_delta(cd.revenue)
    k3.metric("Revenue (TTM)", _rev_val, delta=_rev_delta)

    _ni_val = "N/A"
    _ni_delta = None
    if cd.net_income is not None and len(cd.net_income) > 0:
        _ni_val = format_number(cd.net_income.iloc[0], currency_symbol=cs)
        _ni_delta = _yoy_delta(cd.net_income)
    k4.metric("Net Income", _ni_val, delta=_ni_delta)

    _fcf_val = "N/A"
    _fcf_delta = None
    if cd.free_cashflow_series is not None and len(cd.free_cashflow_series) > 0:
        _fcf_val = format_number(cd.free_cashflow_series.iloc[0], currency_symbol=cs)
        _fcf_delta = _yoy_delta(cd.free_cashflow_series)
    k5.metric("Free Cash Flow", _fcf_val, delta=_fcf_delta)

    # dividend_yield: yfinance may return as decimal (0.009) or already as pct-like (0.9)
    _div_yield = getattr(cd, 'dividend_yield', None)
    if _div_yield and _div_yield > 0.5:
        k6.metric("Dividend Yield", f"{_div_yield:.2f}%")
    else:
        k6.metric("Dividend Yield", format_pct(_div_yield) if _div_yield else "N/A")

    # Sparkline row under KPIs
    _price_hist = getattr(cd, 'price_history', None)
    _spark_price = _make_sparkline(_price_hist, color="#6B5CE7") if _price_hist is not None else None
    _spark_rev = _make_sparkline(cd.revenue[::-1] if cd.revenue is not None and len(cd.revenue) > 1 else None, color="#10B981")
    if _spark_price or _spark_rev:
        sp1, sp2 = st.columns(2)
        if _spark_price:
            with sp1:
                st.markdown('<div style="font-size:0.6rem; color:#8A85AD; text-transform:uppercase; letter-spacing:1px;">Price Trend (30d)</div>', unsafe_allow_html=True)
                st.plotly_chart(_spark_price, use_container_width=True, key="spark_price")
        if _spark_rev:
            with sp2:
                st.markdown('<div style="font-size:0.6rem; color:#8A85AD; text-transform:uppercase; letter-spacing:1px;">Revenue Trend</div>', unsafe_allow_html=True)
                st.plotly_chart(_spark_rev, use_container_width=True, key="spark_rev")

    # ══════════════════════════════════════════════════════
    # 2b. AT A GLANCE CARD
    # ══════════════════════════════════════════════════════
    glance_signals = []
    
    # Analyst rating — derive from recommendations_summary if no direct attribute
    rec = getattr(cd, 'analyst_recommendation', None) or getattr(cd, 'recommendation_key', None)
    if not rec and cd.recommendations_summary is not None and not cd.recommendations_summary.empty:
        try:
            row = cd.recommendations_summary.iloc[0]
            cats = {"strongBuy": 5, "buy": 4, "hold": 3, "sell": 2, "strongSell": 1}
            def _get_rec_val(row, k):
                try:
                    v = row.get(k, 0) if hasattr(row, 'get') else row[k] if k in row.index else 0
                    return int(v) if pd.notna(v) else 0
                except Exception:
                    return 0
            weighted = sum(_get_rec_val(row, k) * v for k, v in cats.items())
            total = sum(_get_rec_val(row, k) for k in cats)
            if total > 0:
                avg = weighted / total
                if avg >= 4.5: rec = "strong_buy"
                elif avg >= 3.5: rec = "buy"
                elif avg >= 2.5: rec = "hold"
                elif avg >= 1.5: rec = "sell"
                else: rec = "strong_sell"
        except Exception:
            pass
    rec_str = rec.replace("_", " ").title() if rec else "N/A"
    rec_color = "#10B981" if rec and "buy" in rec.lower() else "#EF4444" if rec and "sell" in rec.lower() else "#F59E0B"
    
    # Quick valuation check
    pe_status = "N/A"
    pe_color = "#8A85AD"
    if cd.trailing_pe and cd.trailing_pe > 0:
        if cd.trailing_pe < 15:
            pe_status = "Undervalued"
            pe_color = "#10B981"
        elif cd.trailing_pe < 25:
            pe_status = "Fair Value"
            pe_color = "#F59E0B"
        else:
            pe_status = "Premium"
            pe_color = "#E8638B"
    
    # Momentum
    mom_str = "N/A"
    mom_color = "#8A85AD"
    if _cd_price_change_pct:
        if _cd_price_change_pct > 2:
            mom_str = "Strong Bullish"
            mom_color = "#10B981"
        elif _cd_price_change_pct > 0:
            mom_str = "Bullish"
            mom_color = "#34D399"
        elif _cd_price_change_pct > -2:
            mom_str = "Bearish"
            mom_color = "#F97316"
        else:
            mom_str = "Strong Bearish"
            mom_color = "#EF4444"
    
    st.markdown(
        f'<div style="display:grid; grid-template-columns:repeat(4, 1fr); gap:0.8rem; margin:1rem 0;">'
        f'<div style="background:rgba(107,92,231,0.05); border:1px solid rgba(107,92,231,0.1); '
        f'border-radius:12px; padding:0.8rem; text-align:center;">'
        f'<div style="font-size:0.6rem; color:#8A85AD; font-weight:600; text-transform:uppercase; letter-spacing:1px;">Analyst Rating</div>'
        f'<div style="font-size:1rem; font-weight:800; color:{rec_color}; margin-top:0.2rem;">{rec_str}</div>'
        f'</div>'
        f'<div style="background:rgba(107,92,231,0.05); border:1px solid rgba(107,92,231,0.1); '
        f'border-radius:12px; padding:0.8rem; text-align:center;">'
        f'<div style="font-size:0.6rem; color:#8A85AD; font-weight:600; text-transform:uppercase; letter-spacing:1px;">Valuation</div>'
        f'<div style="font-size:1rem; font-weight:800; color:{pe_color}; margin-top:0.2rem;">{pe_status}</div>'
        f'<div style="font-size:0.6rem; color:#8A85AD;">P/E: {cd.trailing_pe:.1f}x</div>'
        f'</div>' if cd.trailing_pe else
        f'<div style="background:rgba(107,92,231,0.05); border:1px solid rgba(107,92,231,0.1); '
        f'border-radius:12px; padding:0.8rem; text-align:center;">'
        f'<div style="font-size:0.6rem; color:#8A85AD; font-weight:600; text-transform:uppercase;">Valuation</div>'
        f'<div style="font-size:1rem; font-weight:800; color:#8A85AD; margin-top:0.2rem;">N/A</div></div>'
        f''
        f'<div style="background:rgba(107,92,231,0.05); border:1px solid rgba(107,92,231,0.1); '
        f'border-radius:12px; padding:0.8rem; text-align:center;">'
        f'<div style="font-size:0.6rem; color:#8A85AD; font-weight:600; text-transform:uppercase; letter-spacing:1px;">Today\'s Momentum</div>'
        f'<div style="font-size:1rem; font-weight:800; color:{mom_color}; margin-top:0.2rem;">{mom_str}</div>'
        f'</div>'
        f'<div style="background:rgba(107,92,231,0.05); border:1px solid rgba(107,92,231,0.1); '
        f'border-radius:12px; padding:0.8rem; text-align:center;">'
        f'<div style="font-size:0.6rem; color:#8A85AD; font-weight:600; text-transform:uppercase; letter-spacing:1px;">Sector</div>'
        f'<div style="font-size:0.85rem; font-weight:700; color:#6B5CE7; margin-top:0.2rem;">{cd.sector or "N/A"}</div>'
        f'</div>'
        f'</div>',
        unsafe_allow_html=True,
    )

    # ══════════════════════════════════════════════════════
    # 2b-ii. EXECUTIVE SUMMARY (IB Pitch Book style)
    # ══════════════════════════════════════════════════════
    with _safe_section("Executive Summary"):
        _es_bullets = []

        # Valuation bullet: EV/EBITDA
        _es_ev = getattr(cd, 'enterprise_value', None) or 0
        _es_ebitda_s = getattr(cd, 'ebitda', None)
        _es_ebitda = None
        if _es_ebitda_s is not None:
            if hasattr(_es_ebitda_s, 'iloc'):
                _es_ebitda = float(_es_ebitda_s.iloc[0]) if len(_es_ebitda_s) > 0 else None
            else:
                try:
                    _es_ebitda = float(_es_ebitda_s)
                except (TypeError, ValueError):
                    _es_ebitda = None
        _es_ev_ebitda = (_es_ev / _es_ebitda) if _es_ebitda and _es_ebitda > 0 else None
        _es_peer_ev_ebitda = getattr(cd, 'peer_ev_ebitda_median', None)
        if _es_ev_ebitda is not None:
            _val_str = f"Trading at **{_es_ev_ebitda:.1f}x EV/EBITDA**"
            if _es_peer_ev_ebitda and _es_peer_ev_ebitda > 0:
                _pct_diff = (_es_ev_ebitda / _es_peer_ev_ebitda - 1) * 100
                _val_tag = "cheap" if _pct_diff < -15 else "rich" if _pct_diff > 15 else "in-line"
                _val_str += f" vs peer median of {_es_peer_ev_ebitda:.1f}x ({_val_tag})"
            _es_bullets.append(("📊", _val_str))
        elif cd.trailing_pe and cd.trailing_pe > 0:
            _es_bullets.append(("📊", f"Trading at **{cd.trailing_pe:.1f}x P/E**"))

        # Growth bullet
        _es_rg = getattr(cd, 'revenue_growth', None)
        if _es_rg is not None:
            _rg_pct = _es_rg * 100 if abs(_es_rg) < 5 else _es_rg
            _es_bullets.append(("📈", f"Revenue growing at **{_rg_pct:.1f}%** YoY"))

        # Profitability bullet
        _es_pm = getattr(cd, 'profit_margins', None)
        _es_om = getattr(cd, 'operating_margins', None)
        if _es_pm is not None:
            _pm_pct = _es_pm * 100 if abs(_es_pm) < 5 else _es_pm
            _pm_qual = "above peers" if _pm_pct > 15 else "below peers" if _pm_pct < 5 else "moderate"
            _es_bullets.append(("💰", f"Net margins **{_pm_pct:.1f}%** ({_pm_qual})"))
        elif _es_om is not None:
            _om_pct = _es_om * 100 if abs(_es_om) < 5 else _es_om
            _es_bullets.append(("💰", f"Operating margins **{_om_pct:.1f}%**"))

        # Balance sheet bullet
        _es_nd = None
        _es_td_s = getattr(cd, 'total_debt', None)
        _es_cash_s = getattr(cd, 'cash_and_equivalents', None)
        _es_td_v = _safe_val(_es_td_s) if callable(_safe_val) else None
        try:
            _es_td_v = float(_es_td_s.iloc[0]) if hasattr(_es_td_s, 'iloc') and len(_es_td_s) > 0 else (float(_es_td_s) if _es_td_s is not None else None)
        except Exception:
            _es_td_v = None
        try:
            _es_cash_v = float(_es_cash_s.iloc[0]) if hasattr(_es_cash_s, 'iloc') and len(_es_cash_s) > 0 else (float(_es_cash_s) if _es_cash_s is not None else None)
        except Exception:
            _es_cash_v = None
        if _es_td_v is not None and _es_cash_v is not None and _es_ebitda and _es_ebitda > 0:
            _es_nd = (_es_td_v - _es_cash_v) / _es_ebitda
            _cr_val = getattr(cd, 'current_ratio', None)
            _bs_str = f"Net debt/EBITDA of **{_es_nd:.1f}x**"
            if _cr_val:
                _bs_str += f", current ratio **{_cr_val:.1f}x**"
            _es_bullets.append(("🏦", _bs_str))

        if _es_bullets:
            _es_html = "".join(
                f'<div style="padding:0.35rem 0; font-size:0.85rem; color:#E0DCF5; line-height:1.7;">'
                f'{icon} {text}</div>'
                for icon, text in _es_bullets
            )
            st.markdown(
                f'<div style="background:linear-gradient(135deg, rgba(107,92,231,0.08), rgba(16,185,129,0.06)); '
                f'border:2px solid transparent; border-image:linear-gradient(135deg, rgba(107,92,231,0.4), rgba(16,185,129,0.3)) 1; '
                f'border-radius:0px; padding:1.2rem 1.5rem; margin:1rem 0;">'
                f'<div style="font-size:0.7rem; font-weight:800; color:#9B8AFF; text-transform:uppercase; '
                f'letter-spacing:2px; margin-bottom:0.6rem;">Executive Summary</div>'
                f'{_es_html}'
                f'</div>',
                unsafe_allow_html=True,
            )

    # ══════════════════════════════════════════════════════
    # 2c. KEY TAKEAWAYS (auto-generated)
    # ══════════════════════════════════════════════════════
    try:
        takeaways = []
        
        # Market cap context
        if cd.market_cap:
            if cd.market_cap >= 200e9:
                takeaways.append(f"**Mega-cap** company (${cd.market_cap/1e12:.1f}T) in {cd.sector}")
            elif cd.market_cap >= 10e9:
                takeaways.append(f"**Large-cap** company (${cd.market_cap/1e9:.0f}B) in {cd.sector}")
            elif cd.market_cap >= 2e9:
                takeaways.append(f"**Mid-cap** company (${cd.market_cap/1e9:.1f}B) in {cd.sector}")
            else:
                takeaways.append(f"**Small-cap** company (${cd.market_cap/1e9:.1f}B) in {cd.sector}")
        
        # Valuation
        if cd.trailing_pe and cd.trailing_pe > 0:
            if cd.trailing_pe > 30:
                takeaways.append(f"Trading at a **premium valuation** ({cd.trailing_pe:.0f}x P/E)")
            elif cd.trailing_pe < 15:
                takeaways.append(f"Trading at an **attractive valuation** ({cd.trailing_pe:.0f}x P/E)")
        
        # Growth
        if cd.revenue_growth:
            if cd.revenue_growth > 20:
                takeaways.append(f"**High-growth** — revenue up {cd.revenue_growth:.0f}% YoY")
            elif cd.revenue_growth > 0:
                takeaways.append(f"Revenue growing at {cd.revenue_growth:.0f}% YoY")
            else:
                takeaways.append(f"Revenue declined {cd.revenue_growth:.0f}% YoY")
        
        # Profitability
        if cd.profit_margins:
            pm = cd.profit_margins * 100
            if pm > 20:
                takeaways.append(f"**Highly profitable** — {pm:.0f}% net margin")
            elif pm > 0:
                takeaways.append(f"Profitable with {pm:.0f}% net margin")
            else:
                takeaways.append(f"Currently unprofitable ({pm:.0f}% net margin)")
        
        # Analyst
        if rec:
            takeaways.append(f"Analyst consensus: **{rec_str}**")
        
        if takeaways:
            bullets = "".join(f'<div style="padding:0.2rem 0; font-size:0.82rem; color:#B8B3D7; line-height:1.6;">• {t}</div>' for t in takeaways[:4])
            st.markdown(
                f'<div style="background:rgba(107,92,231,0.04); border:1px solid rgba(107,92,231,0.12); '
                f'border-radius:12px; padding:1rem 1.2rem; margin:0.5rem 0 1rem 0;">'
                f'<div style="font-size:0.65rem; font-weight:700; color:#6B5CE7; text-transform:uppercase; '
                f'letter-spacing:1.5px; margin-bottom:0.4rem;">Key Takeaways</div>'
                f'{bullets}'
                f'</div>',
                unsafe_allow_html=True,
            )
    except Exception:
        pass

    # ══════════════════════════════════════════════════════
    # 2c-ii. WHAT-IF SIMULATOR
    # ══════════════════════════════════════════════════════
    with _safe_section("What-If Simulator"):
        try:
            _wif_revenue = None
            _wif_ebitda = None
            _wif_ev_ebitda = cd.ev_to_ebitda if cd.ev_to_ebitda and cd.ev_to_ebitda > 0 else None
            _wif_shares = cd.shares_outstanding if hasattr(cd, 'shares_outstanding') and cd.shares_outstanding else None
            if _wif_shares is None and cd.market_cap and cd.current_price and cd.current_price > 0:
                _wif_shares = cd.market_cap / cd.current_price
            if cd.revenue is not None and len(cd.revenue) > 0:
                _wif_revenue = float(cd.revenue.iloc[0])
            if cd.ebitda is not None:
                if hasattr(cd.ebitda, 'iloc') and len(cd.ebitda) > 0:
                    _wif_ebitda = float(cd.ebitda.iloc[0])
                elif isinstance(cd.ebitda, (int, float)):
                    _wif_ebitda = float(cd.ebitda)
            _wif_margin = (_wif_ebitda / _wif_revenue * 100) if _wif_revenue and _wif_ebitda and _wif_revenue > 0 else None

            if _wif_revenue and _wif_shares and _wif_shares > 0 and _wif_ev_ebitda:
                with st.expander("🎯 What-If Simulator", expanded=False):
                    st.markdown(
                        '<div style="font-size:0.75rem; color:#B8B3D7; margin-bottom:0.8rem;">'
                        'Adjust assumptions to see implied share price changes in real time.</div>',
                        unsafe_allow_html=True,
                    )
                    _wif_c1, _wif_c2, _wif_c3 = st.columns(3)
                    with _wif_c1:
                        _wif_rev_growth = st.slider("Revenue Growth (%)", -20, 20, 0, 1, key="wif_rev_growth",
                                                     help="Adjust revenue growth from base")
                    with _wif_c2:
                        _wif_margin_adj = st.slider("Margin Δ (bps)", -500, 500, 0, 25, key="wif_margin_adj",
                                                     help="EBITDA margin expansion/compression in basis points")
                    with _wif_c3:
                        _wif_mult_adj = st.slider("Multiple Δ (x)", -5.0, 5.0, 0.0, 0.5, key="wif_mult_adj",
                                                   help="EV/EBITDA multiple re-rating")

                    # Calculate implied price
                    _wif_adj_revenue = _wif_revenue * (1 + _wif_rev_growth / 100)
                    _wif_base_margin = _wif_margin if _wif_margin else 20.0
                    _wif_adj_margin = (_wif_base_margin + _wif_margin_adj / 100) / 100
                    _wif_adj_ebitda = _wif_adj_revenue * _wif_adj_margin
                    _wif_adj_multiple = _wif_ev_ebitda + _wif_mult_adj
                    _wif_adj_ev = _wif_adj_ebitda * _wif_adj_multiple
                    _wif_net_debt = (cd.enterprise_value or 0) - (cd.market_cap or 0)
                    _wif_eq_value = _wif_adj_ev - _wif_net_debt
                    _wif_implied_price = _wif_eq_value / _wif_shares if _wif_shares > 0 else 0
                    _wif_upside = ((_wif_implied_price / cd.current_price) - 1) * 100 if cd.current_price > 0 else 0

                    # Bull case: +10% rev, +200bps margin, +2x multiple
                    _wif_bull_rev = _wif_revenue * 1.10
                    _wif_bull_ebitda = _wif_bull_rev * ((_wif_base_margin + 2) / 100)
                    _wif_bull_ev = _wif_bull_ebitda * (_wif_ev_ebitda + 2)
                    _wif_bull_price = (_wif_bull_ev - _wif_net_debt) / _wif_shares if _wif_shares > 0 else 0

                    # Bear case: -10% rev, -200bps margin, -2x multiple
                    _wif_bear_rev = _wif_revenue * 0.90
                    _wif_bear_ebitda = _wif_bear_rev * (max(0, _wif_base_margin - 2) / 100)
                    _wif_bear_ev = _wif_bear_ebitda * max(1, _wif_ev_ebitda - 2)
                    _wif_bear_price = (_wif_bear_ev - _wif_net_debt) / _wif_shares if _wif_shares > 0 else 0

                    # Display metrics
                    _wm1, _wm2, _wm3 = st.columns(3)
                    _wif_color = "#10B981" if _wif_upside >= 0 else "#EF4444"
                    _wm1.metric("Current Price", f"{cd.currency_symbol}{cd.current_price:,.2f}")
                    _wm2.metric("Implied Price", f"{cd.currency_symbol}{_wif_implied_price:,.2f}",
                               delta=f"{_wif_upside:+.1f}%")
                    _wm3.metric("Adj. EV/EBITDA", f"{_wif_adj_multiple:.1f}x",
                               delta=f"{_wif_mult_adj:+.1f}x" if _wif_mult_adj != 0 else None)

                    # Plotly bar chart: Current vs Implied vs Bull vs Bear
                    try:
                        _wif_fig = go.Figure()
                        _wif_labels = ["Bear Case", "Current", "Your Scenario", "Bull Case"]
                        _wif_values = [_wif_bear_price, cd.current_price, _wif_implied_price, _wif_bull_price]
                        _wif_colors = ["#EF4444", "#8A85AD", "#6B5CE7", "#10B981"]
                        _wif_fig.add_trace(go.Bar(
                            x=_wif_labels, y=_wif_values,
                            marker_color=_wif_colors,
                            text=[f"{cd.currency_symbol}{v:,.2f}" for v in _wif_values],
                            textposition="outside",
                            textfont=dict(size=11, color="#E0DCF5"),
                        ))
                        _wif_fig.update_layout(
                            **_CHART_LAYOUT_BASE, height=300,
                            margin=dict(t=30, b=30, l=40, r=20),
                            yaxis=dict(title=dict(text="Implied Price", font=dict(size=10, color="#8A85AD")),
                                      tickprefix=cd.currency_symbol, tickfont=dict(size=9, color="#8A85AD"),
                                      showgrid=True, gridcolor="rgba(255,255,255,0.05)"),
                            xaxis=dict(tickfont=dict(size=10, color="#B8B3D7")),
                            showlegend=False,
                        )
                        _apply_space_grid(_wif_fig)
                        st.plotly_chart(_wif_fig, use_container_width=True, key="whatif_bar")
                    except Exception:
                        pass

                    st.markdown(
                        '<div style="font-size:0.65rem; color:#6B6B80; margin-top:0.3rem; line-height:1.5;">'
                        '💡 Bull/Bear cases use ±10% revenue, ±200bps margin, ±2x multiple from current levels. '
                        'Your scenario reflects the slider adjustments above.'
                        '</div>',
                        unsafe_allow_html=True,
                    )
        except Exception:
            pass

    # ══════════════════════════════════════════════════════
    # 2d. BULL / BEAR INVESTMENT THESIS
    # ══════════════════════════════════════════════════════
    with _safe_section("Investment Thesis"):
        bull_points = []
        bear_points = []

        # Revenue Growth
        if cd.revenue_growth:
            if cd.revenue_growth > 0.10:
                bull_points.append(f"Strong revenue growth ({cd.revenue_growth:.0%}) above market average")
            elif cd.revenue_growth < 0:
                bear_points.append(f"Revenue declining ({cd.revenue_growth:.0%}) — top-line contraction risk")

        # Margins
        if cd.profit_margins:
            if cd.profit_margins > 0.20:
                bull_points.append(f"High profit margins ({cd.profit_margins:.0%}) indicate pricing power and moat")
            elif cd.profit_margins < 0.05:
                bear_points.append(f"Thin profit margins ({cd.profit_margins:.0%}) leave little room for error")

        # FCF
        _fcf = cd.free_cashflow_series.iloc[0] if cd.free_cashflow_series is not None and len(cd.free_cashflow_series) > 0 else None
        if _fcf and _fcf > 0:
            _fcf_yield = (_fcf / cd.market_cap * 100) if cd.market_cap else 0
            if _fcf_yield > 5:
                bull_points.append(f"Strong FCF yield ({_fcf_yield:.1f}%) — cash generative business")
            elif _fcf_yield > 3:
                bull_points.append(f"Healthy FCF yield ({_fcf_yield:.1f}%)")
        elif _fcf and _fcf < 0:
            bear_points.append("Negative free cash flow — burning cash")

        # Debt
        if cd.debt_to_equity:
            if cd.debt_to_equity > 200:
                bear_points.append(f"High leverage (D/E: {cd.debt_to_equity:.0f}%) increases financial risk")
            elif cd.debt_to_equity < 30:
                bull_points.append(f"Conservative balance sheet (D/E: {cd.debt_to_equity:.0f}%)")

        # Beta / Risk
        if cd.beta:
            if cd.beta > 1.5:
                bear_points.append(f"High beta ({cd.beta:.2f}) — more volatile than market")
            elif cd.beta < 0.8:
                bull_points.append(f"Low beta ({cd.beta:.2f}) — defensive characteristics")

        # Analyst
        if rec and "buy" in rec.lower():
            bull_points.append(f"Analyst consensus is {rec_str}")
        elif rec and "sell" in rec.lower():
            bear_points.append(f"Analyst consensus is {rec_str}")

        # Valuation
        if cd.trailing_pe and cd.trailing_pe < 15:
            bull_points.append(f"Trading at value multiple ({cd.trailing_pe:.0f}x P/E)")
        elif cd.trailing_pe and cd.trailing_pe > 40:
            bear_points.append(f"Expensive valuation ({cd.trailing_pe:.0f}x P/E) — priced for perfection")

        # Market position
        if cd.market_cap and cd.market_cap > 100e9:
            bull_points.append("Market leader with significant scale advantages")

        # Dividend
        if cd.dividend_yield and cd.dividend_yield > 0.02:
            _dy = cd.dividend_yield * 100 if cd.dividend_yield < 0.2 else cd.dividend_yield
            bull_points.append(f"Attractive dividend yield ({_dy:.1f}%)")

        if bull_points or bear_points:
            bull_html = "".join(
                f'<div style="padding:0.25rem 0; font-size:0.82rem; color:#B8B3D7; line-height:1.6;">'
                f'<span style="color:#10B981;">▲</span> {p}</div>'
                for p in bull_points[:4]
            ) or '<div style="font-size:0.8rem; color:#8A85AD;">No strong bull signals identified</div>'

            bear_html = "".join(
                f'<div style="padding:0.25rem 0; font-size:0.82rem; color:#B8B3D7; line-height:1.6;">'
                f'<span style="color:#EF4444;">▼</span> {p}</div>'
                for p in bear_points[:4]
            ) or '<div style="font-size:0.8rem; color:#8A85AD;">No significant bear signals identified</div>'

            st.markdown(
                f'<div style="display:grid; grid-template-columns:1fr 1fr; gap:1rem; margin:0.5rem 0 1rem 0;">'
                f'<div style="background:rgba(16,185,129,0.04); border:1px solid rgba(16,185,129,0.15); '
                f'border-radius:12px; padding:1rem 1.2rem;">'
                f'<div style="font-size:0.65rem; font-weight:700; color:#10B981; text-transform:uppercase; '
                f'letter-spacing:1.5px; margin-bottom:0.5rem;">🐂 Bull Case</div>'
                f'{bull_html}</div>'
                f'<div style="background:rgba(239,68,68,0.04); border:1px solid rgba(239,68,68,0.15); '
                f'border-radius:12px; padding:1rem 1.2rem;">'
                f'<div style="font-size:0.65rem; font-weight:700; color:#EF4444; text-transform:uppercase; '
                f'letter-spacing:1.5px; margin-bottom:0.5rem;">🐻 Bear Case</div>'
                f'{bear_html}</div>'
                f'</div>',
                unsafe_allow_html=True,
            )

    # ══════════════════════════════════════════════════════
    # 3. BUSINESS OVERVIEW
    # ══════════════════════════════════════════════════════
    _section("Business Overview")
    with st.expander("Company Description", expanded=True):
        if cd.long_business_summary:
            st.markdown(f"<div style='line-height:1.7; color:#B8B3D7; font-size:0.9rem;'>{cd.long_business_summary}</div>", unsafe_allow_html=True)
        else:
            st.info("Business description not available.")
        b1, b2, b3 = st.columns(3)
        with b1:
            emp_val = f"{cd.full_time_employees:,}" if cd.full_time_employees else "N/A"
            st.markdown(f'<div style="background:rgba(255,255,255,0.05); border:1px solid rgba(255,255,255,0.1); border-radius:10px; padding:0.6rem 0.8rem; text-align:center;"><div style="font-size:0.65rem; font-weight:600; text-transform:uppercase; letter-spacing:0.7px; color:#8A85AD; margin-bottom:0.2rem;">Employees</div><div style="font-size:1rem; font-weight:700; color:#E0DCF5;">{emp_val}</div></div>', unsafe_allow_html=True)
        with b2:
            hq = f"{cd.city}, {cd.state}" if cd.city else "N/A"
            if cd.country and cd.country != "United States":
                hq += f", {cd.country}"
            st.markdown(f'<div style="background:rgba(255,255,255,0.05); border:1px solid rgba(255,255,255,0.1); border-radius:10px; padding:0.6rem 0.8rem; text-align:center;"><div style="font-size:0.65rem; font-weight:600; text-transform:uppercase; letter-spacing:0.7px; color:#8A85AD; margin-bottom:0.2rem;">Headquarters</div><div style="font-size:1rem; font-weight:700; color:#E0DCF5;">{hq}</div></div>', unsafe_allow_html=True)
        with b3:
            web_display = cd.website.replace("https://", "").replace("http://", "").rstrip("/") if cd.website else "N/A"
            st.markdown(f'<div style="background:rgba(255,255,255,0.05); border:1px solid rgba(255,255,255,0.1); border-radius:10px; padding:0.6rem 0.8rem; text-align:center;"><div style="font-size:0.65rem; font-weight:600; text-transform:uppercase; letter-spacing:0.7px; color:#8A85AD; margin-bottom:0.2rem;">Website</div><div style="font-size:1rem; font-weight:700; color:#E0DCF5;">{web_display}</div></div>', unsafe_allow_html=True)

    _divider()

    # ══════════════════════════════════════════════════════
    # 3b. REVENUE SEGMENTATION (if available)
    # ══════════════════════════════════════════════════════
    with _safe_section("Revenue Segmentation"):
        _seg_data_found = False
        try:
            _tk_seg = yf.Ticker(cd.ticker)
            # yfinance may expose revenue by segment/geography via get_financials or similar
            _seg_rev = getattr(_tk_seg, 'revenue_estimate', None)
            # Try newer yfinance API for segment data
            _seg_by_product = None
            _seg_by_geo = None
            try:
                _seg_by_product = _tk_seg.get_revenue_estimate() if hasattr(_tk_seg, 'get_revenue_estimate') else None
            except Exception:
                pass
            # Check for quarterly financials that might have segment info
            # yfinance doesn't reliably provide segment data, so we check what's available
            if _seg_by_product is not None and hasattr(_seg_by_product, 'empty') and not _seg_by_product.empty:
                _seg_data_found = True
                _section("Revenue Segmentation", "📊")
                st.dataframe(_seg_by_product, use_container_width=True)
        except Exception:
            pass

        if not _seg_data_found:
            # Silently skip — don't show anything if no segment data
            pass

    _divider()

    # ══════════════════════════════════════════════════════
    # 4. PRICE CHART
    # ══════════════════════════════════════════════════════
    _section("Price History")

    period_choice = st.radio("Period", ["1Y", "3Y", "5Y"], horizontal=True, index=2, label_visibility="collapsed")

    hist = cd.hist_5y if cd.hist_5y is not None and not cd.hist_5y.empty else cd.hist_1y
    if hist is not None and not hist.empty:
        if period_choice == "1Y" and cd.hist_1y is not None:
            plot_hist = cd.hist_1y
        elif period_choice == "3Y" and hist is not None:
            plot_hist = hist.last("3Y") if hasattr(hist, "last") else hist.tail(756)
        else:
            plot_hist = hist

        fig = go.Figure()
        # Glow underlay + main price line
        _glow_line_traces(fig, plot_hist.index, plot_hist["Close"], "#6B5CE7", "Close")
        # Area fill
        fig.add_trace(go.Scatter(
            x=plot_hist.index, y=plot_hist["Close"],
            mode="lines", line=dict(width=0), fill="tozeroy",
            fillcolor="rgba(107,92,231,0.06)",
            showlegend=False, hoverinfo="skip",
        ))
        # Color-coded volume bars
        if "Volume" in plot_hist.columns:
            close_vals = plot_hist["Close"].values
            vol_colors = []
            for i in range(len(close_vals)):
                if i == 0:
                    vol_colors.append("rgba(107,92,231,0.15)")
                elif close_vals[i] >= close_vals[i - 1]:
                    vol_colors.append("rgba(107,92,231,0.15)")
                else:
                    vol_colors.append("rgba(232,99,139,0.12)")
            fig.add_trace(go.Bar(
                x=plot_hist.index, y=plot_hist["Volume"],
                name="Volume", yaxis="y2",
                marker_color=vol_colors,
            ))
            fig.update_layout(
                yaxis2=dict(overlaying="y", side="right", showgrid=False,
                            title=dict(text="Volume", font=dict(size=10, color="#8A85AD")),
                            tickformat=".2s", tickfont=dict(size=8, color="#8A85AD")),
            )
        # Moving Average Overlays
        show_ma = st.checkbox("Show Moving Averages", value=True, key="show_ma")
        if show_ma and len(plot_hist) > 50:
            ma_50 = plot_hist["Close"].rolling(50).mean()
            ma_200 = plot_hist["Close"].rolling(200).mean()
            
            fig.add_trace(go.Scatter(
                x=plot_hist.index, y=ma_50, mode="lines",
                line=dict(color="#F59E0B", width=1.5, dash="dot"),
                name="50-day MA", showlegend=True,
            ))
            if len(plot_hist) > 200:
                fig.add_trace(go.Scatter(
                    x=plot_hist.index, y=ma_200, mode="lines",
                    line=dict(color="#E8638B", width=1.5, dash="dash"),
                    name="200-day MA", showlegend=True,
                ))
                
                # Detect Golden/Death Cross
                if not ma_50.dropna().empty and not ma_200.dropna().empty:
                    recent_50 = ma_50.dropna().iloc[-1]
                    recent_200 = ma_200.dropna().iloc[-1]
                    prev_50 = ma_50.dropna().iloc[-2] if len(ma_50.dropna()) > 1 else recent_50
                    prev_200 = ma_200.dropna().iloc[-2] if len(ma_200.dropna()) > 1 else recent_200
                    
                    if prev_50 <= prev_200 and recent_50 > recent_200:
                        st.markdown(
                            '<div style="text-align:center; padding:0.4rem; background:rgba(16,185,129,0.1); '
                            'border-radius:8px; border:1px solid rgba(16,185,129,0.3); margin-bottom:0.5rem;">'
                            '<span style="font-size:0.8rem; font-weight:700; color:#10B981;">✨ Golden Cross Detected — Bullish Signal</span>'
                            '</div>',
                            unsafe_allow_html=True,
                        )
                    elif prev_50 >= prev_200 and recent_50 < recent_200:
                        st.markdown(
                            '<div style="text-align:center; padding:0.4rem; background:rgba(239,68,68,0.1); '
                            'border-radius:8px; border:1px solid rgba(239,68,68,0.3); margin-bottom:0.5rem;">'
                            '<span style="font-size:0.8rem; font-weight:700; color:#EF4444;">💀 Death Cross Detected — Bearish Signal</span>'
                            '</div>',
                            unsafe_allow_html=True,
                        )
        
        # Fibonacci Retracement Levels
        show_fib = st.checkbox("Show Fibonacci Retracement", value=False, key="show_fib")
        if show_fib and len(plot_hist) > 10:
            fib_high = plot_hist["Close"].max()
            fib_low = plot_hist["Close"].min()
            fib_diff = fib_high - fib_low
            fib_levels = [0, 0.236, 0.382, 0.5, 0.618, 0.786, 1.0]
            fib_colors = ["#EF4444", "#F59E0B", "#F5A623", "#8A85AD", "#10B981", "#3B82F6", "#6B5CE7"]
            for lvl, clr in zip(fib_levels, fib_colors):
                fib_price = fib_high - fib_diff * lvl
                fig.add_hline(
                    y=fib_price, line_dash="dot", line_color=clr, line_width=1,
                    annotation_text=f"Fib {lvl:.1%} ({cs}{fib_price:.2f})",
                    annotation_position="bottom right",
                    annotation_font=dict(size=7, color=clr),
                )

        # Volume Profile (horizontal volume bars)
        show_vol_profile = st.checkbox("Show Volume Profile", value=False, key="show_vol_profile")
        if show_vol_profile and "Volume" in plot_hist.columns and len(plot_hist) > 10:
            price_min = plot_hist["Close"].min()
            price_max = plot_hist["Close"].max()
            n_bins = 30
            bin_edges = np.linspace(price_min, price_max, n_bins + 1)
            vol_profile = np.zeros(n_bins)
            for i in range(n_bins):
                mask = (plot_hist["Close"] >= bin_edges[i]) & (plot_hist["Close"] < bin_edges[i + 1])
                vol_profile[i] = plot_hist.loc[mask, "Volume"].sum()
            # Normalize to fit on right side (max = 15% of date range)
            if vol_profile.max() > 0:
                vol_norm = vol_profile / vol_profile.max()
                x_start = plot_hist.index[-1]
                x_range_days = (plot_hist.index[-1] - plot_hist.index[0]).days
                for i in range(n_bins):
                    bin_mid = (bin_edges[i] + bin_edges[i + 1]) / 2
                    bar_width = (bin_edges[1] - bin_edges[0]) * 0.8
                    opacity = 0.15 + 0.35 * vol_norm[i]
                    fig.add_shape(
                        type="rect",
                        x0=x_start, x1=x_start + pd.Timedelta(days=int(x_range_days * 0.12 * vol_norm[i])),
                        y0=bin_mid - bar_width / 2, y1=bin_mid + bar_width / 2,
                        fillcolor=f"rgba(107,92,231,{opacity:.2f})",
                        line=dict(width=0),
                        layer="below",
                    )

        # 52-week high/low reference lines
        if cd.fifty_two_week_high:
            fig.add_hline(y=cd.fifty_two_week_high, line_dash="dash",
                         line_color="rgba(16,185,129,0.3)", line_width=1,
                         annotation_text="52w High", annotation_position="bottom right",
                         annotation_font=dict(size=8, color="#10B981"))
        if cd.fifty_two_week_low:
            fig.add_hline(y=cd.fifty_two_week_low, line_dash="dash",
                         line_color="rgba(239,68,68,0.3)", line_width=1,
                         annotation_text="52w Low", annotation_position="top right",
                         annotation_font=dict(size=8, color="#EF4444"))
        fig.update_layout(
            **_CHART_LAYOUT_BASE,
            height=550,
            margin=dict(t=20, b=40, l=60, r=60),
            xaxis=dict(showgrid=False, tickfont=dict(size=12, color="#8A85AD"), rangeslider=dict(visible=False)),
            yaxis=dict(
                title=dict(text=f"Price ({cs})", font=dict(size=13, color="#8A85AD")),
                tickfont=dict(size=12, color="#8A85AD"),
                tickprefix=cs,
            ),
            showlegend=True,
            legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1,
                       font=dict(size=9, color="#8A85AD")),
        )
        _apply_space_grid(fig)
        st.markdown('<div class="profile-chart-wrapper">', unsafe_allow_html=True)
        st.plotly_chart(fig, use_container_width=True)
        st.markdown('</div>', unsafe_allow_html=True)
    else:
        st.info("Price history not available.")

    _divider()

    # ══════════════════════════════════════════════════════
    # 4b. TECHNICAL ANALYSIS
    # ══════════════════════════════════════════════════════
    _section("Technical Analysis")
    
    ta_hist = cd.hist_1y if cd.hist_1y is not None and not cd.hist_1y.empty else hist
    if ta_hist is not None and not ta_hist.empty and len(ta_hist) > 20:
        ta_tab1, ta_tab2, ta_tab3 = st.tabs(["RSI", "MACD", "Bollinger Bands"])
        
        close = ta_hist["Close"]
        
        # RSI
        with ta_tab1:
            delta = close.diff()
            gain = delta.where(delta > 0, 0).rolling(14).mean()
            loss = (-delta.where(delta < 0, 0)).rolling(14).mean()
            rs = gain / loss
            rsi = 100 - (100 / (1 + rs))
            
            fig_rsi = go.Figure()
            fig_rsi.add_trace(go.Scatter(
                x=ta_hist.index, y=rsi, mode="lines",
                line=dict(color="#6B5CE7", width=2), name="RSI (14)"
            ))
            fig_rsi.add_hline(y=70, line_dash="dash", line_color="rgba(239,68,68,0.5)", line_width=1,
                             annotation_text="Overbought (70)", annotation_font=dict(size=9, color="#EF4444"))
            fig_rsi.add_hline(y=30, line_dash="dash", line_color="rgba(16,185,129,0.5)", line_width=1,
                             annotation_text="Oversold (30)", annotation_font=dict(size=9, color="#10B981"))
            fig_rsi.add_hrect(y0=70, y1=100, fillcolor="rgba(239,68,68,0.05)", line_width=0)
            fig_rsi.add_hrect(y0=0, y1=30, fillcolor="rgba(16,185,129,0.05)", line_width=0)
            fig_rsi.update_layout(
                **_CHART_LAYOUT_BASE, height=300,
                margin=dict(t=20, b=30, l=50, r=30),
                yaxis=dict(range=[0, 100], tickfont=dict(size=10, color="#8A85AD"),
                          title=dict(text="RSI", font=dict(size=11, color="#8A85AD"))),
                xaxis=dict(showgrid=False, tickfont=dict(size=10, color="#8A85AD")),
                showlegend=False,
            )
            _apply_space_grid(fig_rsi)
            
            # Current RSI value
            current_rsi = rsi.dropna().iloc[-1] if not rsi.dropna().empty else 50
            rsi_color = "#EF4444" if current_rsi > 70 else "#10B981" if current_rsi < 30 else "#F59E0B"
            rsi_label = "Overbought" if current_rsi > 70 else "Oversold" if current_rsi < 30 else "Neutral"
            st.markdown(
                f'<div style="text-align:center; margin-bottom:0.5rem;">'
                f'<span style="font-size:1.5rem; font-weight:800; color:{rsi_color};">{current_rsi:.1f}</span>'
                f'<span style="font-size:0.8rem; color:{rsi_color}; margin-left:0.5rem;">{rsi_label}</span>'
                f'</div>',
                unsafe_allow_html=True,
            )
            st.plotly_chart(fig_rsi, use_container_width=True, key="rsi_chart")
        
        # MACD
        with ta_tab2:
            ema_12 = close.ewm(span=12, adjust=False).mean()
            ema_26 = close.ewm(span=26, adjust=False).mean()
            macd_line = ema_12 - ema_26
            signal_line = macd_line.ewm(span=9, adjust=False).mean()
            macd_hist = macd_line - signal_line
            
            fig_macd = go.Figure()
            # Histogram bars
            colors = ["#10B981" if v >= 0 else "#EF4444" for v in macd_hist.values]
            fig_macd.add_trace(go.Bar(
                x=ta_hist.index, y=macd_hist, name="Histogram",
                marker_color=colors, opacity=0.5
            ))
            fig_macd.add_trace(go.Scatter(
                x=ta_hist.index, y=macd_line, mode="lines",
                line=dict(color="#6B5CE7", width=2), name="MACD"
            ))
            fig_macd.add_trace(go.Scatter(
                x=ta_hist.index, y=signal_line, mode="lines",
                line=dict(color="#E8638B", width=1.5), name="Signal"
            ))
            fig_macd.update_layout(
                **_CHART_LAYOUT_BASE, height=300,
                margin=dict(t=20, b=30, l=50, r=30),
                yaxis=dict(tickfont=dict(size=10, color="#8A85AD"),
                          title=dict(text="MACD", font=dict(size=11, color="#8A85AD"))),
                xaxis=dict(showgrid=False, tickfont=dict(size=10, color="#8A85AD")),
                legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1,
                           font=dict(size=9, color="#8A85AD")),
            )
            _apply_space_grid(fig_macd)
            
            # Signal
            current_macd = macd_line.iloc[-1]
            current_signal = signal_line.iloc[-1]
            macd_verdict = "Bullish" if current_macd > current_signal else "Bearish"
            mv_color = "#10B981" if current_macd > current_signal else "#EF4444"
            st.markdown(
                f'<div style="text-align:center; margin-bottom:0.5rem;">'
                f'<span style="font-size:0.85rem; font-weight:700; color:{mv_color};">Signal: {macd_verdict}</span>'
                f'<span style="font-size:0.7rem; color:#8A85AD; margin-left:0.5rem;">'
                f'MACD: {current_macd:.3f} | Signal: {current_signal:.3f}</span>'
                f'</div>',
                unsafe_allow_html=True,
            )
            st.plotly_chart(fig_macd, use_container_width=True, key="macd_chart")
        
        # Bollinger Bands
        with ta_tab3:
            sma_20 = close.rolling(20).mean()
            std_20 = close.rolling(20).std()
            upper_band = sma_20 + (std_20 * 2)
            lower_band = sma_20 - (std_20 * 2)
            
            fig_bb = go.Figure()
            fig_bb.add_trace(go.Scatter(
                x=ta_hist.index, y=upper_band, mode="lines",
                line=dict(color="rgba(107,92,231,0.4)", width=1), name="Upper Band",
            ))
            fig_bb.add_trace(go.Scatter(
                x=ta_hist.index, y=lower_band, mode="lines",
                line=dict(color="rgba(107,92,231,0.4)", width=1), name="Lower Band",
                fill="tonexty", fillcolor="rgba(107,92,231,0.05)",
            ))
            fig_bb.add_trace(go.Scatter(
                x=ta_hist.index, y=sma_20, mode="lines",
                line=dict(color="#F59E0B", width=1.5, dash="dash"), name="SMA 20",
            ))
            fig_bb.add_trace(go.Scatter(
                x=ta_hist.index, y=close, mode="lines",
                line=dict(color="#E0DCF5", width=2), name="Price",
            ))
            fig_bb.update_layout(
                **_CHART_LAYOUT_BASE, height=400,
                margin=dict(t=20, b=30, l=50, r=30),
                yaxis=dict(tickfont=dict(size=10, color="#8A85AD"), tickprefix=cs,
                          title=dict(text=f"Price ({cs})", font=dict(size=11, color="#8A85AD"))),
                xaxis=dict(showgrid=False, tickfont=dict(size=10, color="#8A85AD")),
                legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1,
                           font=dict(size=9, color="#8A85AD")),
            )
            _apply_space_grid(fig_bb)
            
            # Band position
            if not sma_20.dropna().empty and not upper_band.dropna().empty and not lower_band.dropna().empty:
                curr_price = close.iloc[-1]
                curr_upper = upper_band.dropna().iloc[-1]
                curr_lower = lower_band.dropna().iloc[-1]
                band_width = curr_upper - curr_lower
                band_pct = ((curr_price - curr_lower) / band_width * 100) if band_width > 0 else 50
                bp_color = "#EF4444" if band_pct > 80 else "#10B981" if band_pct < 20 else "#F59E0B"
                st.markdown(
                    f'<div style="text-align:center; margin-bottom:0.5rem;">'
                    f'<span style="font-size:0.85rem; font-weight:700; color:{bp_color};">'
                    f'Band Position: {band_pct:.0f}%</span>'
                    f'<span style="font-size:0.7rem; color:#8A85AD; margin-left:0.5rem;">'
                    f'(0%=Lower, 100%=Upper)</span>'
                    f'</div>',
                    unsafe_allow_html=True,
                )
            st.plotly_chart(fig_bb, use_container_width=True, key="bb_chart")
    else:
        st.info("Insufficient data for technical analysis.")

    _divider()

    # ══════════════════════════════════════════════════════
    # 4c. INSTITUTIONAL & INSIDER SUMMARY
    # ══════════════════════════════════════════════════════
    _section("Ownership Overview")

    with _safe_section("Ownership Overview"):
        try:
            tk_own = yf.Ticker(cd.ticker)
        except Exception:
            tk_own = None

        own_col1, own_col2 = st.columns(2)

        with own_col1:
            # Institutional holders — enhanced
            try:
                inst_holders = tk_own.institutional_holders if tk_own else None
                if inst_holders is not None and not inst_holders.empty:
                    st.markdown(
                        '<div style="font-size:0.8rem; font-weight:700; color:#9B8AFF; margin-bottom:0.5rem;">'
                        '🏛️ Top 10 Institutional Holders</div>',
                        unsafe_allow_html=True,
                    )

                    # Table display
                    for _, row in inst_holders.head(10).iterrows():
                        holder = row.get("Holder", "Unknown")
                        shares = row.get("Shares", 0)
                        value = row.get("Value", 0)
                        pct = row.get("% Out", row.get("pctHeld", 0))
                        pct_str = f"{pct:.2%}" if isinstance(pct, float) and pct < 1 else f"{pct}"
                        date_rep = row.get("Date Reported", "")
                        date_str = date_rep.strftime("%Y-%m-%d") if hasattr(date_rep, 'strftime') else str(date_rep)[:10] if date_rep else ""
                        val_str = f"${value/1e6:,.0f}M" if value and value >= 1e6 else f"${value:,.0f}" if value else "—"
                        st.markdown(
                            f'<div style="display:flex; justify-content:space-between; padding:0.3rem 0; '
                            f'border-bottom:1px solid rgba(255,255,255,0.04); font-size:0.72rem;">'
                            f'<span style="color:#E0DCF5; flex:2.5;">{holder}</span>'
                            f'<span style="color:#8A85AD; flex:1; text-align:right;">{shares:,.0f}</span>'
                            f'<span style="color:#9B8AFF; flex:0.8; text-align:right;">{val_str}</span>'
                            f'<span style="color:#6B5CE7; flex:0.5; text-align:right; font-weight:600;">{pct_str}</span>'
                            f'</div>',
                            unsafe_allow_html=True,
                        )

                    # Horizontal bar chart of top 10 by % ownership
                    _top10 = inst_holders.head(10).copy()
                    _pct_col = "% Out" if "% Out" in _top10.columns else "pctHeld" if "pctHeld" in _top10.columns else None
                    if _pct_col and _top10[_pct_col].notna().any():
                        fig_inst = go.Figure()
                        _holders = _top10["Holder"].tolist()[::-1]
                        _pcts = _top10[_pct_col].fillna(0).tolist()[::-1]
                        _pcts_display = [p * 100 if p < 1 else p for p in _pcts]
                        fig_inst.add_trace(go.Bar(
                            y=_holders, x=_pcts_display, orientation="h",
                            marker_color="rgba(107,92,231,0.7)",
                            text=[f"{p:.1f}%" for p in _pcts_display],
                            textposition="outside",
                            textfont=dict(size=9, color="#B8B3D7"),
                        ))
                        fig_inst.update_layout(
                            **_CHART_LAYOUT_BASE, height=300,
                            margin=dict(t=10, b=20, l=160, r=50),
                            xaxis=dict(title=dict(text="% Outstanding", font=dict(size=9, color="#8A85AD")),
                                      ticksuffix="%", tickfont=dict(size=8, color="#8A85AD"), showgrid=False),
                            yaxis=dict(tickfont=dict(size=8, color="#8A85AD")),
                            showlegend=False,
                        )
                        _apply_space_grid(fig_inst)
                        st.plotly_chart(fig_inst, use_container_width=True, key="inst_holders_bar")

                    # Total institutional ownership
                    try:
                        major = tk_own.major_holders if tk_own else None
                        if major is not None and not major.empty:
                            for _, row in major.iterrows():
                                val = row.iloc[0] if len(row) > 0 else ""
                                label = str(row.iloc[1]) if len(row) > 1 else ""
                                if "institution" in label.lower() and "hold" in label.lower():
                                    st.markdown(
                                        f'<div style="text-align:center; padding:0.5rem; margin-top:0.5rem; '
                                        f'background:rgba(107,92,231,0.08); border-radius:8px;">'
                                        f'<span style="font-size:0.7rem; color:#8A85AD;">Total Institutional Ownership: </span>'
                                        f'<span style="font-size:0.9rem; font-weight:700; color:#6B5CE7;">{val}</span>'
                                        f'</div>',
                                        unsafe_allow_html=True,
                                    )
                                    break
                    except Exception:
                        pass
                else:
                    st.info("No institutional holder data available.")
            except Exception:
                st.info("Could not fetch institutional holders.")

        with own_col2:
            # Major holders summary + Mutual Fund holders + Donut
            try:
                major = tk_own.major_holders if tk_own else None
                if major is not None and not major.empty:
                    st.markdown(
                        '<div style="font-size:0.8rem; font-weight:700; color:#9B8AFF; margin-bottom:0.5rem;">'
                        '📊 Ownership Breakdown</div>',
                        unsafe_allow_html=True,
                    )
                    for _, row in major.iterrows():
                        val = row.iloc[0] if len(row) > 0 else ""
                        label = row.iloc[1] if len(row) > 1 else ""
                        st.markdown(
                            f'<div style="display:flex; justify-content:space-between; padding:0.4rem 0; '
                            f'border-bottom:1px solid rgba(255,255,255,0.04);">'
                            f'<span style="color:#8A85AD; font-size:0.75rem;">{label}</span>'
                            f'<span style="color:#E0DCF5; font-weight:700; font-size:0.85rem;">{val}</span>'
                            f'</div>',
                            unsafe_allow_html=True,
                        )
                else:
                    st.info("No major holder data available.")
            except Exception:
                st.info("Could not fetch major holders.")

            # Mutual Fund Holders
            try:
                mf_holders = tk_own.mutualfund_holders if tk_own else None
                if mf_holders is not None and not mf_holders.empty:
                    st.markdown(
                        '<div style="font-size:0.8rem; font-weight:700; color:#9B8AFF; margin-top:1rem; margin-bottom:0.5rem;">'
                        '💼 Top 5 Mutual Fund Holders</div>',
                        unsafe_allow_html=True,
                    )
                    for _, row in mf_holders.head(5).iterrows():
                        holder = row.get("Holder", "Unknown")
                        pct = row.get("% Out", row.get("pctHeld", 0))
                        pct_str = f"{pct:.2%}" if isinstance(pct, float) and pct < 1 else f"{pct}"
                        st.markdown(
                            f'<div style="display:flex; justify-content:space-between; padding:0.25rem 0; '
                            f'border-bottom:1px solid rgba(255,255,255,0.04); font-size:0.72rem;">'
                            f'<span style="color:#E0DCF5; flex:3;">{holder}</span>'
                            f'<span style="color:#6B5CE7; flex:0.5; text-align:right; font-weight:600;">{pct_str}</span>'
                            f'</div>',
                            unsafe_allow_html=True,
                        )
            except Exception:
                pass

            # Ownership Concentration Donut
            try:
                inst_holders = tk_own.institutional_holders if tk_own else None
                if inst_holders is not None and not inst_holders.empty:
                    _pct_col = "% Out" if "% Out" in inst_holders.columns else "pctHeld" if "pctHeld" in inst_holders.columns else None
                    if _pct_col:
                        _top5_pct = inst_holders.head(5)[_pct_col].fillna(0).tolist()
                        _top5_pct = [p * 100 if p < 1 else p for p in _top5_pct]
                        _top5_names = inst_holders.head(5)["Holder"].tolist()
                        _total_inst = inst_holders[_pct_col].fillna(0).sum()
                        _total_inst = _total_inst * 100 if _total_inst < 1 else _total_inst
                        _rest = max(0, _total_inst - sum(_top5_pct))

                        _labels = [n[:25] for n in _top5_names] + ["Other Institutions"]
                        _values = _top5_pct + [_rest]
                        _colors_donut = ["#6B5CE7", "#E8638B", "#10B981", "#F5A623", "#3B82F6", "#8A85AD"]

                        fig_donut = go.Figure(data=[go.Pie(
                            labels=_labels, values=_values,
                            hole=0.55, marker=dict(colors=_colors_donut),
                            textinfo="percent", textfont=dict(size=9, color="#E0DCF5"),
                            hovertemplate="%{label}<br>%{value:.1f}%<extra></extra>",
                        )])
                        fig_donut.update_layout(
                            **_CHART_LAYOUT_BASE, height=250,
                            margin=dict(t=10, b=10, l=10, r=10),
                            showlegend=True,
                            legend=dict(font=dict(size=8, color="#8A85AD"), orientation="h",
                                       yanchor="bottom", y=-0.2, xanchor="center", x=0.5),
                        )
                        st.markdown(
                            '<div style="font-size:0.75rem; font-weight:700; color:#9B8AFF; margin-top:1rem; margin-bottom:0.3rem;">'
                            '🍩 Ownership Concentration (Top 5)</div>',
                            unsafe_allow_html=True,
                        )
                        st.plotly_chart(fig_donut, use_container_width=True, key="ownership_donut")
            except Exception:
                pass

    _divider()

    # ══════════════════════════════════════════════════════
    # 5. VALUATION DASHBOARD
    # ══════════════════════════════════════════════════════
    _section("Valuation Dashboard")

    vd1, vd2, vd3, vd4, vd5 = st.columns(5)
    vd1.metric("P/E (TTM)", f"{cd.trailing_pe:.1f}x" if cd.trailing_pe else "N/A")
    vd2.metric("Forward P/E", f"{cd.forward_pe:.1f}x" if cd.forward_pe else "N/A")
    vd3.metric("EV/EBITDA", format_multiple(cd.ev_to_ebitda))
    vd4.metric("P/S (TTM)", f"{cd.price_to_sales:.1f}x" if cd.price_to_sales else "N/A")
    vd5.metric("PEG Ratio", f"{cd.peg_ratio:.2f}" if cd.peg_ratio else "N/A")

    # Premium/Discount vs Peers
    if cd.peer_data:
        st.markdown("<p style='font-size:0.75rem; font-weight:600; color:#8A85AD; text-transform:uppercase; letter-spacing:1px; margin:0.8rem 0 0.3rem 0;'>Premium / Discount vs. Peer Median</p>", unsafe_allow_html=True)

        def _calc_premium(company_val, peers, key):
            if company_val is None:
                return None
            vals = [p.get(key) for p in peers if p.get(key) is not None]
            if not vals:
                return None
            median = float(np.median(vals))
            if median == 0:
                return None
            return ((company_val - median) / abs(median)) * 100

        premium_items = [
            ("P/E", _calc_premium(cd.trailing_pe, cd.peer_data, "trailing_pe")),
            ("Fwd P/E", _calc_premium(cd.forward_pe, cd.peer_data, "forward_pe")),
            ("EV/EBITDA", _calc_premium(cd.ev_to_ebitda, cd.peer_data, "ev_to_ebitda")),
            ("P/S", _calc_premium(cd.price_to_sales, cd.peer_data, "price_to_sales")),
        ]

        pc_cols = st.columns(4)
        for col, (label, prem) in zip(pc_cols, premium_items):
            if prem is not None:
                word = "Premium" if prem > 0 else "Discount"
                col.metric(f"{label} vs Peers", f"{prem:+.1f}%", delta=word,
                           delta_color="inverse" if prem > 0 else "normal")
            else:
                col.metric(f"{label} vs Peers", "N/A")

    # ══════════════════════════════════════════════════════
    # 6. PEER COMPARISON
    # ══════════════════════════════════════════════════════
    if cd.peer_data:
        _section("Peer Comparison")

        peer_rows = []
        peer_rows.append({
            "Company": f"{cd.name} \u2605",
            "Ticker": cd.ticker,
            "Mkt Cap": format_number(cd.market_cap, currency_symbol=cs),
            "P/E": f"{cd.trailing_pe:.1f}" if cd.trailing_pe else "N/A",
            "Fwd P/E": f"{cd.forward_pe:.1f}" if cd.forward_pe else "N/A",
            "EV/EBITDA": format_multiple(cd.ev_to_ebitda),
            "P/S": f"{cd.price_to_sales:.1f}" if cd.price_to_sales else "N/A",
            "Gross Margin": format_pct(cd.gross_margins),
            "Op Margin": format_pct(cd.operating_margins),
            "ROE": format_pct(cd.return_on_equity),
        })
        for p in cd.peer_data:
            peer_rows.append({
                "Company": p.get("name", p.get("ticker", "")),
                "Ticker": p.get("ticker", ""),
                "Mkt Cap": format_number(p.get("market_cap"), currency_symbol=cs),
                "P/E": f"{p['trailing_pe']:.1f}" if p.get("trailing_pe") else "N/A",
                "Fwd P/E": f"{p['forward_pe']:.1f}" if p.get("forward_pe") else "N/A",
                "EV/EBITDA": format_multiple(p.get("ev_to_ebitda")),
                "P/S": f"{p['price_to_sales']:.1f}" if p.get("price_to_sales") else "N/A",
                "Gross Margin": format_pct(p.get("gross_margins")),
                "Op Margin": format_pct(p.get("operating_margins")),
                "ROE": format_pct(p.get("return_on_equity")),
            })

        peer_df = pd.DataFrame(peer_rows)

        def _highlight_target(row):
            if row["Ticker"] == cd.ticker:
                return ["background-color: rgba(107,92,231,0.1); font-weight: bold"] * len(row)
            return [""] * len(row)

        styled = peer_df.style.apply(_highlight_target, axis=1)
        st.dataframe(styled, use_container_width=True, hide_index=True, height=300)

        # Radar chart
        rc1, rc2 = st.columns([3, 2])
        with rc1:
            st.markdown('<div class="profile-chart-wrapper">', unsafe_allow_html=True)
            _build_peer_radar_chart(cd)
            st.markdown('</div>', unsafe_allow_html=True)
        with rc2:
            st.markdown("")
            st.markdown("<p style='font-size:0.85rem; font-weight:700; color:#E0DCF5; margin-bottom:0.5rem;'>Peer Group</p>", unsafe_allow_html=True)
            for p in cd.peer_data:
                st.markdown(
                    f"<div style='font-size:0.82rem; color:#B8B3D7; padding:0.2rem 0;'>"
                    f"<span style='font-weight:600; color:#9B8AFF;'>{p['ticker']}</span> &mdash; {p.get('name', '')}"
                    f"</div>",
                    unsafe_allow_html=True,
                )
            st.markdown(f"<div style='font-size:0.7rem; color:#8A85AD; margin-top:0.5rem;'>Industry: {cd.industry}</div>", unsafe_allow_html=True)

        # Percentile Ranking
        try:
            st.markdown(
                '<div style="font-size:0.85rem; font-weight:700; color:#E0DCF5; margin:1rem 0 0.5rem 0;">'
                '📊 Percentile Ranking vs Peers</div>',
                unsafe_allow_html=True,
            )
            
            ranking_metrics = [
                ("Market Cap", cd.market_cap, "market_cap", False),
                ("P/E Ratio", cd.trailing_pe, "trailing_pe", True),  # Lower is better
                ("EV/EBITDA", cd.ev_to_ebitda, "ev_to_ebitda", True),
                ("Gross Margin", cd.gross_margins, "gross_margins", False),
                ("Op Margin", cd.operating_margins, "operating_margins", False),
                ("ROE", cd.return_on_equity, "return_on_equity", False),
                ("Rev Growth", cd.revenue_growth, "revenue_growth", False),
            ]
            
            rank_html = '<div style="display:grid; gap:0.5rem;">'
            for label, company_val, key, lower_better in ranking_metrics:
                if company_val is None:
                    continue
                peer_vals = [p.get(key) for p in cd.peer_data if p.get(key) is not None]
                if not peer_vals:
                    continue
                
                all_vals = sorted(peer_vals + [company_val], reverse=not lower_better)
                rank = all_vals.index(company_val) + 1
                pctile = (1 - (rank - 1) / len(all_vals)) * 100
                
                bar_color = "#10B981" if pctile >= 70 else "#F59E0B" if pctile >= 40 else "#EF4444"
                
                rank_html += (
                    f'<div style="display:flex; align-items:center; gap:0.5rem;">'
                    f'<span style="font-size:0.7rem; color:#8A85AD; width:80px; flex-shrink:0;">{label}</span>'
                    f'<div style="flex:1; background:rgba(255,255,255,0.05); border-radius:4px; height:14px; overflow:hidden;">'
                    f'<div style="width:{pctile}%; height:100%; background:{bar_color}; border-radius:4px;"></div></div>'
                    f'<span style="font-size:0.65rem; color:{bar_color}; font-weight:700; width:40px; text-align:right;">'
                    f'{pctile:.0f}%</span>'
                    f'</div>'
                )
            
            rank_html += '</div>'
            st.markdown(rank_html, unsafe_allow_html=True)
            st.markdown(
                '<div style="font-size:0.6rem; color:#5A567A; margin-top:0.3rem;">Higher = better ranking among peers</div>',
                unsafe_allow_html=True,
            )
        except Exception:
            pass

    # ══════════════════════════════════════════════════════
    # 7. KEY STATISTICS
    # ══════════════════════════════════════════════════════
    _section("Key Statistics")

    st.markdown("<p style='font-size:0.75rem; font-weight:600; color:#8A85AD; text-transform:uppercase; letter-spacing:1px; margin:0.5rem 0 0.3rem 0;'>Valuation</p>", unsafe_allow_html=True)
    v1, v2, v3, v4, v5 = st.columns(5)
    v1.metric("P/E (TTM)", f"{cd.trailing_pe:.1f}" if cd.trailing_pe else "N/A")
    v2.metric("Forward P/E", f"{cd.forward_pe:.1f}" if cd.forward_pe else "N/A")
    v3.metric("PEG Ratio", f"{cd.peg_ratio:.2f}" if cd.peg_ratio else "N/A")
    v4.metric("EV/EBITDA", format_multiple(cd.ev_to_ebitda))
    v5.metric("EV/Revenue", format_multiple(cd.ev_to_revenue))

    st.markdown("<p style='font-size:0.75rem; font-weight:600; color:#8A85AD; text-transform:uppercase; letter-spacing:1px; margin:0.8rem 0 0.3rem 0;'>Profitability</p>", unsafe_allow_html=True)
    p1, p2, p3, p4, p5 = st.columns(5)
    p1.metric("Gross Margin", format_pct(cd.gross_margins))
    p2.metric("Op. Margin", format_pct(cd.operating_margins))
    p3.metric("Net Margin", format_pct(cd.profit_margins))
    p4.metric("ROE", format_pct(cd.return_on_equity))
    p5.metric("ROA", format_pct(cd.return_on_assets))

    st.markdown("<p style='font-size:0.75rem; font-weight:600; color:#8A85AD; text-transform:uppercase; letter-spacing:1px; margin:0.8rem 0 0.3rem 0;'>Financial Health</p>", unsafe_allow_html=True)
    f1, f2, f3, f4, f5 = st.columns(5)
    f1.metric("P/S (TTM)", f"{cd.price_to_sales:.2f}" if cd.price_to_sales else "N/A")
    f2.metric("Price/Book", f"{cd.price_to_book:.2f}" if cd.price_to_book else "N/A")
    f3.metric("Current Ratio", f"{cd.current_ratio:.2f}" if cd.current_ratio else "N/A")
    f4.metric("D/E Ratio", f"{cd.debt_to_equity / 100:.2f}x" if cd.debt_to_equity else "N/A")
    f5.metric("Beta", f"{cd.beta:.2f}" if cd.beta else "N/A")

    _divider()

    # ══════════════════════════════════════════════════════
    # 8. FINANCIAL STATEMENTS (formatted)
    # ══════════════════════════════════════════════════════
    _section("Financial Statements")

    # Financial Summary Trend Table (4 years)
    with _safe_section("Financial Summary Trend"):
        _years_avail = min(4, len(cd.revenue)) if cd.revenue is not None and len(cd.revenue) > 0 else 0
        if _years_avail >= 2:
            _trend_rows = []
            _trend_cols = []

            for i in range(_years_avail):
                yr_label = str(cd.revenue.index[i])[:4] if hasattr(cd.revenue.index[i], 'year') else str(cd.revenue.index[i])[:4]
                _trend_cols.append(yr_label)

            def _extract_trend(series, fmt_fn=None, prefix=""):
                vals = []
                if series is None or len(series) == 0:
                    return ["—"] * _years_avail
                for i in range(_years_avail):
                    if i < len(series):
                        v = float(series.iloc[i])
                        if fmt_fn:
                            vals.append(fmt_fn(v))
                        elif prefix:
                            vals.append(f"{prefix}{v/1e9:.1f}B" if abs(v) >= 1e9 else f"{prefix}{v/1e6:.0f}M")
                        else:
                            vals.append(f"{v:.1f}")
                    else:
                        vals.append("—")
                return vals

            _trend_data = {
                "Revenue": _extract_trend(cd.revenue, prefix=cs),
                "Gross Profit": _extract_trend(cd.gross_profit, prefix=cs),
                "Operating Income": _extract_trend(cd.operating_income, prefix=cs),
                "Net Income": _extract_trend(cd.net_income, prefix=cs),
                "Free Cash Flow": _extract_trend(cd.free_cashflow_series, prefix=cs),
                "Gross Margin": _extract_trend(cd.gross_margin_series, fmt_fn=lambda v: f"{v*100:.1f}%" if abs(v) < 1 else f"{v:.1f}%"),
                "Operating Margin": _extract_trend(cd.operating_margin_series, fmt_fn=lambda v: f"{v*100:.1f}%" if abs(v) < 1 else f"{v:.1f}%"),
                "Net Margin": _extract_trend(cd.net_margin_series, fmt_fn=lambda v: f"{v*100:.1f}%" if abs(v) < 1 else f"{v:.1f}%"),
            }

            _trend_cols.reverse()
            for k in _trend_data:
                _trend_data[k].reverse()

            # Build styled HTML table
            _th_html = "".join(f'<th style="padding:0.4rem 0.6rem; font-weight:700; color:#6B5CE7; font-size:0.75rem; '
                               f'text-align:right; border-bottom:2px solid rgba(107,92,231,0.3);">{yr}</th>'
                               for yr in _trend_cols)

            _tbody_html = ""
            for metric, vals in _trend_data.items():
                _cells = ""
                for j, v in enumerate(vals):
                    # Color negative values red
                    _neg = v.startswith("-") or v.startswith(f"{cs}-")
                    _c = "#EF4444" if _neg else "#E0DCF5"
                    _fw = "600"
                    _cells += f'<td style="padding:0.35rem 0.6rem; text-align:right; font-size:0.78rem; color:{_c}; font-weight:{_fw};">{v}</td>'
                _tbody_html += (
                    f'<tr style="border-bottom:1px solid rgba(255,255,255,0.04);">'
                    f'<td style="padding:0.35rem 0.6rem; font-size:0.78rem; color:#B8B3D7; font-weight:600;">{metric}</td>'
                    f'{_cells}</tr>'
                )

            st.markdown(
                f'<div style="background:rgba(255,255,255,0.02); border:1px solid rgba(107,92,231,0.1); '
                f'border-radius:10px; overflow:hidden; margin-bottom:1rem;">'
                f'<div style="padding:0.6rem 0.8rem; background:rgba(107,92,231,0.06); '
                f'border-bottom:1px solid rgba(107,92,231,0.1);">'
                f'<span style="font-size:0.7rem; font-weight:700; color:#6B5CE7; text-transform:uppercase; '
                f'letter-spacing:1px;">📊 Financial Summary (Annual)</span></div>'
                f'<table style="width:100%; border-collapse:collapse;">'
                f'<thead><tr><th style="padding:0.4rem 0.6rem; text-align:left; font-size:0.7rem; '
                f'color:#8A85AD; border-bottom:2px solid rgba(107,92,231,0.3);">Metric</th>{_th_html}</tr></thead>'
                f'<tbody>{_tbody_html}</tbody></table></div>',
                unsafe_allow_html=True,
            )

    def _display_financial_df(df, label, quarterly=False):
        if df is not None and not df.empty:
            display_df = df.copy()
            fmt = "%b %Y" if quarterly else "%Y"
            new_cols = []
            for c in display_df.columns:
                col_str = c.strftime(fmt) if hasattr(c, "strftime") else str(c)
                base, n = col_str, 1
                while col_str in new_cols:
                    n += 1
                    col_str = f"{base} ({n})"
                new_cols.append(col_str)
            display_df.columns = new_cols

            # Format numeric values
            def _fmt_cell(val):
                if pd.isna(val):
                    return "N/A"
                try:
                    v = float(val)
                    abs_v = abs(v)
                    sign = "-" if v < 0 else ""
                    if abs_v >= 1e9:
                        return f"{sign}{cs}{abs_v / 1e9:.1f}B"
                    elif abs_v >= 1e6:
                        return f"{sign}{cs}{abs_v / 1e6:.1f}M"
                    elif abs_v >= 1e3:
                        return f"{sign}{cs}{abs_v / 1e3:.1f}K"
                    elif abs_v == 0:
                        return f"{cs}0"
                    else:
                        return f"{sign}{cs}{abs_v:,.2f}"
                except (TypeError, ValueError):
                    return str(val)

            formatted_df = display_df.map(_fmt_cell)
            st.dataframe(formatted_df, use_container_width=True, height=400)
        else:
            st.info(f"{label} not available.")

    fin_tab1, fin_tab2, fin_tab3, fin_tab4 = st.tabs([
        "Income Statement", "Balance Sheet", "Cash Flow", "Quarterly Income"
    ])
    with fin_tab1:
        _display_financial_df(cd.income_stmt, "Income Statement")
    with fin_tab2:
        _display_financial_df(cd.balance_sheet, "Balance Sheet")
    with fin_tab3:
        _display_financial_df(cd.cashflow, "Cash Flow Statement")
    with fin_tab4:
        _display_financial_df(cd.quarterly_income_stmt, "Quarterly Income Statement", quarterly=True)

    # Revenue & Income Growth Trend
    if cd.income_stmt is not None and not cd.income_stmt.empty:
        try:
            is_df = cd.income_stmt
            rev_row = None
            ni_row = None
            for idx_name in is_df.index:
                if "total revenue" in str(idx_name).lower() or "revenue" == str(idx_name).lower().strip():
                    rev_row = is_df.loc[idx_name]
                elif "net income" in str(idx_name).lower():
                    ni_row = is_df.loc[idx_name]
            
            if rev_row is not None and len(rev_row) > 1:
                fig_growth = go.Figure()
                
                years = [c.strftime("%Y") if hasattr(c, "strftime") else str(c) for c in rev_row.index]
                rev_vals = [float(v) / 1e9 if pd.notna(v) else 0 for v in rev_row.values]
                
                fig_growth.add_trace(go.Bar(
                    x=years, y=rev_vals, name="Revenue",
                    marker_color="rgba(107,92,231,0.6)",
                    text=[f"${v:.1f}B" for v in rev_vals],
                    textposition="outside", textfont=dict(size=9, color="#B8B3D7"),
                ))
                
                if ni_row is not None:
                    ni_vals = [float(v) / 1e9 if pd.notna(v) else 0 for v in ni_row.values]
                    fig_growth.add_trace(go.Bar(
                        x=years, y=ni_vals, name="Net Income",
                        marker_color="rgba(16,185,129,0.5)",
                        text=[f"${v:.1f}B" for v in ni_vals],
                        textposition="outside", textfont=dict(size=9, color="#B8B3D7"),
                    ))
                
                # Add YoY growth rate line for revenue
                yoy_growth = []
                for i in range(len(rev_vals)):
                    if i < len(rev_vals) - 1 and rev_vals[i+1] != 0:
                        growth = (rev_vals[i] / rev_vals[i+1] - 1) * 100
                        yoy_growth.append(growth)
                    else:
                        yoy_growth.append(None)
                
                fig_growth.add_trace(go.Scatter(
                    x=years, y=yoy_growth, name="YoY Growth %",
                    mode="lines+markers", yaxis="y2",
                    line=dict(color="#F59E0B", width=2),
                    marker=dict(size=8, color="#F59E0B"),
                ))
                
                fig_growth.update_layout(
                    **_CHART_LAYOUT_BASE, height=350,
                    margin=dict(t=30, b=30, l=60, r=60),
                    xaxis=dict(tickfont=dict(size=10, color="#8A85AD"), showgrid=False),
                    yaxis=dict(tickfont=dict(size=9, color="#8A85AD"), title=dict(text="$ Billions", font=dict(size=10, color="#8A85AD"))),
                    yaxis2=dict(overlaying="y", side="right", ticksuffix="%", showgrid=False,
                               tickfont=dict(size=9, color="#F59E0B"),
                               title=dict(text="YoY Growth", font=dict(size=10, color="#F59E0B"))),
                    legend=dict(orientation="h", yanchor="bottom", y=1.02, font=dict(size=9, color="#B8B3D7")),
                    barmode="group",
                )
                _apply_space_grid(fig_growth)
                st.plotly_chart(fig_growth, use_container_width=True, key="revenue_growth_trend")
        except Exception:
            pass

    _divider()

    # ══════════════════════════════════════════════════════
    # 8b. CHART STUDIO
    # ══════════════════════════════════════════════════════
    with st.expander("📊 Chart Studio — Advanced Visualizations", expanded=False):
        with _safe_section("Chart Studio"):
            cs_tab1, cs_tab2, cs_tab3 = st.tabs(["Correlation Matrix", "Margins Waterfall", "Cash Flow Bridge"])

            with cs_tab1:
                # Correlation matrix heatmap of key financial metrics over time
                try:
                    corr_series = {}
                    if cd.revenue is not None and len(cd.revenue) >= 2:
                        corr_series["Revenue"] = cd.revenue.values[::-1].astype(float)
                    if cd.gross_profit is not None and len(cd.gross_profit) >= 2:
                        corr_series["Gross Profit"] = cd.gross_profit.values[::-1].astype(float)
                    if cd.operating_income is not None and len(cd.operating_income) >= 2:
                        corr_series["Op Income"] = cd.operating_income.values[::-1].astype(float)
                    if cd.net_income is not None and len(cd.net_income) >= 2:
                        corr_series["Net Income"] = cd.net_income.values[::-1].astype(float)
                    if cd.free_cashflow_series is not None and len(cd.free_cashflow_series) >= 2:
                        corr_series["FCF"] = cd.free_cashflow_series.values[::-1].astype(float)

                    if len(corr_series) >= 3:
                        min_len = min(len(v) for v in corr_series.values())
                        corr_dict = {k: v[:min_len] for k, v in corr_series.items()}
                        corr_df = pd.DataFrame(corr_dict)
                        corr_mat = corr_df.corr()

                        fig_cm = go.Figure(data=go.Heatmap(
                            z=corr_mat.values,
                            x=corr_mat.columns.tolist(),
                            y=corr_mat.index.tolist(),
                            colorscale=[[0, "#EF4444"], [0.5, "#1a1625"], [1, "#10B981"]],
                            zmin=-1, zmax=1,
                            text=np.round(corr_mat.values, 2),
                            texttemplate="%{text}",
                            textfont=dict(size=11, color="#E0DCF5"),
                        ))
                        fig_cm.update_layout(
                            **_CHART_LAYOUT_BASE, height=400,
                            margin=dict(t=30, b=30, l=80, r=30),
                            xaxis=dict(tickfont=dict(size=10, color="#8A85AD")),
                            yaxis=dict(tickfont=dict(size=10, color="#8A85AD")),
                        )
                        st.plotly_chart(fig_cm, use_container_width=True, key="cs_corr_matrix")
                    else:
                        st.info("Not enough financial series for correlation analysis.")
                except Exception as e:
                    st.info(f"Correlation matrix not available: {str(e)[:80]}")

            with cs_tab2:
                # Margins waterfall: Revenue → Gross Profit → EBITDA → Net Income
                try:
                    rev_latest = float(cd.revenue.iloc[0]) if cd.revenue is not None and len(cd.revenue) > 0 else 0
                    gp_latest = float(cd.gross_profit.iloc[0]) if cd.gross_profit is not None and len(cd.gross_profit) > 0 else 0
                    oi_latest = float(cd.operating_income.iloc[0]) if cd.operating_income is not None and len(cd.operating_income) > 0 else 0
                    ni_latest = float(cd.net_income.iloc[0]) if cd.net_income is not None and len(cd.net_income) > 0 else 0

                    if rev_latest > 0:
                        cogs = rev_latest - gp_latest
                        opex = gp_latest - oi_latest
                        below_line = oi_latest - ni_latest

                        fig_mw = go.Figure(go.Waterfall(
                            x=["Revenue", "COGS", "Gross Profit", "OpEx", "Operating Income", "Tax/Int/Other", "Net Income"],
                            y=[rev_latest, -cogs, 0, -opex, 0, -below_line, 0],
                            measure=["absolute", "relative", "total", "relative", "total", "relative", "total"],
                            text=[format_number(v, currency_symbol=cs) for v in [rev_latest, -cogs, gp_latest, -opex, oi_latest, -below_line, ni_latest]],
                            textposition="outside",
                            textfont=dict(size=9, color="#B8B3D7"),
                            connector=dict(line=dict(color="rgba(107,92,231,0.2)", width=1, dash="dot")),
                            increasing=dict(marker=dict(color="#10B981")),
                            decreasing=dict(marker=dict(color="#EF4444")),
                            totals=dict(marker=dict(color="#6B5CE7")),
                        ))
                        fig_mw.update_layout(
                            **_CHART_LAYOUT_BASE, height=400,
                            margin=dict(t=30, b=40, l=60, r=30),
                            xaxis=dict(tickfont=dict(size=9, color="#8A85AD"), showgrid=False),
                            yaxis=dict(tickfont=dict(size=9, color="#8A85AD")),
                        )
                        _apply_space_grid(fig_mw)
                        st.plotly_chart(fig_mw, use_container_width=True, key="cs_margins_waterfall")
                    else:
                        st.info("Revenue data not available for margins waterfall.")
                except Exception as e:
                    st.info(f"Margins waterfall not available: {str(e)[:80]}")

            with cs_tab3:
                # Cash flow bridge: Operating CF → CapEx → FCF → Dividends → Net Cash Change
                try:
                    if cd.cashflow is not None and not cd.cashflow.empty:
                        cf = cd.cashflow
                        def _cf_val(names):
                            for n in names:
                                for idx in cf.index:
                                    if n.lower() in str(idx).lower():
                                        v = cf.loc[idx].iloc[0]
                                        return float(v) if pd.notna(v) else 0
                            return 0

                        op_cf = _cf_val(["operating cash flow", "total cash from operating", "cash flow from operations"])
                        capex = _cf_val(["capital expenditure", "capital expenditures"])
                        fcf_val = op_cf + capex  # capex is typically negative
                        dividends = _cf_val(["dividends paid", "cash dividends paid"])
                        buybacks = _cf_val(["repurchase", "buyback", "stock repurchase"])
                        net_change = fcf_val + dividends + buybacks

                        if op_cf != 0:
                            bridge_labels = ["Operating CF", "CapEx", "Free Cash Flow", "Dividends", "Buybacks", "Net Cash"]
                            bridge_values = [op_cf, capex, 0, dividends, buybacks, 0]
                            bridge_measures = ["absolute", "relative", "total", "relative", "relative", "total"]
                            bridge_display = [op_cf, capex, fcf_val, dividends, buybacks, net_change]

                            fig_cfb = go.Figure(go.Waterfall(
                                x=bridge_labels,
                                y=bridge_values,
                                measure=bridge_measures,
                                text=[format_number(v, currency_symbol=cs) for v in bridge_display],
                                textposition="outside",
                                textfont=dict(size=9, color="#B8B3D7"),
                                connector=dict(line=dict(color="rgba(107,92,231,0.2)", width=1, dash="dot")),
                                increasing=dict(marker=dict(color="#10B981")),
                                decreasing=dict(marker=dict(color="#EF4444")),
                                totals=dict(marker=dict(color="#6B5CE7")),
                            ))
                            fig_cfb.update_layout(
                                **_CHART_LAYOUT_BASE, height=400,
                                margin=dict(t=30, b=40, l=60, r=30),
                                xaxis=dict(tickfont=dict(size=9, color="#8A85AD"), showgrid=False),
                                yaxis=dict(tickfont=dict(size=9, color="#8A85AD")),
                            )
                            _apply_space_grid(fig_cfb)
                            st.plotly_chart(fig_cfb, use_container_width=True, key="cs_cashflow_bridge")
                        else:
                            st.info("Operating cash flow data not available.")
                    else:
                        st.info("Cash flow statement not available.")
                except Exception as e:
                    st.info(f"Cash flow bridge not available: {str(e)[:80]}")

    _divider()

    # ══════════════════════════════════════════════════════
    # 9. ANALYST CONSENSUS
    # ══════════════════════════════════════════════════════
    _section("Analyst Consensus")
    a1, a2 = st.columns([3, 2])

    with a1:
        if cd.recommendations_summary is not None and not cd.recommendations_summary.empty:
            try:
                row = cd.recommendations_summary.iloc[0]
                cats = ["Strong Buy", "Buy", "Hold", "Sell", "Strong Sell"]
                keys = ["strongBuy", "buy", "hold", "sell", "strongSell"]
                vals = []
                for k in keys:
                    try:
                        v = row.get(k, 0) if hasattr(row, 'get') else row[k] if k in row.index else 0
                        vals.append(int(v) if pd.notna(v) else 0)
                    except Exception:
                        vals.append(0)
                colors = ["#10B981", "#34D399", "#F59E0B", "#EF4444", "#991B1B"]
                total = sum(vals)

                # Wider bar for the majority category
                max_idx = vals.index(max(vals)) if vals else -1
                widths = [0.7 if i == max_idx else 0.5 for i in range(len(vals))]
                fig_rec = go.Figure(go.Bar(
                    x=vals, y=cats, orientation="h",
                    marker=dict(color=colors, line=dict(color="rgba(255,255,255,0.15)", width=1)),
                    width=widths,
                    text=[f"  {v} ({v/total*100:.0f}%)" if total > 0 else f"  {v}" for v in vals],
                    textposition="outside",
                    textfont=dict(size=11, color="#B8B3D7", family="Inter"),
                ))
                fig_rec.update_layout(
                    **_CHART_LAYOUT_BASE,
                    height=400, margin=dict(t=50, b=30, l=130, r=70),
                    title=dict(text="Analyst Recommendation Distribution",
                               font=dict(size=16, color="#E0DCF5", family="Inter")),
                    xaxis=dict(title=dict(text="# Analysts", font=dict(size=13, color="#8A85AD")),
                               tickfont=dict(size=12, color="#8A85AD")),
                    yaxis=dict(autorange="reversed", tickfont=dict(size=13, color="#8A85AD")),
                    bargap=0.3,
                )
                _apply_space_grid(fig_rec, show_x_grid=True)
                st.markdown('<div class="profile-chart-wrapper">', unsafe_allow_html=True)
                st.plotly_chart(fig_rec, use_container_width=True)
                st.markdown('</div>', unsafe_allow_html=True)
            except Exception:
                st.info("Recommendation data not available.")
        else:
            st.info("Analyst recommendation data not available.")

    with a2:
        if cd.analyst_price_targets:
            pt = cd.analyst_price_targets
            st.markdown("<p style='font-size:0.85rem; font-weight:700; color:#E0DCF5; margin-bottom:0.5rem;'>Price Targets</p>", unsafe_allow_html=True)
            
            # Visual price target range
            pt_low = pt.get('low', 0) or 0
            pt_mean = pt.get('mean', 0) or 0
            pt_median = pt.get('median', 0) or 0
            pt_high = pt.get('high', 0) or 0
            curr = cd.current_price or 0
            
            if pt_low and pt_high and pt_high > pt_low:
                fig_pt = go.Figure()
                
                # Range bar (low to high)
                fig_pt.add_trace(go.Scatter(
                    x=[pt_low, pt_high], y=["Target", "Target"],
                    mode="lines", line=dict(color="rgba(107,92,231,0.4)", width=20),
                    showlegend=False, hoverinfo="skip",
                ))
                
                # Low marker
                fig_pt.add_trace(go.Scatter(
                    x=[pt_low], y=["Target"], mode="markers+text",
                    marker=dict(size=14, color="#EF4444", symbol="diamond"),
                    text=[f"{cs}{pt_low:,.0f}"], textposition="bottom center",
                    textfont=dict(size=10, color="#EF4444"),
                    name="Low", showlegend=False,
                ))
                
                # Mean marker
                fig_pt.add_trace(go.Scatter(
                    x=[pt_mean], y=["Target"], mode="markers+text",
                    marker=dict(size=18, color="#6B5CE7", symbol="star"),
                    text=[f"Mean: {cs}{pt_mean:,.0f}"], textposition="top center",
                    textfont=dict(size=11, color="#6B5CE7", weight="bold" if hasattr(dict, 'weight') else None),
                    name="Mean Target", showlegend=False,
                ))
                
                # High marker
                fig_pt.add_trace(go.Scatter(
                    x=[pt_high], y=["Target"], mode="markers+text",
                    marker=dict(size=14, color="#10B981", symbol="diamond"),
                    text=[f"{cs}{pt_high:,.0f}"], textposition="bottom center",
                    textfont=dict(size=10, color="#10B981"),
                    name="High", showlegend=False,
                ))
                
                # Current price line
                fig_pt.add_vline(x=curr, line_dash="dash", line_color="#F59E0B", line_width=2)
                fig_pt.add_annotation(
                    x=curr, y="Target", text=f"Current: {cs}{curr:,.0f}",
                    showarrow=True, arrowhead=2, arrowcolor="#F59E0B",
                    font=dict(size=10, color="#F59E0B"),
                    ax=0, ay=-40,
                )
                
                fig_pt.update_layout(
                    **_CHART_LAYOUT_BASE, height=150,
                    margin=dict(t=40, b=30, l=20, r=20),
                    xaxis=dict(tickprefix=cs, tickfont=dict(size=9, color="#8A85AD"), showgrid=False),
                    yaxis=dict(visible=False),
                    showlegend=False,
                )
                st.plotly_chart(fig_pt, use_container_width=True, key="price_target_range")
            
            # Metrics row
            pt1, pt2 = st.columns(2)
            pt1.metric("Mean", f"{cs}{pt_mean:,.2f}" if pt_mean else "N/A")
            pt2.metric("Median", f"{cs}{pt_median:,.2f}" if pt_median else "N/A")
            
            if pt_mean and curr:
                upside = (pt_mean - curr) / curr * 100
                color = "#10B981" if upside >= 0 else "#EF4444"
                st.markdown(
                    f'<div style="text-align:center; margin-top:0.5rem; padding:0.5rem; '
                    f'background:{"rgba(16,185,129,0.08)" if upside >= 0 else "rgba(239,68,68,0.08)"}; '
                    f'border-radius:10px;">'
                    f'<span style="font-size:0.75rem; color:#8A85AD; font-weight:600;">IMPLIED UPSIDE</span><br>'
                    f'<span style="font-size:1.3rem; font-weight:800; color:{color};">{upside:+.1f}%</span>'
                    f'</div>',
                    unsafe_allow_html=True,
                )
        else:
            st.info("Price target data not available.")

    # ══════════════════════════════════════════════════════
    # 10. EARNINGS HISTORY
    # ══════════════════════════════════════════════════════
    _section("Earnings History")
    if cd.earnings_dates is not None and not cd.earnings_dates.empty:
        st.dataframe(cd.earnings_dates.head(8), use_container_width=True)
    else:
        st.info("Earnings data not available.")

    _divider()

    # ══════════════════════════════════════════════════════
    # 10b. OPTIONS OVERVIEW
    # ══════════════════════════════════════════════════════
    _section("Options Overview")
    try:
        tk_opt = yf.Ticker(cd.ticker)
        exp_dates = tk_opt.options
        if exp_dates and len(exp_dates) > 0:
            # Show nearest expiry options summary
            nearest_exp = exp_dates[0]
            opt_chain = tk_opt.option_chain(nearest_exp)
            
            opt_col1, opt_col2 = st.columns(2)
            
            with opt_col1:
                calls = opt_chain.calls
                if not calls.empty:
                    total_call_vol = calls["volume"].sum() if "volume" in calls.columns else 0
                    total_call_oi = calls["openInterest"].sum() if "openInterest" in calls.columns else 0
                    st.markdown(
                        f'<div style="background:rgba(16,185,129,0.08); border:1px solid rgba(16,185,129,0.2); '
                        f'border-radius:12px; padding:1rem;">'
                        f'<div style="font-size:0.75rem; font-weight:700; color:#10B981; text-transform:uppercase; '
                        f'letter-spacing:1px; margin-bottom:0.5rem;">📈 Calls</div>'
                        f'<div style="display:flex; justify-content:space-between; padding:0.2rem 0;">'
                        f'<span style="color:#8A85AD; font-size:0.75rem;">Volume</span>'
                        f'<span style="color:#E0DCF5; font-weight:700;">{total_call_vol:,.0f}</span></div>'
                        f'<div style="display:flex; justify-content:space-between; padding:0.2rem 0;">'
                        f'<span style="color:#8A85AD; font-size:0.75rem;">Open Interest</span>'
                        f'<span style="color:#E0DCF5; font-weight:700;">{total_call_oi:,.0f}</span></div>'
                        f'<div style="display:flex; justify-content:space-between; padding:0.2rem 0;">'
                        f'<span style="color:#8A85AD; font-size:0.75rem;">Contracts</span>'
                        f'<span style="color:#E0DCF5; font-weight:700;">{len(calls)}</span></div>'
                        f'</div>',
                        unsafe_allow_html=True,
                    )
            
            with opt_col2:
                puts = opt_chain.puts
                if not puts.empty:
                    total_put_vol = puts["volume"].sum() if "volume" in puts.columns else 0
                    total_put_oi = puts["openInterest"].sum() if "openInterest" in puts.columns else 0
                    st.markdown(
                        f'<div style="background:rgba(239,68,68,0.08); border:1px solid rgba(239,68,68,0.2); '
                        f'border-radius:12px; padding:1rem;">'
                        f'<div style="font-size:0.75rem; font-weight:700; color:#EF4444; text-transform:uppercase; '
                        f'letter-spacing:1px; margin-bottom:0.5rem;">📉 Puts</div>'
                        f'<div style="display:flex; justify-content:space-between; padding:0.2rem 0;">'
                        f'<span style="color:#8A85AD; font-size:0.75rem;">Volume</span>'
                        f'<span style="color:#E0DCF5; font-weight:700;">{total_put_vol:,.0f}</span></div>'
                        f'<div style="display:flex; justify-content:space-between; padding:0.2rem 0;">'
                        f'<span style="color:#8A85AD; font-size:0.75rem;">Open Interest</span>'
                        f'<span style="color:#E0DCF5; font-weight:700;">{total_put_oi:,.0f}</span></div>'
                        f'<div style="display:flex; justify-content:space-between; padding:0.2rem 0;">'
                        f'<span style="color:#8A85AD; font-size:0.75rem;">Contracts</span>'
                        f'<span style="color:#E0DCF5; font-weight:700;">{len(puts)}</span></div>'
                        f'</div>',
                        unsafe_allow_html=True,
                    )
            
            # Put/Call Ratio
            if total_call_vol and total_call_vol > 0:
                pcr = total_put_vol / total_call_vol
                pcr_color = "#EF4444" if pcr > 1.2 else "#10B981" if pcr < 0.7 else "#F59E0B"
                pcr_label = "Bearish Sentiment" if pcr > 1.2 else "Bullish Sentiment" if pcr < 0.7 else "Neutral"
                st.markdown(
                    f'<div style="text-align:center; margin-top:0.8rem; padding:0.6rem; '
                    f'background:rgba(107,92,231,0.05); border-radius:10px;">'
                    f'<span style="font-size:0.7rem; color:#8A85AD; font-weight:600;">PUT/CALL RATIO</span><br>'
                    f'<span style="font-size:1.5rem; font-weight:800; color:{pcr_color};">{pcr:.2f}</span>'
                    f'<span style="font-size:0.75rem; color:{pcr_color}; margin-left:0.5rem;">{pcr_label}</span>'
                    f'<div style="font-size:0.6rem; color:#8A85AD; margin-top:0.2rem;">Expiry: {nearest_exp}</div>'
                    f'</div>',
                    unsafe_allow_html=True,
                )
            
            # Available expiration dates
            st.markdown(
                f'<div style="font-size:0.7rem; color:#8A85AD; margin-top:0.5rem; text-align:center;">'
                f'{len(exp_dates)} expiration dates available: {", ".join(exp_dates[:6])}'
                f'{"..." if len(exp_dates) > 6 else ""}</div>',
                unsafe_allow_html=True,
            )
        else:
            st.info("Options data not available for this ticker.")
    except Exception:
        st.info("Options data not available for this ticker.")

    _divider()

    # ══════════════════════════════════════════════════════
    # 10c. DIVIDEND ANALYSIS
    # ══════════════════════════════════════════════════════
    if cd.dividend_yield and cd.dividend_yield > 0:
        _section("Dividend Analysis")
        
        div_col1, div_col2, div_col3, div_col4 = st.columns(4)
        
        with div_col1:
            # yfinance may return dividendYield as decimal (0.009) or pct-like (0.9)
            dy = cd.dividend_yield * 100 if cd.dividend_yield < 0.2 else cd.dividend_yield
            dy_color = "#10B981" if dy > 3 else "#F59E0B" if dy > 1 else "#8A85AD"
            st.markdown(
                f'<div style="text-align:center; padding:0.8rem; background:rgba(107,92,231,0.05); '
                f'border-radius:12px; border:1px solid rgba(107,92,231,0.1);">'
                f'<div style="font-size:0.65rem; color:#8A85AD; font-weight:600; text-transform:uppercase;">Dividend Yield</div>'
                f'<div style="font-size:1.4rem; font-weight:800; color:{dy_color};">{dy:.2f}%</div>'
                f'</div>',
                unsafe_allow_html=True,
            )
        
        with div_col2:
            try:
                tk_div = yf.Ticker(cd.ticker)
                info_div = tk_div.info or {}
                payout = info_div.get("payoutRatio", 0)
                payout_pct = payout * 100 if payout and payout < 5 else payout or 0
                po_color = "#EF4444" if payout_pct > 80 else "#10B981" if payout_pct < 60 else "#F59E0B"
                st.markdown(
                    f'<div style="text-align:center; padding:0.8rem; background:rgba(107,92,231,0.05); '
                    f'border-radius:12px; border:1px solid rgba(107,92,231,0.1);">'
                    f'<div style="font-size:0.65rem; color:#8A85AD; font-weight:600; text-transform:uppercase;">Payout Ratio</div>'
                    f'<div style="font-size:1.4rem; font-weight:800; color:{po_color};">{payout_pct:.0f}%</div>'
                    f'</div>',
                    unsafe_allow_html=True,
                )
            except Exception:
                st.markdown(
                    '<div style="text-align:center; padding:0.8rem; background:rgba(107,92,231,0.05); '
                    'border-radius:12px;"><div style="color:#8A85AD;">Payout N/A</div></div>',
                    unsafe_allow_html=True,
                )
        
        with div_col3:
            try:
                fwd_div = info_div.get("dividendRate", 0) or 0
                st.markdown(
                    f'<div style="text-align:center; padding:0.8rem; background:rgba(107,92,231,0.05); '
                    f'border-radius:12px; border:1px solid rgba(107,92,231,0.1);">'
                    f'<div style="font-size:0.65rem; color:#8A85AD; font-weight:600; text-transform:uppercase;">Annual Dividend</div>'
                    f'<div style="font-size:1.4rem; font-weight:800; color:#E0DCF5;">{cs}{fwd_div:.2f}</div>'
                    f'</div>',
                    unsafe_allow_html=True,
                )
            except Exception:
                pass
        
        with div_col4:
            try:
                ex_date = info_div.get("exDividendDate")
                if ex_date:
                    from datetime import date as date_type
                    ex_dt = datetime.fromtimestamp(ex_date) if isinstance(ex_date, (int, float)) else ex_date
                    ex_str = ex_dt.strftime("%b %d, %Y") if hasattr(ex_dt, 'strftime') else str(ex_dt)
                else:
                    ex_str = "N/A"
                st.markdown(
                    f'<div style="text-align:center; padding:0.8rem; background:rgba(107,92,231,0.05); '
                    f'border-radius:12px; border:1px solid rgba(107,92,231,0.1);">'
                    f'<div style="font-size:0.65rem; color:#8A85AD; font-weight:600; text-transform:uppercase;">Ex-Dividend Date</div>'
                    f'<div style="font-size:0.95rem; font-weight:700; color:#E0DCF5; margin-top:0.2rem;">{ex_str}</div>'
                    f'</div>',
                    unsafe_allow_html=True,
                )
            except Exception:
                pass
        
        # Dividend history chart
        try:
            divs = tk_div.dividends
            if divs is not None and not divs.empty:
                # Last 5 years
                divs_recent = divs.last("5Y") if hasattr(divs, "last") else divs.tail(20)
                if len(divs_recent) > 2:
                    fig_div = go.Figure()
                    fig_div.add_trace(go.Bar(
                        x=divs_recent.index, y=divs_recent.values,
                        marker_color="rgba(107,92,231,0.6)",
                        name="Dividend",
                    ))
                    fig_div.update_layout(
                        **_CHART_LAYOUT_BASE, height=250,
                        margin=dict(t=20, b=30, l=50, r=30),
                        yaxis=dict(tickprefix=cs, tickfont=dict(size=10, color="#8A85AD"),
                                  title=dict(text="Dividend/Share", font=dict(size=11, color="#8A85AD"))),
                        xaxis=dict(showgrid=False, tickfont=dict(size=10, color="#8A85AD")),
                        showlegend=False,
                    )
                    _apply_space_grid(fig_div)
                    st.plotly_chart(fig_div, use_container_width=True, key="dividend_history_chart")
        except Exception:
            pass
        
        _divider()

    # ══════════════════════════════════════════════════════
    # 10d. FINANCIAL HEALTH SCORECARD
    # ══════════════════════════════════════════════════════
    _section("Financial Health Scorecard")
    
    # Helper to safely extract scalar from Series or scalar
    def _safe_val(v):
        if v is None:
            return None
        if hasattr(v, 'iloc'):
            try:
                return float(v.iloc[0]) if len(v) > 0 else None
            except Exception:
                return None
        try:
            return float(v)
        except (TypeError, ValueError):
            return None
    
    _ni = _safe_val(getattr(cd, 'net_income', None))
    _ocf = _safe_val(getattr(cd, 'operating_cashflow_series', getattr(cd, 'operating_cash_flow', None)))
    _fcf = _safe_val(getattr(cd, 'free_cashflow_series', getattr(cd, 'free_cash_flow', None)))
    _roa = _safe_val(getattr(cd, 'return_on_assets', getattr(cd, 'roa', None)))
    _cr = _safe_val(getattr(cd, 'current_ratio', None))
    _gm = _safe_val(getattr(cd, 'gross_margins', getattr(cd, 'gross_margin', None)))
    _rg = _safe_val(getattr(cd, 'revenue_growth', None))
    
    # Calculate a simplified Piotroski-inspired score
    score_items = []
    total_score = 0
    
    # 1. Profitability: Net Income > 0
    if _ni is not None and _ni > 0:
        score_items.append(("Net Income Positive", True, "Profitability"))
        total_score += 1
    else:
        score_items.append(("Net Income Positive", False, "Profitability"))
    
    # 2. ROA > 0
    if _roa is not None and _roa > 0:
        score_items.append(("ROA Positive", True, "Profitability"))
        total_score += 1
    else:
        score_items.append(("ROA Positive", False, "Profitability"))
    
    # 3. Operating Cash Flow > 0
    if _ocf is not None and _ocf > 0:
        score_items.append(("Operating Cash Flow Positive", True, "Profitability"))
        total_score += 1
    else:
        score_items.append(("Operating Cash Flow Positive", False, "Profitability"))
    
    # 4. Cash Flow > Net Income (quality of earnings)
    if _ocf is not None and _ni is not None and _ocf > _ni:
        score_items.append(("Cash Flow > Net Income", True, "Profitability"))
        total_score += 1
    else:
        score_items.append(("Cash Flow > Net Income", False, "Profitability"))
    
    # 5. Current Ratio > 1
    if _cr is not None and _cr > 1:
        score_items.append(("Current Ratio > 1", True, "Leverage"))
        total_score += 1
    else:
        score_items.append(("Current Ratio > 1", False, "Leverage"))
    
    # 6. Gross Margin Positive
    if _gm is not None and _gm > 0:
        score_items.append(("Gross Margin Positive", True, "Efficiency"))
        total_score += 1
    else:
        score_items.append(("Gross Margin Positive", False, "Efficiency"))
    
    # 7. Revenue Growth
    if _rg is not None and _rg > 0:
        score_items.append(("Revenue Growing", True, "Efficiency"))
        total_score += 1
    else:
        score_items.append(("Revenue Growing", False, "Efficiency"))
    
    # 8. Positive Free Cash Flow
    if _fcf is not None and _fcf > 0:
        score_items.append(("Free Cash Flow Positive", True, "Profitability"))
        total_score += 1
    else:
        score_items.append(("Free Cash Flow Positive", False, "Profitability"))
    
    max_score = len(score_items) if score_items else 8
    score_pct = (total_score / max_score * 100) if max_score > 0 else 0
    
    if score_pct >= 75:
        grade = "A"
        grade_color = "#10B981"
        grade_label = "Strong"
    elif score_pct >= 50:
        grade = "B"
        grade_color = "#34D399"
        grade_label = "Good"
    elif score_pct >= 25:
        grade = "C"
        grade_color = "#F59E0B"
        grade_label = "Fair"
    else:
        grade = "D"
        grade_color = "#EF4444"
        grade_label = "Weak"
    
    # Score display
    sc_col1, sc_col2 = st.columns([1, 2])
    
    with sc_col1:
        st.markdown(
            f'<div style="text-align:center; padding:1.5rem; background:rgba(107,92,231,0.05); '
            f'border-radius:16px; border:1px solid rgba(107,92,231,0.15);">'
            f'<div style="font-size:3rem; font-weight:900; color:{grade_color};">{grade}</div>'
            f'<div style="font-size:0.85rem; font-weight:700; color:{grade_color};">{grade_label}</div>'
            f'<div style="font-size:0.7rem; color:#8A85AD; margin-top:0.3rem;">{total_score}/{max_score} criteria met</div>'
            f'<div style="margin-top:0.8rem; background:rgba(255,255,255,0.05); border-radius:8px; '
            f'height:8px; overflow:hidden;">'
            f'<div style="width:{score_pct}%; height:100%; background:{grade_color}; border-radius:8px;"></div>'
            f'</div>'
            f'</div>',
            unsafe_allow_html=True,
        )
    
    with sc_col2:
        for item_name, passed, category in score_items:
            icon = "✅" if passed else "❌"
            st.markdown(
                f'<div style="display:flex; align-items:center; gap:0.5rem; padding:0.25rem 0; '
                f'border-bottom:1px solid rgba(255,255,255,0.03);">'
                f'<span style="font-size:0.8rem;">{icon}</span>'
                f'<span style="color:#E0DCF5; font-size:0.78rem; flex:1;">{item_name}</span>'
                f'<span style="color:#8A85AD; font-size:0.65rem; font-weight:600;">{category}</span>'
                f'</div>',
                unsafe_allow_html=True,
            )
    
    # ── Piotroski F-Score (from data_engine) ──
    try:
        pio_result = calculate_piotroski_score(cd)
        if pio_result and pio_result.get("score") is not None:
            pio_score = pio_result["score"]
            pio_max = pio_result.get("max_score", 9)
            if pio_score >= 7:
                pio_color, pio_label = "#10B981", "Strong"
            elif pio_score >= 4:
                pio_color, pio_label = "#F59E0B", "Moderate"
            else:
                pio_color, pio_label = "#EF4444", "Weak"
            st.markdown(
                f'<div style="text-align:center; padding:0.8rem; margin-top:0.5rem; '
                f'background:rgba(107,92,231,0.05); border-radius:12px; '
                f'border:1px solid rgba(107,92,231,0.15);">'
                f'<span style="font-size:0.75rem; color:#8A85AD;">Piotroski F-Score</span><br>'
                f'<span style="font-size:1.8rem; font-weight:900; color:{pio_color};">{pio_score}</span>'
                f'<span style="font-size:0.8rem; color:#8A85AD;">/{pio_max}</span>'
                f'<span style="font-size:0.75rem; color:{pio_color}; margin-left:0.5rem;">{pio_label}</span>'
                f'</div>',
                unsafe_allow_html=True,
            )
    except Exception:
        pass

    _divider()

    # ══════════════════════════════════════════════════════
    # 10d-ii. ALTMAN Z-SCORE
    # ══════════════════════════════════════════════════════
    with _safe_section("Altman Z-Score"):
        _section("Altman Z-Score", "🔬")
        
        _z_ta = _safe_val(getattr(cd, 'total_assets', None))
        _z_tl = _safe_val(getattr(cd, 'total_liabilities', None))
        _z_te = _safe_val(getattr(cd, 'total_equity', None))
        _z_rev = _safe_val(getattr(cd, 'revenue', None))
        _z_oi = _safe_val(getattr(cd, 'operating_income', None))
        _z_mcap = cd.market_cap or 0
        _z_td = _safe_val(getattr(cd, 'total_debt', None))
        _z_cash = _safe_val(getattr(cd, 'cash_and_equivalents', None))
        
        # Working Capital = Current Assets - Current Liabilities
        # Approximate: WC = (Total Assets - non-current) but simpler: TA - TL is net worth, not WC
        # Better: WC ≈ (Cash + current_ratio proxy). Use balance sheet if available.
        _z_ca = None
        _z_cl = None
        if cd.balance_sheet is not None and not cd.balance_sheet.empty:
            bs = cd.balance_sheet
            _z_ca = float(bs.loc["Current Assets"].iloc[0]) if "Current Assets" in bs.index else None
            _z_cl = float(bs.loc["Current Liabilities"].iloc[0]) if "Current Liabilities" in bs.index else None
            _z_re = float(bs.loc["Retained Earnings"].iloc[0]) if "Retained Earnings" in bs.index else None
        else:
            _z_re = None
        
        # Fallback for WC if no current assets/liabilities
        if _z_ca is not None and _z_cl is not None:
            _z_wc = _z_ca - _z_cl
        elif cd.current_ratio and _z_tl:
            # rough approximation
            _z_wc = None
        else:
            _z_wc = None
        
        if _z_re is None:
            _z_re = (_z_te or 0) * 0.6 if _z_te else None  # rough fallback
        
        if _z_ta and _z_ta > 0 and _z_tl and _z_tl > 0 and _z_wc is not None:
            _z_a = (_z_wc / _z_ta) if _z_ta else 0
            _z_b = ((_z_re or 0) / _z_ta)
            _z_c = ((_z_oi or 0) / _z_ta)
            _z_d = (_z_mcap / _z_tl) if _z_tl > 0 else 0
            _z_e = ((_z_rev or 0) / _z_ta)
            
            z_score = 1.2 * _z_a + 1.4 * _z_b + 3.3 * _z_c + 0.6 * _z_d + 1.0 * _z_e
            
            if z_score > 2.99:
                z_zone = "Safe Zone"
                z_color = "#10B981"
                z_bg = "rgba(16,185,129,0.1)"
            elif z_score >= 1.81:
                z_zone = "Grey Zone"
                z_color = "#F59E0B"
                z_bg = "rgba(245,158,11,0.1)"
            else:
                z_zone = "Distress Zone"
                z_color = "#EF4444"
                z_bg = "rgba(239,68,68,0.1)"
            
            # Gauge visualization
            z_gauge = go.Figure(go.Indicator(
                mode="gauge+number",
                value=z_score,
                number=dict(font=dict(size=36, color="#E0DCF5"), suffix=""),
                title=dict(text=z_zone, font=dict(size=16, color=z_color)),
                gauge=dict(
                    axis=dict(range=[0, 5], tickfont=dict(size=10, color="#8A85AD"),
                              tickcolor="rgba(255,255,255,0.1)"),
                    bar=dict(color=z_color, thickness=0.3),
                    bgcolor="rgba(0,0,0,0)",
                    bordercolor="rgba(107,92,231,0.2)",
                    steps=[
                        dict(range=[0, 1.81], color="rgba(239,68,68,0.15)"),
                        dict(range=[1.81, 2.99], color="rgba(245,158,11,0.15)"),
                        dict(range=[2.99, 5], color="rgba(16,185,129,0.15)"),
                    ],
                    threshold=dict(line=dict(color="#E0DCF5", width=2), thickness=0.8, value=z_score),
                ),
            ))
            z_gauge.update_layout(
                paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
                font=dict(family="Inter", color="#B8B3D7"),
                height=280, margin=dict(t=60, b=20, l=40, r=40),
            )
            
            z_col1, z_col2 = st.columns([1, 1])
            with z_col1:
                st.plotly_chart(z_gauge, use_container_width=True, key="altman_z_gauge")
            
            with z_col2:
                # Component breakdown
                z_components = [
                    ("1.2 × WC/TA", 1.2 * _z_a, "Working Capital / Total Assets"),
                    ("1.4 × RE/TA", 1.4 * _z_b, "Retained Earnings / Total Assets"),
                    ("3.3 × EBIT/TA", 3.3 * _z_c, "Operating Income / Total Assets"),
                    ("0.6 × MVE/TL", 0.6 * _z_d, "Market Cap / Total Liabilities"),
                    ("1.0 × Sales/TA", 1.0 * _z_e, "Revenue / Total Assets"),
                ]
                
                st.markdown(
                    '<div style="font-size:0.75rem; font-weight:700; color:#9B8AFF; margin-bottom:0.5rem;">Component Breakdown</div>',
                    unsafe_allow_html=True,
                )
                for comp_label, comp_val, comp_desc in z_components:
                    comp_pct = (comp_val / z_score * 100) if z_score != 0 else 0
                    st.markdown(
                        f'<div style="display:flex; justify-content:space-between; align-items:center; '
                        f'padding:0.3rem 0; border-bottom:1px solid rgba(255,255,255,0.03); font-size:0.75rem;">'
                        f'<span style="color:#B8B3D7;">{comp_label}</span>'
                        f'<span style="color:#E0DCF5; font-weight:600;">{comp_val:.3f}</span>'
                        f'</div>',
                        unsafe_allow_html=True,
                    )
                
                st.markdown(
                    f'<div style="margin-top:0.5rem; padding:0.5rem; background:{z_bg}; border-radius:8px; text-align:center;">'
                    f'<span style="font-size:0.75rem; color:{z_color}; font-weight:700;">'
                    f'Z = {z_score:.2f} → {z_zone}</span><br>'
                    f'<span style="font-size:0.6rem; color:#8A85AD;">'
                    f'>2.99 Safe | 1.81-2.99 Grey | <1.81 Distress</span>'
                    f'</div>',
                    unsafe_allow_html=True,
                )
        else:
            st.info("Insufficient balance sheet data for Altman Z-Score calculation.")
    
    _divider()

    # ══════════════════════════════════════════════════════
    # 10d-iii. DUPONT ANALYSIS
    # ══════════════════════════════════════════════════════
    with _safe_section("DuPont Analysis"):
        _section("DuPont Analysis", "🔍")
        
        # Calculate DuPont components for available years
        _dp_years = []
        _dp_npm = []  # Net Profit Margin
        _dp_at = []   # Asset Turnover
        _dp_em = []   # Equity Multiplier
        _dp_tb = []   # Tax Burden (5-factor)
        _dp_ib = []   # Interest Burden (5-factor)
        _dp_roe3 = [] # 3-factor ROE
        _dp_roe5 = [] # 5-factor ROE
        
        _dp_rev_s = getattr(cd, 'revenue', None)
        _dp_ni_s = getattr(cd, 'net_income', None)
        _dp_ta_s = getattr(cd, 'total_assets', None)
        _dp_te_s = getattr(cd, 'total_equity', None)
        _dp_oi_s = getattr(cd, 'operating_income', None)
        _dp_ie_s = getattr(cd, 'interest_expense', None)
        _dp_tp_s = getattr(cd, 'tax_provision', None)
        
        _dp_n = min(4, *[len(s) for s in [_dp_rev_s, _dp_ni_s, _dp_ta_s, _dp_te_s] if s is not None and len(s) > 0]) if all(
            s is not None and len(s) > 0 for s in [_dp_rev_s, _dp_ni_s, _dp_ta_s, _dp_te_s]
        ) else 0
        
        for i in range(_dp_n):
            try:
                _rev = float(_dp_rev_s.iloc[i])
                _ni = float(_dp_ni_s.iloc[i])
                _ta = float(_dp_ta_s.iloc[i])
                _te = float(_dp_te_s.iloc[i])
                
                if _rev == 0 or _ta == 0 or _te == 0:
                    continue
                
                npm = _ni / _rev
                at = _rev / _ta
                em = _ta / _te
                roe3 = npm * at * em
                
                # Year label from index
                yr_label = str(_dp_rev_s.index[i])[:4] if hasattr(_dp_rev_s.index[i], 'year') else str(_dp_rev_s.index[i])[:4]
                
                _dp_years.append(yr_label)
                _dp_npm.append(npm)
                _dp_at.append(at)
                _dp_em.append(em)
                _dp_roe3.append(roe3)
                
                # 5-factor
                _oi = float(_dp_oi_s.iloc[i]) if _dp_oi_s is not None and len(_dp_oi_s) > i else None
                _ie = abs(float(_dp_ie_s.iloc[i])) if _dp_ie_s is not None and len(_dp_ie_s) > i else None
                _tp = abs(float(_dp_tp_s.iloc[i])) if _dp_tp_s is not None and len(_dp_tp_s) > i else None
                
                if _oi and _oi != 0:
                    _pretax = _oi - (_ie or 0)
                    _ib_val = _pretax / _oi if _oi != 0 else 1.0
                    _tb_val = _ni / _pretax if _pretax != 0 else 1.0
                else:
                    _ib_val = 1.0
                    _tb_val = 1.0
                
                _dp_tb.append(_tb_val)
                _dp_ib.append(_ib_val)
                _dp_roe5.append(_tb_val * _ib_val * ((_oi or 0) / _rev if _rev else 0) * at * em)
            except Exception:
                continue
        
        if _dp_years:
            _dp_years.reverse()
            _dp_npm.reverse()
            _dp_at.reverse()
            _dp_em.reverse()
            _dp_roe3.reverse()
            _dp_tb.reverse()
            _dp_ib.reverse()
            _dp_roe5.reverse()
            
            dp_tab1, dp_tab2 = st.tabs(["3-Factor DuPont", "5-Factor DuPont"])
            
            with dp_tab1:
                # 3-Factor: ROE = NPM × AT × EM
                st.markdown(
                    '<div style="font-size:0.8rem; color:#B8B3D7; margin-bottom:0.8rem;">'
                    '<b>ROE = Net Profit Margin × Asset Turnover × Equity Multiplier</b></div>',
                    unsafe_allow_html=True,
                )
                
                fig_dp3 = go.Figure()
                fig_dp3.add_trace(go.Bar(
                    x=_dp_years, y=[v * 100 for v in _dp_npm], name="Net Profit Margin %",
                    marker_color="rgba(107,92,231,0.7)",
                ))
                fig_dp3.add_trace(go.Bar(
                    x=_dp_years, y=[v * 100 for v in _dp_at], name="Asset Turnover × 100",
                    marker_color="rgba(16,185,129,0.7)",
                ))
                fig_dp3.add_trace(go.Bar(
                    x=_dp_years, y=[v * 100 for v in _dp_em], name="Equity Multiplier × 100",
                    marker_color="rgba(232,99,139,0.7)",
                ))
                fig_dp3.add_trace(go.Scatter(
                    x=_dp_years, y=[v * 100 for v in _dp_roe3], name="ROE %",
                    mode="lines+markers", line=dict(color="#F5A623", width=3),
                    marker=dict(size=8),
                ))
                fig_dp3.update_layout(
                    **_CHART_LAYOUT_BASE, height=350, barmode="group",
                    margin=dict(t=30, b=40, l=50, r=30),
                    legend=dict(font=dict(size=9, color="#B8B3D7"), orientation="h", yanchor="bottom", y=1.02),
                )
                _apply_space_grid(fig_dp3)
                st.plotly_chart(fig_dp3, use_container_width=True, key="dupont_3factor")
                
                # Current year metrics
                dp_mc1, dp_mc2, dp_mc3, dp_mc4 = st.columns(4)
                dp_mc1.metric("Net Profit Margin", f"{_dp_npm[-1]*100:.1f}%")
                dp_mc2.metric("Asset Turnover", f"{_dp_at[-1]:.2f}x")
                dp_mc3.metric("Equity Multiplier", f"{_dp_em[-1]:.2f}x")
                dp_mc4.metric("ROE (3-Factor)", f"{_dp_roe3[-1]*100:.1f}%")
            
            with dp_tab2:
                st.markdown(
                    '<div style="font-size:0.8rem; color:#B8B3D7; margin-bottom:0.8rem;">'
                    '<b>ROE = Tax Burden × Interest Burden × Operating Margin × AT × EM</b></div>',
                    unsafe_allow_html=True,
                )
                
                if _dp_tb and _dp_ib:
                    fig_dp5 = go.Figure()
                    fig_dp5.add_trace(go.Bar(x=_dp_years, y=[v * 100 for v in _dp_tb], name="Tax Burden %",
                                             marker_color="rgba(107,92,231,0.7)"))
                    fig_dp5.add_trace(go.Bar(x=_dp_years, y=[v * 100 for v in _dp_ib], name="Interest Burden %",
                                             marker_color="rgba(16,185,129,0.7)"))
                    fig_dp5.add_trace(go.Scatter(x=_dp_years, y=[v * 100 for v in _dp_roe5], name="ROE %",
                                                  mode="lines+markers", line=dict(color="#F5A623", width=3),
                                                  marker=dict(size=8)))
                    fig_dp5.update_layout(
                        **_CHART_LAYOUT_BASE, height=350, barmode="group",
                        margin=dict(t=30, b=40, l=50, r=30),
                        legend=dict(font=dict(size=9, color="#B8B3D7"), orientation="h", yanchor="bottom", y=1.02),
                    )
                    _apply_space_grid(fig_dp5)
                    st.plotly_chart(fig_dp5, use_container_width=True, key="dupont_5factor")
                    
                    dp5_c1, dp5_c2, dp5_c3 = st.columns(3)
                    dp5_c1.metric("Tax Burden", f"{_dp_tb[-1]*100:.1f}%")
                    dp5_c2.metric("Interest Burden", f"{_dp_ib[-1]*100:.1f}%")
                    dp5_c3.metric("ROE (5-Factor)", f"{_dp_roe5[-1]*100:.1f}%")
                else:
                    st.info("Insufficient data for 5-factor DuPont analysis.")
        else:
            st.info("Insufficient financial data for DuPont Analysis.")
    
    _divider()

    # ══════════════════════════════════════════════════════
    # 10e. RISK METRICS
    # ══════════════════════════════════════════════════════
    _section("Risk Metrics")
    
    risk_col1, risk_col2, risk_col3, risk_col4 = st.columns(4)
    
    with risk_col1:
        beta_val = cd.beta or 0
        beta_color = "#EF4444" if beta_val > 1.5 else "#F59E0B" if beta_val > 1 else "#10B981"
        beta_label = "High" if beta_val > 1.5 else "Moderate" if beta_val > 1 else "Low"
        st.markdown(
            f'<div style="text-align:center; padding:0.8rem; background:rgba(107,92,231,0.05); '
            f'border-radius:12px; border:1px solid rgba(107,92,231,0.1);">'
            f'<div style="font-size:0.65rem; color:#8A85AD; font-weight:600; text-transform:uppercase;">Beta</div>'
            f'<div style="font-size:1.4rem; font-weight:800; color:{beta_color};">{beta_val:.2f}</div>'
            f'<div style="font-size:0.6rem; color:{beta_color};">{beta_label} Volatility</div>'
            f'</div>',
            unsafe_allow_html=True,
        )
    
    # Calculate Sharpe, Max Drawdown, Volatility from 1Y data
    try:
        risk_hist = cd.hist_1y if cd.hist_1y is not None and not cd.hist_1y.empty else None
        if risk_hist is not None and len(risk_hist) > 20:
            returns = risk_hist["Close"].pct_change().dropna()
            
            # Annualized volatility
            ann_vol = returns.std() * np.sqrt(252) * 100
            vol_color = "#EF4444" if ann_vol > 40 else "#F59E0B" if ann_vol > 25 else "#10B981"
            
            with risk_col2:
                st.markdown(
                    f'<div style="text-align:center; padding:0.8rem; background:rgba(107,92,231,0.05); '
                    f'border-radius:12px; border:1px solid rgba(107,92,231,0.1);">'
                    f'<div style="font-size:0.65rem; color:#8A85AD; font-weight:600; text-transform:uppercase;">Volatility (1Y)</div>'
                    f'<div style="font-size:1.4rem; font-weight:800; color:{vol_color};">{ann_vol:.1f}%</div>'
                    f'<div style="font-size:0.6rem; color:#8A85AD;">Annualized</div>'
                    f'</div>',
                    unsafe_allow_html=True,
                )
            
            # Sharpe ratio (assume 5% risk-free rate)
            ann_return = returns.mean() * 252 * 100
            sharpe = (ann_return - 5) / ann_vol if ann_vol > 0 else 0
            sharpe_color = "#10B981" if sharpe > 1 else "#F59E0B" if sharpe > 0.5 else "#EF4444"
            
            with risk_col3:
                st.markdown(
                    f'<div style="text-align:center; padding:0.8rem; background:rgba(107,92,231,0.05); '
                    f'border-radius:12px; border:1px solid rgba(107,92,231,0.1);">'
                    f'<div style="font-size:0.65rem; color:#8A85AD; font-weight:600; text-transform:uppercase;">Sharpe Ratio</div>'
                    f'<div style="font-size:1.4rem; font-weight:800; color:{sharpe_color};">{sharpe:.2f}</div>'
                    f'<div style="font-size:0.6rem; color:#8A85AD;">Rf=5%</div>'
                    f'</div>',
                    unsafe_allow_html=True,
                )
            
            # Max Drawdown
            cum_returns = (1 + returns).cumprod()
            running_max = cum_returns.cummax()
            drawdown = (cum_returns / running_max - 1) * 100
            max_dd = drawdown.min()
            dd_color = "#EF4444" if max_dd < -30 else "#F59E0B" if max_dd < -15 else "#10B981"
            
            with risk_col4:
                st.markdown(
                    f'<div style="text-align:center; padding:0.8rem; background:rgba(107,92,231,0.05); '
                    f'border-radius:12px; border:1px solid rgba(107,92,231,0.1);">'
                    f'<div style="font-size:0.65rem; color:#8A85AD; font-weight:600; text-transform:uppercase;">Max Drawdown</div>'
                    f'<div style="font-size:1.4rem; font-weight:800; color:{dd_color};">{max_dd:.1f}%</div>'
                    f'<div style="font-size:0.6rem; color:#8A85AD;">1Y Period</div>'
                    f'</div>',
                    unsafe_allow_html=True,
                )
            
            # Drawdown chart
            fig_dd = go.Figure()
            fig_dd.add_trace(go.Scatter(
                x=drawdown.index, y=drawdown.values,
                mode="lines", fill="tozeroy",
                line=dict(color="#EF4444", width=1.5),
                fillcolor="rgba(239,68,68,0.1)",
                name="Drawdown",
            ))
            fig_dd.update_layout(
                **_CHART_LAYOUT_BASE, height=200,
                margin=dict(t=10, b=25, l=50, r=30),
                yaxis=dict(ticksuffix="%", tickfont=dict(size=9, color="#8A85AD"),
                          title=dict(text="Drawdown", font=dict(size=10, color="#8A85AD"))),
                xaxis=dict(showgrid=False, tickfont=dict(size=9, color="#8A85AD")),
                showlegend=False,
            )
            _apply_space_grid(fig_dd)
            st.plotly_chart(fig_dd, use_container_width=True, key="drawdown_chart")
        else:
            with risk_col2:
                st.info("Insufficient data")
    except Exception:
        pass

    _divider()

    # ══════════════════════════════════════════════════════
    # 10f. ESG SCORES
    # ══════════════════════════════════════════════════════
    try:
        if cd.esg_scores is not None and not cd.esg_scores.empty:
            _section("ESG Scores", "🌱")
            
            esg = cd.esg_scores
            
            # Try to extract key scores
            esg_total = None
            esg_env = None
            esg_social = None
            esg_gov = None
            
            for idx_name in esg.index:
                idx_lower = str(idx_name).lower()
                if "total" in idx_lower and "esg" in idx_lower:
                    esg_total = float(esg.loc[idx_name].iloc[0]) if pd.notna(esg.loc[idx_name].iloc[0]) else None
                elif "environment" in idx_lower:
                    esg_env = float(esg.loc[idx_name].iloc[0]) if pd.notna(esg.loc[idx_name].iloc[0]) else None
                elif "social" in idx_lower:
                    esg_social = float(esg.loc[idx_name].iloc[0]) if pd.notna(esg.loc[idx_name].iloc[0]) else None
                elif "governance" in idx_lower:
                    esg_gov = float(esg.loc[idx_name].iloc[0]) if pd.notna(esg.loc[idx_name].iloc[0]) else None
            
            if esg_total is not None or esg_env is not None:
                esg_cols = st.columns(4)
                
                esg_items = [
                    ("Total ESG", esg_total, "🌍"),
                    ("Environmental", esg_env, "🌿"),
                    ("Social", esg_social, "👥"),
                    ("Governance", esg_gov, "⚖️"),
                ]
                
                for i, (label, score, icon) in enumerate(esg_items):
                    with esg_cols[i]:
                        if score is not None:
                            # Lower ESG risk score = better (Sustainalytics scale)
                            if score < 15:
                                color = "#10B981"
                                rating = "Low Risk"
                            elif score < 25:
                                color = "#34D399"
                                rating = "Medium"
                            elif score < 35:
                                color = "#F59E0B"
                                rating = "High"
                            else:
                                color = "#EF4444"
                                rating = "Severe"
                            
                            st.markdown(
                                f'<div style="text-align:center; padding:0.8rem; background:rgba(107,92,231,0.05); '
                                f'border-radius:12px; border:1px solid rgba(107,92,231,0.1);">'
                                f'<div style="font-size:1.2rem;">{icon}</div>'
                                f'<div style="font-size:0.6rem; color:#8A85AD; font-weight:600; text-transform:uppercase;">{label}</div>'
                                f'<div style="font-size:1.3rem; font-weight:800; color:{color};">{score:.1f}</div>'
                                f'<div style="font-size:0.6rem; color:{color};">{rating}</div>'
                                f'</div>',
                                unsafe_allow_html=True,
                            )
                        else:
                            st.markdown(
                                f'<div style="text-align:center; padding:0.8rem; background:rgba(107,92,231,0.05); '
                                f'border-radius:12px;"><div style="font-size:1.2rem;">{icon}</div>'
                                f'<div style="font-size:0.6rem; color:#8A85AD;">{label}</div>'
                                f'<div style="color:#8A85AD;">N/A</div></div>',
                                unsafe_allow_html=True,
                            )
                
                st.markdown(
                    '<div style="font-size:0.6rem; color:#5A567A; text-align:center; margin-top:0.5rem;">'
                    'Lower scores = lower ESG risk (Sustainalytics methodology)</div>',
                    unsafe_allow_html=True,
                )
            
            _divider()
    except Exception:
        pass

    # ══════════════════════════════════════════════════════
    # 11. M&A HISTORY
    # ══════════════════════════════════════════════════════
    _section("M&A History")
    if cd.ma_deals:
        deal_count = len(cd.ma_deals)
        source_link = f' &middot; <a href="{cd.ma_source}" target="_blank" style="color:#6B5CE7; text-decoration:none; font-weight:500;">View on Wikipedia &rarr;</a>' if cd.ma_source else ""
        st.markdown(
            f'<div style="margin-bottom:0.8rem;">'
            f'<span class="pill pill-purple">{deal_count} Acquisitions</span>'
            f'{source_link}'
            f'</div>',
            unsafe_allow_html=True,
        )
        ma_df = pd.DataFrame([
            {
                "Date": d.get("date", ""),
                "Target": d.get("company", ""),
                "Business": d.get("business", ""),
                "Country": d.get("country", ""),
                "Value (USD)": d.get("value", "Undisclosed"),
            }
            for d in cd.ma_deals[:30]
        ])
        st.dataframe(ma_df, use_container_width=True, hide_index=True, height=400)
        if deal_count > 30:
            st.caption(f"Showing 30 of {deal_count} deals.")
    elif cd.ma_history:
        st.markdown(cd.ma_history)
    else:
        st.info("No public M&A history found for this company.")

    # ══════════════════════════════════════════════════════
    # 12. MANAGEMENT
    # ══════════════════════════════════════════════════════
    _section("Management Team")
    mgmt_col1, mgmt_col2 = st.columns([3, 2])
    with mgmt_col1:
        if cd.officers:
            mgmt_data = []
            for o in cd.officers[:10]:
                mgmt_data.append({
                    "Name": o.get("name", "N/A"),
                    "Title": o.get("title", "N/A"),
                    "Age": o.get("age", ""),
                    "Total Pay": format_number(o.get("totalPay"), currency_symbol=cs) if o.get("totalPay") else "\u2014",
                })
            st.dataframe(pd.DataFrame(mgmt_data), use_container_width=True, hide_index=True)
        else:
            st.info("Management data not available.")
    with mgmt_col2:
        if cd.mgmt_sentiment:
            st.markdown("<p style='font-size:0.85rem; font-weight:700; color:#E0DCF5; margin-bottom:0.3rem;'>Management Assessment</p>", unsafe_allow_html=True)
            for line in cd.mgmt_sentiment.split("\n"):
                line = line.strip()
                if line.startswith("- "):
                    line = line[2:]
                if line:
                    st.markdown(f"<div style='font-size:0.82rem; color:#B8B3D7; line-height:1.7; padding:0.15rem 0;'>&bull; {line}</div>", unsafe_allow_html=True)

    # ══════════════════════════════════════════════════════
    # 13. NEWS
    # ══════════════════════════════════════════════════════
    _section("Recent News")
    if cd.news:
        for n in cd.news[:10]:
            title = n.get("title", "")
            publisher = n.get("publisher", "")
            link = n.get("link", "")
            if link:
                st.markdown(
                    f'<div class="news-item">'
                    f'<a href="{link}" target="_blank" class="news-title">{title}</a>'
                    f'<span class="news-pub"> &mdash; {publisher}</span>'
                    f'</div>',
                    unsafe_allow_html=True,
                )
            else:
                st.markdown(
                    f'<div class="news-item">'
                    f'<span class="news-title">{title}</span>'
                    f'<span class="news-pub"> &mdash; {publisher}</span>'
                    f'</div>',
                    unsafe_allow_html=True,
                )
    else:
        st.info("No recent news available.")

    # ══════════════════════════════════════════════════════
    # 14a. EARNINGS SURPRISE CHART (Alpha Vantage)
    # ══════════════════════════════════════════════════════
    if getattr(cd, "earnings_history", None) and len(cd.earnings_history) > 0:
        _section("Earnings Surprise")
        earnings = cd.earnings_history[:8]  # last 8 quarters, most recent first
        earnings = list(reversed(earnings))  # oldest first for chart
        eq_dates = [e.get("date", "")[:7] for e in earnings]
        eq_actual = [e.get("actual_eps") for e in earnings]
        eq_estimate = [e.get("estimated_eps") for e in earnings]

        # Build colors: green for beat, red for miss
        bar_colors = []
        for a, est in zip(eq_actual, eq_estimate):
            if a is not None and est is not None:
                bar_colors.append("#10B981" if a >= est else "#EF4444")
            else:
                bar_colors.append("#8A85AD")

        fig_earn = go.Figure()
        # Ghost bar for estimates (transparent fill + outline)
        fig_earn.add_trace(go.Bar(
            x=eq_dates, y=eq_estimate, name="Estimate",
            marker=dict(
                color="rgba(138,133,173,0.08)",
                line=dict(color="rgba(138,133,173,0.5)", width=1.5),
            ),
            text=[f"{v:.2f}" if v is not None else "" for v in eq_estimate],
            textposition="outside", textfont=dict(size=9, color="#8A85AD"),
        ))
        # Solid actuals with white outline
        fig_earn.add_trace(go.Bar(
            x=eq_dates, y=eq_actual, name="Actual",
            marker=dict(
                color=bar_colors,
                line=dict(color="rgba(255,255,255,0.2)", width=1),
            ),
            text=[f"{v:.2f}" if v is not None else "" for v in eq_actual],
            textposition="outside", textfont=dict(size=9, color="#B8B3D7"),
        ))
        fig_earn.update_layout(
            **_CHART_LAYOUT_BASE,
            height=500, barmode="group",
            margin=dict(t=40, b=40, l=60, r=40),
            xaxis=dict(tickfont=dict(size=12, color="#8A85AD"), showgrid=False),
            yaxis=dict(title=dict(text="EPS", font=dict(size=13, color="#8A85AD")),
                       tickfont=dict(size=12, color="#8A85AD"),
                       tickprefix=cd.currency_symbol),
            legend=dict(font=dict(size=12, color="#B8B3D7"), orientation="h",
                        yanchor="bottom", y=1.02),
        )
        _apply_space_grid(fig_earn)
        st.markdown('<div class="profile-chart-wrapper">', unsafe_allow_html=True)
        st.plotly_chart(fig_earn, use_container_width=True, key="earnings_surprise_chart")
        st.markdown('</div>', unsafe_allow_html=True)

    # ══════════════════════════════════════════════════════
    # 14b. NEWS SENTIMENT (Alpha Vantage)
    # ══════════════════════════════════════════════════════
    if getattr(cd, "news_sentiment", None) and len(cd.news_sentiment) > 0:
        _section("News Sentiment")
        for ns in cd.news_sentiment[:10]:
            title = ns.get("title", "")
            url = ns.get("url", "")
            source = ns.get("source", "")
            published = ns.get("published", "")
            sentiment = ns.get("overall_sentiment", "Neutral").lower()
            score = ns.get("overall_score")
            score_str = f"{score:.2f}" if score is not None else ""

            if "bullish" in sentiment or "positive" in sentiment:
                css_class = "news-sentiment-bullish"
                badge = '<span style="color:#10B981; font-weight:700; font-size:0.7rem;">BULLISH</span>'
            elif "bearish" in sentiment or "negative" in sentiment:
                css_class = "news-sentiment-bearish"
                badge = '<span style="color:#EF4444; font-weight:700; font-size:0.7rem;">BEARISH</span>'
            else:
                css_class = "news-sentiment-neutral"
                badge = '<span style="color:#8A85AD; font-weight:700; font-size:0.7rem;">NEUTRAL</span>'

            link_html = f'<a href="{url}" target="_blank" style="color:#E0DCF5; text-decoration:none; font-weight:600; font-size:0.85rem;">{title}</a>' if url else f'<span style="color:#E0DCF5; font-weight:600; font-size:0.85rem;">{title}</span>'
            st.markdown(
                f'<div class="news-card {css_class}">'
                f'{link_html}'
                f'<div style="margin-top:0.3rem; display:flex; gap:0.8rem; align-items:center;">'
                f'{badge}'
                f'<span style="color:#8A85AD; font-size:0.7rem;">{source} &middot; {published}</span>'
                f'</div></div>',
                unsafe_allow_html=True,
            )

    # ══════════════════════════════════════════════════════
    # 14c. INSIDER ACTIVITY (Alpha Vantage)
    # ══════════════════════════════════════════════════════
    av_insiders = getattr(cd, "av_insider_transactions", None)
    if av_insiders and len(av_insiders) > 0:
        _section("Insider Activity")
        rows_html = ""
        for t in av_insiders[:20]:
            date = t.get("date", "")
            insider = t.get("insider", "")
            title = t.get("title", "")
            txn_type = t.get("type", "")
            shares = t.get("shares")
            value = t.get("value")

            if txn_type == "A":
                type_label = '<span style="color:#10B981; font-weight:700;">Buy</span>'
                row_bg = "rgba(16,185,129,0.04)"
            elif txn_type == "D":
                type_label = '<span style="color:#EF4444; font-weight:700;">Sell</span>'
                row_bg = "rgba(239,68,68,0.04)"
            else:
                type_label = txn_type
                row_bg = "transparent"

            shares_str = f"{shares:,.0f}" if shares else "—"
            value_str = f"{cd.currency_symbol}{value:,.0f}" if value else "—"
            rows_html += (
                f'<tr style="background:{row_bg};">'
                f'<td>{date}</td><td>{insider}</td><td style="font-size:0.72rem;">{title}</td>'
                f'<td>{type_label}</td><td style="text-align:right;">{shares_str}</td>'
                f'<td style="text-align:right;">{value_str}</td></tr>'
            )
        st.markdown(
            f'<table class="insider-table">'
            f'<thead><tr><th>Date</th><th>Insider</th><th>Title</th>'
            f'<th>Type</th><th style="text-align:right;">Shares</th>'
            f'<th style="text-align:right;">Value</th></tr></thead>'
            f'<tbody>{rows_html}</tbody></table>'.replace("$", "&#36;"),
            unsafe_allow_html=True,
        )

    # ══════════════════════════════════════════════════════
    # 15. INSIGHTS — 7 Rich Tabs
    # ══════════════════════════════════════════════════════
    _section("Insights")
    ai_tab1, ai_tab2, ai_tab3, ai_tab4, ai_tab5, ai_tab6, ai_tab7 = st.tabs([
        "Executive Summary", "Financial Trends", "SWOT Analysis",
        "Growth Outlook", "Capital Allocation", "Industry Analysis", "Risk Factors"
    ])

    # ── Tab 1: Executive Summary ──────────────────────
    with ai_tab1:
        es_left, es_right = st.columns([3, 2])
        with es_left:
            if cd.executive_summary_bullets:
                for b in cd.executive_summary_bullets:
                    st.markdown(f"<div style='font-size:0.88rem; color:#B8B3D7; line-height:1.7; padding:0.2rem 0;'>&bull; {b}</div>", unsafe_allow_html=True)
            else:
                st.info("Executive summary not available.")
            if cd.product_overview:
                st.markdown('<div style="margin-top:1rem;"><div style="font-size:0.8rem; font-weight:700; color:#9B8AFF; text-transform:uppercase; letter-spacing:0.5px; margin-bottom:0.3rem;">Product Overview</div></div>', unsafe_allow_html=True)
                for line in cd.product_overview.split("\n"):
                    line = line.strip()
                    if line.startswith("- "):
                        line = line[2:]
                    if line:
                        st.markdown(f"<div style='font-size:0.84rem; color:#B8B3D7; line-height:1.7; padding:0.15rem 0;'>&bull; {line}</div>", unsafe_allow_html=True)
        with es_right:
            st.markdown('<div class="profile-chart-wrapper">', unsafe_allow_html=True)
            _build_peer_valuation_chart(cd)
            _build_earnings_surprise_chart(cd)
            st.markdown('</div>', unsafe_allow_html=True)

    # ── Tab 2: Financial Trends ───────────────────────
    with ai_tab2:
        ft_c1, ft_c2 = st.columns(2)
        with ft_c1:
            st.markdown('<div style="font-size:0.8rem; font-weight:700; color:#9B8AFF; text-transform:uppercase; letter-spacing:0.5px; margin-bottom:0.3rem;">Revenue & Margins</div>', unsafe_allow_html=True)
            st.markdown('<div class="profile-chart-wrapper">', unsafe_allow_html=True)
            _build_revenue_margin_chart(cd)
            st.markdown('</div>', unsafe_allow_html=True)
        with ft_c2:
            st.markdown('<div style="font-size:0.8rem; font-weight:700; color:#9B8AFF; text-transform:uppercase; letter-spacing:0.5px; margin-bottom:0.3rem;">Cash Flow</div>', unsafe_allow_html=True)
            st.markdown('<div class="profile-chart-wrapper">', unsafe_allow_html=True)
            _build_cashflow_chart(cd)
            st.markdown('</div>', unsafe_allow_html=True)
        st.markdown('<div style="font-size:0.8rem; font-weight:700; color:#9B8AFF; text-transform:uppercase; letter-spacing:0.5px; margin-bottom:0.3rem; margin-top:0.5rem;">Balance Sheet</div>', unsafe_allow_html=True)
        st.markdown('<div class="profile-chart-wrapper">', unsafe_allow_html=True)
        _build_balance_sheet_chart(cd)
        st.markdown('</div>', unsafe_allow_html=True)

    # ── Tab 3: SWOT Analysis ─────────────────────────
    with ai_tab3:
        _render_swot_grid(cd.swot_analysis)

    # ── Tab 4: Growth Outlook ────────────────────────
    with ai_tab4:
        go_left, go_right = st.columns([3, 2])
        with go_left:
            _render_growth_outlook(cd.growth_outlook, cd)
        with go_right:
            st.markdown('<div style="font-size:0.8rem; font-weight:700; color:#9B8AFF; text-transform:uppercase; letter-spacing:0.5px; margin-bottom:0.3rem;">Revenue & Margin Trends</div>', unsafe_allow_html=True)
            st.markdown('<div class="profile-chart-wrapper">', unsafe_allow_html=True)
            _build_revenue_margin_chart(cd, key="rev_margin_growth")
            st.markdown('</div>', unsafe_allow_html=True)

    # ── Tab 5: Capital Allocation ────────────────────
    with ai_tab5:
        ca_left, ca_right = st.columns([3, 2])
        with ca_left:
            _render_capital_allocation(cd.capital_allocation_analysis, cd)
        with ca_right:
            st.markdown('<div style="font-size:0.8rem; font-weight:700; color:#9B8AFF; text-transform:uppercase; letter-spacing:0.5px; margin-bottom:0.3rem;">Cash Flow Trends</div>', unsafe_allow_html=True)
            st.markdown('<div class="profile-chart-wrapper">', unsafe_allow_html=True)
            _build_cashflow_chart(cd, key="cashflow_capalloc")
            st.markdown('</div>', unsafe_allow_html=True)

    # ── Tab 6: Industry Analysis ─────────────────────
    with ai_tab6:
        if cd.industry_analysis:
            for line in cd.industry_analysis.split("\n"):
                line = line.strip()
                if line.startswith("- "):
                    line = line[2:]
                if line:
                    st.markdown(f"<div style='font-size:0.88rem; color:#B8B3D7; line-height:1.7; padding:0.2rem 0;'>&bull; {line}</div>", unsafe_allow_html=True)
        else:
            st.info("Industry analysis not available.")

    # ── Tab 7: Risk Factors (color-coded severity) ───
    with ai_tab7:
        if cd.risk_factors:
            for line in cd.risk_factors.split("\n"):
                line = line.strip()
                if line.startswith("- "):
                    line = line[2:]
                if not line:
                    continue
                # Detect severity tag
                severity_color = "#8A85AD"
                severity_bg = "rgba(138,133,173,0.05)"
                severity_border = "rgba(138,133,173,0.2)"
                if line.startswith("[HIGH]"):
                    line = line[6:].strip()
                    severity_color = "#EF4444"
                    severity_bg = "rgba(239,68,68,0.06)"
                    severity_border = "rgba(239,68,68,0.3)"
                elif line.startswith("[MEDIUM]"):
                    line = line[8:].strip()
                    severity_color = "#F5A623"
                    severity_bg = "rgba(245,166,35,0.06)"
                    severity_border = "rgba(245,166,35,0.3)"
                elif line.startswith("[LOW]"):
                    line = line[5:].strip()
                    severity_color = "#10B981"
                    severity_bg = "rgba(16,185,129,0.06)"
                    severity_border = "rgba(16,185,129,0.3)"
                st.markdown(
                    f'<div style="border-left:3px solid {severity_border}; background:{severity_bg}; '
                    f'padding:0.5rem 0.8rem; margin-bottom:0.4rem; border-radius:0 8px 8px 0;">'
                    f'<div style="font-size:0.86rem; color:#B8B3D7; line-height:1.7;">{line}</div>'
                    f'</div>',
                    unsafe_allow_html=True,
                )
        else:
            st.info("Risk factors not available.")

    _divider()

    # ══════════════════════════════════════════════════════
    # 15. DOWNLOAD PPTX
    # ══════════════════════════════════════════════════════
    st.markdown("")
    st.markdown("")
    _section("Download Tear Sheet")

    if not os.path.exists("assets/template.pptx"):
        with st.spinner("Creating template..."):
            from create_template import build
            build()

    with st.spinner("Building 8-slide PowerPoint presentation..."):
        pptx_buf = generate_presentation(cd)

    dl1, dl2, dl3 = st.columns([1, 2, 1])
    with dl2:
        st.download_button(
            label=f"Download {cd.ticker} Orbital Profile  (3 slides)",
            data=pptx_buf,
            file_name=f"{cd.ticker}_Orbital_Profile.pptx",
            mime="application/vnd.openxmlformats-officedocument.presentationml.presentation",
            use_container_width=True,
        )
        st.markdown(
            "<p style='text-align:center; font-size:0.72rem; color:#8A85AD; margin-top:0.3rem;'>"
            "Professional IB-grade presentation &middot; Editable charts &middot; Navy/Gold palette"
            "</p>",
            unsafe_allow_html=True,
        )
    
    # Excel Export
    _divider()
    _section("Export Financial Data", "📊")
    
    ex1, ex2, ex3 = st.columns([1, 2, 1])
    with ex2:
        try:
            excel_data = _export_to_excel(cd)
            st.download_button(
                label=f"📥 Download {cd.ticker} Financial Data (Excel)",
                data=excel_data,
                file_name=f"{cd.ticker}_Financial_Data.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
            st.markdown(
                "<p style='text-align:center; font-size:0.72rem; color:#8A85AD; margin-top:0.3rem;'>"
                "Multi-sheet Excel workbook &middot; Income Statement &middot; Balance Sheet &middot; Cash Flow &middot; Peer Data"
                "</p>",
                unsafe_allow_html=True,
            )
        except Exception as e:
            st.warning(f"Excel export not available: {e}")
        
        # CSV Quick Export
        csv_data = _export_to_csv(cd)
        st.download_button(
            label=f"📄 Quick Export (CSV)",
            data=csv_data,
            file_name=f"{cd.ticker}_Summary.csv",
            mime="text/csv",
            use_container_width=True,
        )
        
        # JSON Export
        try:
            json_export = {
                "ticker": cd.ticker,
                "name": cd.name,
                "sector": cd.sector,
                "industry": cd.industry,
                "market_cap": cd.market_cap,
                "enterprise_value": cd.enterprise_value,
                "current_price": cd.current_price,
                "currency": cd.currency_code,
                "trailing_pe": cd.trailing_pe,
                "forward_pe": cd.forward_pe,
                "ev_to_ebitda": cd.ev_to_ebitda,
                "ev_to_revenue": cd.ev_to_revenue,
                "price_to_book": cd.price_to_book,
                "gross_margins": cd.gross_margins,
                "operating_margins": cd.operating_margins,
                "profit_margins": cd.profit_margins,
                "return_on_equity": cd.return_on_equity,
                "return_on_assets": getattr(cd, 'return_on_assets', None),
                "revenue_growth": cd.revenue_growth,
                "dividend_yield": cd.dividend_yield,
                "beta": cd.beta,
                "52w_high": cd.fifty_two_week_high,
                "52w_low": cd.fifty_two_week_low,
                "analyst_price_targets": cd.analyst_price_targets,
                "employees": cd.full_time_employees,
                "export_date": datetime.now().isoformat(),
            }
            json_str = json.dumps(json_export, indent=2, default=str)
            st.download_button(
                label=f"🔗 Export as JSON (API-Ready)",
                data=json_str,
                file_name=f"{cd.ticker}_data.json",
                mime="application/json",
                use_container_width=True,
            )
        except Exception:
            pass
    
    # ══════════════════════════════════════════════════════
    # 🔔 INTELLIGENCE FEED
    # ══════════════════════════════════════════════════════
    with _safe_section("Intelligence Feed"):
        _divider()
        st.markdown(
            '<div style="font-size:1.1rem; font-weight:800; color:#E0DCF5; margin-bottom:0.8rem;">'
            '🔔 Intelligence</div>',
            unsafe_allow_html=True,
        )
        try:
            _intel_ticker = yf.Ticker(cd.ticker)
            _intel_info = _intel_ticker.info or {}

            _intel_cols = st.columns(2)

            # Unusual Volume
            with _intel_cols[0]:
                try:
                    _vol_today = _intel_info.get("volume", 0) or 0
                    _vol_avg = _intel_info.get("averageVolume", 1) or 1
                    _vol_ratio = _vol_today / _vol_avg if _vol_avg > 0 else 0
                    _vol_flag = "🔴 UNUSUAL" if _vol_ratio > 2.0 else "🟡 ELEVATED" if _vol_ratio > 1.5 else "🟢 NORMAL"
                    st.markdown(
                        f'<div style="background:rgba(107,92,231,0.06); border-radius:10px; padding:0.8rem; margin-bottom:0.5rem;">'
                        f'<div style="font-size:0.75rem; font-weight:700; color:#9B8AFF; margin-bottom:0.4rem;">📊 Volume Analysis</div>'
                        f'<div style="font-size:0.78rem; color:#B8B3D7;">Today: <b>{_vol_today:,.0f}</b></div>'
                        f'<div style="font-size:0.78rem; color:#B8B3D7;">30d Avg: <b>{_vol_avg:,.0f}</b></div>'
                        f'<div style="font-size:0.78rem; color:#B8B3D7;">Ratio: <b>{_vol_ratio:.2f}x</b> {_vol_flag}</div>'
                        f'</div>',
                        unsafe_allow_html=True,
                    )
                except Exception:
                    pass

            # Short Interest
            with _intel_cols[1]:
                try:
                    _short_pct = _intel_info.get("shortPercentOfFloat", None)
                    _short_ratio = _intel_info.get("shortRatio", None)
                    _si_html = ""
                    if _short_pct is not None:
                        _sp = _short_pct * 100 if _short_pct < 1 else _short_pct
                        _si_level = "🔴 HIGH" if _sp > 20 else "🟡 MODERATE" if _sp > 10 else "🟢 LOW"
                        _si_html += f'<div style="font-size:0.78rem; color:#B8B3D7;">Short % Float: <b>{_sp:.1f}%</b> {_si_level}</div>'
                    if _short_ratio is not None:
                        _si_html += f'<div style="font-size:0.78rem; color:#B8B3D7;">Days to Cover: <b>{_short_ratio:.1f}</b></div>'
                    if _si_html:
                        st.markdown(
                            f'<div style="background:rgba(107,92,231,0.06); border-radius:10px; padding:0.8rem; margin-bottom:0.5rem;">'
                            f'<div style="font-size:0.75rem; font-weight:700; color:#9B8AFF; margin-bottom:0.4rem;">📉 Short Interest</div>'
                            f'{_si_html}</div>',
                            unsafe_allow_html=True,
                        )
                    else:
                        st.markdown(
                            '<div style="background:rgba(107,92,231,0.06); border-radius:10px; padding:0.8rem;">'
                            '<div style="font-size:0.75rem; font-weight:700; color:#9B8AFF;">📉 Short Interest</div>'
                            '<div style="font-size:0.78rem; color:#8A85AD;">Data not available</div></div>',
                            unsafe_allow_html=True,
                        )
                except Exception:
                    pass

            # Insider Transactions
            try:
                _insider_txns = _intel_ticker.insider_transactions
                if _insider_txns is not None and not _insider_txns.empty:
                    _buys = len(_insider_txns[_insider_txns.get("Text", _insider_txns.columns[0] if len(_insider_txns.columns) > 0 else "").str.contains("Buy|Purchase", case=False, na=False)] if "Text" in _insider_txns.columns else [])
                    _sells = len(_insider_txns) - _buys
                    _sentiment = "🟢 BULLISH" if _buys > _sells else "🔴 BEARISH" if _sells > _buys else "🟡 NEUTRAL"
                    st.markdown(
                        f'<div style="background:rgba(107,92,231,0.06); border-radius:10px; padding:0.8rem; margin-bottom:0.5rem;">'
                        f'<div style="font-size:0.75rem; font-weight:700; color:#9B8AFF; margin-bottom:0.4rem;">👤 Insider Activity</div>'
                        f'<div style="font-size:0.78rem; color:#B8B3D7;">Buys: <b style="color:#10B981;">{_buys}</b> | Sells: <b style="color:#EF4444;">{_sells}</b></div>'
                        f'<div style="font-size:0.78rem; color:#B8B3D7;">Net Sentiment: <b>{_sentiment}</b></div>'
                        f'</div>',
                        unsafe_allow_html=True,
                    )
                else:
                    st.markdown(
                        '<div style="background:rgba(107,92,231,0.06); border-radius:10px; padding:0.8rem; margin-bottom:0.5rem;">'
                        '<div style="font-size:0.75rem; font-weight:700; color:#9B8AFF;">👤 Insider Activity</div>'
                        '<div style="font-size:0.78rem; color:#8A85AD;">No recent insider transactions</div></div>',
                        unsafe_allow_html=True,
                    )
            except Exception:
                st.markdown(
                    '<div style="background:rgba(107,92,231,0.06); border-radius:10px; padding:0.8rem; margin-bottom:0.5rem;">'
                    '<div style="font-size:0.75rem; font-weight:700; color:#9B8AFF;">👤 Insider Activity</div>'
                    '<div style="font-size:0.78rem; color:#8A85AD;">Data not available</div></div>',
                    unsafe_allow_html=True,
                )

            # Earnings Surprise
            try:
                _earnings = _intel_ticker.earnings_dates
                if _earnings is not None and not _earnings.empty:
                    _recent = _earnings.head(4)
                    _earn_html = ""
                    for _idx, _row in _recent.iterrows():
                        _est = _row.get("EPS Estimate", None)
                        _act = _row.get("Reported EPS", None)
                        if _est is not None and _act is not None:
                            try:
                                _est_f = float(_est)
                                _act_f = float(_act)
                                _surprise = ((_act_f - _est_f) / abs(_est_f) * 100) if _est_f != 0 else 0
                                _s_color = "#10B981" if _surprise >= 0 else "#EF4444"
                                _s_icon = "✅" if _surprise >= 0 else "❌"
                                _date_str = str(_idx.date()) if hasattr(_idx, 'date') else str(_idx)[:10]
                                _earn_html += (
                                    f'<div style="display:flex; justify-content:space-between; font-size:0.72rem; padding:0.15rem 0;">'
                                    f'<span style="color:#8A85AD;">{_date_str}</span>'
                                    f'<span style="color:#B8B3D7;">Est: ${_est_f:.2f}</span>'
                                    f'<span style="color:#B8B3D7;">Act: ${_act_f:.2f}</span>'
                                    f'<span style="color:{_s_color};">{_s_icon} {_surprise:+.1f}%</span>'
                                    f'</div>'
                                )
                            except (ValueError, TypeError):
                                pass
                    if _earn_html:
                        st.markdown(
                            f'<div style="background:rgba(107,92,231,0.06); border-radius:10px; padding:0.8rem;">'
                            f'<div style="font-size:0.75rem; font-weight:700; color:#9B8AFF; margin-bottom:0.4rem;">📅 Earnings Surprises (Last 4)</div>'
                            f'{_earn_html}</div>',
                            unsafe_allow_html=True,
                        )
            except Exception:
                pass

        except Exception:
            st.markdown(
                '<div style="font-size:0.8rem; color:#8A85AD; padding:0.5rem;">Intelligence data temporarily unavailable.</div>',
                unsafe_allow_html=True,
            )

    # ══════════════════════════════════════════════════════
    # EXPORT SUMMARY (copy-paste friendly)
    # ══════════════════════════════════════════════════════
    with _safe_section("Export Summary"):
        _divider()
        with st.expander("📋 Export Summary (copy-paste friendly)", expanded=False):
            _ex_lines = []
            _ex_lines.append(f"{getattr(cd, 'name', 'N/A')} ({getattr(cd, 'ticker', 'N/A')})")
            _ex_lines.append(f"{'='*50}")
            _ex_price = getattr(cd, 'current_price', None)
            _ex_cs = getattr(cd, 'currency_symbol', '$')
            _ex_lines.append(f"Price: {_ex_cs}{_ex_price:,.2f}" if _ex_price else "Price: N/A")
            _ex_mcap = getattr(cd, 'market_cap', None)
            _ex_lines.append(f"Market Cap: {format_number(_ex_mcap, currency_symbol=_ex_cs)}" if _ex_mcap else "Market Cap: N/A")
            _ex_lines.append(f"Sector: {getattr(cd, 'sector', 'N/A')} | Industry: {getattr(cd, 'industry', 'N/A')}")
            _ex_lines.append("")
            _ex_lines.append("KEY MULTIPLES")
            _ex_lines.append(f"-" * 30)
            _ex_pe = getattr(cd, 'trailing_pe', None)
            _ex_lines.append(f"P/E Ratio: {_ex_pe:.1f}x" if _ex_pe else "P/E Ratio: N/A")
            _ex_fpe = getattr(cd, 'forward_pe', None)
            _ex_lines.append(f"Forward P/E: {_ex_fpe:.1f}x" if _ex_fpe else "Forward P/E: N/A")
            _ex_ps = getattr(cd, 'price_to_sales', None)
            _ex_lines.append(f"P/S: {_ex_ps:.1f}x" if _ex_ps else "P/S: N/A")
            # EV/EBITDA
            _ex_ev = getattr(cd, 'enterprise_value', None) or 0
            _ex_ebitda_s = getattr(cd, 'ebitda', None)
            _ex_ebitda_v = None
            try:
                _ex_ebitda_v = float(_ex_ebitda_s.iloc[0]) if hasattr(_ex_ebitda_s, 'iloc') and len(_ex_ebitda_s) > 0 else (float(_ex_ebitda_s) if _ex_ebitda_s is not None else None)
            except Exception:
                pass
            if _ex_ebitda_v and _ex_ebitda_v > 0:
                _ex_lines.append(f"EV/EBITDA: {_ex_ev/_ex_ebitda_v:.1f}x")
            else:
                _ex_lines.append("EV/EBITDA: N/A")
            _ex_lines.append("")
            # Valuation assessment
            _ex_lines.append("VALUATION ASSESSMENT")
            _ex_lines.append(f"-" * 30)
            if _ex_pe and _ex_pe > 0:
                if _ex_pe < 15:
                    _ex_lines.append("Appears undervalued relative to historical P/E norms.")
                elif _ex_pe < 25:
                    _ex_lines.append("Valuation appears fair relative to market averages.")
                else:
                    _ex_lines.append("Trading at a premium valuation; priced for strong growth.")
            else:
                _ex_lines.append("Insufficient data for valuation assessment.")
            _ex_lines.append("")
            # Key risks
            _ex_lines.append("KEY RISKS")
            _ex_lines.append(f"-" * 30)
            _ex_risks = []
            _ex_dte = getattr(cd, 'debt_to_equity', None)
            if _ex_dte and _ex_dte > 150:
                _ex_risks.append(f"High leverage (D/E: {_ex_dte:.0f}%)")
            _ex_pm = getattr(cd, 'profit_margins', None)
            if _ex_pm is not None and _ex_pm < 0.05:
                _pm_d = _ex_pm * 100 if abs(_ex_pm) < 5 else _ex_pm
                _ex_risks.append(f"Thin/declining margins ({_pm_d:.1f}%)")
            _ex_rg = getattr(cd, 'revenue_growth', None)
            if _ex_rg is not None and _ex_rg < 0:
                _rg_d = _ex_rg * 100 if abs(_ex_rg) < 5 else _ex_rg
                _ex_risks.append(f"Revenue declining ({_rg_d:.1f}% YoY)")
            _ex_beta = getattr(cd, 'beta', None)
            if _ex_beta and _ex_beta > 1.5:
                _ex_risks.append(f"High volatility (beta: {_ex_beta:.2f})")
            if not _ex_risks:
                _ex_risks.append("No major red flags identified.")
            for r in _ex_risks:
                _ex_lines.append(f"- {r}")
            _ex_text = "\n".join(_ex_lines)
            st.code(_ex_text, language=None)

    # Add to Watchlist
    _divider()
    if not _is_in_watchlist(cd.ticker):
        wl1, wl2, wl3 = st.columns([1, 2, 1])
        with wl2:
            if st.button(f"⭐ Add {cd.ticker} to Watchlist", use_container_width=True):
                _add_to_watchlist(cd.ticker)
                st.success(f"Added {cd.ticker} to watchlist!")
                st.rerun()
    else:
        st.markdown(
            f'<div style="text-align:center; padding:0.5rem; background:rgba(16,185,129,0.1); '
            f'border-radius:10px; color:#10B981; font-size:0.85rem;">'
            f'⭐ {cd.ticker} is in your watchlist</div>',
            unsafe_allow_html=True,
        )

elif analysis_mode == "Company Profile" and generate_btn and not ticker_input:
    st.warning("Please enter a ticker symbol in the sidebar.")

elif analysis_mode == "Comps Analysis" and comps_btn and comps_ticker_input:
    # ══════════════════════════════════════════════════════
    # COMPARABLE COMPANY ANALYSIS
    # ══════════════════════════════════════════════════════
    
    # Loading animation
    progress_placeholder = st.empty()
    status_placeholder = st.empty()
    
    def update_progress(pct, msg):
        progress_placeholder.progress(pct, text=msg)
    
    with st.spinner(f"Running comps analysis for {comps_ticker_input}..."):
        comps_analysis = run_comps_analysis(
            ticker=comps_ticker_input,
            max_peers=max_peers,
            include_saas=include_saas,
            progress_callback=update_progress
        )
    
    progress_placeholder.empty()
    status_placeholder.empty()
    
    if not comps_analysis.target_comps or not comps_analysis.target_comps.valid:
        st.error(f"Could not fetch data for {comps_ticker_input}. Please check the ticker and try again.")
    elif not comps_analysis.peers:
        st.warning(f"No comparable companies found for {comps_ticker_input}.")
    else:
        tc = comps_analysis.target_comps
        
        # Header
        st.markdown(
            f'<div style="text-align:center; padding:1.5rem 0;">'
            f'<div style="font-size:2.5rem; font-weight:800; color:#E0DCF5; margin-bottom:0.3rem;">'
            f'{tc.name}</div>'
            f'<div style="font-size:1rem; color:#8A85AD;">'
            f'{tc.sector} · {tc.industry}</div>'
            f'</div>',
            unsafe_allow_html=True,
        )
        
        # Key metrics cards
        col1, col2, col3, col4 = st.columns(4)
        
        def format_num(x, prefix="$", suffix=""):
            if x is None or x == 0:
                return "—"
            if abs(x) >= 1e12:
                return f"{prefix}{x/1e12:.1f}T{suffix}"
            if abs(x) >= 1e9:
                return f"{prefix}{x/1e9:.1f}B{suffix}"
            if abs(x) >= 1e6:
                return f"{prefix}{x/1e6:.0f}M{suffix}"
            return f"{prefix}{x:,.0f}{suffix}"
        
        def format_mult(x):
            if x is None or x == 0:
                return "—"
            return f"{x:.1f}x"
        
        with col1:
            st.markdown(
                f'<div style="background:rgba(107,92,231,0.1); border:1px solid rgba(107,92,231,0.3); '
                f'border-radius:12px; padding:1rem; text-align:center;">'
                f'<div style="font-size:0.75rem; color:#8A85AD; text-transform:uppercase;">Market Cap</div>'
                f'<div style="font-size:1.5rem; font-weight:700; color:#E0DCF5;">{format_num(tc.market_cap)}</div>'
                f'</div>',
                unsafe_allow_html=True,
            )
        
        with col2:
            st.markdown(
                f'<div style="background:rgba(232,99,139,0.1); border:1px solid rgba(232,99,139,0.3); '
                f'border-radius:12px; padding:1rem; text-align:center;">'
                f'<div style="font-size:0.75rem; color:#8A85AD; text-transform:uppercase;">EV/EBITDA</div>'
                f'<div style="font-size:1.5rem; font-weight:700; color:#E0DCF5;">{format_mult(tc.ev_ebitda)}</div>'
                f'</div>',
                unsafe_allow_html=True,
            )
        
        with col3:
            st.markdown(
                f'<div style="background:rgba(16,185,129,0.1); border:1px solid rgba(16,185,129,0.3); '
                f'border-radius:12px; padding:1rem; text-align:center;">'
                f'<div style="font-size:0.75rem; color:#8A85AD; text-transform:uppercase;">EV/Revenue</div>'
                f'<div style="font-size:1.5rem; font-weight:700; color:#E0DCF5;">{format_mult(tc.ev_revenue)}</div>'
                f'</div>',
                unsafe_allow_html=True,
            )
        
        with col4:
            st.markdown(
                f'<div style="background:rgba(245,166,35,0.1); border:1px solid rgba(245,166,35,0.3); '
                f'border-radius:12px; padding:1rem; text-align:center;">'
                f'<div style="font-size:0.75rem; color:#8A85AD; text-transform:uppercase;">P/E Ratio</div>'
                f'<div style="font-size:1.5rem; font-weight:700; color:#E0DCF5;">{format_mult(tc.pe_ratio)}</div>'
                f'</div>',
                unsafe_allow_html=True,
            )
        
        st.markdown('<div style="height:1.5rem;"></div>', unsafe_allow_html=True)
        
        # Peer Median Comparison
        st.markdown(
            '<div style="font-size:1.2rem; font-weight:700; color:#E0DCF5; margin-bottom:1rem;">'
            '📊 Valuation vs Peer Median</div>',
            unsafe_allow_html=True,
        )
        
        comp_col1, comp_col2, comp_col3 = st.columns(3)
        
        with comp_col1:
            target_val = tc.ev_ebitda or 0
            median_val = comps_analysis.median_ev_ebitda or 0
            if median_val > 0:
                diff_pct = ((target_val - median_val) / median_val) * 100
                diff_color = "#10B981" if diff_pct < 0 else "#EF4444"
                diff_text = f"{diff_pct:+.1f}%"
            else:
                diff_color = "#8A85AD"
                diff_text = "—"
            st.markdown(
                f'<div style="background:rgba(255,255,255,0.03); border-radius:12px; padding:1rem;">'
                f'<div style="font-size:0.8rem; color:#8A85AD;">EV/EBITDA vs Median</div>'
                f'<div style="font-size:1.3rem; font-weight:700; color:{diff_color};">{diff_text}</div>'
                f'<div style="font-size:0.7rem; color:#6B6B80;">Target: {format_mult(target_val)} · Median: {format_mult(median_val)}</div>'
                f'</div>',
                unsafe_allow_html=True,
            )
        
        with comp_col2:
            target_val = tc.ev_revenue or 0
            median_val = comps_analysis.median_ev_revenue or 0
            if median_val > 0:
                diff_pct = ((target_val - median_val) / median_val) * 100
                diff_color = "#10B981" if diff_pct < 0 else "#EF4444"
                diff_text = f"{diff_pct:+.1f}%"
            else:
                diff_color = "#8A85AD"
                diff_text = "—"
            st.markdown(
                f'<div style="background:rgba(255,255,255,0.03); border-radius:12px; padding:1rem;">'
                f'<div style="font-size:0.8rem; color:#8A85AD;">EV/Revenue vs Median</div>'
                f'<div style="font-size:1.3rem; font-weight:700; color:{diff_color};">{diff_text}</div>'
                f'<div style="font-size:0.7rem; color:#6B6B80;">Target: {format_mult(target_val)} · Median: {format_mult(median_val)}</div>'
                f'</div>',
                unsafe_allow_html=True,
            )
        
        with comp_col3:
            target_val = tc.pe_ratio or 0
            median_val = comps_analysis.median_pe or 0
            if median_val > 0:
                diff_pct = ((target_val - median_val) / median_val) * 100
                diff_color = "#10B981" if diff_pct < 0 else "#EF4444"
                diff_text = f"{diff_pct:+.1f}%"
            else:
                diff_color = "#8A85AD"
                diff_text = "—"
            st.markdown(
                f'<div style="background:rgba(255,255,255,0.03); border-radius:12px; padding:1rem;">'
                f'<div style="font-size:0.8rem; color:#8A85AD;">P/E vs Median</div>'
                f'<div style="font-size:1.3rem; font-weight:700; color:{diff_color};">{diff_text}</div>'
                f'<div style="font-size:0.7rem; color:#6B6B80;">Target: {format_mult(target_val)} · Median: {format_mult(median_val)}</div>'
                f'</div>',
                unsafe_allow_html=True,
            )
        
        st.markdown('<div style="height:1.5rem;"></div>', unsafe_allow_html=True)
        
        # Implied Valuation
        if comps_analysis.implied_ev_from_ebitda or comps_analysis.implied_ev_from_revenue:
            st.markdown(
                '<div style="font-size:1.2rem; font-weight:700; color:#E0DCF5; margin-bottom:1rem;">'
                '💰 Implied Enterprise Value (at Peer Median Multiples)</div>',
                unsafe_allow_html=True,
            )
            
            iv_col1, iv_col2, iv_col3 = st.columns(3)
            
            with iv_col1:
                st.markdown(
                    f'<div style="background:rgba(107,92,231,0.1); border:1px solid rgba(107,92,231,0.3); '
                    f'border-radius:12px; padding:1rem; text-align:center;">'
                    f'<div style="font-size:0.75rem; color:#8A85AD;">Current EV</div>'
                    f'<div style="font-size:1.3rem; font-weight:700; color:#E0DCF5;">{format_num(tc.enterprise_value)}</div>'
                    f'</div>',
                    unsafe_allow_html=True,
                )
            
            with iv_col2:
                st.markdown(
                    f'<div style="background:rgba(16,185,129,0.1); border:1px solid rgba(16,185,129,0.3); '
                    f'border-radius:12px; padding:1rem; text-align:center;">'
                    f'<div style="font-size:0.75rem; color:#8A85AD;">Implied EV (EBITDA)</div>'
                    f'<div style="font-size:1.3rem; font-weight:700; color:#E0DCF5;">{format_num(comps_analysis.implied_ev_from_ebitda)}</div>'
                    f'</div>',
                    unsafe_allow_html=True,
                )
            
            with iv_col3:
                st.markdown(
                    f'<div style="background:rgba(245,166,35,0.1); border:1px solid rgba(245,166,35,0.3); '
                    f'border-radius:12px; padding:1rem; text-align:center;">'
                    f'<div style="font-size:0.75rem; color:#8A85AD;">Implied EV (Revenue)</div>'
                    f'<div style="font-size:1.3rem; font-weight:700; color:#E0DCF5;">{format_num(comps_analysis.implied_ev_from_revenue)}</div>'
                    f'</div>',
                    unsafe_allow_html=True,
                )
        
        # Implied Valuation Football Field
        try:
            ff_methods = []
            current_ev = tc.enterprise_value or 0
            
            if comps_analysis.implied_ev_from_ebitda and comps_analysis.implied_ev_from_ebitda > 0:
                ff_methods.append(("EV/EBITDA\n(Peer Median)", comps_analysis.implied_ev_from_ebitda))
            if comps_analysis.implied_ev_from_revenue and comps_analysis.implied_ev_from_revenue > 0:
                ff_methods.append(("EV/Revenue\n(Peer Median)", comps_analysis.implied_ev_from_revenue))
            
            # Add 25th/75th percentile valuations if we can calculate
            if comps_analysis.peers and tc.ev_ebitda:
                peer_ebitda_mults = [p.ev_ebitda for p in comps_analysis.peers if p.ev_ebitda and p.ev_ebitda > 0]
                if peer_ebitda_mults and tc.ev_ebitda > 0:
                    base_ebitda = current_ev / tc.ev_ebitda if tc.ev_ebitda > 0 else 0
                    if base_ebitda > 0:
                        p25 = np.percentile(peer_ebitda_mults, 25) * base_ebitda
                        p75 = np.percentile(peer_ebitda_mults, 75) * base_ebitda
                        ff_methods.append(("EV/EBITDA\n(25th pctile)", p25))
                        ff_methods.append(("EV/EBITDA\n(75th pctile)", p75))
            
            if len(ff_methods) >= 2:
                fig_ff = go.Figure()
                
                names = [m[0] for m in ff_methods]
                values = [m[1] for m in ff_methods]
                colors = ["#6B5CE7", "#E8638B", "#10B981", "#F5A623", "#3B82F6"][:len(ff_methods)]
                
                fig_ff.add_trace(go.Bar(
                    y=names, x=values, orientation="h",
                    marker_color=colors,
                    text=[format_num(v) for v in values],
                    textposition="outside",
                    textfont=dict(size=10, color="#B8B3D7"),
                ))
                
                # Current EV line
                if current_ev > 0:
                    fig_ff.add_vline(x=current_ev, line_dash="dash", line_color="#F59E0B", line_width=2)
                    fig_ff.add_annotation(
                        x=current_ev, y=names[0], text=f"Current EV: {format_num(current_ev)}",
                        showarrow=False, font=dict(size=9, color="#F59E0B"),
                        yshift=30,
                    )
                
                fig_ff.update_layout(
                    **_CHART_LAYOUT_BASE, height=250,
                    margin=dict(t=30, b=30, l=100, r=80),
                    xaxis=dict(tickfont=dict(size=9, color="#8A85AD"), showgrid=False),
                    yaxis=dict(tickfont=dict(size=9, color="#8A85AD")),
                    showlegend=False,
                )
                _apply_space_grid(fig_ff)
                st.plotly_chart(fig_ff, use_container_width=True, key="comps_football_field")
        except Exception:
            pass
        
        st.markdown('<div style="height:1.5rem;"></div>', unsafe_allow_html=True)
        
        # Full Comps Table
        st.markdown(
            '<div style="font-size:1.2rem; font-weight:700; color:#E0DCF5; margin-bottom:1rem;">'
            f'📋 Comparable Companies ({len(comps_analysis.peers)} peers)</div>',
            unsafe_allow_html=True,
        )
        
        comps_df = generate_comps_table(comps_analysis)
        display_df = format_comps_for_display(comps_df)
        
        # Style the dataframe
        st.dataframe(
            display_df,
            use_container_width=True,
            hide_index=True,
            column_config={
                "Company": st.column_config.TextColumn("Company", width="medium"),
                "Ticker": st.column_config.TextColumn("Ticker", width="small"),
                "Market Cap": st.column_config.TextColumn("Mkt Cap", width="small"),
                "EV": st.column_config.TextColumn("EV", width="small"),
                "Revenue": st.column_config.TextColumn("Revenue", width="small"),
                "EBITDA": st.column_config.TextColumn("EBITDA", width="small"),
                "EV/Rev": st.column_config.TextColumn("EV/Rev", width="small"),
                "EV/EBITDA": st.column_config.TextColumn("EV/EBITDA", width="small"),
                "P/E": st.column_config.TextColumn("P/E", width="small"),
                "Rev Growth": st.column_config.TextColumn("Growth", width="small"),
                "EBITDA Margin": st.column_config.TextColumn("Margin", width="small"),
                "Rule of 40": st.column_config.TextColumn("Ro40", width="small"),
            }
        )
        
        st.markdown(
            '<div style="font-size:0.7rem; color:#6B6B80; margin-top:0.5rem;">'
            'Note: Peers selected based on sector, industry, and market cap proximity. '
            'Multiples based on LTM financials from Yahoo Finance.</div>',
            unsafe_allow_html=True,
        )

        # ── Export Comps Table ──
        try:
            _comps_excel = _export_comps_to_excel(comps_analysis)
            st.download_button(
                label=f"📥 Export Comps Table (.xlsx)",
                data=_comps_excel,
                file_name=f"{comps_ticker_input}_Comps_Analysis.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="export_comps_xlsx",
            )
        except Exception:
            pass

        # ── Enhanced: Implied Valuation Range (Football Field by Share Price) ──
        with _safe_section("Implied Valuation Range"):
            st.markdown('<div style="height:1rem;"></div>', unsafe_allow_html=True)
            st.markdown(
                '<div style="font-size:1.2rem; font-weight:700; color:#E0DCF5; margin-bottom:1rem;">'
                '🏈 Implied Share Price — Football Field</div>',
                unsafe_allow_html=True,
            )

            shares_out = tc.market_cap / tc.pe_ratio / (tc.net_income_ltm / tc.market_cap * tc.pe_ratio) if tc.pe_ratio and tc.pe_ratio > 0 and tc.net_income_ltm else None
            # Simpler: shares = market_cap / price_per_share. Use price_to_sales as proxy if available
            if tc.market_cap and tc.pe_ratio and tc.net_income_ltm and tc.net_income_ltm > 0:
                _eps = tc.net_income_ltm / (tc.market_cap / (tc.pe_ratio * (tc.net_income_ltm / tc.market_cap * tc.pe_ratio))) if tc.pe_ratio else 0
            # Actually: shares_outstanding = market_cap / (EV/Revenue * Revenue / EV * market_cap)... let's just derive from PE
            # Current price = market_cap / shares => shares = market_cap / current_price
            # We don't have current_price in CompanyComps directly, but we can infer:
            # P/E = price / EPS => price = P/E * EPS; EPS = net_income / shares => price = P/E * net_income / shares
            # market_cap = price * shares = P/E * net_income => shares = market_cap / (P/E * EPS) = market_cap^2 / (P/E * net_income * market_cap)
            # Simpler: if we have P/S = price/sales_per_share, then price = P/S * revenue / shares
            # Let's try to get current price from yfinance for the target
            _target_price = None
            _target_shares = None
            try:
                _tk_comps = yf.Ticker(tc.ticker)
                _tk_info = _tk_comps.info or {}
                _target_price = _tk_info.get("currentPrice") or _tk_info.get("regularMarketPrice")
                _target_shares = _tk_info.get("sharesOutstanding")
                if not _target_shares and _target_price and _target_price > 0:
                    _target_shares = tc.market_cap / _target_price
            except Exception:
                pass

            if _target_price and _target_shares and _target_shares > 0:
                ff_price_data = {}
                _net_debt = (tc.enterprise_value - tc.market_cap) if tc.enterprise_value and tc.market_cap else 0

                # P/E implied
                if comps_analysis.median_pe and tc.net_income_ltm and tc.net_income_ltm > 0:
                    _pe_vals = [p.pe_ratio for p in comps_analysis.peers if p.pe_ratio and p.pe_ratio > 0]
                    if _pe_vals:
                        _eps = tc.net_income_ltm / _target_shares
                        _low = np.percentile(_pe_vals, 25) * _eps
                        _high = np.percentile(_pe_vals, 75) * _eps
                        _med = comps_analysis.median_pe * _eps
                        if _low > 0 and _high > 0:
                            ff_price_data["P/E"] = {"low": _low, "high": _high, "mid": _med}

                # EV/EBITDA implied
                if comps_analysis.median_ev_ebitda and tc.ebitda_ltm and tc.ebitda_ltm > 0:
                    _ev_ebitda_vals = [p.ev_ebitda for p in comps_analysis.peers if p.ev_ebitda and p.ev_ebitda > 0]
                    if _ev_ebitda_vals:
                        _low_ev = np.percentile(_ev_ebitda_vals, 25) * tc.ebitda_ltm
                        _high_ev = np.percentile(_ev_ebitda_vals, 75) * tc.ebitda_ltm
                        _med_ev = comps_analysis.median_ev_ebitda * tc.ebitda_ltm
                        _low_p = (_low_ev - _net_debt) / _target_shares
                        _high_p = (_high_ev - _net_debt) / _target_shares
                        _med_p = (_med_ev - _net_debt) / _target_shares
                        if _low_p > 0 and _high_p > 0:
                            ff_price_data["EV/EBITDA"] = {"low": _low_p, "high": _high_p, "mid": _med_p}

                # EV/Revenue implied
                if comps_analysis.median_ev_revenue and tc.revenue_ltm and tc.revenue_ltm > 0:
                    _ev_rev_vals = [p.ev_revenue for p in comps_analysis.peers if p.ev_revenue and p.ev_revenue > 0]
                    if _ev_rev_vals:
                        _low_ev = np.percentile(_ev_rev_vals, 25) * tc.revenue_ltm
                        _high_ev = np.percentile(_ev_rev_vals, 75) * tc.revenue_ltm
                        _med_ev = comps_analysis.median_ev_revenue * tc.revenue_ltm
                        _low_p = (_low_ev - _net_debt) / _target_shares
                        _high_p = (_high_ev - _net_debt) / _target_shares
                        _med_p = (_med_ev - _net_debt) / _target_shares
                        if _low_p > 0 and _high_p > 0:
                            ff_price_data["EV/Revenue"] = {"low": _low_p, "high": _high_p, "mid": _med_p}

                # P/S implied
                _ps_vals = [p.price_to_sales for p in comps_analysis.peers if p.price_to_sales and p.price_to_sales > 0]
                if _ps_vals and tc.revenue_ltm and tc.revenue_ltm > 0:
                    _rev_per_share = tc.revenue_ltm / _target_shares
                    _low = np.percentile(_ps_vals, 25) * _rev_per_share
                    _high = np.percentile(_ps_vals, 75) * _rev_per_share
                    _med = np.median(_ps_vals) * _rev_per_share
                    if _low > 0 and _high > 0:
                        ff_price_data["P/S"] = {"low": _low, "high": _high, "mid": _med}

                if ff_price_data:
                    fig_ff2 = go.Figure()
                    labels = list(ff_price_data.keys())
                    colors_ff = ["#6B5CE7", "#E8638B", "#10B981", "#F5A623"]

                    for i, label in enumerate(labels):
                        d = ff_price_data[label]
                        fig_ff2.add_trace(go.Bar(
                            y=[label], x=[d["high"] - d["low"]],
                            base=[d["low"]], orientation="h",
                            marker=dict(color=colors_ff[i % len(colors_ff)], opacity=0.85,
                                        line=dict(color="rgba(255,255,255,0.15)", width=1)),
                            name=label,
                            text=[f"${d['low']:,.0f} — ${d['high']:,.0f}"],
                            textposition="inside",
                            textfont=dict(size=9, color="#fff"),
                            showlegend=False,
                        ))
                        # Median marker
                        fig_ff2.add_trace(go.Scatter(
                            x=[d["mid"]], y=[label], mode="markers",
                            marker=dict(color="#fff", size=8, symbol="diamond",
                                        line=dict(color=colors_ff[i % len(colors_ff)], width=2)),
                            showlegend=False, hovertext=f"Median: ${d['mid']:,.0f}",
                        ))

                    # Current price line
                    fig_ff2.add_vline(x=_target_price, line_dash="dash", line_color="#F59E0B", line_width=2)
                    fig_ff2.add_annotation(
                        x=_target_price, y=labels[0], text=f"Current: ${_target_price:,.2f}",
                        showarrow=False, font=dict(size=9, color="#F59E0B"), yshift=25,
                    )

                    fig_ff2.update_layout(
                        **_CHART_LAYOUT_BASE, height=250,
                        margin=dict(t=30, b=30, l=80, r=60),
                        xaxis=dict(tickprefix="$", tickfont=dict(size=9, color="#8A85AD"), showgrid=False),
                        yaxis=dict(tickfont=dict(size=10, color="#8A85AD")),
                        showlegend=False, barmode="stack",
                    )
                    _apply_space_grid(fig_ff2, show_x_grid=True)
                    st.plotly_chart(fig_ff2, use_container_width=True, key="comps_implied_price_ff")

                    st.markdown(
                        '<div style="font-size:0.7rem; color:#6B6B80; margin-top:0.3rem;">'
                        'Range = 25th–75th percentile of peer multiples applied to target financials. '
                        'Diamond = peer median implied price.</div>',
                        unsafe_allow_html=True,
                    )

        # ── Enhanced: Percentile Ranking Within Peer Group ──
        with _safe_section("Percentile Ranking"):
            st.markdown('<div style="height:1rem;"></div>', unsafe_allow_html=True)
            st.markdown(
                '<div style="font-size:1.2rem; font-weight:700; color:#E0DCF5; margin-bottom:1rem;">'
                '📊 Percentile Ranking vs Peers</div>',
                unsafe_allow_html=True,
            )

            _pctile_metrics = {}
            def _calc_pctile(metric_name, target_val, peer_vals):
                if target_val is not None and peer_vals:
                    valid = [v for v in peer_vals if v is not None]
                    if valid:
                        pctile = sum(1 for v in valid if v < target_val) / len(valid) * 100
                        _pctile_metrics[metric_name] = pctile

            _calc_pctile("EV/Revenue", tc.ev_revenue, [p.ev_revenue for p in comps_analysis.peers if p.ev_revenue])
            _calc_pctile("EV/EBITDA", tc.ev_ebitda, [p.ev_ebitda for p in comps_analysis.peers if p.ev_ebitda])
            _calc_pctile("P/E Ratio", tc.pe_ratio, [p.pe_ratio for p in comps_analysis.peers if p.pe_ratio])
            _calc_pctile("Revenue Growth", tc.revenue_growth, [p.revenue_growth for p in comps_analysis.peers if p.revenue_growth is not None])
            _calc_pctile("Gross Margin", tc.gross_margin, [p.gross_margin for p in comps_analysis.peers if p.gross_margin is not None])
            _calc_pctile("EBITDA Margin", tc.ebitda_margin, [p.ebitda_margin for p in comps_analysis.peers if p.ebitda_margin is not None])
            _calc_pctile("Net Margin", tc.net_margin, [p.net_margin for p in comps_analysis.peers if p.net_margin is not None])

            if _pctile_metrics:
                for metric_name, pctile in _pctile_metrics.items():
                    bar_color = "#10B981" if pctile >= 50 else "#F59E0B" if pctile >= 25 else "#EF4444"
                    # For valuation multiples, lower percentile is better (cheaper)
                    if metric_name in ("EV/Revenue", "EV/EBITDA", "P/E Ratio"):
                        bar_color = "#10B981" if pctile <= 50 else "#F59E0B" if pctile <= 75 else "#EF4444"
                    st.markdown(
                        f'<div style="margin-bottom:0.6rem;">'
                        f'<div style="display:flex; justify-content:space-between; margin-bottom:0.2rem;">'
                        f'<span style="font-size:0.75rem; color:#B8B3D7;">{metric_name}</span>'
                        f'<span style="font-size:0.75rem; font-weight:700; color:{bar_color};">{pctile:.0f}th percentile</span>'
                        f'</div>'
                        f'<div style="background:rgba(255,255,255,0.05); border-radius:4px; height:8px; overflow:hidden;">'
                        f'<div style="width:{pctile:.0f}%; height:100%; background:{bar_color}; border-radius:4px;"></div>'
                        f'</div></div>',
                        unsafe_allow_html=True,
                    )

elif analysis_mode == "Comps Analysis" and comps_btn and not comps_ticker_input:
    st.warning("Please enter a ticker symbol in the sidebar.")

elif analysis_mode == "Merger Analysis" and merger_btn and acquirer_input and target_input:
    # ══════════════════════════════════════════════════════
    # MERGER ANALYSIS DASHBOARD
    # ══════════════════════════════════════════════════════

    # ── Mission Control animated loading ─────────────────
    mission = st.empty()
    acq_label = acquirer_input.upper()
    tgt_label = target_input.upper()

    # Phase 0 → fetch acquirer
    mission.markdown(_render_mission_control(acq_label, tgt_label, 0), unsafe_allow_html=True)
    try:
        acq_cd = fetch_company_data(acquirer_input)
    except Exception as e:
        mission.empty()
        st.error(f"Failed to fetch data for **{acquirer_input}**: {e}")
        st.stop()

    # Ensure acquirer has shares_outstanding (calculate from market_cap/price if needed)
    if not acq_cd.shares_outstanding:
        if acq_cd.market_cap and acq_cd.current_price and acq_cd.current_price > 0:
            acq_cd.shares_outstanding = acq_cd.market_cap / acq_cd.current_price
        elif acq_cd.enterprise_value and acq_cd.current_price and acq_cd.current_price > 0:
            # Rough estimate from EV
            acq_cd.shares_outstanding = acq_cd.enterprise_value / acq_cd.current_price
    if not acq_cd.shares_outstanding:
        mission.empty()
        st.error(f"Unable to determine shares outstanding for **{acquirer_input}**. Market cap: {acq_cd.market_cap}, Price: {acq_cd.current_price}")
        st.stop()

    # Phase 1 → fetch target (with rate limit delay)
    mission.markdown(_render_mission_control(acq_label, tgt_label, 1), unsafe_allow_html=True)
    time.sleep(1)
    try:
        tgt_cd = fetch_company_data(target_input)
    except Exception as e:
        mission.empty()
        st.error(f"Failed to fetch data for **{target_input}**: {e}")
        st.stop()

    # Ensure target has shares_outstanding (calculate from market_cap/price if needed)
    if not tgt_cd.shares_outstanding:
        if tgt_cd.market_cap and tgt_cd.current_price and tgt_cd.current_price > 0:
            tgt_cd.shares_outstanding = tgt_cd.market_cap / tgt_cd.current_price
        elif tgt_cd.enterprise_value and tgt_cd.current_price and tgt_cd.current_price > 0:
            tgt_cd.shares_outstanding = tgt_cd.enterprise_value / tgt_cd.current_price
    if not tgt_cd.shares_outstanding:
        mission.empty()
        st.error(f"Unable to determine shares outstanding for **{target_input}**. Market cap: {tgt_cd.market_cap}, Price: {tgt_cd.current_price}")
        st.stop()

    # Phase 2 → fetch peers
    mission.markdown(_render_mission_control(acq_label, tgt_label, 2), unsafe_allow_html=True)
    try:
        tgt_cd = fetch_peer_data(tgt_cd)
    except Exception:
        pass

    # Phase 3 → compute pro forma + precedent transactions
    mission.markdown(_render_mission_control(acq_label, tgt_label, 3), unsafe_allow_html=True)
    try:
        pro_forma = calculate_pro_forma(acq_cd, tgt_cd, merger_assumptions)
    except Exception as e:
        mission.empty()
        st.error(f"Failed to calculate pro forma: {e}")
        import traceback
        st.code(traceback.format_exc())
        st.stop()

    # Fetch precedent transactions
    precedent = None
    try:
        from precedent_deals import fetch_precedent_transactions
        precedent = fetch_precedent_transactions(
            target_input, getattr(tgt_cd, "cik", ""), tgt_cd.sector
        )
    except Exception as e:
        print(f"Precedent transactions fetch failed: {e}")

    try:
        pro_forma.football_field = build_football_field(acq_cd, tgt_cd, pro_forma, precedent)
    except Exception as e:
        st.warning(f"Football field build failed: {e}")
        pro_forma.football_field = {}

    # Phase 4 → generate insights
    mission.markdown(_render_mission_control(acq_label, tgt_label, 4), unsafe_allow_html=True)
    try:
        merger_insights = generate_merger_insights(acq_cd, tgt_cd, pro_forma, merger_assumptions)
    except Exception as e:
        st.warning(f"Merger insights generation failed: {e}")
        from ai_insights import MergerInsights
        merger_insights = MergerInsights()

    # Phase 5 → mission complete, rocket launches
    mission.markdown(_render_mission_control(acq_label, tgt_label, 5), unsafe_allow_html=True)
    time.sleep(1.5)
    mission.empty()

    acq_cs = acq_cd.currency_symbol
    tgt_cs = tgt_cd.currency_symbol

    # ── Warnings ──────────────────────────────────────────
    for warn in pro_forma.warnings:
        st.warning(warn)

    # Helper: escape $ to prevent Streamlit LaTeX rendering in markdown
    def _mhtml(html_str):
        """Render HTML via st.markdown with $ escaped to prevent LaTeX."""
        st.markdown(html_str.replace("$", "&#36;"), unsafe_allow_html=True)

    # ══════════════════════════════════════════════════════
    # M1. DEAL HEADER
    # ══════════════════════════════════════════════════════
    acq_logo = ""
    if acq_cd.logo_url:
        _ald = getattr(acq_cd, 'logo_domain', '')
        acq_fallback = f"this.onerror=null; this.src='https://logo.clearbit.com/{_ald}';" if _ald else "this.style.display='none';"
        acq_logo = (
            f'<img src="{acq_cd.logo_url}" '
            f'style="width:48px; height:48px; border-radius:10px; object-fit:contain; '
            f'background:white; padding:4px;" onerror="{acq_fallback}">'
        )
    tgt_logo = ""
    if tgt_cd.logo_url:
        _tld = getattr(tgt_cd, 'logo_domain', '')
        tgt_fallback = f"this.onerror=null; this.src='https://logo.clearbit.com/{_tld}';" if _tld else "this.style.display='none';"
        tgt_logo = (
            f'<img src="{tgt_cd.logo_url}" '
            f'style="width:48px; height:48px; border-radius:10px; object-fit:contain; '
            f'background:white; padding:4px;" onerror="{tgt_fallback}">'
        )

    st.markdown(
        f'<div class="company-card">'
        f'<div style="display:flex; align-items:center; gap:1.2rem; position:relative;">'
        f'{acq_logo}'
        f'<div>'
        f'<p class="company-name" style="font-size:1.5rem;">{acq_cd.name}</p>'
        f'<p class="company-meta"><span>{acq_cd.ticker}</span> &middot; {acq_cd.sector}</p>'
        f'</div>'
        f'<div style="font-size:2rem; font-weight:300; color:#6B5CE7; margin:0 1rem;">+</div>'
        f'{tgt_logo}'
        f'<div>'
        f'<p class="company-name" style="font-size:1.5rem;">{tgt_cd.name}</p>'
        f'<p class="company-meta"><span>{tgt_cd.ticker}</span> &middot; {tgt_cd.sector}</p>'
        f'</div>'
        f'</div>'
        f'</div>',
        unsafe_allow_html=True,
    )

    # ══════════════════════════════════════════════════════
    # M2. COMPANY COMPARISON
    # ══════════════════════════════════════════════════════
    _section("Company Comparison")

    cc1, cc2, cc3, cc4, cc5, cc6 = st.columns(6)
    cc1.metric(f"{acq_cd.ticker} Mkt Cap", format_number(acq_cd.market_cap, currency_symbol=acq_cs))
    cc2.metric(f"{tgt_cd.ticker} Mkt Cap", format_number(tgt_cd.market_cap, currency_symbol=tgt_cs))
    cc3.metric(f"{acq_cd.ticker} Revenue", format_number(pro_forma.acq_revenue, currency_symbol=acq_cs))
    cc4.metric(f"{tgt_cd.ticker} Revenue", format_number(pro_forma.tgt_revenue, currency_symbol=tgt_cs))
    cc5.metric(f"{acq_cd.ticker} EBITDA", format_number(pro_forma.acq_ebitda, currency_symbol=acq_cs))
    cc6.metric(f"{tgt_cd.ticker} EBITDA", format_number(pro_forma.tgt_ebitda, currency_symbol=tgt_cs))

    # Company comparison bars
    _mhtml('<div class="merger-chart-wrapper">')
    _build_company_comparison_bars(acq_cd, tgt_cd)
    _mhtml('</div>')

    _divider()

    # ══════════════════════════════════════════════════════
    # M3. DEAL TERMS
    # ══════════════════════════════════════════════════════
    _section("Deal Terms")

    dt1, dt2, dt3, dt4, dt5 = st.columns(5)
    dt1.metric("Purchase Price", format_number(pro_forma.purchase_price, currency_symbol=acq_cs))
    dt2.metric("Offer Premium", f"{merger_assumptions.offer_premium_pct:.0f}%")
    dt3.metric("Implied EV/EBITDA", f"{pro_forma.implied_ev_ebitda:.1f}x" if pro_forma.implied_ev_ebitda else "N/A")
    dt4.metric("Implied P/E", f"{pro_forma.implied_pe:.1f}x" if pro_forma.implied_pe else "N/A")
    dt5.metric("Transaction Fees", format_number(pro_forma.transaction_fees, currency_symbol=acq_cs))

    # Deal structure donut + enhanced consideration detail
    deal_col1, deal_col2 = st.columns([2, 3])
    with deal_col1:
        _build_deal_structure_donut(merger_assumptions)
    with deal_col2:
        cash_pct = merger_assumptions.pct_cash
        stock_pct = 100 - cash_pct
        premium_pct = ((pro_forma.offer_price_per_share / tgt_cd.current_price) - 1) * 100 if tgt_cd.current_price else 0
        _mhtml(
            f'<div class="deal-consideration-card">'
            f'<div class="deal-header"><span class="deal-header-icon">💰</span> Consideration Structure</div>'
            f'<div class="deal-consideration-row cash" style="animation-delay:0.1s;">'
            f'<div><div class="deal-label"><span class="emoji">💵</span> Cash Component</div>'
            f'<div class="deal-sub">Debt-funded • {cash_pct:.0f}% of deal</div></div>'
            f'<div class="deal-value">{format_number(pro_forma.cash_consideration, currency_symbol=acq_cs)}</div>'
            f'</div>'
            f'<div class="deal-consideration-row stock" style="animation-delay:0.2s;">'
            f'<div><div class="deal-label"><span class="emoji">📈</span> Stock Component</div>'
            f'<div class="deal-sub">{pro_forma.new_shares_issued / 1e6:,.1f}M shares @ {acq_cs}{acq_cd.current_price:,.2f} • {stock_pct:.0f}% of deal</div></div>'
            f'<div class="deal-value">{format_number(pro_forma.stock_consideration, currency_symbol=acq_cs)}</div>'
            f'</div>'
            f'<div class="deal-consideration-row offer" style="animation-delay:0.3s;">'
            f'<div><div class="deal-label"><span class="emoji">🎯</span> Offer Price</div>'
            f'<div class="deal-sub">+{premium_pct:.1f}% premium vs. current {tgt_cs}{tgt_cd.current_price:,.2f}</div></div>'
            f'<div class="deal-value">{acq_cs}{pro_forma.offer_price_per_share:,.2f}</div>'
            f'</div>'
            f'</div>'
        )

    _divider()

    # ══════════════════════════════════════════════════════
    # M3b. IMPLIED MULTIPLES AT OFFER PRICE
    # ══════════════════════════════════════════════════════
    _section("Implied Multiples at Offer Price", "📐")
    
    # Calculate implied multiples
    offer_equity_value = pro_forma.purchase_price if pro_forma.purchase_price else 0
    tgt_net_debt = (getattr(tgt_cd, 'total_debt', 0) or 0) - (getattr(tgt_cd, 'total_cash', 0) or 0)
    offer_ev = offer_equity_value + tgt_net_debt
    
    tgt_revenue_val = pro_forma.tgt_revenue or 0
    tgt_ebitda_val = pro_forma.tgt_ebitda or 0
    tgt_ni_val = pro_forma.tgt_net_income or 0
    
    implied_ev_rev = offer_ev / tgt_revenue_val if tgt_revenue_val > 0 else 0
    implied_ev_ebitda = offer_ev / tgt_ebitda_val if tgt_ebitda_val > 0 else 0
    implied_pe = offer_equity_value / tgt_ni_val if tgt_ni_val > 0 else 0
    
    # Current multiples
    curr_ev = tgt_cd.enterprise_value or 0
    curr_ev_rev = curr_ev / tgt_revenue_val if tgt_revenue_val > 0 else 0
    curr_ev_ebitda = curr_ev / tgt_ebitda_val if tgt_ebitda_val > 0 else 0
    curr_pe = tgt_cd.trailing_pe or 0
    
    mult_table = (
        '<table class="pf-table">'
        '<thead><tr><th>Multiple</th><th>Current</th><th>At Offer Price</th><th>Premium Paid</th></tr></thead>'
        '<tbody>'
    )
    
    multiples_data = [
        ("EV / Revenue", curr_ev_rev, implied_ev_rev),
        ("EV / EBITDA", curr_ev_ebitda, implied_ev_ebitda),
        ("P / E", curr_pe, implied_pe),
    ]
    
    for name, curr, implied in multiples_data:
        prem = ((implied / curr - 1) * 100) if curr > 0 and implied > 0 else 0
        prem_color = "#EF4444" if prem > 50 else "#F59E0B" if prem > 20 else "#10B981"
        mult_table += (
            f'<tr>'
            f'<td style="font-weight:600;">{name}</td>'
            f'<td>{curr:.1f}x</td>'
            f'<td style="color:#6B5CE7; font-weight:700;">{implied:.1f}x</td>'
            f'<td style="color:{prem_color}; font-weight:700;">+{prem:.0f}%</td>'
            f'</tr>'
        )
    
    mult_table += '</tbody></table>'
    _mhtml(f'<div class="pf-table-wrapper">{mult_table}</div>')
    
    _divider()

    # ══════════════════════════════════════════════════════
    # M3c. CONTRIBUTION ANALYSIS
    # ══════════════════════════════════════════════════════
    _section("Contribution Analysis", "📊")
    st.markdown(
        '<div style="font-size:0.8rem; color:#B8B3D7; margin-bottom:0.8rem;">'
        'What percentage does each company contribute to the combined entity?</div>',
        unsafe_allow_html=True,
    )
    
    acq_rev = pro_forma.acq_revenue or 0
    tgt_rev = pro_forma.tgt_revenue or 0
    acq_ebitda = pro_forma.acq_ebitda or 0
    tgt_ebitda = pro_forma.tgt_ebitda or 0
    acq_ni = pro_forma.acq_net_income or 0
    tgt_ni = pro_forma.tgt_net_income or 0
    
    contrib_metrics = []
    for label, acq_v, tgt_v in [
        ("Revenue", acq_rev, tgt_rev),
        ("EBITDA", acq_ebitda, tgt_ebitda),
        ("Net Income", acq_ni, tgt_ni),
        ("Market Cap", acq_cd.market_cap or 0, tgt_cd.market_cap or 0),
    ]:
        total = acq_v + tgt_v
        acq_pct = (acq_v / total * 100) if total > 0 else 0
        tgt_pct = 100 - acq_pct
        contrib_metrics.append((label, acq_pct, tgt_pct))
    
    # Ownership split
    acq_own_pct = (pro_forma.acq_shares / pro_forma.pf_shares_outstanding * 100) if pro_forma.pf_shares_outstanding else 0
    tgt_own_pct = 100 - acq_own_pct
    contrib_metrics.append(("Pro Forma Ownership", acq_own_pct, tgt_own_pct))
    
    contrib_html = '<div style="display:grid; gap:0.8rem;">'
    for label, acq_pct, tgt_pct in contrib_metrics:
        contrib_html += (
            f'<div>'
            f'<div style="display:flex; justify-content:space-between; margin-bottom:0.2rem;">'
            f'<span style="font-size:0.72rem; color:#8A85AD; font-weight:600;">{label}</span>'
            f'<span style="font-size:0.65rem; color:#8A85AD;">'
            f'{acq_cd.ticker}: {acq_pct:.0f}% | {tgt_cd.ticker}: {tgt_pct:.0f}%</span>'
            f'</div>'
            f'<div style="display:flex; height:24px; border-radius:6px; overflow:hidden; '
            f'border:1px solid rgba(255,255,255,0.05);">'
            f'<div style="width:{acq_pct}%; background:linear-gradient(90deg, #6B5CE7, #9B8AFF); '
            f'display:flex; align-items:center; justify-content:center; font-size:0.6rem; color:#fff; font-weight:700;">'
            f'{acq_pct:.0f}%</div>'
            f'<div style="width:{tgt_pct}%; background:linear-gradient(90deg, #E8638B, #F5A0B8); '
            f'display:flex; align-items:center; justify-content:center; font-size:0.6rem; color:#fff; font-weight:700;">'
            f'{tgt_pct:.0f}%</div>'
            f'</div></div>'
        )
    contrib_html += '</div>'
    _mhtml(contrib_html)
    
    _divider()
    
    # ══════════════════════════════════════════════════════
    # M3d. GOODWILL & PURCHASE PRICE ALLOCATION
    # ══════════════════════════════════════════════════════
    _section("Purchase Price Allocation", "🏷️")
    
    tgt_book_value = getattr(tgt_cd, 'book_value', None)
    tgt_total_equity = None
    if tgt_cd.balance_sheet is not None and not tgt_cd.balance_sheet.empty:
        for idx_name in tgt_cd.balance_sheet.index:
            if "stockholder" in str(idx_name).lower() or "total equity" in str(idx_name).lower():
                try:
                    tgt_total_equity = float(tgt_cd.balance_sheet.loc[idx_name].iloc[0])
                except Exception:
                    pass
                break
    
    if not tgt_total_equity and tgt_book_value and tgt_cd.shares_outstanding:
        tgt_total_equity = tgt_book_value * tgt_cd.shares_outstanding
    
    if tgt_total_equity and offer_equity_value:
        goodwill = offer_equity_value - tgt_total_equity
        goodwill_pct = (goodwill / offer_equity_value * 100) if offer_equity_value > 0 else 0
        
        ppa_col1, ppa_col2, ppa_col3 = st.columns(3)
        ppa_col1.metric("Purchase Price", format_number(offer_equity_value, currency_symbol=acq_cs))
        ppa_col2.metric("Target Book Value", format_number(tgt_total_equity, currency_symbol=acq_cs))
        ppa_col3.metric("Implied Goodwill", format_number(goodwill, currency_symbol=acq_cs))
        
        # Goodwill waterfall
        fig_gw = go.Figure(go.Waterfall(
            x=["Book Value", "Intangibles &<br>Goodwill", "Purchase Price"],
            y=[tgt_total_equity, goodwill, 0],
            measure=["absolute", "relative", "total"],
            connector=dict(line=dict(color="rgba(107,92,231,0.3)")),
            increasing=dict(marker_color="#E8638B"),
            totals=dict(marker_color="#6B5CE7"),
            text=[format_number(tgt_total_equity, currency_symbol=acq_cs),
                  format_number(goodwill, currency_symbol=acq_cs),
                  format_number(offer_equity_value, currency_symbol=acq_cs)],
            textposition="outside",
            textfont=dict(size=10, color="#B8B3D7"),
        ))
        fig_gw.update_layout(
            **_CHART_LAYOUT_BASE, height=300,
            margin=dict(t=30, b=30, l=50, r=30),
            yaxis=dict(tickfont=dict(size=9, color="#8A85AD"), visible=False),
            xaxis=dict(tickfont=dict(size=10, color="#8A85AD")),
            showlegend=False,
        )
        _apply_space_grid(fig_gw)
        st.plotly_chart(fig_gw, use_container_width=True, key="goodwill_waterfall")
        
        st.markdown(
            f'<div style="text-align:center; font-size:0.75rem; color:#8A85AD;">'
            f'Goodwill represents {goodwill_pct:.0f}% of total purchase price</div>',
            unsafe_allow_html=True,
        )
    else:
        st.info("Insufficient balance sheet data for purchase price allocation.")
    
    _divider()

    # ══════════════════════════════════════════════════════
    # M3e. SYNERGY NPV ANALYSIS
    # ══════════════════════════════════════════════════════
    _section("Synergy Value Analysis", "⚡")
    
    total_syn = pro_forma.total_synergies or 0
    cost_syn_val = pro_forma.cost_synergies or 0
    rev_syn_val = pro_forma.revenue_synergies or 0
    tax_rate = merger_assumptions.tax_rate / 100
    wacc = merger_assumptions.cost_of_debt / 100  # Using cost of debt as proxy
    
    # Assume synergies phase in over 3 years (33%, 66%, 100%)
    phase_in = [0.33, 0.66, 1.0, 1.0, 1.0]
    syn_pv = 0
    syn_timeline = []
    for yr, pct in enumerate(phase_in, 1):
        yr_syn = total_syn * pct * (1 - tax_rate)
        pv = yr_syn / (1 + wacc) ** yr
        syn_pv += pv
        syn_timeline.append({"year": yr, "synergy": yr_syn, "pv": pv, "phase_pct": pct * 100})
    
    syn_col1, syn_col2, syn_col3, syn_col4 = st.columns(4)
    syn_col1.metric("Cost Synergies (Annual)", format_number(cost_syn_val, currency_symbol=acq_cs))
    syn_col2.metric("Revenue Synergies (Annual)", format_number(rev_syn_val, currency_symbol=acq_cs))
    syn_col3.metric("Total AT Synergies", format_number(total_syn * (1 - tax_rate), currency_symbol=acq_cs))
    syn_col4.metric("NPV of Synergies (5Y)", format_number(syn_pv, currency_symbol=acq_cs))
    
    # Synergy phase-in chart
    fig_syn = go.Figure()
    fig_syn.add_trace(go.Bar(
        x=[f"Year {s['year']}" for s in syn_timeline],
        y=[s["synergy"] for s in syn_timeline],
        name="AT Synergies",
        marker_color="rgba(16,185,129,0.6)",
        text=[f"{s['phase_pct']:.0f}%" for s in syn_timeline],
        textposition="outside",
        textfont=dict(size=10, color="#10B981"),
    ))
    fig_syn.add_trace(go.Scatter(
        x=[f"Year {s['year']}" for s in syn_timeline],
        y=[s["pv"] for s in syn_timeline],
        mode="lines+markers",
        name="PV of Synergies",
        line=dict(color="#6B5CE7", width=2),
        marker=dict(size=8, color="#6B5CE7"),
    ))
    fig_syn.update_layout(
        **_CHART_LAYOUT_BASE, height=300,
        margin=dict(t=20, b=30, l=50, r=30),
        yaxis=dict(tickfont=dict(size=9, color="#8A85AD")),
        xaxis=dict(tickfont=dict(size=10, color="#8A85AD"), showgrid=False),
        legend=dict(orientation="h", yanchor="bottom", y=1.02, font=dict(size=9, color="#B8B3D7")),
        barmode="group",
    )
    _apply_space_grid(fig_syn)
    st.plotly_chart(fig_syn, use_container_width=True, key="synergy_phasing")
    
    # Premium paid vs synergy value
    if syn_pv > 0 and offer_equity_value > 0:
        premium_paid = offer_equity_value - (tgt_cd.market_cap or 0)
        syn_coverage = (syn_pv / premium_paid * 100) if premium_paid > 0 else 0
        cov_color = "#10B981" if syn_coverage > 100 else "#F59E0B" if syn_coverage > 60 else "#EF4444"
        st.markdown(
            f'<div style="text-align:center; padding:0.8rem; background:rgba(107,92,231,0.05); '
            f'border-radius:12px; margin-top:0.5rem;">'
            f'<span style="font-size:0.7rem; color:#8A85AD;">Premium Paid: {format_number(premium_paid, currency_symbol=acq_cs)} | '
            f'Synergy NPV covers </span>'
            f'<span style="font-size:1.2rem; font-weight:800; color:{cov_color};">{syn_coverage:.0f}%</span>'
            f'<span style="font-size:0.7rem; color:#8A85AD;"> of premium</span>'
            f'</div>',
            unsafe_allow_html=True,
        )

    _divider()

    # ══════════════════════════════════════════════════════
    # M4. PRO FORMA FINANCIALS
    # ══════════════════════════════════════════════════════
    _section("Pro Forma Financials")

    tax_r = merger_assumptions.tax_rate / 100
    ats = pro_forma.total_synergies * (1 - tax_r)
    ati = pro_forma.incremental_interest * (1 - tax_r)

    # Enhanced visual Pro Forma table
    pf_rows = [
        ("📊 Revenue",
         format_number(pro_forma.acq_revenue, currency_symbol=acq_cs),
         format_number(pro_forma.tgt_revenue, currency_symbol=tgt_cs),
         format_number(pro_forma.revenue_synergies, currency_symbol=acq_cs),
         format_number(pro_forma.pf_revenue, currency_symbol=acq_cs)),
        ("💹 EBITDA",
         format_number(pro_forma.acq_ebitda, currency_symbol=acq_cs),
         format_number(pro_forma.tgt_ebitda, currency_symbol=tgt_cs),
         format_number(pro_forma.total_synergies, currency_symbol=acq_cs),
         format_number(pro_forma.pf_ebitda, currency_symbol=acq_cs)),
        ("💰 Net Income",
         format_number(pro_forma.acq_net_income, currency_symbol=acq_cs),
         format_number(pro_forma.tgt_net_income, currency_symbol=tgt_cs),
         format_number(ats - ati, currency_symbol=acq_cs),
         format_number(pro_forma.pf_net_income, currency_symbol=acq_cs)),
        ("📈 Shares (M)",
         f"{pro_forma.acq_shares / 1e6:,.0f}" if pro_forma.acq_shares else "N/A",
         "—",
         f"+{pro_forma.new_shares_issued / 1e6:,.0f}" if pro_forma.new_shares_issued else "—",
         f"{pro_forma.pf_shares_outstanding / 1e6:,.0f}" if pro_forma.pf_shares_outstanding else "N/A"),
        ("🎯 EPS",
         f"{acq_cs}{pro_forma.acq_eps:.2f}" if pro_forma.acq_eps else "N/A",
         "—",
         "—",
         f"{acq_cs}{pro_forma.pf_eps:.2f}" if pro_forma.pf_eps else "N/A"),
    ]

    pf_table_html = (
        f'<div class="pf-table-wrapper">'
        f'<table class="pf-table">'
        f'<thead><tr>'
        f'<th>Metric</th>'
        f'<th>{acq_cd.ticker}</th>'
        f'<th>{tgt_cd.ticker}</th>'
        f'<th>Adjustments</th>'
        f'<th>✨ Pro Forma</th>'
        f'</tr></thead>'
        f'<tbody>'
    )
    for metric, acq_val, tgt_val, adj_val, pf_val in pf_rows:
        adj_class = ' class="pf-adj"' if adj_val not in ["—", "N/A"] else ''
        pf_table_html += (
            f'<tr>'
            f'<td>{metric}</td>'
            f'<td>{acq_val}</td>'
            f'<td>{tgt_val}</td>'
            f'<td{adj_class}>{adj_val}</td>'
            f'<td>{pf_val}</td>'
            f'</tr>'
        )
    pf_table_html += '</tbody></table></div>'
    _mhtml(pf_table_html)

    _divider()

    # ══════════════════════════════════════════════════════
    # M5. ACCRETION / DILUTION
    # ══════════════════════════════════════════════════════
    _section("Accretion / Dilution Analysis")

    acc_color = "#10B981" if pro_forma.is_accretive else "#EF4444"
    acc_word = "ACCRETIVE" if pro_forma.is_accretive else "DILUTIVE"
    acc_bg = "rgba(16,185,129,0.08)" if pro_forma.is_accretive else "rgba(239,68,68,0.08)"

    _mhtml(
        f'<div style="text-align:center; padding:1rem; background:{acc_bg}; border-radius:14px; margin-bottom:1rem;">'
        f'<div style="font-size:0.7rem; font-weight:600; color:#8A85AD; text-transform:uppercase; letter-spacing:1px;">EPS Impact</div>'
        f'<div style="font-size:2.5rem; font-weight:800; color:{acc_color};">{pro_forma.accretion_dilution_pct:+.1f}%</div>'
        f'<div style="font-size:1rem; font-weight:700; color:{acc_color};">{acc_word}</div>'
        f'<div style="font-size:0.8rem; color:#B8B3D7; margin-top:0.3rem;">'
        f'Standalone: {acq_cs}{pro_forma.acq_eps:.2f} &rarr; Pro Forma: {acq_cs}{pro_forma.pf_eps:.2f}</div>'
        f'</div>'
    )

    _mhtml('<div class="merger-chart-wrapper">')
    _build_accretion_waterfall(pro_forma)
    _mhtml('</div>')

    # Premium Sensitivity on Accretion/Dilution
    try:
        _section("Premium Sensitivity", "🎚️")
        st.markdown(
            '<div style="font-size:0.8rem; color:#B8B3D7; margin-bottom:0.8rem;">'
            'How does the offer premium affect EPS accretion/dilution?</div>',
            unsafe_allow_html=True,
        )
        
        premiums = [10, 15, 20, 25, 30, 35, 40, 50, 60]
        ad_results = []
        for p in premiums:
            test_assumptions = MergerAssumptions(
                offer_premium_pct=p,
                pct_cash=merger_assumptions.pct_cash,
                pct_stock=merger_assumptions.pct_stock,
                cost_synergies_pct=merger_assumptions.cost_synergies_pct,
                revenue_synergies_pct=merger_assumptions.revenue_synergies_pct,
                transaction_fees_pct=merger_assumptions.transaction_fees_pct,
                tax_rate=merger_assumptions.tax_rate,
                cost_of_debt=merger_assumptions.cost_of_debt,
            )
            try:
                test_pf = calculate_pro_forma(acq_cd, tgt_cd, test_assumptions)
                ad_results.append({"premium": p, "ad_pct": test_pf.accretion_dilution_pct or 0})
            except Exception:
                ad_results.append({"premium": p, "ad_pct": 0})
        
        if ad_results:
            fig_prem = go.Figure()
            colors = ["#10B981" if r["ad_pct"] >= 0 else "#EF4444" for r in ad_results]
            fig_prem.add_trace(go.Bar(
                x=[f"{r['premium']}%" for r in ad_results],
                y=[r["ad_pct"] for r in ad_results],
                marker_color=colors,
                text=[f"{r['ad_pct']:+.1f}%" for r in ad_results],
                textposition="outside",
                textfont=dict(size=9, color="#B8B3D7"),
            ))
            # Highlight current premium
            curr_idx = None
            for i, r in enumerate(ad_results):
                if r["premium"] == merger_assumptions.offer_premium_pct:
                    curr_idx = i
                    break
            if curr_idx is not None:
                fig_prem.add_annotation(
                    x=f"{merger_assumptions.offer_premium_pct}%", y=ad_results[curr_idx]["ad_pct"],
                    text="Current", showarrow=True, arrowhead=2, arrowcolor="#F59E0B",
                    font=dict(size=10, color="#F59E0B"), ax=0, ay=-30,
                )
            
            fig_prem.add_hline(y=0, line_dash="dash", line_color="rgba(255,255,255,0.2)", line_width=1)
            fig_prem.update_layout(
                **_CHART_LAYOUT_BASE, height=300,
                margin=dict(t=30, b=30, l=50, r=30),
                xaxis=dict(title=dict(text="Offer Premium", font=dict(size=11, color="#8A85AD")),
                          tickfont=dict(size=10, color="#8A85AD"), showgrid=False),
                yaxis=dict(title=dict(text="EPS Accretion/Dilution %", font=dict(size=11, color="#8A85AD")),
                          ticksuffix="%", tickfont=dict(size=9, color="#8A85AD")),
                showlegend=False,
            )
            _apply_space_grid(fig_prem)
            st.plotly_chart(fig_prem, use_container_width=True, key="premium_sensitivity")
    except Exception:
        pass  # Premium sensitivity is non-critical

    _divider()

    # ══════════════════════════════════════════════════════
    # M5b. DEAL FINANCING MIX SENSITIVITY
    # ══════════════════════════════════════════════════════
    with _safe_section("Deal Financing Mix"):
        _section("Deal Financing Mix Sensitivity", "💸")
        st.markdown(
            '<div style="font-size:0.8rem; color:#B8B3D7; margin-bottom:0.8rem;">'
            'Explore how different cash/debt/stock mixes and synergy levels affect EPS accretion/dilution.</div>',
            unsafe_allow_html=True,
        )

        fmix_c1, fmix_c2, fmix_c3 = st.columns(3)
        with fmix_c1:
            fm_cash = st.slider("% Cash", 0, 100, int(merger_assumptions.pct_cash), 5, key="fm_cash_slider")
        with fmix_c2:
            fm_debt = st.slider("% Debt", 0, 100 - fm_cash, min(100 - fm_cash, 0), 5, key="fm_debt_slider")
        with fmix_c3:
            fm_stock = 100 - fm_cash - fm_debt
            st.markdown(
                f'<div style="text-align:center; padding:1.5rem; background:rgba(255,255,255,0.04); border-radius:10px;">'
                f'<div style="font-size:0.65rem; font-weight:600; color:#8A85AD; text-transform:uppercase;">Stock (remainder)</div>'
                f'<div style="font-size:2rem; font-weight:800; color:#6B5CE7;">{fm_stock}%</div></div>',
                unsafe_allow_html=True,
            )

        # Financing mix vs Synergies sensitivity matrix
        synergy_levels = [0, 5, 10, 15, 20, 25]
        cash_levels = [0, 25, 50, 75, 100]

        matrix_data = []
        for cash_pct_test in cash_levels:
            row = {"Cash %": f"{cash_pct_test}%"}
            for syn_pct in synergy_levels:
                try:
                    test_a = MergerAssumptions(
                        offer_premium_pct=merger_assumptions.offer_premium_pct,
                        pct_cash=cash_pct_test,
                        pct_stock=100 - cash_pct_test,
                        cost_synergies_pct=syn_pct,
                        revenue_synergies_pct=merger_assumptions.revenue_synergies_pct,
                        transaction_fees_pct=merger_assumptions.transaction_fees_pct,
                        tax_rate=merger_assumptions.tax_rate,
                        cost_of_debt=merger_assumptions.cost_of_debt,
                    )
                    test_pf = calculate_pro_forma(acq_cd, tgt_cd, test_a)
                    row[f"Syn {syn_pct}%"] = f"{test_pf.accretion_dilution_pct:+.1f}%"
                except Exception:
                    row[f"Syn {syn_pct}%"] = "—"
            matrix_data.append(row)

        if matrix_data:
            fm_df = pd.DataFrame(matrix_data)
            # Build styled HTML table
            fm_th = "".join(
                f'<th style="padding:0.4rem 0.5rem; font-weight:700; color:#6B5CE7; font-size:0.7rem; '
                f'text-align:center; border-bottom:2px solid rgba(107,92,231,0.3);">{c}</th>'
                for c in fm_df.columns
            )
            fm_tbody = ""
            for _, r in fm_df.iterrows():
                cells = ""
                for j, c in enumerate(fm_df.columns):
                    val = r[c]
                    if j == 0:
                        cells += f'<td style="padding:0.35rem 0.5rem; font-size:0.75rem; color:#B8B3D7; font-weight:600; text-align:center;">{val}</td>'
                    else:
                        _c = "#10B981" if val.startswith("+") else "#EF4444" if val.startswith("-") else "#8A85AD"
                        cells += f'<td style="padding:0.35rem 0.5rem; font-size:0.75rem; color:{_c}; font-weight:600; text-align:center;">{val}</td>'
                fm_tbody += f'<tr style="border-bottom:1px solid rgba(255,255,255,0.04);">{cells}</tr>'

            _mhtml(
                f'<div style="background:rgba(255,255,255,0.02); border:1px solid rgba(107,92,231,0.1); '
                f'border-radius:10px; overflow:hidden; margin-bottom:1rem;">'
                f'<div style="padding:0.6rem 0.8rem; background:rgba(107,92,231,0.06); '
                f'border-bottom:1px solid rgba(107,92,231,0.1);">'
                f'<span style="font-size:0.7rem; font-weight:700; color:#6B5CE7; text-transform:uppercase; '
                f'letter-spacing:1px;">Cash % vs Cost Synergy % → EPS Accretion/Dilution</span></div>'
                f'<table style="width:100%; border-collapse:collapse;">'
                f'<thead><tr>{fm_th}</tr></thead>'
                f'<tbody>{fm_tbody}</tbody></table></div>'
            )

    _divider()

    # ══════════════════════════════════════════════════════
    # M5c. COMPARABLE TRANSACTIONS (from M&A history)
    # ══════════════════════════════════════════════════════
    with _safe_section("Comparable Transactions"):
        # Show acquirer/target M&A history as comparable transactions if available
        acq_ma = getattr(acq_cd, 'ma_history', None) or []
        tgt_ma = getattr(tgt_cd, 'ma_history', None) or []
        all_ma = acq_ma + tgt_ma
        if all_ma:
            _section("Comparable Transactions (M&A History)", "🔄")
            ct_rows = ""
            for deal in all_ma[:20]:
                d_date = deal.get("date", deal.get("year", ""))
                d_acquirer = deal.get("acquirer", deal.get("buyer", ""))
                d_target = deal.get("target", deal.get("name", ""))
                d_value = deal.get("deal_value", deal.get("value", None))
                d_ev_ebitda = deal.get("ev_ebitda", None)
                d_ev_rev = deal.get("ev_revenue", None)
                d_premium = deal.get("premium", None)
                val_str = format_number(d_value, currency_symbol=tgt_cs) if d_value else "—"
                ev_eb_str = f"{d_ev_ebitda:.1f}x" if d_ev_ebitda else "—"
                ev_rev_str = f"{d_ev_rev:.1f}x" if d_ev_rev else "—"
                prem_str = f"{d_premium:.0f}%" if d_premium else "—"
                ct_rows += (
                    f"<tr><td>{d_date}</td><td>{d_acquirer}</td><td>{d_target}</td>"
                    f"<td>{val_str}</td><td>{ev_eb_str}</td><td>{ev_rev_str}</td><td>{prem_str}</td></tr>"
                )
            if ct_rows:
                _mhtml(
                    f'<table class="precedent-table">'
                    f'<thead><tr><th>Date</th><th>Acquirer</th><th>Target</th>'
                    f'<th>Deal Value</th><th>EV/EBITDA</th><th>EV/Revenue</th><th>Premium</th></tr></thead>'
                    f'<tbody>{ct_rows}</tbody></table>'
                )

    _divider()

    # ══════════════════════════════════════════════════════
    # M6. FOOTBALL FIELD VALUATION
    # ══════════════════════════════════════════════════════
    if pro_forma.football_field and len([k for k in pro_forma.football_field if not k.startswith("_")]) > 0:
        _section("Football Field Valuation")
        _mhtml('<div class="merger-chart-wrapper">')
        _build_football_field_chart(pro_forma.football_field, acq_cs)
        _mhtml('</div>')
        _divider()

    # ══════════════════════════════════════════════════════
    # M6b. PRECEDENT TRANSACTIONS TABLE
    # ══════════════════════════════════════════════════════
    if precedent and precedent.deals:
        _section("Precedent Transactions")
        rows_html = ""
        for d in precedent.deals[:15]:
            name = d.get("name", d.get("target", ""))
            date = d.get("date", "")
            ev_eb = d.get("ev_ebitda")
            ev_rev = d.get("ev_revenue")
            dval = d.get("deal_value")
            ev_eb_str = f"{ev_eb:.1f}x" if ev_eb else "—"
            ev_rev_str = f"{ev_rev:.1f}x" if ev_rev else "—"
            dval_str = format_number(dval, currency_symbol=tgt_cs) if dval else "—"
            rows_html += (
                f"<tr><td>{date}</td><td>{name}</td>"
                f"<td>{dval_str}</td><td>{ev_eb_str}</td><td>{ev_rev_str}</td></tr>"
            )
        source_note = ""
        if precedent.source_url:
            source_note = f'<div style="font-size:0.7rem; color:#8A85AD; margin-top:0.5rem;">Source: {precedent.source} — <a href="{precedent.source_url}" style="color:#9B8AFF;" target="_blank">Filing</a></div>'
        elif precedent.source:
            source_note = f'<div style="font-size:0.7rem; color:#8A85AD; margin-top:0.5rem;">Source: {precedent.source}</div>'
        _mhtml(
            f'<table class="precedent-table">'
            f'<thead><tr><th>Date</th><th>Transaction</th>'
            f'<th>Deal Value</th><th>EV/EBITDA</th><th>EV/Revenue</th></tr></thead>'
            f'<tbody>{rows_html}</tbody></table>'
            f'{source_note}'
        )
        _divider()

    # ══════════════════════════════════════════════════════
    # M7. SOURCES & USES
    # ══════════════════════════════════════════════════════
    _section("Sources & Uses")

    # Calculate totals for bar percentages
    sources_total = sum(v for k, v in pro_forma.sources.items() if not k.startswith("Total") and v)
    uses_total = sum(v for k, v in pro_forma.uses.items() if not k.startswith("Total") and v)

    def _build_su_rows(items, total):
        rows_html = ""
        delay = 0.1
        for k, v in items.items():
            is_total = k.startswith("Total")
            pct = (v / total * 100) if total and v else 0
            total_class = " total" if is_total else ""
            val_str = format_number(v, currency_symbol=acq_cs)
            rows_html += (
                f'<div class="su-row{total_class}" style="animation-delay:{delay:.2f}s;">'
                f'<div class="su-row-header">'
                f'<span class="su-row-label">{k}</span>'
                f'<span class="su-row-value">{val_str}</span>'
                f'</div>'
                f'<div class="su-bar">'
                f'<div class="su-bar-fill" style="width:{pct:.1f}%; animation-delay:{delay + 0.2:.2f}s;"></div>'
                f'</div>'
                f'</div>'
            )
            delay += 0.08
        return rows_html

    sources_rows = _build_su_rows(pro_forma.sources, sources_total)
    uses_rows = _build_su_rows(pro_forma.uses, uses_total)

    su_html = (
        f'<div class="su-container">'
        f'<div class="su-panel sources">'
        f'<div class="su-panel-header"><span class="su-icon">💵</span> Sources of Funds</div>'
        f'{sources_rows}'
        f'</div>'
        f'<div class="su-panel uses">'
        f'<div class="su-panel-header"><span class="su-icon">💸</span> Uses of Funds</div>'
        f'{uses_rows}'
        f'</div>'
        f'</div>'
    )
    _mhtml(su_html)

    _divider()

    # ══════════════════════════════════════════════════════
    # M8. PRO FORMA CREDIT
    # ══════════════════════════════════════════════════════
    _section("Pro Forma Credit Profile")

    cr1, cr2, cr3, cr4 = st.columns(4)

    def _lev_color(val):
        if val is None: return "#8A85AD"
        if val < 2: return "#10B981"
        if val < 4: return "#F5A623"
        return "#EF4444"

    def _cov_color(val):
        if val is None: return "#8A85AD"
        if val > 5: return "#10B981"
        if val > 2.5: return "#F5A623"
        return "#EF4444"

    lev_c = _lev_color(pro_forma.pf_leverage_ratio)
    cov_c = _cov_color(pro_forma.pf_interest_coverage)

    cr1.metric("PF Debt / EBITDA", f"{pro_forma.pf_leverage_ratio:.1f}x" if pro_forma.pf_leverage_ratio else "N/A")
    cr2.metric("PF Interest Coverage", f"{pro_forma.pf_interest_coverage:.1f}x" if pro_forma.pf_interest_coverage else "N/A")
    cr3.metric("PF Total Debt", format_number(pro_forma.pf_total_debt, currency_symbol=acq_cs))
    cr4.metric("PF Net Debt", format_number(pro_forma.pf_net_debt, currency_symbol=acq_cs))

    _mhtml(
        f'<div style="display:flex; gap:1rem; margin-top:0.5rem;">'
        f'<div style="flex:1; text-align:center; padding:0.6rem; background:rgba(255,255,255,0.04); border-radius:10px; border-left:3px solid {lev_c};">'
        f'<div style="font-size:0.65rem; font-weight:600; color:#8A85AD; text-transform:uppercase;">Leverage</div>'
        f'<div style="font-size:1.1rem; font-weight:700; color:{lev_c};">'
        f'{"Conservative" if (pro_forma.pf_leverage_ratio or 0) < 2 else "Moderate" if (pro_forma.pf_leverage_ratio or 0) < 4 else "Aggressive"}</div></div>'
        f'<div style="flex:1; text-align:center; padding:0.6rem; background:rgba(255,255,255,0.04); border-radius:10px; border-left:3px solid {cov_c};">'
        f'<div style="font-size:0.65rem; font-weight:600; color:#8A85AD; text-transform:uppercase;">Coverage</div>'
        f'<div style="font-size:1.1rem; font-weight:700; color:{cov_c};">'
        f'{"Strong" if (pro_forma.pf_interest_coverage or 0) > 5 else "Adequate" if (pro_forma.pf_interest_coverage or 0) > 2.5 else "Tight"}</div></div>'
        f'<div style="flex:1; text-align:center; padding:0.6rem; background:rgba(255,255,255,0.04); border-radius:10px;">'
        f'<div style="font-size:0.65rem; font-weight:600; color:#8A85AD; text-transform:uppercase;">Goodwill</div>'
        f'<div style="font-size:1.1rem; font-weight:700; color:#E0DCF5;">{format_number(pro_forma.goodwill, currency_symbol=acq_cs)}</div></div>'
        f'</div>'
    )

    # Debt Paydown Visualization
    try:
        if pro_forma.cash_consideration and pro_forma.pf_ebitda and pro_forma.pf_ebitda > 0:
            new_debt = pro_forma.cash_consideration
            annual_fcf = pro_forma.pf_ebitda * 0.4  # Assume 40% of EBITDA available for debt paydown
            
            if annual_fcf > 0:
                years_data = []
                remaining = new_debt
                for yr in range(1, 11):
                    remaining = max(0, remaining - annual_fcf)
                    years_data.append({"year": yr, "remaining": remaining})
                    if remaining == 0:
                        break
                
                payoff_years = next((d["year"] for d in years_data if d["remaining"] == 0), 10)
                py_color = "#10B981" if payoff_years <= 4 else "#F59E0B" if payoff_years <= 7 else "#EF4444"
                
                st.markdown(
                    f'<div style="text-align:center; padding:0.5rem; margin-top:0.5rem;">'
                    f'<span style="font-size:0.75rem; color:#8A85AD;">Est. Debt Paydown: </span>'
                    f'<span style="font-size:1.1rem; font-weight:800; color:{py_color};">'
                    f'~{payoff_years} years</span>'
                    f'<span style="font-size:0.65rem; color:#8A85AD;"> (assuming 40% EBITDA allocation)</span>'
                    f'</div>',
                    unsafe_allow_html=True,
                )
    except Exception:
        pass

    _divider()

    # ══════════════════════════════════════════════════════
    # M8b. SYNERGY REALIZATION SCHEDULE
    # ══════════════════════════════════════════════════════
    with _safe_section("Synergy Realization"):
        _section("Synergy Realization Schedule", "📅")

        if hasattr(pro_forma, 'synergy_schedule') and pro_forma.synergy_schedule:
            syn_cols = st.columns(3)
            for i, s in enumerate(pro_forma.synergy_schedule):
                with syn_cols[i]:
                    pct_color = "#EF4444" if s["pct"] < 50 else "#F5A623" if s["pct"] < 100 else "#10B981"
                    st.markdown(
                        f'<div style="text-align:center; padding:0.8rem; background:rgba(255,255,255,0.04); border-radius:10px;">'
                        f'<div style="font-size:0.65rem; font-weight:600; color:#8A85AD; text-transform:uppercase;">Year {s["year"]}</div>'
                        f'<div style="font-size:1.5rem; font-weight:800; color:{pct_color};">{s["pct"]}%</div>'
                        f'<div style="font-size:0.8rem; color:#B8B3D7;">{format_number(s["amount"], currency_symbol=acq_cs)}</div>'
                        f'</div>',
                        unsafe_allow_html=True,
                    )

            # Synergy ramp chart
            fig_syn = go.Figure()
            years = [f"Year {s['year']}" for s in pro_forma.synergy_schedule]
            amounts = [s["amount"] for s in pro_forma.synergy_schedule]
            fig_syn.add_trace(go.Bar(
                x=years, y=amounts,
                marker_color=["#EF4444", "#F5A623", "#10B981"],
                text=[format_number(a, currency_symbol=acq_cs) for a in amounts],
                textposition="outside", textfont=dict(size=10, color="#B8B3D7"),
            ))
            fig_syn.add_hline(y=pro_forma.total_synergies, line_dash="dash",
                             line_color="rgba(107,92,231,0.5)", annotation_text="Full Run-Rate",
                             annotation_font=dict(size=10, color="#8A85AD"))
            fig_syn.update_layout(**_CHART_LAYOUT_BASE, height=250,
                                  margin=dict(t=30, b=30, l=50, r=30), showlegend=False,
                                  xaxis=dict(tickfont=dict(size=10, color="#8A85AD")),
                                  yaxis=dict(tickfont=dict(size=9, color="#8A85AD")))
            _apply_space_grid(fig_syn)
            st.plotly_chart(fig_syn, use_container_width=True, key="synergy_ramp")

    # ══════════════════════════════════════════════════════
    # M8c. DEBT PAYDOWN SCHEDULE
    # ══════════════════════════════════════════════════════
    with _safe_section("Debt Paydown Schedule"):
        if hasattr(pro_forma, 'debt_paydown_schedule') and pro_forma.debt_paydown_schedule:
            _section("Debt Paydown Schedule", "📉")

            fig_dp = go.Figure()
            dp_years = [f"Year {d['year']}" for d in pro_forma.debt_paydown_schedule]
            dp_remaining = [d["remaining"] for d in pro_forma.debt_paydown_schedule]
            dp_leverage = [d["leverage"] for d in pro_forma.debt_paydown_schedule]

            fig_dp.add_trace(go.Bar(
                x=dp_years, y=dp_remaining, name="Remaining Debt",
                marker_color="#6B5CE7",
                text=[format_number(r, currency_symbol=acq_cs) for r in dp_remaining],
                textposition="outside", textfont=dict(size=9, color="#B8B3D7"),
            ))
            fig_dp.add_trace(go.Scatter(
                x=dp_years, y=dp_leverage, name="Leverage (Debt/EBITDA)",
                yaxis="y2", mode="lines+markers",
                line=dict(color="#F5A623", width=2),
                marker=dict(size=8, color="#F5A623"),
            ))
            fig_dp.update_layout(
                **_CHART_LAYOUT_BASE, height=300,
                margin=dict(t=30, b=30, l=60, r=60), showlegend=True,
                legend=dict(orientation="h", yanchor="bottom", y=1.02, font=dict(size=9, color="#8A85AD")),
                xaxis=dict(tickfont=dict(size=10, color="#8A85AD")),
                yaxis=dict(title=dict(text="Debt Remaining", font=dict(size=10, color="#8A85AD")),
                          tickfont=dict(size=9, color="#8A85AD")),
                yaxis2=dict(title=dict(text="Leverage (x)", font=dict(size=10, color="#8A85AD")),
                           tickfont=dict(size=9, color="#8A85AD"), overlaying="y", side="right",
                           tickformat=".1f", ticksuffix="x"),
            )
            _apply_space_grid(fig_dp)
            st.plotly_chart(fig_dp, use_container_width=True, key="debt_paydown")

    # ══════════════════════════════════════════════════════
    # M8d. BREAK-EVEN SYNERGIES & DEAL IRR
    # ══════════════════════════════════════════════════════
    with _safe_section("Break-Even & IRR"):
        _section("Break-Even Analysis & Deal IRR", "🎯")

        be_cols = st.columns(3)
        with be_cols[0]:
            be_syn = getattr(pro_forma, 'breakeven_synergies', 0)
            be_pct = getattr(pro_forma, 'breakeven_synergies_pct_of_target_rev', 0)
            achievable = be_syn <= pro_forma.total_synergies if be_syn else True
            be_color = "#10B981" if achievable else "#EF4444"
            st.markdown(
                f'<div style="text-align:center; padding:0.8rem; background:rgba(255,255,255,0.04); border-radius:10px;">'
                f'<div style="font-size:0.65rem; font-weight:600; color:#8A85AD; text-transform:uppercase;">Break-Even Synergies</div>'
                f'<div style="font-size:1.3rem; font-weight:800; color:{be_color};">{format_number(be_syn, currency_symbol=acq_cs)}</div>'
                f'<div style="font-size:0.75rem; color:#8A85AD;">{be_pct:.1f}% of target revenue</div>'
                f'<div style="font-size:0.7rem; color:{be_color}; margin-top:0.3rem;">{"✓ Below assumed synergies" if achievable else "⚠ Above assumed synergies"}</div>'
                f'</div>',
                unsafe_allow_html=True,
            )
        with be_cols[1]:
            st.markdown(
                f'<div style="text-align:center; padding:0.8rem; background:rgba(255,255,255,0.04); border-radius:10px;">'
                f'<div style="font-size:0.65rem; font-weight:600; color:#8A85AD; text-transform:uppercase;">Synergy Cushion</div>'
                f'<div style="font-size:1.3rem; font-weight:800; color:{"#10B981" if achievable else "#EF4444"};">'
                f'{format_number(pro_forma.total_synergies - be_syn, currency_symbol=acq_cs)}</div>'
                f'<div style="font-size:0.75rem; color:#8A85AD;">{"Margin of safety" if achievable else "Shortfall"}</div>'
                f'</div>',
                unsafe_allow_html=True,
            )
        with be_cols[2]:
            deal_irr = getattr(pro_forma, 'deal_irr', None)
            if deal_irr is not None:
                irr_color = "#10B981" if deal_irr > 0.15 else "#F5A623" if deal_irr > 0.08 else "#EF4444"
                irr_label = "Attractive" if deal_irr > 0.15 else "Acceptable" if deal_irr > 0.08 else "Below Hurdle"
                st.markdown(
                    f'<div style="text-align:center; padding:0.8rem; background:rgba(255,255,255,0.04); border-radius:10px;">'
                    f'<div style="font-size:0.65rem; font-weight:600; color:#8A85AD; text-transform:uppercase;">Est. Deal IRR (5yr)</div>'
                    f'<div style="font-size:1.3rem; font-weight:800; color:{irr_color};">{deal_irr:.1%}</div>'
                    f'<div style="font-size:0.75rem; color:{irr_color};">{irr_label}</div>'
                    f'</div>',
                    unsafe_allow_html=True,
                )
            else:
                st.markdown(
                    f'<div style="text-align:center; padding:0.8rem; background:rgba(255,255,255,0.04); border-radius:10px;">'
                    f'<div style="font-size:0.65rem; font-weight:600; color:#8A85AD;">Est. Deal IRR</div>'
                    f'<div style="font-size:1.1rem; color:#8A85AD;">N/A</div></div>',
                    unsafe_allow_html=True,
                )

    _divider()

    # ══════════════════════════════════════════════════════
    # M9. AI STRATEGIC RATIONALE
    # ══════════════════════════════════════════════════════
    _section("Strategic Rationale")

    _sr_tag_config = [
        ("[DEAL LOGIC]", "Deal Logic", "#6B5CE7", "rgba(107,92,231,0.06)", "rgba(107,92,231,0.3)"),
        ("[FINANCIAL MERIT]", "Financial Merit", "#E8638B", "rgba(232,99,139,0.06)", "rgba(232,99,139,0.3)"),
        ("[STRATEGIC FIT]", "Strategic Fit", "#10B981", "rgba(16,185,129,0.06)", "rgba(16,185,129,0.3)"),
        ("[COMPETITIVE POSITIONING]", "Competitive Positioning", "#F5A623", "rgba(245,166,35,0.06)", "rgba(245,166,35,0.3)"),
    ]

    for line in merger_insights.strategic_rationale.split("\n"):
        line = line.strip()
        if line.startswith("- "):
            line = line[2:]
        if not line:
            continue
        matched_tag = False
        for tag, label, color, bg, border in _sr_tag_config:
            if line.startswith(tag):
                line = line[len(tag):].strip().replace("$", "&#36;")
                st.markdown(
                    f'<div style="border-left:3px solid {border}; background:{bg}; '
                    f'padding:0.5rem 0.8rem; margin-bottom:0.5rem; border-radius:0 8px 8px 0;">'
                    f'<div style="font-size:0.7rem; font-weight:700; color:{color}; text-transform:uppercase; '
                    f'letter-spacing:0.5px; margin-bottom:0.2rem;">{label}</div>'
                    f'<div style="font-size:0.86rem; color:#B8B3D7; line-height:1.7;">{line}</div>'
                    f'</div>',
                    unsafe_allow_html=True,
                )
                matched_tag = True
                break
        if not matched_tag and line:
            line = line.replace("$", "&#36;")
            st.markdown(f"<div style='font-size:0.88rem; color:#B8B3D7; line-height:1.7; padding:0.2rem 0;'>&bull; {line}</div>", unsafe_allow_html=True)

    _divider()

    # ══════════════════════════════════════════════════════
    # M10. AI DEAL RISKS
    # ══════════════════════════════════════════════════════
    _section("Deal Risk Assessment")

    _risk_tag_config = [
        ("[VALUATION]", "Valuation", "#EF4444", "rgba(239,68,68,0.06)", "rgba(239,68,68,0.3)"),
        ("[FINANCIAL]", "Financial", "#E8638B", "rgba(232,99,139,0.06)", "rgba(232,99,139,0.3)"),
        ("[INTEGRATION]", "Integration", "#F5A623", "rgba(245,166,35,0.06)", "rgba(245,166,35,0.3)"),
        ("[EXECUTION]", "Execution", "#6B5CE7", "rgba(107,92,231,0.06)", "rgba(107,92,231,0.3)"),
        ("[MARKET]", "Market", "#10B981", "rgba(16,185,129,0.06)", "rgba(16,185,129,0.3)"),
        # Legacy tag support
        ("[ANTITRUST]", "Antitrust", "#EF4444", "rgba(239,68,68,0.06)", "rgba(239,68,68,0.3)"),
    ]

    # Severity keyword tinting — override base colors for high-severity language
    _high_severity_words = {"distressed", "unsustainable", "aggressive", "concerning", "substantial", "significant", "elevated", "transformative"}
    _low_severity_words = {"manageable", "adequate", "comfortable", "low", "conservative", "modest", "contained"}

    for line in merger_insights.deal_risks.split("\n"):
        line = line.strip()
        if line.startswith("- "):
            line = line[2:]
        if not line:
            continue

        tag_label = ""
        tag_color = "#8A85AD"
        tag_bg = "rgba(138,133,173,0.05)"
        tag_border = "rgba(138,133,173,0.2)"

        for tag, label, color, bg, border in _risk_tag_config:
            if line.startswith(tag):
                line = line[len(tag):].strip()
                tag_label = label
                tag_color = color
                tag_bg = bg
                tag_border = border
                break

        # Severity-based tint adjustment (before escaping)
        line_lower = line.lower()
        has_high = any(w in line_lower for w in _high_severity_words)
        has_low = any(w in line_lower for w in _low_severity_words)

        if has_high and not has_low:
            severity_indicator = '<span style="color:#EF4444; font-size:0.7rem; margin-left:0.4rem;">&#9650; ELEVATED</span>'
        elif has_low and not has_high:
            severity_indicator = '<span style="color:#10B981; font-size:0.7rem; margin-left:0.4rem;">&#9660; LOW</span>'
        else:
            severity_indicator = ""

        header_html = ""
        if tag_label:
            header_html = (
                f'<div style="font-size:0.7rem; font-weight:700; color:{tag_color}; text-transform:uppercase; '
                f'letter-spacing:0.5px; margin-bottom:0.2rem;">{tag_label}{severity_indicator}</div>'
            )

        line = line.replace("$", "&#36;")
        st.markdown(
            f'<div style="border-left:3px solid {tag_border}; background:{tag_bg}; '
            f'padding:0.5rem 0.8rem; margin-bottom:0.5rem; border-radius:0 8px 8px 0;">'
            f'{header_html}'
            f'<div style="font-size:0.86rem; color:#B8B3D7; line-height:1.7;">{line}</div>'
            f'</div>',
            unsafe_allow_html=True,
        )

    _divider()

    # ══════════════════════════════════════════════════════
    # M11. AI DEAL VERDICT
    # ══════════════════════════════════════════════════════
    _section("Deal Verdict")

    grade_colors = {"A": "#10B981", "B": "#6B5CE7", "C": "#F5A623", "D": "#EF4444", "F": "#EF4444"}
    grade_c = grade_colors.get(merger_insights.deal_grade, "#8A85AD")
    grade_bg = {"A": "rgba(16,185,129,0.12)", "B": "rgba(107,92,231,0.12)",
                "C": "rgba(245,166,35,0.12)", "D": "rgba(239,68,68,0.12)", "F": "rgba(239,68,68,0.12)"}

    st.markdown(
        f'<div style="display:inline-block; background:{grade_bg.get(merger_insights.deal_grade, "rgba(138,133,173,0.12)")}; '
        f'color:{grade_c}; padding:0.5rem 1.5rem; border-radius:20px; font-weight:800; '
        f'font-size:1.2rem; letter-spacing:1px; margin-bottom:1rem;">Deal Grade: {merger_insights.deal_grade}</div>',
        unsafe_allow_html=True,
    )

    _verdict_tag_config = {
        "[OVERALL]": ("Overall Assessment", None, "rgba(255,255,255,0.04)", "rgba(138,133,173,0.3)"),
        "[BULL CASE]": ("Bull Case", "#10B981", "rgba(16,185,129,0.06)", "rgba(16,185,129,0.35)"),
        "[BEAR CASE]": ("Bear Case", "#EF4444", "rgba(239,68,68,0.06)", "rgba(239,68,68,0.35)"),
        "[KEY CONDITION]": ("Key Condition", "#F5A623", "rgba(245,166,35,0.08)", "rgba(245,166,35,0.35)"),
    }

    for line in merger_insights.deal_verdict.split("\n"):
        line = line.strip()
        if line.startswith("- "):
            line = line[2:]
        if not line:
            continue

        matched_tag = False
        for tag, (label, color, bg, border) in _verdict_tag_config.items():
            if line.startswith(tag):
                line = line[len(tag):].strip().replace("$", "&#36;")
                header_color = color or "#B8B3D7"
                st.markdown(
                    f'<div style="border-left:3px solid {border}; background:{bg}; '
                    f'padding:0.6rem 0.8rem; margin-bottom:0.5rem; border-radius:0 8px 8px 0;">'
                    f'<div style="font-size:0.7rem; font-weight:700; color:{header_color}; text-transform:uppercase; '
                    f'letter-spacing:0.5px; margin-bottom:0.2rem;">{label}</div>'
                    f'<div style="font-size:0.86rem; color:#B8B3D7; line-height:1.7;">{line}</div>'
                    f'</div>',
                    unsafe_allow_html=True,
                )
                matched_tag = True
                break

        if not matched_tag and line:
            line = line.replace("$", "&#36;")
            st.markdown(f"<div style='font-size:0.88rem; color:#B8B3D7; line-height:1.7; padding:0.2rem 0;'>&bull; {line}</div>", unsafe_allow_html=True)

    _divider()

    # ══════════════════════════════════════════════════════
    # M11b. QUANTITATIVE DEAL SCORECARD
    # ══════════════════════════════════════════════════════
    _section("Deal Scorecard", "📋")
    
    # Compute quantitative deal metrics
    deal_scores = []
    
    # 1. Accretion/Dilution
    if pro_forma.accretion_dilution_pct:
        ad = pro_forma.accretion_dilution_pct
        if ad > 5: deal_scores.append(("EPS Accretion", 10, "#10B981", f"+{ad:.1f}%"))
        elif ad > 0: deal_scores.append(("EPS Accretion", 7, "#34D399", f"+{ad:.1f}%"))
        elif ad > -5: deal_scores.append(("EPS Accretion", 4, "#F59E0B", f"{ad:.1f}%"))
        else: deal_scores.append(("EPS Accretion", 2, "#EF4444", f"{ad:.1f}%"))
    
    # 2. Premium reasonableness
    prem = merger_assumptions.offer_premium_pct
    if prem < 20: deal_scores.append(("Premium", 9, "#10B981", f"{prem:.0f}%"))
    elif prem < 35: deal_scores.append(("Premium", 7, "#34D399", f"{prem:.0f}%"))
    elif prem < 50: deal_scores.append(("Premium", 5, "#F59E0B", f"{prem:.0f}%"))
    else: deal_scores.append(("Premium", 2, "#EF4444", f"{prem:.0f}%"))
    
    # 3. Synergy coverage of premium
    if 'syn_pv' in dir() and 'premium_paid' in dir() and premium_paid > 0:
        syn_cov = syn_pv / premium_paid * 100
        if syn_cov > 100: deal_scores.append(("Synergy Coverage", 10, "#10B981", f"{syn_cov:.0f}%"))
        elif syn_cov > 60: deal_scores.append(("Synergy Coverage", 7, "#34D399", f"{syn_cov:.0f}%"))
        elif syn_cov > 30: deal_scores.append(("Synergy Coverage", 4, "#F59E0B", f"{syn_cov:.0f}%"))
        else: deal_scores.append(("Synergy Coverage", 2, "#EF4444", f"{syn_cov:.0f}%"))
    
    # 4. Strategic fit (same sector?)
    same_sector = (acq_cd.sector or "").lower() == (tgt_cd.sector or "").lower()
    if same_sector:
        deal_scores.append(("Strategic Fit", 8, "#10B981", "Same Sector"))
    else:
        deal_scores.append(("Strategic Fit", 5, "#F59E0B", "Cross-Sector"))
    
    # 5. Size ratio (target should be <50% of acquirer for clean integration)
    if acq_cd.market_cap and tgt_cd.market_cap:
        size_ratio = tgt_cd.market_cap / acq_cd.market_cap * 100
        if size_ratio < 15: deal_scores.append(("Size Ratio", 9, "#10B981", f"{size_ratio:.0f}%"))
        elif size_ratio < 30: deal_scores.append(("Size Ratio", 7, "#34D399", f"{size_ratio:.0f}%"))
        elif size_ratio < 50: deal_scores.append(("Size Ratio", 5, "#F59E0B", f"{size_ratio:.0f}%"))
        else: deal_scores.append(("Size Ratio", 3, "#EF4444", f"{size_ratio:.0f}%"))
    
    # 6. Financing mix (balanced is better)
    cash_pct_score = merger_assumptions.pct_cash
    if 30 <= cash_pct_score <= 70: deal_scores.append(("Financing Mix", 8, "#10B981", f"{cash_pct_score}% Cash"))
    elif 20 <= cash_pct_score <= 80: deal_scores.append(("Financing Mix", 6, "#F59E0B", f"{cash_pct_score}% Cash"))
    else: deal_scores.append(("Financing Mix", 4, "#EF4444", f"{cash_pct_score}% Cash"))
    
    if deal_scores:
        avg_score = sum(s[1] for s in deal_scores) / len(deal_scores)
        overall_color = "#10B981" if avg_score >= 7 else "#F59E0B" if avg_score >= 5 else "#EF4444"
        overall_label = "Strong" if avg_score >= 7 else "Moderate" if avg_score >= 5 else "Weak"
        
        # Overall score
        sc_left, sc_right = st.columns([1, 3])
        with sc_left:
            st.markdown(
                f'<div style="text-align:center; padding:1.5rem; background:rgba(107,92,231,0.05); '
                f'border-radius:16px; border:1px solid rgba(107,92,231,0.15);">'
                f'<div style="font-size:2.5rem; font-weight:900; color:{overall_color};">{avg_score:.1f}</div>'
                f'<div style="font-size:0.75rem; font-weight:700; color:{overall_color};">{overall_label}</div>'
                f'<div style="font-size:0.6rem; color:#8A85AD; margin-top:0.2rem;">out of 10</div>'
                f'</div>',
                unsafe_allow_html=True,
            )
        
        with sc_right:
            for name, score, color, detail in deal_scores:
                bar_width = score * 10
                st.markdown(
                    f'<div style="display:flex; align-items:center; gap:0.5rem; padding:0.3rem 0; '
                    f'border-bottom:1px solid rgba(255,255,255,0.03);">'
                    f'<span style="color:#8A85AD; font-size:0.72rem; width:120px; flex-shrink:0;">{name}</span>'
                    f'<div style="flex:1; background:rgba(255,255,255,0.05); border-radius:4px; height:16px; overflow:hidden;">'
                    f'<div style="width:{bar_width}%; height:100%; background:{color}; border-radius:4px; '
                    f'transition:width 0.5s ease;"></div></div>'
                    f'<span style="color:{color}; font-size:0.72rem; font-weight:700; width:80px; text-align:right;">{detail}</span>'
                    f'</div>',
                    unsafe_allow_html=True,
                )

    _divider()

    # ══════════════════════════════════════════════════════
    # M12. DOWNLOAD DEAL BOOK
    # ══════════════════════════════════════════════════════
    _section("Download Deal Book")

    if not os.path.exists("assets/template.pptx"):
        with st.spinner("Creating template..."):
            from create_template import build
            build()

    with st.spinner("Building 10-slide Deal Book..."):
        deal_book_buf = generate_deal_book(acq_cd, tgt_cd, pro_forma, merger_insights, merger_assumptions)

    dl1, dl2, dl3 = st.columns([1, 2, 1])
    with dl2:
        st.download_button(
            label=f"Download {acq_cd.ticker}+{tgt_cd.ticker} Deal Book  (3 slides)",
            data=deal_book_buf,
            file_name=f"{acq_cd.ticker}_{tgt_cd.ticker}_Deal_Book.pptx",
            mime="application/vnd.openxmlformats-officedocument.presentationml.presentation",
            use_container_width=True,
        )
        st.markdown(
            "<p style='text-align:center; font-size:0.72rem; color:#8A85AD; margin-top:0.3rem;'>"
            "Professional deal book &middot; Pro forma analysis &middot; AI-powered insights"
            "</p>",
            unsafe_allow_html=True,
        )

elif analysis_mode == "Merger Analysis" and merger_btn and (not acquirer_input or not target_input):
    st.warning("Please enter both Acquirer and Target tickers in the sidebar.")

# ══════════════════════════════════════════════════════════════
# DCF VALUATION MODE
# ══════════════════════════════════════════════════════════════
elif analysis_mode == "DCF Valuation" and dcf_btn and dcf_ticker_input:
    st.markdown(
        '<div class="hero-header">'
        '<div class="orbital-brand">'
        f'{_orbital_logo()}'
        '<p class="orbital-tagline">DCF Valuation Analysis</p>'
        '</div>'
        '<span class="hero-tagline">Discounted Cash Flow Model</span>'
        '</div>',
        unsafe_allow_html=True,
    )
    
    with st.spinner(f"Fetching data for {dcf_ticker_input}..."):
        try:
            dcf_cd = fetch_company_data(dcf_ticker_input)
        except Exception as e:
            st.error(f"Failed to fetch data for {dcf_ticker_input}: {e}")
            st.stop()
    
    cs = dcf_cd.currency_symbol
    
    # Company Header
    st.markdown(
        f'<div class="company-card">'
        f'<div><p class="company-name">{dcf_cd.name}</p>'
        f'<p class="company-meta"><span>{dcf_cd.ticker}</span> &nbsp;&middot;&nbsp; {dcf_cd.sector} &rarr; {dcf_cd.industry}</p></div>'
        f'<div style="margin-top:0.8rem;">'
        f'<span style="font-size:1.5rem; font-weight:700; color:#E0DCF5;">{cs}{dcf_cd.current_price:,.2f}</span>'
        f'</div></div>',
        unsafe_allow_html=True,
    )
    
    # Calculate DCF
    dcf_result = _calculate_dcf(
        dcf_cd,
        growth_rate=dcf_growth_rate,
        terminal_growth=dcf_terminal_growth,
        discount_rate=dcf_discount_rate,
        projection_years=dcf_years
    )
    
    if "error" in dcf_result:
        st.error(dcf_result["error"])
    else:
        # DCF Results Summary
        _section("DCF Valuation Results", "💰")
        
        # Key metrics
        r1, r2, r3, r4 = st.columns(4)
        r1.metric("Base FCF", format_number(dcf_result["base_fcf"], currency_symbol=cs))
        r2.metric("DCF Enterprise Value", format_number(dcf_result["enterprise_value"], currency_symbol=cs))
        r3.metric("DCF Equity Value", format_number(dcf_result["equity_value"], currency_symbol=cs))
        
        # Implied share price with upside/downside
        upside = dcf_result["upside_pct"]
        upside_color = "#10B981" if upside >= 0 else "#EF4444"
        upside_text = f"+{upside:.1f}%" if upside >= 0 else f"{upside:.1f}%"
        r4.metric("Implied Share Price", f"{cs}{dcf_result['implied_share_price']:,.2f}", delta=upside_text)
        
        _divider()
        
        # Valuation Summary Card
        st.markdown(
            f'<div style="background:linear-gradient(135deg, rgba(107,92,231,0.1), rgba(16,185,129,0.05)); '
            f'border:1px solid rgba(107,92,231,0.25); border-radius:16px; padding:1.5rem; margin:1rem 0;">'
            f'<div style="display:flex; justify-content:space-between; align-items:center;">'
            f'<div>'
            f'<div style="font-size:0.7rem; color:#8A85AD; text-transform:uppercase; letter-spacing:1px;">Current Price</div>'
            f'<div style="font-size:1.8rem; font-weight:700; color:#E0DCF5;">{cs}{dcf_result["current_price"]:,.2f}</div>'
            f'</div>'
            f'<div style="font-size:2rem; color:#8A85AD;">→</div>'
            f'<div>'
            f'<div style="font-size:0.7rem; color:#8A85AD; text-transform:uppercase; letter-spacing:1px;">Implied Value</div>'
            f'<div style="font-size:1.8rem; font-weight:700; color:{upside_color};">{cs}{dcf_result["implied_share_price"]:,.2f}</div>'
            f'</div>'
            f'<div style="background:{"rgba(16,185,129,0.15)" if upside >= 0 else "rgba(239,68,68,0.15)"}; '
            f'padding:0.8rem 1.5rem; border-radius:12px; text-align:center;">'
            f'<div style="font-size:0.7rem; color:#8A85AD; text-transform:uppercase;">{"Upside" if upside >= 0 else "Downside"}</div>'
            f'<div style="font-size:1.5rem; font-weight:800; color:{upside_color};">{upside_text}</div>'
            f'</div></div></div>',
            unsafe_allow_html=True,
        )
        
        _divider()
        
        # Assumptions Used
        _section("Model Assumptions", "📊")
        a1, a2, a3, a4 = st.columns(4)
        a1.metric("FCF Growth Rate", f"{dcf_result['growth_rate']*100:.1f}%")
        a2.metric("Terminal Growth", f"{dcf_result['terminal_growth']*100:.1f}%")
        a3.metric("Discount Rate (WACC)", f"{dcf_result['discount_rate']*100:.1f}%")
        a4.metric("Projection Years", f"{dcf_result['projection_years']}")
        
        _divider()
        
        # Projected FCF Chart
        _section("Projected Free Cash Flow", "📈")
        _build_dcf_chart(dcf_result, currency_symbol=cs, key="dcf_main_chart")
        
        _divider()
        
        # Value Bridge
        _section("Value Bridge", "🌉")
        st.markdown(
            f'<div style="display:grid; grid-template-columns:repeat(4,1fr); gap:1rem;">'
            f'<div style="background:rgba(107,92,231,0.1); border-radius:12px; padding:1rem; text-align:center;">'
            f'<div style="font-size:0.7rem; color:#8A85AD; margin-bottom:0.3rem;">Sum of PV (FCF)</div>'
            f'<div style="font-size:1.2rem; font-weight:700; color:#6B5CE7;">{format_number(sum(dcf_result["pv_fcf"]), currency_symbol=cs)}</div></div>'
            f'<div style="background:rgba(232,99,139,0.1); border-radius:12px; padding:1rem; text-align:center;">'
            f'<div style="font-size:0.7rem; color:#8A85AD; margin-bottom:0.3rem;">PV of Terminal Value</div>'
            f'<div style="font-size:1.2rem; font-weight:700; color:#E8638B;">{format_number(dcf_result["pv_terminal"], currency_symbol=cs)}</div></div>'
            f'<div style="background:rgba(245,166,35,0.1); border-radius:12px; padding:1rem; text-align:center;">'
            f'<div style="font-size:0.7rem; color:#8A85AD; margin-bottom:0.3rem;">Less: Net Debt</div>'
            f'<div style="font-size:1.2rem; font-weight:700; color:#F5A623;">({format_number(abs(dcf_result["net_debt"]), currency_symbol=cs)})</div></div>'
            f'<div style="background:rgba(16,185,129,0.1); border-radius:12px; padding:1rem; text-align:center;">'
            f'<div style="font-size:0.7rem; color:#8A85AD; margin-bottom:0.3rem;">= Equity Value</div>'
            f'<div style="font-size:1.2rem; font-weight:700; color:#10B981;">{format_number(dcf_result["equity_value"], currency_symbol=cs)}</div></div>'
            f'</div>',
            unsafe_allow_html=True,
        )
        
        _divider()
        
        # Sensitivity Analysis
        _section("Sensitivity Analysis", "📐")
        st.markdown(
            '<div style="font-size:0.85rem; color:#B8B3D7; margin-bottom:1rem;">'
            'How does the implied share price change with different growth and discount rate assumptions?'
            '</div>',
            unsafe_allow_html=True,
        )
        
        sens_col1, sens_col2 = st.columns(2)
        
        with sens_col1:
            st.markdown("**Growth Rate vs. WACC Matrix**")
            _build_dcf_sensitivity(dcf_cd, dcf_result, key="dcf_sens_matrix")
        
        with sens_col2:
            st.markdown("**Terminal Growth Impact**")
            _build_terminal_value_sensitivity(dcf_cd, dcf_result, key="dcf_tv_sens")
        
        _divider()
        
        # Reverse DCF
        _section("Reverse DCF — Implied Growth Rate", "🔄")
        st.markdown(
            '<div style="font-size:0.85rem; color:#B8B3D7; margin-bottom:1rem;">'
            'What FCF growth rate does the market currently imply at the current share price?'
            '</div>',
            unsafe_allow_html=True,
        )
        
        try:
            target_price = dcf_result["current_price"]
            shares = dcf_result.get("shares_outstanding", 1)
            target_equity = target_price * shares
            net_debt = dcf_result["net_debt"]
            target_ev = target_equity + net_debt
            base_fcf = dcf_result["base_fcf"]
            wacc = dcf_result["discount_rate"]
            tg = dcf_result["terminal_growth"]
            years = dcf_result["projection_years"]
            
            # Binary search for implied growth rate
            low_g, high_g = -0.10, 0.40
            implied_growth = None
            for _ in range(50):  # Binary search iterations
                mid_g = (low_g + high_g) / 2
                fcf = base_fcf
                pv_sum = 0
                for yr in range(1, years + 1):
                    fcf = fcf * (1 + mid_g)
                    pv_sum += fcf / (1 + wacc) ** yr
                
                if wacc > tg:
                    tv = (fcf * (1 + tg)) / (wacc - tg)
                    pv_tv = tv / (1 + wacc) ** years
                else:
                    pv_tv = 0
                
                calc_ev = pv_sum + pv_tv
                
                if calc_ev < target_ev:
                    low_g = mid_g
                else:
                    high_g = mid_g
                
                if abs(calc_ev - target_ev) / target_ev < 0.001:
                    implied_growth = mid_g
                    break
            
            if implied_growth is None:
                implied_growth = (low_g + high_g) / 2
            
            ig_pct = implied_growth * 100
            model_g_pct = dcf_result["growth_rate"] * 100
            
            ig_color = "#10B981" if ig_pct < model_g_pct else "#EF4444"
            verdict = "Market expects LESS growth than your model → potentially undervalued" if ig_pct < model_g_pct else "Market expects MORE growth than your model → potentially overvalued"
            
            rdcf_c1, rdcf_c2, rdcf_c3 = st.columns(3)
            rdcf_c1.metric("Your Growth Assumption", f"{model_g_pct:.1f}%")
            rdcf_c2.metric("Market-Implied Growth", f"{ig_pct:.1f}%")
            rdcf_c3.metric("Difference", f"{ig_pct - model_g_pct:+.1f}%")
            
            st.markdown(
                f'<div style="text-align:center; padding:0.6rem; background:rgba(107,92,231,0.05); '
                f'border-radius:10px; margin-top:0.5rem;">'
                f'<span style="font-size:0.8rem; color:{ig_color}; font-weight:600;">{verdict}</span>'
                f'</div>',
                unsafe_allow_html=True,
            )
        except Exception:
            st.info("Could not calculate reverse DCF.")
        
        _divider()
        
        # Monte Carlo Simulation
        _section("Monte Carlo Simulation", "🎲")
        st.markdown(
            '<div style="font-size:0.85rem; color:#B8B3D7; margin-bottom:1rem;">'
            '1,000 simulations with randomized growth rate and WACC to generate a probability distribution of fair value.'
            '</div>',
            unsafe_allow_html=True,
        )
        
        try:
            n_sims = 1000
            base_growth = dcf_result["growth_rate"]
            base_wacc = dcf_result["discount_rate"]
            base_fcf = dcf_result["base_fcf"]
            term_growth = dcf_result["terminal_growth"]
            years = dcf_result["projection_years"]
            shares = dcf_result.get("shares_outstanding", 1)
            net_debt = dcf_result["net_debt"]
            
            mc_results = []
            for _ in range(n_sims):
                sim_growth = np.random.normal(base_growth, 0.03)  # ±3% std
                sim_wacc = np.random.normal(base_wacc, 0.015)  # ±1.5% std
                sim_wacc = max(sim_wacc, 0.04)  # Floor at 4%
                
                # Project FCF
                sim_fcfs = []
                fcf = base_fcf
                for yr in range(1, years + 1):
                    fcf = fcf * (1 + sim_growth)
                    sim_fcfs.append(fcf / (1 + sim_wacc) ** yr)
                
                # Terminal value
                if sim_wacc > term_growth:
                    tv = (fcf * (1 + term_growth)) / (sim_wacc - term_growth)
                    pv_tv = tv / (1 + sim_wacc) ** years
                else:
                    pv_tv = 0
                
                ev = sum(sim_fcfs) + pv_tv
                eq = ev - net_debt
                price = eq / shares if shares > 0 else 0
                mc_results.append(price)
            
            mc_results = [p for p in mc_results if 0 < p < dcf_result["implied_share_price"] * 5]
            
            if mc_results:
                mc_arr = np.array(mc_results)
                p10 = np.percentile(mc_arr, 10)
                p25 = np.percentile(mc_arr, 25)
                p50 = np.percentile(mc_arr, 50)
                p75 = np.percentile(mc_arr, 75)
                p90 = np.percentile(mc_arr, 90)
                
                # Stats
                mc_c1, mc_c2, mc_c3, mc_c4, mc_c5 = st.columns(5)
                mc_c1.metric("10th %ile", f"{cs}{p10:,.2f}")
                mc_c2.metric("25th %ile", f"{cs}{p25:,.2f}")
                mc_c3.metric("Median", f"{cs}{p50:,.2f}")
                mc_c4.metric("75th %ile", f"{cs}{p75:,.2f}")
                mc_c5.metric("90th %ile", f"{cs}{p90:,.2f}")
                
                # Histogram
                fig_mc = go.Figure()
                fig_mc.add_trace(go.Histogram(
                    x=mc_results, nbinsx=50,
                    marker_color="rgba(107,92,231,0.5)",
                    marker_line=dict(color="rgba(107,92,231,0.8)", width=1),
                ))
                # Add current price line
                fig_mc.add_vline(x=dcf_result["current_price"], line_dash="dash",
                                line_color="#EF4444", line_width=2,
                                annotation_text=f"Current: {cs}{dcf_result['current_price']:,.2f}",
                                annotation_font=dict(size=10, color="#EF4444"))
                # Add median line
                fig_mc.add_vline(x=p50, line_dash="dash",
                                line_color="#10B981", line_width=2,
                                annotation_text=f"Median: {cs}{p50:,.2f}",
                                annotation_font=dict(size=10, color="#10B981"),
                                annotation_position="top left")
                
                fig_mc.update_layout(
                    **_CHART_LAYOUT_BASE, height=350,
                    margin=dict(t=30, b=40, l=50, r=30),
                    xaxis=dict(title=dict(text="Implied Share Price", font=dict(size=11, color="#8A85AD")),
                              tickprefix=cs, tickfont=dict(size=10, color="#8A85AD"), showgrid=False),
                    yaxis=dict(title=dict(text="Frequency", font=dict(size=11, color="#8A85AD")),
                              tickfont=dict(size=10, color="#8A85AD")),
                    showlegend=False,
                    bargap=0.05,
                )
                _apply_space_grid(fig_mc)
                st.plotly_chart(fig_mc, use_container_width=True, key="mc_simulation")
                
                # Probability of upside
                upside_prob = sum(1 for p in mc_results if p > dcf_result["current_price"]) / len(mc_results) * 100
                prob_color = "#10B981" if upside_prob > 60 else "#EF4444" if upside_prob < 40 else "#F59E0B"
                st.markdown(
                    f'<div style="text-align:center; padding:0.8rem; background:rgba(107,92,231,0.05); '
                    f'border-radius:12px; margin-top:0.5rem;">'
                    f'<span style="font-size:0.75rem; color:#8A85AD;">Probability stock is undervalued: </span>'
                    f'<span style="font-size:1.3rem; font-weight:800; color:{prob_color};">{upside_prob:.0f}%</span>'
                    f'<div style="font-size:0.6rem; color:#8A85AD; margin-top:0.2rem;">'
                    f'Based on {len(mc_results):,} simulations (growth ±3%, WACC ±1.5%)</div>'
                    f'</div>',
                    unsafe_allow_html=True,
                )
        except Exception:
            st.info("Could not run Monte Carlo simulation.")
        
        _divider()

        # ══════════════════════════════════════════════════════
        # SCENARIO ANALYSIS (Bull / Base / Bear)
        # ══════════════════════════════════════════════════════
        with _safe_section("Scenario Analysis"):
            _section("Scenario Analysis", "📊")

            st.markdown(
                '<div style="font-size:0.8rem; color:#B8B3D7; margin-bottom:0.8rem;">'
                'Three scenarios with different growth and discount rate assumptions.</div>',
                unsafe_allow_html=True,
            )

            scenarios = {
                "🐻 Bear": {"growth": max(dcf_growth_rate - 0.03, 0.0), "wacc": dcf_discount_rate + 0.015, "weight": 0.25},
                "📊 Base": {"growth": dcf_growth_rate, "wacc": dcf_discount_rate, "weight": 0.50},
                "🐂 Bull": {"growth": dcf_growth_rate + 0.03, "wacc": max(dcf_discount_rate - 0.01, dcf_terminal_growth + 0.01), "weight": 0.25},
            }

            scen_results = {}
            for name, params in scenarios.items():
                try:
                    scen_dcf = _calculate_dcf(dcf_cd, growth_rate=params["growth"],
                                               terminal_growth=dcf_terminal_growth,
                                               discount_rate=params["wacc"],
                                               projection_years=dcf_years)
                    scen_results[name] = {**scen_dcf, "weight": params["weight"],
                                          "growth": params["growth"], "wacc": params["wacc"]}
                except Exception:
                    pass

            if scen_results:
                scen_cols = st.columns(len(scen_results))
                for col, (name, res) in zip(scen_cols, scen_results.items()):
                    with col:
                        if "error" not in res:
                            price = res["implied_share_price"]
                            upside = res["upside_pct"]
                            color = "#10B981" if upside > 0 else "#EF4444"
                            st.markdown(
                                f'<div style="text-align:center; padding:0.8rem; background:rgba(255,255,255,0.04); '
                                f'border-radius:10px; border-top:3px solid {color};">'
                                f'<div style="font-size:0.7rem; font-weight:700; color:#8A85AD;">{name}</div>'
                                f'<div style="font-size:1.5rem; font-weight:800; color:{color};">{cs}{price:,.2f}</div>'
                                f'<div style="font-size:0.8rem; color:{color};">{upside:+.1f}%</div>'
                                f'<div style="font-size:0.65rem; color:#8A85AD; margin-top:0.4rem;">'
                                f'Growth: {res["growth"]:.0%} | WACC: {res["wacc"]:.0%}</div>'
                                f'<div style="font-size:0.6rem; color:#8A85AD;">Weight: {res["weight"]:.0%}</div>'
                                f'</div>',
                                unsafe_allow_html=True,
                            )

                # Probability-weighted fair value
                weighted_price = sum(
                    r.get("implied_share_price", 0) * r.get("weight", 0)
                    for r in scen_results.values() if "error" not in r
                )
                weighted_upside = ((weighted_price / dcf_cd.current_price) - 1) * 100 if dcf_cd.current_price else 0
                w_color = "#10B981" if weighted_upside > 0 else "#EF4444"

                st.markdown(
                    f'<div style="text-align:center; padding:1rem; margin-top:0.8rem; '
                    f'background:rgba(107,92,231,0.06); border:1px solid rgba(107,92,231,0.15); border-radius:12px;">'
                    f'<div style="font-size:0.65rem; font-weight:700; color:#6B5CE7; text-transform:uppercase; '
                    f'letter-spacing:1.5px;">Probability-Weighted Fair Value</div>'
                    f'<div style="font-size:2rem; font-weight:800; color:{w_color};">{cs}{weighted_price:,.2f}</div>'
                    f'<div style="font-size:0.9rem; color:{w_color};">{weighted_upside:+.1f}% vs Current ({cs}{dcf_cd.current_price:,.2f})</div>'
                    f'</div>',
                    unsafe_allow_html=True,
                )

                # Scenario comparison bar chart
                fig_scen = go.Figure()
                scen_names = list(scen_results.keys())
                scen_prices = [scen_results[n].get("implied_share_price", 0) for n in scen_names if "error" not in scen_results[n]]
                scen_colors = ["#EF4444", "#6B5CE7", "#10B981"][:len(scen_prices)]
                fig_scen.add_trace(go.Bar(
                    x=scen_names[:len(scen_prices)], y=scen_prices,
                    marker_color=scen_colors,
                    text=[f"{cs}{p:,.2f}" for p in scen_prices],
                    textposition="outside", textfont=dict(size=10, color="#B8B3D7"),
                ))
                fig_scen.add_hline(y=dcf_cd.current_price, line_dash="dash",
                                   line_color="rgba(255,255,255,0.3)",
                                   annotation_text=f"Current: {cs}{dcf_cd.current_price:,.2f}",
                                   annotation_font=dict(size=10, color="#8A85AD"))
                fig_scen.update_layout(**_CHART_LAYOUT_BASE, height=300,
                                       margin=dict(t=30, b=30, l=50, r=30), showlegend=False,
                                       xaxis=dict(tickfont=dict(size=10, color="#8A85AD")),
                                       yaxis=dict(tickprefix=cs, tickfont=dict(size=9, color="#8A85AD")))
                _apply_space_grid(fig_scen)
                st.plotly_chart(fig_scen, use_container_width=True, key="scenario_chart")

        _divider()
        
        # ══════════════════════════════════════════════════════
        # WACC CALCULATOR BREAKDOWN
        # ══════════════════════════════════════════════════════
        with _safe_section("WACC Calculator"):
            _section("WACC Calculator", "🧮")
            
            _wacc_data = st.session_state.get("_auto_wacc_data", None)
            
            # Calculate WACC components from dcf_cd
            if _wacc_data is None:
                # Compute from scratch
                _wc_beta = dcf_cd.beta if dcf_cd.beta else 1.0
                _wc_rf = 0.0425
                _wc_erp = 0.055
                _wc_ke = _wc_rf + _wc_beta * _wc_erp
                
                _wc_ie = abs(float(dcf_cd.interest_expense.iloc[0])) if dcf_cd.interest_expense is not None and len(dcf_cd.interest_expense) > 0 else 0
                _wc_td = float(dcf_cd.total_debt.iloc[0]) if dcf_cd.total_debt is not None and len(dcf_cd.total_debt) > 0 else 0
                _wc_kd = (_wc_ie / _wc_td) if _wc_td > 0 else 0.05
                
                _wc_tp = abs(float(dcf_cd.tax_provision.iloc[0])) if dcf_cd.tax_provision is not None and len(dcf_cd.tax_provision) > 0 else 0
                _wc_oi = float(dcf_cd.operating_income.iloc[0]) if dcf_cd.operating_income is not None and len(dcf_cd.operating_income) > 0 else 0
                _wc_pretax = _wc_oi - _wc_ie
                _wc_tax = (abs(_wc_tp) / abs(_wc_pretax)) if abs(_wc_pretax) > 0 else 0.21
                _wc_tax = max(0, min(_wc_tax, 0.50))
                
                _wc_mcap = dcf_cd.market_cap or 0
                _wc_ev = _wc_mcap + _wc_td
                _wc_we = (_wc_mcap / _wc_ev) if _wc_ev > 0 else 0.7
                _wc_wd = (_wc_td / _wc_ev) if _wc_ev > 0 else 0.3
                _wc_wacc = _wc_we * _wc_ke + _wc_wd * _wc_kd * (1 - _wc_tax)
                
                _wacc_data = {
                    "rf": _wc_rf, "beta": _wc_beta, "erp": _wc_erp, "ke": _wc_ke,
                    "kd": _wc_kd, "tax_rate": _wc_tax, "we": _wc_we, "wd": _wc_wd,
                    "wacc": _wc_wacc, "mcap": _wc_mcap, "debt": _wc_td,
                }
            
            # Breakdown table
            wacc_items = [
                ("Risk-Free Rate (Rf)", f"{_wacc_data['rf']*100:.2f}%"),
                ("Beta (β)", f"{_wacc_data['beta']:.2f}"),
                ("Equity Risk Premium", f"{_wacc_data['erp']*100:.2f}%"),
                ("Cost of Equity (Ke)", f"{_wacc_data['ke']*100:.2f}%"),
                ("Cost of Debt (Kd)", f"{_wacc_data['kd']*100:.2f}%"),
                ("Tax Rate", f"{_wacc_data['tax_rate']*100:.1f}%"),
                ("Equity Weight (E/V)", f"{_wacc_data['we']*100:.1f}%"),
                ("Debt Weight (D/V)", f"{_wacc_data['wd']*100:.1f}%"),
                ("WACC", f"{_wacc_data['wacc']*100:.2f}%"),
            ]
            
            wc_c1, wc_c2 = st.columns([1, 1])
            
            with wc_c1:
                for label, val in wacc_items:
                    is_wacc = label == "WACC"
                    bg = "rgba(107,92,231,0.15)" if is_wacc else "transparent"
                    fw = "800" if is_wacc else "600"
                    fc = "#9B8AFF" if is_wacc else "#E0DCF5"
                    st.markdown(
                        f'<div style="display:flex; justify-content:space-between; padding:0.35rem 0.5rem; '
                        f'background:{bg}; border-radius:6px; border-bottom:1px solid rgba(255,255,255,0.03);">'
                        f'<span style="font-size:0.78rem; color:#B8B3D7;">{label}</span>'
                        f'<span style="font-size:0.78rem; font-weight:{fw}; color:{fc};">{val}</span>'
                        f'</div>',
                        unsafe_allow_html=True,
                    )
            
            with wc_c2:
                # Waterfall chart
                _wf_labels = ["Ke × E/V", "Kd × D/V × (1-T)", "WACC"]
                _wf_ke_contrib = _wacc_data["we"] * _wacc_data["ke"] * 100
                _wf_kd_contrib = _wacc_data["wd"] * _wacc_data["kd"] * (1 - _wacc_data["tax_rate"]) * 100
                _wf_wacc = _wacc_data["wacc"] * 100
                
                fig_wf = go.Figure(go.Waterfall(
                    x=_wf_labels,
                    y=[_wf_ke_contrib, _wf_kd_contrib, 0],
                    measure=["relative", "relative", "total"],
                    text=[f"{_wf_ke_contrib:.2f}%", f"{_wf_kd_contrib:.2f}%", f"{_wf_wacc:.2f}%"],
                    textposition="outside",
                    textfont=dict(size=11, color="#E0DCF5"),
                    connector=dict(line=dict(color="rgba(107,92,231,0.3)", width=1)),
                    increasing=dict(marker=dict(color="rgba(107,92,231,0.7)")),
                    totals=dict(marker=dict(color="rgba(16,185,129,0.7)")),
                ))
                fig_wf.update_layout(
                    **_CHART_LAYOUT_BASE, height=300,
                    margin=dict(t=30, b=40, l=40, r=30),
                    showlegend=False,
                )
                _apply_space_grid(fig_wf)
                st.plotly_chart(fig_wf, use_container_width=True, key="wacc_waterfall")
            
            _divider()
            
            # Sensitivity: WACC vs Beta vs ERP
            st.markdown(
                '<div style="font-size:0.85rem; color:#B8B3D7; margin-bottom:0.8rem;">'
                '<b>WACC Sensitivity — Beta vs. Equity Risk Premium</b></div>',
                unsafe_allow_html=True,
            )
            
            _s_betas = [0.6, 0.8, 1.0, 1.2, 1.4, 1.6, 1.8]
            _s_erps = [0.04, 0.045, 0.05, 0.055, 0.06, 0.065, 0.07]
            
            _sens_rows = []
            for b in _s_betas:
                row = {"Beta": f"{b:.1f}"}
                for erp in _s_erps:
                    ke = _wacc_data["rf"] + b * erp
                    wacc = _wacc_data["we"] * ke + _wacc_data["wd"] * _wacc_data["kd"] * (1 - _wacc_data["tax_rate"])
                    row[f"ERP {erp*100:.1f}%"] = f"{wacc*100:.1f}%"
                _sens_rows.append(row)
            
            import pandas as _pd_wacc
            _sens_df = _pd_wacc.DataFrame(_sens_rows).set_index("Beta")
            
            # Highlight the cell closest to actual values
            st.dataframe(_sens_df, use_container_width=True)
        
        _divider()

        # ── Scenario Analysis (Bull / Base / Bear) ──
        _section("Scenario Analysis", "🎯")
        with _safe_section("Scenario Analysis"):
            st.markdown(
                '<div style="font-size:0.85rem; color:#B8B3D7; margin-bottom:1rem;">'
                'Three scenarios varying growth rate and discount rate assumptions.</div>',
                unsafe_allow_html=True,
            )

            _base_gr = dcf_result["growth_rate"]
            _base_dr = dcf_result["discount_rate"]

            _scenarios = {
                "🐂 Bull Case": {"growth_rate": _base_gr + 0.03, "discount_rate": max(_base_dr - 0.01, 0.03)},
                "📊 Base Case": {"growth_rate": _base_gr, "discount_rate": _base_dr},
                "🐻 Bear Case": {"growth_rate": max(_base_gr - 0.03, -0.05), "discount_rate": _base_dr + 0.015},
            }

            _scenario_results = {}
            for _sname, _sparams in _scenarios.items():
                _sr = _calculate_dcf(
                    dcf_cd,
                    growth_rate=_sparams["growth_rate"],
                    terminal_growth=dcf_result["terminal_growth"],
                    discount_rate=_sparams["discount_rate"],
                    projection_years=dcf_result["projection_years"],
                )
                _scenario_results[_sname] = _sr

            # Three column display
            _sc1, _sc2, _sc3 = st.columns(3)
            _sc_cols = [_sc1, _sc2, _sc3]
            _sc_colors = ["#10B981", "#6B5CE7", "#EF4444"]
            _sc_bg = ["rgba(16,185,129,0.1)", "rgba(107,92,231,0.1)", "rgba(239,68,68,0.1)"]
            _sc_border = ["rgba(16,185,129,0.3)", "rgba(107,92,231,0.3)", "rgba(239,68,68,0.3)"]

            for _idx_sc, (_sname, _sr) in enumerate(_scenario_results.items()):
                with _sc_cols[_idx_sc]:
                    if "error" in _sr:
                        st.markdown(f'<div style="padding:1rem; text-align:center; color:#8A85AD;">{_sname}<br>N/A</div>', unsafe_allow_html=True)
                    else:
                        _sp = _sr["implied_share_price"]
                        _su = _sr["upside_pct"]
                        _su_text = f"+{_su:.1f}%" if _su >= 0 else f"{_su:.1f}%"
                        _sc_params = _scenarios[_sname]
                        st.markdown(
                            f'<div style="background:{_sc_bg[_idx_sc]}; border:1px solid {_sc_border[_idx_sc]}; '
                            f'border-radius:12px; padding:1.2rem; text-align:center;">'
                            f'<div style="font-size:0.85rem; font-weight:700; color:{_sc_colors[_idx_sc]}; margin-bottom:0.5rem;">{_sname}</div>'
                            f'<div style="font-size:1.6rem; font-weight:800; color:#E0DCF5;">{cs}{_sp:,.2f}</div>'
                            f'<div style="font-size:1rem; font-weight:700; color:{_sc_colors[_idx_sc]}; margin:0.3rem 0;">{_su_text}</div>'
                            f'<div style="font-size:0.65rem; color:#8A85AD; line-height:1.4;">'
                            f'Growth: {_sc_params["growth_rate"]*100:.1f}% · WACC: {_sc_params["discount_rate"]*100:.1f}%</div>'
                            f'</div>',
                            unsafe_allow_html=True,
                        )

            # Probability-weighted price
            st.markdown('<div style="height:1rem;"></div>', unsafe_allow_html=True)
            _weight_cols = st.columns([1, 1, 1, 2])
            with _weight_cols[0]:
                _w_bull = st.number_input("Bull Weight %", min_value=0, max_value=100, value=25, step=5, key="scenario_w_bull")
            with _weight_cols[1]:
                _w_base = st.number_input("Base Weight %", min_value=0, max_value=100, value=50, step=5, key="scenario_w_base")
            with _weight_cols[2]:
                _w_bear = st.number_input("Bear Weight %", min_value=0, max_value=100, value=25, step=5, key="scenario_w_bear")

            _w_total = _w_bull + _w_base + _w_bear
            _sr_list = list(_scenario_results.values())
            if _w_total > 0 and all("error" not in _sr_item for _sr_item in _sr_list):
                _weighted_price = (
                    _sr_list[0]["implied_share_price"] * _w_bull +
                    _sr_list[1]["implied_share_price"] * _w_base +
                    _sr_list[2]["implied_share_price"] * _w_bear
                ) / _w_total
                _weighted_upside = ((_weighted_price / dcf_result["current_price"]) - 1) * 100 if dcf_result["current_price"] > 0 else 0
                _wu_color = "#10B981" if _weighted_upside >= 0 else "#EF4444"
                _wu_text = f"+{_weighted_upside:.1f}%" if _weighted_upside >= 0 else f"{_weighted_upside:.1f}%"

                with _weight_cols[3]:
                    st.markdown(
                        f'<div style="background:rgba(107,92,231,0.08); border:1px solid rgba(107,92,231,0.2); '
                        f'border-radius:12px; padding:1rem; text-align:center; margin-top:0.5rem;">'
                        f'<div style="font-size:0.7rem; color:#8A85AD; text-transform:uppercase;">Probability-Weighted Price</div>'
                        f'<div style="font-size:1.5rem; font-weight:800; color:#E0DCF5;">{cs}{_weighted_price:,.2f}</div>'
                        f'<div style="font-size:0.9rem; font-weight:700; color:{_wu_color};">{_wu_text} vs current</div>'
                        f'</div>',
                        unsafe_allow_html=True,
                    )

            # Grouped bar chart
            if all("error" not in _sr_item for _sr_item in _sr_list):
                fig_scenario = go.Figure()
                _snames = list(_scenario_results.keys())
                fig_scenario.add_trace(go.Bar(
                    x=_snames,
                    y=[_sr_item["implied_share_price"] for _sr_item in _sr_list],
                    marker_color=_sc_colors,
                    text=[f"{cs}{_sr_item['implied_share_price']:,.2f}" for _sr_item in _sr_list],
                    textposition="outside",
                    textfont=dict(size=11, color="#B8B3D7"),
                ))
                fig_scenario.add_hline(
                    y=dcf_result["current_price"], line_dash="dash", line_color="#F59E0B", line_width=2,
                    annotation_text=f"Current: {cs}{dcf_result['current_price']:,.2f}",
                    annotation_font=dict(size=10, color="#F59E0B"),
                )
                fig_scenario.update_layout(
                    **_CHART_LAYOUT_BASE, height=300,
                    margin=dict(t=30, b=30, l=50, r=30),
                    xaxis=dict(tickfont=dict(size=11, color="#B8B3D7")),
                    yaxis=dict(tickprefix=cs, tickfont=dict(size=9, color="#8A85AD"), showgrid=False),
                    showlegend=False,
                )
                _apply_space_grid(fig_scenario)
                st.plotly_chart(fig_scenario, use_container_width=True, key="scenario_bar_chart")

        _divider()

        # ── Export DCF Model ──
        try:
            _dcf_assumptions = {
                "growth_rate": dcf_growth_rate,
                "terminal_growth": dcf_terminal_growth,
                "discount_rate": dcf_discount_rate,
                "years": dcf_years,
            }
            _dcf_excel = _export_dcf_to_excel(dcf_cd, dcf_result, _dcf_assumptions)
            st.download_button(
                label=f"📥 Export DCF Model (.xlsx)",
                data=_dcf_excel,
                file_name=f"{dcf_ticker_input}_DCF_Model.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="export_dcf_xlsx",
            )
        except Exception:
            pass
        
        st.markdown(
            '<div style="background:rgba(245,166,35,0.1); border:1px solid rgba(245,166,35,0.3); '
            'border-radius:12px; padding:1rem; margin-top:1rem;">'
            '<div style="font-size:0.75rem; font-weight:700; color:#F5A623; margin-bottom:0.3rem;">⚠️ DCF Disclaimer</div>'
            '<div style="font-size:0.8rem; color:#B8B3D7; line-height:1.6;">'
            'This DCF model uses simplified assumptions and historical data. Actual valuations depend on many factors '
            'including future growth trajectories, capital structure changes, and market conditions. '
            'This tool is for educational and research purposes only — not investment advice.'
            '</div></div>',
            unsafe_allow_html=True,
        )

# ══════════════════════════════════════════════════════════════
# QUICK COMPARE MODE
# ══════════════════════════════════════════════════════════════
elif analysis_mode == "Quick Compare" and compare_btn and compare_tickers:
    st.markdown(
        '<div class="hero-header">'
        '<div class="orbital-brand">'
        f'{_orbital_logo()}'
        '<p class="orbital-tagline">Company Comparison Analysis</p>'
        '</div>'
        '<span class="hero-tagline">Side-by-Side Intelligence</span>'
        '</div>',
        unsafe_allow_html=True,
    )
    
    with st.spinner(f"Fetching data for {len(compare_tickers)} companies..."):
        companies = _fetch_comparison_data(compare_tickers[:10])  # Max 10
    
    if not companies:
        st.error("Could not fetch data for any of the specified tickers.")
    else:
        st.success(f"✅ Loaded {len(companies)} companies: {', '.join([c.ticker for c in companies])}")
        
        _divider()
        
        # Comparison Table
        _section("Key Metrics Comparison", "📊")
        
        comp_df = _build_comparison_table(companies)
        
        # Style the dataframe
        st.dataframe(
            comp_df,
            use_container_width=True,
            hide_index=True,
            height=400,
        )
        
        # Download as CSV
        csv_data = comp_df.to_csv(index=False)
        st.download_button(
            "📥 Download Comparison (CSV)",
            data=csv_data,
            file_name=f"comparison_{'_'.join([c.ticker for c in companies[:5]])}.csv",
            mime="text/csv",
        )
        
        _divider()
        
        # Radar Chart Comparison
        if len(companies) >= 2:
            _section("Valuation Radar", "🎯")
            _build_comparison_radar(companies, key="compare_radar")
        
        _divider()
        
        # Price Performance Comparison
        _section("Market Cap Comparison", "💰")
        
        mc_data = [(c.ticker, c.market_cap or 0) for c in companies]
        mc_data.sort(key=lambda x: x[1], reverse=True)
        
        fig = go.Figure(go.Bar(
            x=[d[0] for d in mc_data],
            y=[d[1] for d in mc_data],
            marker=dict(
                color=["#6B5CE7", "#E8638B", "#10B981", "#F5A623", "#3B82F6", 
                       "#8B5CF6", "#EC4899", "#14B8A6", "#F59E0B", "#6366F1"][:len(mc_data)],
                line=dict(color="rgba(255,255,255,0.15)", width=1),
            ),
            text=[format_number(d[1], currency_symbol="$") for d in mc_data],
            textposition="outside",
            textfont=dict(size=10, color="#B8B3D7"),
        ))
        
        fig.update_layout(
            paper_bgcolor="rgba(0,0,0,0)",
            plot_bgcolor="rgba(0,0,0,0)",
            font=dict(family="Inter", size=14, color="#B8B3D7"),
            height=400,
            margin=dict(t=40, b=40, l=60, r=60),
            xaxis=dict(tickfont=dict(size=11, color="#8A85AD"), showgrid=False),
            yaxis=dict(tickfont=dict(size=9, color="#8A85AD"), gridcolor="rgba(107,92,231,0.1)", griddash="dot"),
        )
        
        st.plotly_chart(fig, use_container_width=True, key="mc_comparison")
        
        _divider()
        
        # Profitability Comparison
        _section("Profitability Comparison", "📈")
        
        prof_metrics = ["Gross Margin", "Op Margin", "Net Margin", "ROE"]
        prof_data = []
        for c in companies:
            prof_data.append({
                "Company": c.ticker,
                "Gross Margin": (c.gross_margins or 0) * 100,
                "Op Margin": (c.operating_margins or 0) * 100,
                "Net Margin": (c.profit_margins or 0) * 100,
                "ROE": (c.return_on_equity or 0) * 100,
            })
        
        prof_df = pd.DataFrame(prof_data)
        
        fig2 = go.Figure()
        colors = ["#6B5CE7", "#E8638B", "#10B981", "#F5A623"]
        for i, metric in enumerate(prof_metrics):
            fig2.add_trace(go.Bar(
                x=prof_df["Company"],
                y=prof_df[metric],
                name=metric,
                marker=dict(color=colors[i], line=dict(color="rgba(255,255,255,0.15)", width=1)),
            ))
        
        fig2.update_layout(
            paper_bgcolor="rgba(0,0,0,0)",
            plot_bgcolor="rgba(0,0,0,0)",
            font=dict(family="Inter", size=14, color="#B8B3D7"),
            height=400,
            margin=dict(t=40, b=40, l=60, r=60),
            xaxis=dict(tickfont=dict(size=11, color="#8A85AD"), showgrid=False),
            yaxis=dict(tickfont=dict(size=9, color="#8A85AD"), gridcolor="rgba(107,92,231,0.1)", 
                      griddash="dot", ticksuffix="%"),
            legend=dict(font=dict(size=10, color="#B8B3D7"), orientation="h", yanchor="bottom", y=1.02),
            barmode="group",
        )
        
        st.plotly_chart(fig2, use_container_width=True, key="prof_comparison")
        
        _divider()
        
        # Price Performance
        _section("Price Performance (1 Year)", "📉")
        
        perf_period = st.selectbox(
            "Select Period",
            ["1mo", "3mo", "6mo", "1y", "2y", "5y"],
            index=3,
            label_visibility="collapsed",
            key="perf_period"
        )
        
        _build_price_performance_chart([c.ticker for c in companies], period=perf_period, key="price_perf_chart")
        
        _divider()
        
        # Valuation Multiples Comparison
        _section("Valuation Multiples", "💹")
        
        val_metrics = ["P/E", "EV/EBITDA", "EV/Revenue", "P/B"]
        val_data = []
        for c in companies:
            val_data.append({
                "Company": c.ticker,
                "P/E": c.trailing_pe if c.trailing_pe and c.trailing_pe > 0 else 0,
                "EV/EBITDA": c.ev_to_ebitda if c.ev_to_ebitda and c.ev_to_ebitda > 0 else 0,
                "EV/Revenue": c.ev_to_revenue if c.ev_to_revenue and c.ev_to_revenue > 0 else 0,
                "P/B": c.price_to_book if c.price_to_book and c.price_to_book > 0 else 0,
            })
        
        val_df = pd.DataFrame(val_data)
        
        fig3 = go.Figure()
        colors_val = ["#6B5CE7", "#E8638B", "#10B981", "#F5A623"]
        for i, metric in enumerate(val_metrics):
            fig3.add_trace(go.Bar(
                x=val_df["Company"],
                y=val_df[metric],
                name=metric,
                marker=dict(color=colors_val[i], line=dict(color="rgba(255,255,255,0.15)", width=1)),
                text=[f"{v:.1f}x" if v > 0 else "N/A" for v in val_df[metric]],
                textposition="outside",
                textfont=dict(size=9, color="#B8B3D7"),
            ))
        
        fig3.update_layout(
            paper_bgcolor="rgba(0,0,0,0)",
            plot_bgcolor="rgba(0,0,0,0)",
            font=dict(family="Inter", size=14, color="#B8B3D7"),
            height=400,
            margin=dict(t=40, b=40, l=60, r=60),
            xaxis=dict(tickfont=dict(size=11, color="#8A85AD"), showgrid=False),
            yaxis=dict(tickfont=dict(size=9, color="#8A85AD"), gridcolor="rgba(107,92,231,0.1)", 
                      griddash="dot", ticksuffix="x"),
            legend=dict(font=dict(size=10, color="#B8B3D7"), orientation="h", yanchor="bottom", y=1.02),
            barmode="group",
        )
        
        st.plotly_chart(fig3, use_container_width=True, key="val_comparison")
        
        _divider()
        
        # Correlation Matrix
        if len(companies) >= 3:
            _section("Price Correlation Matrix", "🔗")
            
            st.markdown(
                '<div style="font-size:0.75rem; color:#8A85AD; margin-bottom:0.5rem;">'
                'Shows how closely stock prices move together (1Y daily returns)</div>',
                unsafe_allow_html=True,
            )
            
            try:
                corr_data = {}
                for c in companies:
                    try:
                        tk_c = yf.Ticker(c.ticker)
                        h = tk_c.history(period="1y")
                        if not h.empty:
                            corr_data[c.ticker] = h["Close"].pct_change().dropna()
                    except Exception:
                        pass
                
                if len(corr_data) >= 3:
                    corr_df = pd.DataFrame(corr_data)
                    corr_matrix = corr_df.corr()
                    
                    fig_corr = go.Figure(data=go.Heatmap(
                        z=corr_matrix.values,
                        x=corr_matrix.columns,
                        y=corr_matrix.index,
                        colorscale=[
                            [0, "#EF4444"],
                            [0.5, "#1a1625"],
                            [1, "#10B981"]
                        ],
                        zmin=-1, zmax=1,
                        text=np.round(corr_matrix.values, 2),
                        texttemplate="%{text}",
                        textfont=dict(size=12, color="#E0DCF5"),
                        hovertemplate="%{x} vs %{y}: %{z:.3f}<extra></extra>",
                    ))
                    
                    fig_corr.update_layout(
                        paper_bgcolor="rgba(0,0,0,0)",
                        plot_bgcolor="rgba(0,0,0,0)",
                        font=dict(family="Inter", color="#B8B3D7"),
                        height=400,
                        margin=dict(t=20, b=40, l=60, r=30),
                        xaxis=dict(tickfont=dict(size=11, color="#8A85AD")),
                        yaxis=dict(tickfont=dict(size=11, color="#8A85AD")),
                    )
                    
                    st.plotly_chart(fig_corr, use_container_width=True, key="corr_matrix")
            except Exception:
                st.info("Could not generate correlation matrix.")

        # ══════════════════════════════════════════════════
        # MULTI-DIMENSIONAL RADAR + WINNER SUMMARY
        # ══════════════════════════════════════════════════
        if len(companies) >= 2:
            _divider()
            with _safe_section("Multi-Dimensional Radar"):
                _section("Multi-Dimensional Comparison", "🕸️")
                st.markdown(
                    '<div style="font-size:0.8rem; color:#B8B3D7; margin-bottom:0.8rem;">'
                    'Companies scored 0–100 across five dimensions: Valuation, Growth, Profitability, Leverage, Size.</div>',
                    unsafe_allow_html=True,
                )

                dimensions = ["Valuation", "Growth", "Profitability", "Leverage", "Size"]
                company_scores = {}

                for c in companies[:5]:
                    scores = {}
                    # Valuation (lower P/E = better, invert & normalize)
                    pe = c.trailing_pe if c.trailing_pe and 0 < c.trailing_pe < 200 else 50
                    scores["Valuation"] = max(0, min(100, 100 - (pe / 50 * 50)))

                    # Growth (revenue growth)
                    rg = (c.revenue_growth or 0) * 100
                    scores["Growth"] = max(0, min(100, rg * 2 + 50))

                    # Profitability (net margin)
                    nm = (c.profit_margins or 0) * 100
                    scores["Profitability"] = max(0, min(100, nm * 2.5 + 25))

                    # Leverage (lower D/E = better)
                    de = c.debt_to_equity if c.debt_to_equity and c.debt_to_equity > 0 else 50
                    scores["Leverage"] = max(0, min(100, 100 - min(de / 2, 100)))

                    # Size (log market cap normalized)
                    mc = c.market_cap or 1e9
                    import math
                    scores["Size"] = max(0, min(100, (math.log10(max(mc, 1)) - 8) / 5 * 100))

                    company_scores[c.ticker] = scores

                # Build radar chart
                fig_radar = go.Figure()
                radar_colors = ["#6B5CE7", "#E8638B", "#10B981", "#F5A623", "#3B82F6"]

                for i, (ticker, scores) in enumerate(company_scores.items()):
                    vals = [scores[d] for d in dimensions]
                    color = radar_colors[i % len(radar_colors)]
                    fig_radar.add_trace(go.Scatterpolar(
                        r=vals + [vals[0]],
                        theta=dimensions + [dimensions[0]],
                        fill='toself',
                        name=ticker,
                        fillcolor=f"rgba({int(color[1:3], 16)},{int(color[3:5], 16)},{int(color[5:7], 16)},0.1)",
                        line=dict(color=color, width=2),
                        marker=dict(size=6),
                    ))

                fig_radar.update_layout(
                    paper_bgcolor="rgba(0,0,0,0)",
                    plot_bgcolor="rgba(0,0,0,0)",
                    font=dict(family="Inter", size=12, color="#B8B3D7"),
                    polar=dict(
                        radialaxis=dict(visible=True, range=[0, 100], tickfont=dict(size=8, color="#8A85AD"),
                                       gridcolor="rgba(107,92,231,0.1)"),
                        angularaxis=dict(tickfont=dict(size=11, color="#8A85AD"),
                                        gridcolor="rgba(107,92,231,0.08)"),
                        bgcolor="rgba(0,0,0,0)",
                    ),
                    showlegend=True,
                    height=500,
                    margin=dict(t=50, b=50, l=80, r=80),
                    legend=dict(font=dict(size=11, color="#B8B3D7")),
                )
                st.plotly_chart(fig_radar, use_container_width=True, key="multi_dim_radar")

                # Winner summary
                _section("🏆 Winner Summary")
                total_scores = {t: sum(s.values()) for t, s in company_scores.items()}
                sorted_companies = sorted(total_scores.items(), key=lambda x: x[1], reverse=True)
                winner = sorted_companies[0]

                winner_cols = st.columns(min(len(sorted_companies), 5))
                for i, (ticker, total) in enumerate(sorted_companies[:5]):
                    with winner_cols[i]:
                        medal = "🥇" if i == 0 else "🥈" if i == 1 else "🥉" if i == 2 else f"#{i+1}"
                        border_c = radar_colors[i % len(radar_colors)]
                        # Find which dimension this company leads in
                        best_dim = max(company_scores[ticker].items(), key=lambda x: x[1])
                        st.markdown(
                            f'<div style="text-align:center; padding:0.8rem; background:rgba(255,255,255,0.04); '
                            f'border-radius:10px; border-top:3px solid {border_c};">'
                            f'<div style="font-size:1.2rem;">{medal}</div>'
                            f'<div style="font-size:1rem; font-weight:800; color:#E0DCF5;">{ticker}</div>'
                            f'<div style="font-size:1.5rem; font-weight:800; color:{border_c};">{total:.0f}</div>'
                            f'<div style="font-size:0.65rem; color:#8A85AD; text-transform:uppercase;">Total Score</div>'
                            f'<div style="font-size:0.7rem; color:#B8B3D7; margin-top:0.3rem;">Best: {best_dim[0]} ({best_dim[1]:.0f})</div>'
                            f'</div>',
                            unsafe_allow_html=True,
                        )

                # Per-dimension winners
                dim_winners_html = ""
                for dim in dimensions:
                    dim_best = max(company_scores.items(), key=lambda x: x[1][dim])
                    dim_winners_html += (
                        f'<span style="display:inline-block; padding:0.3rem 0.6rem; margin:0.2rem; '
                        f'background:rgba(107,92,231,0.08); border-radius:6px; font-size:0.75rem;">'
                        f'<span style="color:#8A85AD;">{dim}:</span> '
                        f'<span style="color:#E0DCF5; font-weight:700;">{dim_best[0]}</span> '
                        f'<span style="color:#6B5CE7;">({dim_best[1][dim]:.0f})</span></span>'
                    )
                st.markdown(
                    f'<div style="text-align:center; margin-top:0.8rem;">{dim_winners_html}</div>',
                    unsafe_allow_html=True,
                )

elif analysis_mode == "VMS Screener" and vms_screen_btn:
    # ══════════════════════════════════════════════════════
    # VMS SCREENER RESULTS
    # ══════════════════════════════════════════════════════
    import pandas as pd

    st.markdown(
        '<div class="hero-header">'
        '<div class="orbital-brand">'
        f'{_orbital_logo()}'
        '<p class="orbital-subtitle" style="font-size:1.1rem;">VMS Acquisition Screener</p>'
        '</div></div>',
        unsafe_allow_html=True,
    )

    with _safe_section("VMS Screening Philosophy"):
        st.markdown(
            '<div style="background:rgba(107,92,231,0.06); border-radius:12px; padding:1.2rem; '
            'border-left:4px solid #6B5CE7; margin-bottom:1rem;">'
            '<div style="font-size:0.85rem; font-weight:700; color:#9B8AFF; margin-bottom:0.5rem;">🏛️ Constellation Software Philosophy</div>'
            '<div style="font-size:0.78rem; color:#B8B3D7; line-height:1.6;">'
            'Vertical Market Software (VMS) companies serve niche industries with mission-critical software. '
            'These businesses exhibit high switching costs, recurring revenue, and durable competitive moats. '
            'Ideal acquisition targets have: <b>stable revenue ($5M-$200M)</b>, <b>healthy EBITDA margins (15%+)</b>, '
            '<b>low churn</b>, and <b>dominant positions</b> in their vertical. '
            'This screener identifies public VMS companies matching your criteria.</div></div>',
            unsafe_allow_html=True,
        )

    with _safe_section("Screening Results"):
        with st.spinner("Fetching data for VMS universe..."):
            # Convert to tuple of tuples for caching
            _vms_tuple = tuple(tuple(sorted(d.items())) for d in VMS_UNIVERSE)
            raw_results = _fetch_vms_screening_data(tuple(dict(t) for t in _vms_tuple))

        if raw_results:
            df = pd.DataFrame(raw_results)

            # Apply filters
            mask = (
                (df["Revenue ($M)"] >= vms_rev_min) &
                (df["Revenue ($M)"] <= vms_rev_max) &
                (df["EBITDA Margin (%)"] >= vms_ebitda_min) &
                (df["Revenue Growth (%)"] >= vms_growth_min)
            )
            if vms_industries:
                mask &= df["Vertical"].isin(vms_industries)
            if vms_geographies:
                mask &= df["Geography"].isin(vms_geographies)

            df["Pass"] = mask
            df_display = df[["Company", "Ticker", "Vertical", "Revenue ($M)", "EBITDA Margin (%)",
                             "Revenue Growth (%)", "EV/Revenue", "EV/EBITDA", "Pass"]].copy()

            # Summary
            n_pass = mask.sum()
            st.markdown(
                f'<div style="text-align:center; margin-bottom:1rem;">'
                f'<span style="font-size:1.5rem; font-weight:800; color:#6B5CE7;">{n_pass}</span>'
                f'<span style="font-size:0.85rem; color:#8A85AD;"> / {len(df)} companies pass your criteria</span></div>',
                unsafe_allow_html=True,
            )

            def _color_pass(row):
                if row["Pass"]:
                    return ["background-color: rgba(16,185,129,0.12)"] * len(row)
                return [""] * len(row)

            styled = df_display.style.apply(_color_pass, axis=1).format({
                "Revenue ($M)": "${:,.1f}M",
                "EBITDA Margin (%)": "{:.1f}%",
                "Revenue Growth (%)": "{:.1f}%",
                "EV/Revenue": "{:.2f}x",
                "EV/EBITDA": "{:.1f}x",
            })
            st.dataframe(styled, use_container_width=True, height=500)

            # Download CSV
            csv_data = df_display[df_display["Pass"]].drop(columns=["Pass"]).to_csv(index=False)
            st.download_button(
                "📄 Generate Target List (CSV)",
                data=csv_data,
                file_name="vms_target_list.csv",
                mime="text/csv",
                use_container_width=True,
            )

            # Add to watchlist button
            if n_pass > 0:
                if st.button("⭐ Add Passing Companies to Watchlist", use_container_width=True):
                    for _, row in df_display[df_display["Pass"]].iterrows():
                        _add_to_watchlist(row["Ticker"])
                    st.success(f"Added {n_pass} companies to watchlist!")
                    st.rerun()
        else:
            st.warning("Could not fetch data for VMS universe. Please try again.")

elif analysis_mode == "VMS Screener" and not vms_screen_btn:
    # VMS Screener splash/landing
    st.markdown(
        '<div class="splash-hero">'
        '<div class="star-layer-1">&#8203;</div>'
        '<div class="star-layer-2">&#8203;</div>'
        '<div class="star-layer-3">&#8203;</div>'
        '<div class="nebula-overlay">&#8203;</div>'
        '<div class="orb orb-1">&#8203;</div>'
        '<div class="orb orb-2">&#8203;</div>'
        '<div class="orb orb-3">&#8203;</div>'
        '<div class="orb orb-4">&#8203;</div>'
        '<div class="orb orb-5">&#8203;</div>'
        '<div class="shooting-star shooting-star-1">&#8203;</div>'
        '<div class="shooting-star shooting-star-2">&#8203;</div>'
        '<div class="noise-overlay">&#8203;</div>'
        '<div class="title-glow">&#8203;</div>'
        '<div class="splash-content">'
        '<div class="orbital-logo orbital-logo-lg">'
        '<span class="orbital-text">ORBITAL</span>'
        '<div class="orbital-ring orbital-ring-1"></div>'
        '<div class="orbital-ring orbital-ring-2"></div>'
        '<div class="orbital-ring orbital-ring-3"></div>'
        '<div class="orbital-particle orbital-particle-1"></div>'
        '<div class="orbital-particle orbital-particle-2"></div>'
        '<div class="orbital-particle orbital-particle-3"></div>'
        '</div>'
        '<p class="splash-subtitle" style="font-size:1.4rem; margin-top:1rem;">VMS Acquisition Screener</p>'
        '<div class="pill-row">'
        '<span class="feature-pill">Constellation Style</span>'
        '<span class="feature-pill">Vertical Market Software</span>'
        '<span class="feature-pill">Niche Dominators</span>'
        '<span class="feature-pill">Recurring Revenue</span>'
        '</div>'
        '<div class="splash-stats">'
        '<div class="splash-stat"><div class="splash-stat-value">20</div><div class="splash-stat-label">VMS Companies</div></div>'
        '<div class="splash-stat"><div class="splash-stat-value">10</div><div class="splash-stat-label">Industry Verticals</div></div>'
        '<div class="splash-stat"><div class="splash-stat-value">7</div><div class="splash-stat-label">Screening Metrics</div></div>'
        '</div>'
        '</div>'
        '</div>',
        unsafe_allow_html=True,
    )

    st.markdown(
        '<div class="space-section">'
        '<div class="space-section-title">How It Works</div>'
        '<div class="step-grid">'
        '<div class="step-card"><div class="step-num">1</div><div class="step-label">Set Criteria</div><div class="step-detail">Revenue, margins, growth thresholds</div></div>'
        '<div class="step-card"><div class="step-num">2</div><div class="step-label">Filter Verticals</div><div class="step-detail">Healthcare IT, GovTech, and more</div></div>'
        '<div class="step-card"><div class="step-num">3</div><div class="step-label">Run Screen</div><div class="step-detail">Scan 20 public VMS companies</div></div>'
        '<div class="step-card"><div class="step-num">4</div><div class="step-label">Export Targets</div><div class="step-detail">Download CSV target list</div></div>'
        '</div>'
        '<div class="space-section-title">VMS Verticals Covered</div>'
        '<div class="feature-grid">'
        '<div class="feature-card"><div class="feature-icon">&#127973;</div><div class="feature-title">Healthcare IT</div><div class="feature-desc">EHR, clinical, billing systems</div></div>'
        '<div class="feature-card"><div class="feature-icon">&#127963;</div><div class="feature-title">GovTech</div><div class="feature-desc">Municipal, courts, public safety</div></div>'
        '<div class="feature-card"><div class="feature-icon">&#9878;</div><div class="feature-title">Legal Tech</div><div class="feature-desc">Case management, e-discovery</div></div>'
        '<div class="feature-card"><div class="feature-icon">&#127979;</div><div class="feature-title">Education Tech</div><div class="feature-desc">LMS, SIS, campus management</div></div>'
        '<div class="feature-card"><div class="feature-icon">&#127968;</div><div class="feature-title">Real Estate Tech</div><div class="feature-desc">Property, lease, facilities mgmt</div></div>'
        '<div class="feature-card"><div class="feature-icon">&#128679;</div><div class="feature-title">Construction Tech</div><div class="feature-desc">Project, BIM, field management</div></div>'
        '<div class="feature-card"><div class="feature-icon">&#9889;</div><div class="feature-title">Utilities</div><div class="feature-desc">Grid, metering, asset management</div></div>'
        '<div class="feature-card"><div class="feature-icon">&#128666;</div><div class="feature-title">Transportation</div><div class="feature-desc">Fleet, logistics, routing</div></div>'
        '</div>'
        '<p style="font-size:0.72rem; color:#8A85AD; margin-top:2rem; text-align:center;">'
        'Set your screening criteria in the sidebar and click Run Screen<br>'
        'Inspired by Constellation Software&#39;s acquisition philosophy'
        '</p>'
        '</div>',
        unsafe_allow_html=True,
    )

else:
    # ══════════════════════════════════════════════════════
    # SPLASH / LANDING PAGE — Immersive space experience
    # ══════════════════════════════════════════════════════
    
    # DCF Valuation Splash
    if analysis_mode == "DCF Valuation":
        st.markdown(
            '<div class="splash-hero">'
            '<div class="star-layer-1">&#8203;</div>'
            '<div class="star-layer-2">&#8203;</div>'
            '<div class="star-layer-3">&#8203;</div>'
            '<div class="nebula-overlay">&#8203;</div>'
            '<div class="orb orb-1">&#8203;</div>'
            '<div class="orb orb-2">&#8203;</div>'
            '<div class="orb orb-3">&#8203;</div>'
            '<div class="orb orb-4">&#8203;</div>'
            '<div class="orb orb-5">&#8203;</div>'
            '<div class="shooting-star shooting-star-1">&#8203;</div>'
            '<div class="shooting-star shooting-star-2">&#8203;</div>'
            '<div class="noise-overlay">&#8203;</div>'
            '<div class="title-glow">&#8203;</div>'
            '<div class="splash-content">'
            '<div class="orbital-logo orbital-logo-lg">'
            '<span class="orbital-text">ORBITAL</span>'
            '<div class="orbital-ring orbital-ring-1"></div>'
            '<div class="orbital-ring orbital-ring-2"></div>'
            '<div class="orbital-ring orbital-ring-3"></div>'
            '<div class="orbital-particle orbital-particle-1"></div>'
            '<div class="orbital-particle orbital-particle-2"></div>'
            '<div class="orbital-particle orbital-particle-3"></div>'
            '</div>'
            '<p class="splash-subtitle" style="font-size:1.4rem; margin-top:1rem;">DCF Valuation Engine</p>'
            '<div class="pill-row">'
            '<span class="feature-pill">Free Cash Flow Projection</span>'
            '<span class="feature-pill">Terminal Value</span>'
            '<span class="feature-pill">Sensitivity Analysis</span>'
            '<span class="feature-pill">WACC Modeling</span>'
            '</div>'
            '<div class="splash-stats">'
            '<div class="splash-stat"><div class="splash-stat-value">5-10</div><div class="splash-stat-label">Projection Years</div></div>'
            '<div class="splash-stat"><div class="splash-stat-value">25</div><div class="splash-stat-label">Sensitivity Scenarios</div></div>'
            '<div class="splash-stat"><div class="splash-stat-value">∞</div><div class="splash-stat-label">Terminal Value</div></div>'
            '</div>'
            '</div>'
            '</div>',
            unsafe_allow_html=True,
        )
        
        st.markdown(
            '<div class="space-section">'
            '<div class="space-section-title">How It Works</div>'
            '<div class="step-grid">'
            '<div class="step-card"><div class="step-num">1</div><div class="step-label">Enter Ticker</div><div class="step-detail">Company with positive free cash flow</div></div>'
            '<div class="step-card"><div class="step-num">2</div><div class="step-label">Set Assumptions</div><div class="step-detail">Growth rate, WACC, terminal growth</div></div>'
            '<div class="step-card"><div class="step-num">3</div><div class="step-label">Calculate DCF</div><div class="step-detail">Project FCF &amp; discount to present value</div></div>'
            '<div class="step-card"><div class="step-num">4</div><div class="step-label">Sensitivity</div><div class="step-detail">Test different scenarios</div></div>'
            '</div>'
            '<div class="space-section-title">Model Features</div>'
            '<div class="feature-grid">'
            '<div class="feature-card"><div class="feature-icon">&#128200;</div><div class="feature-title">FCF Projection</div><div class="feature-desc">Project free cash flows based on growth assumptions</div></div>'
            '<div class="feature-card"><div class="feature-icon">&#128202;</div><div class="feature-title">Terminal Value</div><div class="feature-desc">Gordon Growth Model for perpetuity value</div></div>'
            '<div class="feature-card"><div class="feature-icon">&#127919;</div><div class="feature-title">Present Value</div><div class="feature-desc">Discount future cash flows at WACC</div></div>'
            '<div class="feature-card"><div class="feature-icon">&#128176;</div><div class="feature-title">Equity Bridge</div><div class="feature-desc">Enterprise value to equity value</div></div>'
            '<div class="feature-card"><div class="feature-icon">&#128161;</div><div class="feature-title">Implied Price</div><div class="feature-desc">Per-share intrinsic value estimate</div></div>'
            '<div class="feature-card"><div class="feature-icon">&#9888;</div><div class="feature-title">Sensitivity Matrix</div><div class="feature-desc">Growth vs. WACC scenario analysis</div></div>'
            '<div class="feature-card"><div class="feature-icon">&#127942;</div><div class="feature-title">Upside/Downside</div><div class="feature-desc">Compare to current market price</div></div>'
            '<div class="feature-card"><div class="feature-icon">&#128196;</div><div class="feature-title">Visual Charts</div><div class="feature-desc">Interactive valuation visualizations</div></div>'
            '</div>'
            '<p style="font-size:0.72rem; color:#8A85AD; margin-top:2rem; text-align:center;">'
            'Enter a ticker and set your DCF assumptions in the sidebar<br>'
            'Works best for companies with positive, predictable free cash flow'
            '</p>'
            '</div>',
            unsafe_allow_html=True,
        )
    
    # Quick Compare Splash
    elif analysis_mode == "Quick Compare":
        st.markdown(
            '<div class="splash-hero">'
            '<div class="star-layer-1">&#8203;</div>'
            '<div class="star-layer-2">&#8203;</div>'
            '<div class="star-layer-3">&#8203;</div>'
            '<div class="nebula-overlay">&#8203;</div>'
            '<div class="orb orb-1">&#8203;</div>'
            '<div class="orb orb-2">&#8203;</div>'
            '<div class="orb orb-3">&#8203;</div>'
            '<div class="orb orb-4">&#8203;</div>'
            '<div class="orb orb-5">&#8203;</div>'
            '<div class="shooting-star shooting-star-1">&#8203;</div>'
            '<div class="shooting-star shooting-star-2">&#8203;</div>'
            '<div class="noise-overlay">&#8203;</div>'
            '<div class="title-glow">&#8203;</div>'
            '<div class="splash-content">'
            '<div class="orbital-logo orbital-logo-lg">'
            '<span class="orbital-text">ORBITAL</span>'
            '<div class="orbital-ring orbital-ring-1"></div>'
            '<div class="orbital-ring orbital-ring-2"></div>'
            '<div class="orbital-ring orbital-ring-3"></div>'
            '<div class="orbital-particle orbital-particle-1"></div>'
            '<div class="orbital-particle orbital-particle-2"></div>'
            '<div class="orbital-particle orbital-particle-3"></div>'
            '</div>'
            '<p class="splash-subtitle" style="font-size:1.4rem; margin-top:1rem;">Company Comparison Tool</p>'
            '<div class="pill-row">'
            '<span class="feature-pill">Side-by-Side Analysis</span>'
            '<span class="feature-pill">Multiple Metrics</span>'
            '<span class="feature-pill">Price Performance</span>'
            '<span class="feature-pill">Radar Charts</span>'
            '</div>'
            '<div class="splash-stats">'
            '<div class="splash-stat"><div class="splash-stat-value">10</div><div class="splash-stat-label">Max Companies</div></div>'
            '<div class="splash-stat"><div class="splash-stat-value">15+</div><div class="splash-stat-label">Comparison Metrics</div></div>'
            '<div class="splash-stat"><div class="splash-stat-value">6</div><div class="splash-stat-label">Preset Groups</div></div>'
            '</div>'
            '</div>'
            '</div>',
            unsafe_allow_html=True,
        )
        
        st.markdown(
            '<div class="space-section">'
            '<div class="space-section-title">How It Works</div>'
            '<div class="step-grid">'
            '<div class="step-card"><div class="step-num">1</div><div class="step-label">Enter Tickers</div><div class="step-detail">Comma-separated list or load preset</div></div>'
            '<div class="step-card"><div class="step-num">2</div><div class="step-label">Compare</div><div class="step-detail">Side-by-side metrics comparison</div></div>'
            '<div class="step-card"><div class="step-num">3</div><div class="step-label">Visualize</div><div class="step-detail">Radar charts, bar charts, price performance</div></div>'
            '<div class="step-card"><div class="step-num">4</div><div class="step-label">Export</div><div class="step-detail">Download comparison as CSV</div></div>'
            '</div>'
            '<div class="space-section-title">Preset Comparisons</div>'
            '<div class="feature-grid">'
            '<div class="feature-card"><div class="feature-icon">&#128187;</div><div class="feature-title">FAANG</div><div class="feature-desc">META, AAPL, AMZN, NFLX, GOOGL</div></div>'
            '<div class="feature-card"><div class="feature-icon">&#127760;</div><div class="feature-title">Big Tech</div><div class="feature-desc">AAPL, MSFT, GOOGL, AMZN, META, NVDA</div></div>'
            '<div class="feature-card"><div class="feature-icon">&#127464;</div><div class="feature-title">Canadian Banks</div><div class="feature-desc">RY.TO, TD.TO, BNS.TO, BMO.TO, CM.TO</div></div>'
            '<div class="feature-card"><div class="feature-icon">&#128187;</div><div class="feature-title">Software/SaaS</div><div class="feature-desc">CRM, ADBE, NOW, WDAY, TEAM</div></div>'
            '<div class="feature-card"><div class="feature-icon">&#9889;</div><div class="feature-title">Semiconductors</div><div class="feature-desc">NVDA, AMD, INTC, QCOM, AVGO</div></div>'
            '<div class="feature-card"><div class="feature-icon">&#128138;</div><div class="feature-title">Healthcare Giants</div><div class="feature-desc">JNJ, UNH, PFE, ABBV, MRK</div></div>'
            '</div>'
            '<p style="font-size:0.72rem; color:#8A85AD; margin-top:2rem; text-align:center;">'
            'Enter multiple tickers separated by commas or select a preset in the sidebar<br>'
            'Compare up to 10 companies at once'
            '</p>'
            '</div>',
            unsafe_allow_html=True,
        )
    
    elif analysis_mode == "Merger Analysis":
        # Merger-specific splash
        st.markdown(
            '<div class="splash-hero">'
            '<div class="star-layer-1">&#8203;</div>'
            '<div class="star-layer-2">&#8203;</div>'
            '<div class="star-layer-3">&#8203;</div>'
            '<div class="nebula-overlay">&#8203;</div>'
            '<div class="orb orb-1">&#8203;</div>'
            '<div class="orb orb-2">&#8203;</div>'
            '<div class="orb orb-3">&#8203;</div>'
            '<div class="orb orb-4">&#8203;</div>'
            '<div class="orb orb-5">&#8203;</div>'
            '<div class="shooting-star shooting-star-1">&#8203;</div>'
            '<div class="shooting-star shooting-star-2">&#8203;</div>'
            '<div class="shooting-star shooting-star-3">&#8203;</div>'
            '<div class="shooting-star shooting-star-4">&#8203;</div>'
            '<div class="shooting-star shooting-star-5">&#8203;</div>'
            '<div class="noise-overlay">&#8203;</div>'
            '<div class="title-glow">&#8203;</div>'
            '<div class="splash-content">'
            '<div class="orbital-logo orbital-logo-lg">'
            '<span class="orbital-text">ORBITAL</span>'
            '<div class="orbital-ring orbital-ring-1"></div>'
            '<div class="orbital-ring orbital-ring-2"></div>'
            '<div class="orbital-ring orbital-ring-3"></div>'
            '<div class="orbital-particle orbital-particle-1"></div>'
            '<div class="orbital-particle orbital-particle-2"></div>'
            '<div class="orbital-particle orbital-particle-3"></div>'
            '</div>'
            '<p class="splash-subtitle" style="font-size:1.4rem; margin-top:1rem;">M&amp;A Simulator &amp; Deal Intelligence</p>'
            '<div class="pill-row">'
            '<span class="feature-pill">Pro Forma Analysis</span>'
            '<span class="feature-pill">Accretion/Dilution</span>'
            '<span class="feature-pill">Football Field</span>'
            '<span class="feature-pill">AI Insights</span>'
            '<span class="feature-pill">Deal Book PPTX</span>'
            '</div>'
            '<div class="splash-stats">'
            '<div class="splash-stat"><div class="splash-stat-value">12</div><div class="splash-stat-label">Dashboard Sections</div></div>'
            '<div class="splash-stat"><div class="splash-stat-value">10</div><div class="splash-stat-label">Deal Book Slides</div></div>'
            '<div class="splash-stat"><div class="splash-stat-value">4</div><div class="splash-stat-label">AI Analyses</div></div>'
            '</div>'
            '</div>'
            '</div>',
            unsafe_allow_html=True,
        )

        st.markdown(
            '<div class="space-section">'
            '<div class="space-section-title">How It Works</div>'
            '<div class="step-grid">'
            '<div class="step-card"><div class="step-num">1</div><div class="step-label">Enter Tickers</div><div class="step-detail">Acquirer + Target company tickers</div></div>'
            '<div class="step-card"><div class="step-num">2</div><div class="step-label">Set Assumptions</div><div class="step-detail">Premium, cash/stock mix, synergies</div></div>'
            '<div class="step-card"><div class="step-num">3</div><div class="step-label">Analyze Deal</div><div class="step-detail">Pro forma financials &amp; AI insights</div></div>'
            '<div class="step-card"><div class="step-num">4</div><div class="step-label">Download Book</div><div class="step-detail">10-slide deal book PowerPoint</div></div>'
            '</div>'
            '<div class="space-section-title">Analysis Features</div>'
            '<div class="feature-grid">'
            '<div class="feature-card"><div class="feature-icon">&#128200;</div><div class="feature-title">Pro Forma P&amp;L</div><div class="feature-desc">Combined income statement with synergy adjustments</div></div>'
            '<div class="feature-card"><div class="feature-icon">&#128202;</div><div class="feature-title">Accretion/Dilution</div><div class="feature-desc">Waterfall chart showing EPS bridge</div></div>'
            '<div class="feature-card"><div class="feature-icon">&#127919;</div><div class="feature-title">Football Field</div><div class="feature-desc">Multi-method valuation range analysis</div></div>'
            '<div class="feature-card"><div class="feature-icon">&#128176;</div><div class="feature-title">Sources &amp; Uses</div><div class="feature-desc">Classic IB deal structure breakdown</div></div>'
            '<div class="feature-card"><div class="feature-icon">&#128161;</div><div class="feature-title">AI Rationale</div><div class="feature-desc">Strategic fit and synergy assessment</div></div>'
            '<div class="feature-card"><div class="feature-icon">&#9888;</div><div class="feature-title">Risk Analysis</div><div class="feature-desc">Antitrust, integration, financial risks</div></div>'
            '<div class="feature-card"><div class="feature-icon">&#127942;</div><div class="feature-title">Deal Grade</div><div class="feature-desc">AI-powered A-F deal verdict</div></div>'
            '<div class="feature-card"><div class="feature-icon">&#128196;</div><div class="feature-title">Deal Book</div><div class="feature-desc">10-slide professional PPTX export</div></div>'
            '</div>'
            '<p style="font-size:0.72rem; color:#8A85AD; margin-top:2rem; text-align:center;">'
            'Enter Acquirer &amp; Target tickers in the sidebar to begin<br>'
            'Set <code style="color:#9B8AFF;">OPENAI_API_KEY</code> for AI-powered deal insights'
            '</p>'
            '</div>',
            unsafe_allow_html=True,
        )
    else:
        st.markdown(
            '<div class="splash-hero">'
            '<div class="star-layer-1">&#8203;</div>'
            '<div class="star-layer-2">&#8203;</div>'
            '<div class="star-layer-3">&#8203;</div>'
            '<div class="nebula-overlay">&#8203;</div>'
            '<div class="orb orb-1">&#8203;</div>'
            '<div class="orb orb-2">&#8203;</div>'
            '<div class="orb orb-3">&#8203;</div>'
            '<div class="orb orb-4">&#8203;</div>'
            '<div class="orb orb-5">&#8203;</div>'
            '<div class="shooting-star shooting-star-1">&#8203;</div>'
            '<div class="shooting-star shooting-star-2">&#8203;</div>'
            '<div class="shooting-star shooting-star-3">&#8203;</div>'
            '<div class="shooting-star shooting-star-4">&#8203;</div>'
            '<div class="shooting-star shooting-star-5">&#8203;</div>'
            '<div class="noise-overlay">&#8203;</div>'
            '<div class="title-glow">&#8203;</div>'
            '<div class="splash-content">'
            '<div class="orbital-logo orbital-logo-lg">'
            '<span class="orbital-text">ORBITAL</span>'
            '<div class="orbital-ring orbital-ring-1"></div>'
            '<div class="orbital-ring orbital-ring-2"></div>'
            '<div class="orbital-ring orbital-ring-3"></div>'
            '<div class="orbital-particle orbital-particle-1"></div>'
            '<div class="orbital-particle orbital-particle-2"></div>'
            '<div class="orbital-particle orbital-particle-3"></div>'
            '</div>'
            '<p class="splash-subtitle" style="font-size:1.4rem; margin-top:1rem;">Company Intelligence &amp; Tear Sheet Generator</p>'
            '<div class="pill-row">'
            '<span class="feature-pill">Live Market Data</span>'
            '<span class="feature-pill">Wikipedia M&amp;A</span>'
            '<span class="feature-pill">Peer Analysis</span>'
            '<span class="feature-pill">AI Powered</span>'
            '<span class="feature-pill">Global Exchanges</span>'
            '</div>'
            '<div class="splash-stats">'
            '<div class="splash-stat"><div class="splash-stat-value">100+</div><div class="splash-stat-label">Data Points</div></div>'
            '<div class="splash-stat"><div class="splash-stat-value">8</div><div class="splash-stat-label">PPTX Slides</div></div>'
            '<div class="splash-stat"><div class="splash-stat-value">20+</div><div class="splash-stat-label">Exchanges</div></div>'
            '</div>'
            '</div>'
            '</div>',
            unsafe_allow_html=True,
        )

        # Step cards and feature grid in dark space-section
        st.markdown(
            '<div class="space-section">'
            '<div class="space-section-title">How It Works</div>'
            '<div class="step-grid">'
            '<div class="step-card"><div class="step-num">1</div><div class="step-label">Enter Ticker</div><div class="step-detail">Any global exchange &mdash; AAPL, RY.TO, NVDA.L</div></div>'
            '<div class="step-card"><div class="step-num">2</div><div class="step-label">Generate Profile</div><div class="step-detail">100+ data points pulled in real-time</div></div>'
            '<div class="step-card"><div class="step-num">3</div><div class="step-label">Explore Dashboard</div><div class="step-detail">Charts, peer comparison &amp; insights</div></div>'
            '<div class="step-card"><div class="step-num">4</div><div class="step-label">Download PPTX</div><div class="step-detail">8-slide IB-grade PowerPoint</div></div>'
            '</div>'
            '<div class="space-section-title">Platform Features</div>'
            '<div class="feature-grid">'
            '<div class="feature-card"><div class="feature-icon">&#128200;</div><div class="feature-title">Price &amp; Valuation</div><div class="feature-desc">Live prices, multiples, and historical charts</div></div>'
            '<div class="feature-card"><div class="feature-icon">&#128101;</div><div class="feature-title">Peer Comparison</div><div class="feature-desc">Side-by-side valuation vs industry peers</div></div>'
            '<div class="feature-card"><div class="feature-icon">&#128202;</div><div class="feature-title">Financial Statements</div><div class="feature-desc">Income, balance sheet, cash flow analysis</div></div>'
            '<div class="feature-card"><div class="feature-icon">&#129309;</div><div class="feature-title">M&amp;A History</div><div class="feature-desc">Deal history scraped from Wikipedia</div></div>'
            '<div class="feature-card"><div class="feature-icon">&#127919;</div><div class="feature-title">Analyst Consensus</div><div class="feature-desc">Recommendations &amp; price targets</div></div>'
            '<div class="feature-card"><div class="feature-icon">&#128161;</div><div class="feature-title">AI Insights</div><div class="feature-desc">Powered by GPT (optional API key)</div></div>'
            '<div class="feature-card"><div class="feature-icon">&#127760;</div><div class="feature-title">Global Exchanges</div><div class="feature-desc">TSX, LSE, JPX and more with local currencies</div></div>'
            '<div class="feature-card"><div class="feature-icon">&#128196;</div><div class="feature-title">PowerPoint Export</div><div class="feature-desc">8-slide professional presentation</div></div>'
            '</div>'
            '<p style="font-size:0.72rem; color:#8A85AD; margin-top:2rem; text-align:center;">'
            'M&amp;A history scraped from Wikipedia &mdash; no API key needed<br>'
            'Set <code style="color:#9B8AFF;">OPENAI_API_KEY</code> for enhanced insights'
            '</p>'
            '</div>',
            unsafe_allow_html=True,
        )
        
        # Market Overview Section
        st.markdown('<div style="height:1rem;"></div>', unsafe_allow_html=True)
        
        # Fetch indices once (cached)
        try:
            indices = _fetch_market_indices()
        except Exception:
            indices = None
        
        # Scrolling Market Ticker
        try:
            if indices:
                _render_market_ticker(indices)
        except Exception:
            pass
        
        st.markdown('<div style="height:0.5rem;"></div>', unsafe_allow_html=True)
        
        # Market Overview Cards
        try:
            if indices:
                st.markdown(
                    '<div style="background:rgba(107,92,231,0.05); border-radius:16px; padding:1.5rem; '
                    'border:1px solid rgba(107,92,231,0.15);">'
                    '<div style="font-size:0.8rem; font-weight:700; color:#9B8AFF; text-transform:uppercase; '
                    'letter-spacing:1.5px; margin-bottom:1rem; text-align:center;">📊 Market Overview</div>',
                    unsafe_allow_html=True,
                )
                
                idx_cols = st.columns(len(indices))
                for i, idx in enumerate(indices):
                    with idx_cols[i]:
                        color = "#10B981" if idx["change_pct"] >= 0 else "#EF4444"
                        arrow = "▲" if idx["change_pct"] >= 0 else "▼"
                        st.markdown(
                            f'<div style="text-align:center;">'
                            f'<div style="font-size:0.7rem; color:#8A85AD; font-weight:600;">{idx["name"]}</div>'
                            f'<div style="font-size:1.1rem; font-weight:700; color:#E0DCF5;">{idx["price"]:,.2f}</div>'
                            f'<div style="font-size:0.8rem; color:{color};">{arrow} {idx["change_pct"]:+.2f}%</div>'
                            f'</div>',
                            unsafe_allow_html=True,
                        )
                
                st.markdown('</div>', unsafe_allow_html=True)
        except Exception:
            pass  # Market overview is non-critical
        
        # Top Movers Section
        try:
            movers = _fetch_top_movers()
            if movers and (movers.get("gainers") or movers.get("losers")):
                _render_movers_cards(movers)
        except Exception:
            pass  # Top movers is non-critical
        
        # Sentiment Gauge + Earnings Calendar (side by side)
        sent_col, earn_col = st.columns(2)
        
        with sent_col:
            try:
                sentiment = _calculate_market_sentiment()
                _render_sentiment_gauge(sentiment)
            except Exception:
                pass
        
        with earn_col:
            try:
                earnings = _fetch_earnings_calendar()
                _render_earnings_calendar(earnings)
            except Exception:
                pass
        
        # Sector Heatmap + News Feed (side by side)
        heatmap_col, news_col = st.columns(2)
        
        with heatmap_col:
            try:
                sector_perf = _fetch_sector_performance()
                _render_sector_heatmap(sector_perf)
            except Exception:
                pass
        
        with news_col:
            try:
                news = _fetch_news_feed()
                _render_news_feed(news)
            except Exception:
                pass

# ══════════════════════════════════════════════════════════════
# FOOTER
# ══════════════════════════════════════════════════════════════
st.markdown(
    '<div class="orbital-footer">'
    '<div class="orbital-footer-brand">ORBITAL</div>'
    '<div style="font-size:0.7rem; color:#8A85AD; margin-top:0.2rem;">M&A Intelligence Platform</div>'
    '<div class="orbital-footer-links">'
    '<a href="https://github.com/rajkcho/profilebuilder" target="_blank">GitHub</a>'
    '<a href="#">Documentation</a>'
    '<a href="#">API</a>'
    '</div>'
    '<div class="orbital-footer-version">v5.0 · Built with Streamlit · Data from Yahoo Finance & Alpha Vantage</div>'
    '</div>',
    unsafe_allow_html=True,
)
